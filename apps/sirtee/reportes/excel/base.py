from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class ExcelBase:

    titulo = "Reporte SIRTEE"


    def crear_workbook(self):

        wb = Workbook()

        ws = wb.active

        ws.title = "SIRTEE"

        return wb



    def titulo_hoja(
        self,
        ws,
        titulo
    ):

        ws["A1"] = titulo

        ws["A1"].font = Font(
            bold=True,
            size=16
        )

        ws.merge_cells(
            "A1:H1"
        )

        ws["A1"].alignment = Alignment(
            horizontal="center"
        )



    def encabezados(
        self,
        ws,
        fila,
        columnas
    ):

        for col, valor in enumerate(
            columnas,
            1
        ):

            cell = ws.cell(
                fila,
                col,
                valor
            )

            cell.font = Font(
                bold=True
            )

            cell.fill = PatternFill(
                "solid",
                fgColor="CCCCCC"
            )



    def ajustar_columnas(
        self,
        ws
    ):

        for columna in ws.columns:

            largo = max(
                len(str(c.value))
                if c.value
                else 0
                for c in columna
            )

            ws.column_dimensions[
                get_column_letter(
                    columna[0].column
                )
            ].width = largo + 3