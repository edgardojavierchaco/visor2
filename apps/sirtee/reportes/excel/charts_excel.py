from openpyxl.chart import (
    BarChart,
    PieChart,
    LineChart,
    Reference,
)



class ExcelCharts:



    def grafico_estados(
        self,
        ws
    ):


        chart = PieChart()


        labels = Reference(
            ws,
            min_col=1,
            min_row=2,
            max_row=ws.max_row
        )


        data = Reference(
            ws,
            min_col=2,
            min_row=1,
            max_row=ws.max_row
        )


        chart.add_data(
            data,
            titles_from_data=True
        )


        chart.set_categories(
            labels
        )


        chart.title = (
            "Estado de intervenciones"
        )


        ws.add_chart(
            chart,
            "D3"
        )





    def grafico_empresas(
        self,
        ws
    ):


        chart = BarChart()


        data = Reference(
            ws,
            min_col=2,
            min_row=1,
            max_row=ws.max_row
        )


        categorias = Reference(
            ws,
            min_col=1,
            min_row=2,
            max_row=ws.max_row
        )


        chart.add_data(
            data,
            titles_from_data=True
        )


        chart.set_categories(
            categorias
        )


        chart.title = (
            "Intervenciones por empresa"
        )


        ws.add_chart(
            chart,
            "D3"
        )


