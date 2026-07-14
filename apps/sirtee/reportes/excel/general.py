from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from apps.sirtee.reportes.metrics import (
    obtener_kpis_generales,
)

from apps.sirtee.reportes.queries import (
    intervenciones_por_estado,
    intervenciones_por_empresa,
)


class ReporteGeneralExcel:


    def generar(self):


        wb = Workbook()



        # ==================================================
        # HOJA RESUMEN
        # ==================================================

        ws = wb.active

        ws.title = "Resumen"



        ws["A1"] = (
            "Reporte Institucional SIRTEE"
        )

        ws["A1"].font = Font(
            bold=True,
            size=16
        )


        ws.merge_cells(
            "A1:D1"
        )


        ws["A3"] = "Indicador"

        ws["B3"] = "Valor"



        for cell in ws[3]:

            cell.font = Font(
                bold=True
            )



        kpis = obtener_kpis_generales()



        fila = 4


        for nombre, valor in kpis.items():


            ws.cell(
                fila,
                1,
                nombre.replace(
                    "_",
                    " "
                ).title()
            )


            ws.cell(
                fila,
                2,
                valor
            )


            fila += 1



        self.ajustar_columnas(
            ws
        )



        # ==================================================
        # ESTADOS
        # ==================================================

        ws_estado = wb.create_sheet(
            "Estados"
        )


        ws_estado.append(
            [
                "Estado",
                "Cantidad"
            ]
        )


        for dato in intervenciones_por_estado():


            ws_estado.append(
                [
                    dato[
                        "estado__nombre"
                    ],
                    dato[
                        "cantidad"
                    ]
                ]
            )



        self.ajustar_columnas(
            ws_estado
        )



        # ==================================================
        # EMPRESAS
        # ==================================================

        ws_empresa = wb.create_sheet(
            "Empresas"
        )


        ws_empresa.append(
            [
                "Empresa",
                "Intervenciones"
            ]
        )


        for empresa in intervenciones_por_empresa():


            ws_empresa.append(
                [
                    empresa.razon_social,

                    empresa.cantidad
                ]
            )


        self.ajustar_columnas(
            ws_empresa
        )



        return wb



    # ==================================================
    # AJUSTE COLUMNAS
    # ==================================================

    def ajustar_columnas(
        self,
        ws
    ):


        for columna in ws.columns:


            largo = max(

                len(
                    str(celda.value)
                )
                if celda.value
                else 0

                for celda in columna

            )


            ws.column_dimensions[
                get_column_letter(
                    columna[0].column
                )
            ].width = largo + 4