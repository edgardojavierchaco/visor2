import json
import logging
import re
from decimal import Decimal
from io import BytesIO
from urllib.parse import urlencode

from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist, ValidationError
from django.core.paginator import Paginator
from django.db import DatabaseError, IntegrityError, connection, transaction
from django.db.models import Max
from django.db.utils import OperationalError, ProgrammingError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from .services.filtros_pof_service import MENSAJE_FILTROS_INVALIDOS, validar_anio_filtro

from .forms import (
    DUPLICADO_REUNIDA_POF,
    ProyectoEspecialPofForm,
    ReunidaPofForm,
    validar_payload_guardar_carga,
    validar_payload_guardar_carga_proyecto_especial,
)
from .models import (
    CargoPof,
    LocalizacionPof,
    ProyectosEspecialesPof,
    ReunidaPof,
    usuario_tiene_acceso_completo_pof,
)
from .permisos import (
    pof_api_required,
    pof_required,
    pof_visualizacion_api_required,
    pof_visualizacion_required,
)
from .services.carga_service import construir_contexto_carga, validar_cabecera_reunida
from .services.consulta_service import construir_contexto_consulta
from .services.exportacion_reunida import (
    COLUMNAS_BUSQUEDA_EXPORTACION,
    construir_contexto_exportacion,
    _obtener_cue_grupo_visual_reunida,
    normalizar_cueanexo_exportacion,
    obtener_clave_grupo_visual_reunida,
    obtener_clave_total_general_cueanexo_reunida,
)
from .services.exportacion_rows import obtener_clave_total_general
from .services.exportacion_columnas_config import REPETIR_POR_CUE, REPETIR_POR_CUEANEXO
from .services.excel_estilos import COLOR_SEPARADOR_CUE, COLOR_SEPARADOR_CUEANEXO
from .services.grilla_pof import construir_grilla_pof_desde_cargos, obtener_cargos_grilla_reunida
from .services.grilla_pof.detalle_rows import (
    obtener_grupo_operativo_detalle,
    serializar_grupo_operativo_detalle,
)
from .services.guardado_pof_service import (
    cambiar_estado_cargo_pof,
    eliminar_cargo_pof,
    eliminar_proyecto_especial_pof,
    eliminar_reunida_pof,
    guardar_carga_pof as guardar_carga_pof_service,
    modificar_cargo_pof,
    obtener_detalle_cargo_pof,
)
from .services.historial_service import (
    construir_contexto_historial,
    obtener_detalle_movimiento_pof,
    obtener_historial_cantidad_cargos_pof,
    obtener_historial_observacion_cargos_pof,
    obtener_historial_estado_cargos_pof,
)
from .services.herencia_reunida_service import (
    heredar_estado_inicial_proyecto_especial,
    heredar_estado_inicial_reunida,
)
from .services.padron_materializadas_service import (
    buscar_ofertas_padron,
    construir_cueanexo_sin_guion,
    obtener_catalogos_padron_ingreso_manual_pof,
    obtener_opciones_filtro_detalle_padron,
)
from .services.proyecto_especial_manual_service import (
    buscar_cuof_manual_proyecto_especial as buscar_cuof_manual_proyecto_especial_service,
)
from .services.niveles import obtener_niveles_ceic_para_reunida
from .services.niveles_service import normalizar_nivel
from .services.reunidas_service import (
    construir_contexto_detalle_reunida,
    construir_contexto_reunidas,
    construir_payload_cargos_detalle_proyecto_localizacion,
)
from .services.visualizacion_cargos_localizacion_service import (
    construir_contexto_visualizacion_cargos_localizacion,
    construir_excel_visualizacion_cargos_localizacion,
    construir_opciones_filtros_visualizacion_cargos_localizacion,
    construir_payload_visualizacion_cargos_localizacion,
)
from .utils.api_responses import (
    api_error_interno,
    api_error_no_encontrado,
    api_error_permiso,
    api_error_sesion,
    api_error_sin_cambios,
    api_error_validacion,
    api_ok,
)


logger = logging.getLogger(__name__)

_EXCEL_DOWNLOAD_TOKEN_PARAMETER = "download_token"
_EXCEL_DOWNLOAD_READY_COOKIE = "pof_excel_download_ready"
_EXCEL_DOWNLOAD_TOKEN_PATTERN = re.compile(r"^[A-Za-z0-9_-]{20,128}$")
_EXCEL_DOWNLOAD_COOKIE_PATH = "/"
_EXCEL_DOWNLOAD_COOKIE_MAX_AGE = 120


def _obtener_token_descarga_excel(request):
    """Devuelve un token de correlación válido, sin usarlo como autorización."""
    token = request.GET.get(_EXCEL_DOWNLOAD_TOKEN_PARAMETER, "")
    if not isinstance(token, str) or not (20 <= len(token) <= 128):
        return ""
    if _EXCEL_DOWNLOAD_TOKEN_PATTERN.fullmatch(token) is None:
        return ""
    return token


def _marcar_descarga_excel_lista(request, response):
    """Marca la respuesta XLSX lista mediante una cookie legible por el navegador."""
    token = _obtener_token_descarga_excel(request)
    if token:
        response.set_cookie(
            _EXCEL_DOWNLOAD_READY_COOKIE,
            token,
            max_age=_EXCEL_DOWNLOAD_COOKIE_MAX_AGE,
            path=_EXCEL_DOWNLOAD_COOKIE_PATH,
            secure=request.is_secure(),
            httponly=False,
            samesite="Lax",
        )
    return response


def _obtener_niveles_ceic_validos(nivel_reunida):
    """
    Resuelve los niveles CEIC permitidos para una Reunida validada.

    - Normaliza el nivel recibido antes de consultar la matriz central.
    - Devuelve una lista estable para usarla en filtros SQL parametrizados.
    - No devuelve resultados cuando el nivel es inválido o no está definido.
    """
    nivel_normalizado = normalizar_nivel(nivel_reunida)
    if not nivel_normalizado:
        return ""

    niveles_ceic = sorted(obtener_niveles_ceic_para_reunida(nivel_normalizado))
    return niveles_ceic if niveles_ceic else ""


def _validar_busqueda_ceic_codigo(q):
    """
    Valida el código CEIC usado por el autocompletado de alta.

    - Acepta únicamente números, sin espacios ni símbolos.
    - Limita la búsqueda a códigos de hasta 3 dígitos.
    - Devuelve un mensaje claro para mantener consistente frontend y backend.
    """
    if not q:
        return "Debe ingresar un CEIC."
    if not q.isdigit() or len(q) > 3:
        return "Ingresá solo números, hasta 3 dígitos."
    return ""


def _consultar_ceic_puntos(niveles_ceic=None, q="", limite=500):
    """
    Consulta CEIC activos filtrando por nivel permitido y prefijo numérico.

    - Usa filtros parametrizados para evitar concatenar valores en SQL.
    - Permite omitir el límite solo para catálogos completos explícitamente requeridos.
    - Exige CEIC activos y, si se indican niveles, restringidos al nivel de la Reunida.
    """
    clausulas = [
        "estado = TRUE",
    ]
    parametros = []

    if niveles_ceic:
        placeholders = ", ".join(["%s"] * len(niveles_ceic))
        clausulas.append(f"nivel IN ({placeholders})")
        parametros.extend(niveles_ceic)

    if q:
        clausulas.append("CAST(ceic_id AS TEXT) LIKE %s")
        parametros.append(f"{q}%")

    limite_sql = " LIMIT %s" if limite is not None else ""
    sql = f"""
        SELECT ceic_id AS ceic, descripcion_ceic, puntos, nivel
        FROM cenpe.ceic_puntos
        WHERE {" AND ".join(clausulas)}
        ORDER BY ceic_id{limite_sql};
    """
    if limite is not None:
        parametros.append(limite)

    with connection.cursor() as cursor:
        cursor.execute(sql, parametros)
        return cursor.fetchall()

COLUMNAS_TEXTO_EXCEL = {
    "cue",
    "subcue",
    "cueanexo",
    "cui",
    "cuof",
    "sub_cuof",
    "cue_cui",
    "cue_bloque_final",
    "cue_anexo",
    "cui_bloque_final",
    "cui_anexo",
    "anexo",
    "numero_establecimiento",
}

COLUMNAS_NUMERICAS_EXCEL = {
    "cantidad",
    "cantidad_cargos",
    "cantidad_horas",
    "puntos",
    "total",
    "total_general",
    "total_general_exportacion",
    "total_horas_catedra",
    "puntos_horas_catedra",
    "total_puntos",
}

COLUMNAS_CANTIDAD_ENTERA_EXCEL = {
    "cantidad",
    "cantidad_cargos",
    "cantidad_horas",
    "total_horas_catedra",
}


def _normalizar_titulo_hoja(valor):
    titulo = "".join(
        caracter if caracter not in '[]:*?/\\' else " "
        for caracter in str(valor or "Reunida")
    ).strip()
    return (titulo or "Reunida")[:31]


def _ajustar_ancho_columnas_excel(ws, fila_inicio=1):
    from openpyxl.utils import get_column_letter

    for indice_columna in range(1, ws.max_column + 1):
        ancho = 10
        letra = get_column_letter(indice_columna)
        for fila in ws.iter_rows(
            min_row=fila_inicio,
            min_col=indice_columna,
            max_col=indice_columna,
        ):
            celda = fila[0]
            valor = "" if celda.value is None else str(celda.value)
            ancho = max(ancho, min(len(valor) + 4, 42))
        ws.column_dimensions[letra].width = ancho


def _valor_excel_exportacion(valor):
    """
    Convierte valores internos a tipos compatibles con celdas Excel.

    - Decimal: se exporta como número.
    - Listas y tuplas: se serializan como texto separado por comas.
    - Otros tipos escalares: se conservan sin cambios.
    """
    if isinstance(valor, Decimal):
        return float(valor)

    if isinstance(valor, (list, tuple)):
        return ", ".join(str(item) for item in valor)

    return valor


def _escribir_filas_excel_optimizado(
    ws,
    filas,
    columnas,
    mensaje="",
    fila_encabezado=4,
):
    """
    Escribe las mismas filas Excel y combina formato y anchos en una pasada.

    - Conserva la conversion de `_valor_excel_exportacion`.
    - Replica los formatos de texto, cantidades, puntos y totales vigentes.
    - Mantiene los anchos minimo 10 y maximo 42 usados por el exportador.
    - Aplica formatos de datos desde la fila siguiente al encabezado recibido.
    - No altera filas, columnas, orden ni los bordes aplicados posteriormente.
    """
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    metadata_columnas = [
        (
            _es_columna_texto_excel(columna),
            _es_columna_numerica_excel(columna),
            _es_columna_cantidad_entera_excel(columna),
        )
        for columna in columnas
    ]
    cuerpo_alignment = Alignment(
        vertical="top",
        wrap_text=True,
    )

    for fila in filas:
        ws.append([_valor_excel_exportacion(valor) for valor in fila])

    if mensaje and not filas:
        ws.append([mensaje])

    cantidad_columnas = max(ws.max_column, 1)
    anchos_columnas = [10] * cantidad_columnas
    fila_inicio_datos = fila_encabezado + 1
    for numero_fila, fila in enumerate(
        ws.iter_rows(min_row=fila_encabezado),
        start=fila_encabezado,
    ):
        for indice_columna, celda in enumerate(fila):
            if (
                numero_fila >= fila_inicio_datos
                and indice_columna < len(metadata_columnas)
            ):
                celda.alignment = cuerpo_alignment
                es_texto, es_numerica, es_cantidad_entera = metadata_columnas[
                    indice_columna
                ]
                if es_texto:
                    if celda.value not in (None, ""):
                        celda.value = str(celda.value)
                    celda.number_format = "@"
                elif es_numerica and isinstance(celda.value, (int, float)):
                    celda.number_format = (
                        "#,##0" if es_cantidad_entera else "#,##0.00"
                    )

            valor = "" if celda.value is None else str(celda.value)
            anchos_columnas[indice_columna] = max(
                anchos_columnas[indice_columna],
                min(len(valor) + 4, 42),
            )

    for indice_columna, ancho in enumerate(anchos_columnas, start=1):
        ws.column_dimensions[get_column_letter(indice_columna)].width = ancho


def _escribir_excel_reunida_filtrable(
    ws,
    filas_normalizadas,
    columnas_config,
    schema,
    mensaje="",
    fila_encabezado=4,
):
    """Escribe la Reunida común con datos físicos y presentación filtrable.

    Las columnas visibles conservan cada valor de cargo. Las siete auxiliares
    ocultas propagan el último grupo visible mediante referencias O(1), para
    que la compactación y los separadores sigan las filas filtradas por Excel.
    Los auxiliares adicionales se crean solo para los resúmenes derivados
    visibles y encadenan sus aportes por grupo sin rangos crecientes.
    """
    from openpyxl.formatting.rule import FormulaRule, Rule
    from openpyxl.styles import Alignment, Border, Side
    from openpyxl.styles.differential import DifferentialStyle
    from openpyxl.styles.numbers import NumberFormat
    from openpyxl.utils import get_column_letter

    columnas_config = columnas_config or []
    columnas = [columna["titulo"] for columna in columnas_config]
    keys_schema = {
        columna.get("key")
        for columna in (schema.get("columnas") or [])
    }
    metricas_tecnicas = {
        "total_horas_catedra",
        "puntos_horas_catedra",
        "total_puntos",
    }
    schema_totales_tecnicos = metricas_tecnicas.issubset(keys_schema)
    acumulador_por_source = {
        "total_general": "total",
        "total_general_exportacion": "total",
    }
    if schema_totales_tecnicos:
        acumulador_por_source.update({
            "total_horas_catedra": "horas",
            "puntos_horas_catedra": "puntos_horas",
            "total_puntos": "total",
        })
    indices_resumen_dinamico = [
        indice
        for indice, columna in enumerate(columnas_config, start=1)
        if columna.get("source") in acumulador_por_source
    ]
    fuentes_resumen_dinamico = {
        columnas_config[indice - 1]["source"]
        for indice in indices_resumen_dinamico
    }
    fuentes_total_general_estandar = fuentes_resumen_dinamico & {
        "total_general",
        "total_general_exportacion",
    }
    fuentes_totales_tecnicos = fuentes_resumen_dinamico & metricas_tecnicas
    requiere_resumen_dinamico = bool(indices_resumen_dinamico)
    requiere_total_cargo = bool(
        fuentes_resumen_dinamico
        & {
            "total_general",
            "total_general_exportacion",
            "puntos_horas_catedra",
            "total_puntos",
        }
    )
    requiere_cantidad_horas = bool(
        fuentes_resumen_dinamico
        & {"total_horas_catedra", "puntos_horas_catedra"}
    )
    requiere_total_visible = any(
        acumulador_por_source[source] == "total"
        for source in fuentes_resumen_dinamico
    )
    requiere_horas_visibles = "total_horas_catedra" in fuentes_resumen_dinamico
    requiere_puntos_horas_visibles = (
        "puntos_horas_catedra" in fuentes_resumen_dinamico
    )
    total_columnas = len(columnas)
    fila_inicio = fila_encabezado + 1
    cuerpo_alignment = Alignment(vertical="top", wrap_text=True)
    borde_superior_cue = Side(style="medium", color=COLOR_SEPARADOR_CUE)
    borde_superior_anexo = Side(style="medium", color=COLOR_SEPARADOR_CUEANEXO)

    fila_visible_columna = total_columnas + 1
    clave_grupo_columna = fila_visible_columna + 1
    clave_cue_columna = clave_grupo_columna + 1
    tiene_anexo_columna = clave_cue_columna + 1
    cue_con_separador_columna = tiene_anexo_columna + 1
    ultimo_grupo_visible_columna = cue_con_separador_columna + 1
    ultimo_cue_visible_columna = ultimo_grupo_visible_columna + 1
    siguiente_columna_auxiliar = ultimo_cue_visible_columna + 1

    def reservar_columna_auxiliar(requerida):
        nonlocal siguiente_columna_auxiliar
        if not requerida:
            return None
        columna = siguiente_columna_auxiliar
        siguiente_columna_auxiliar += 1
        return columna

    es_afectado_columna = reservar_columna_auxiliar(requiere_resumen_dinamico)
    total_cargo_columna = reservar_columna_auxiliar(requiere_total_cargo)
    cantidad_horas_columna = reservar_columna_auxiliar(requiere_cantidad_horas)
    total_visible_grupo_columna = reservar_columna_auxiliar(requiere_total_visible)
    horas_visible_grupo_columna = reservar_columna_auxiliar(
        requiere_horas_visibles
    )
    puntos_horas_visible_grupo_columna = reservar_columna_auxiliar(
        requiere_puntos_horas_visibles
    )
    grupo_visto_visible_columna = reservar_columna_auxiliar(
        requiere_resumen_dinamico
    )

    letra_fila_visible = get_column_letter(fila_visible_columna)
    letra_clave_grupo = get_column_letter(clave_grupo_columna)
    letra_clave_cue = get_column_letter(clave_cue_columna)
    letra_tiene_anexo = get_column_letter(tiene_anexo_columna)
    letra_cue_con_separador = get_column_letter(cue_con_separador_columna)
    letra_ultimo_grupo_visible = get_column_letter(ultimo_grupo_visible_columna)
    letra_ultimo_cue_visible = get_column_letter(ultimo_cue_visible_columna)
    letra_es_afectado = (
        get_column_letter(es_afectado_columna) if es_afectado_columna else None
    )
    letra_total_cargo = (
        get_column_letter(total_cargo_columna) if total_cargo_columna else None
    )
    letra_cantidad_horas = (
        get_column_letter(cantidad_horas_columna) if cantidad_horas_columna else None
    )
    letra_total_visible_grupo = (
        get_column_letter(total_visible_grupo_columna)
        if total_visible_grupo_columna
        else None
    )
    letra_horas_visible_grupo = (
        get_column_letter(horas_visible_grupo_columna)
        if horas_visible_grupo_columna
        else None
    )
    letra_puntos_horas_visible_grupo = (
        get_column_letter(puntos_horas_visible_grupo_columna)
        if puntos_horas_visible_grupo_columna
        else None
    )
    letra_grupo_visto_visible = (
        get_column_letter(grupo_visto_visible_columna)
        if grupo_visto_visible_columna
        else None
    )

    encabezados_auxiliares = (
        (fila_visible_columna, "_fila_visible"),
        (clave_grupo_columna, "_clave_grupo"),
        (clave_cue_columna, "_clave_cue"),
        (tiene_anexo_columna, "_tiene_anexo"),
        (cue_con_separador_columna, "_cue_con_separador"),
        (ultimo_grupo_visible_columna, "_ultimo_grupo_visible"),
        (ultimo_cue_visible_columna, "_ultimo_cue_visible"),
        (es_afectado_columna, "_es_afectado"),
        (total_cargo_columna, "_total_cargo"),
        (cantidad_horas_columna, "_cantidad_horas"),
        (total_visible_grupo_columna, "_total_visible_grupo"),
        (horas_visible_grupo_columna, "_horas_visible_grupo"),
        (
            puntos_horas_visible_grupo_columna,
            "_puntos_horas_visible_grupo",
        ),
        (grupo_visto_visible_columna, "_grupo_visto_visible"),
    )
    for indice, encabezado in encabezados_auxiliares:
        if indice is None:
            continue
        ws.cell(row=fila_encabezado, column=indice, value=encabezado)
        ws.column_dimensions[get_column_letter(indice)].hidden = True

    def expresion_total_afectado_visible(numero_fila):
        return (
            f"IF(${letra_es_afectado}{numero_fila}=1,"
            f"SUBTOTAL(109,${letra_total_cargo}{numero_fila}),0)"
        )

    def expresion_horas_afectadas_visibles(numero_fila):
        return (
            f"IF(${letra_es_afectado}{numero_fila}=1,"
            f"SUBTOTAL(109,${letra_cantidad_horas}{numero_fila}),0)"
        )

    def expresion_puntos_horas_afectados_visibles(numero_fila):
        return (
            f"IF(AND(${letra_es_afectado}{numero_fila}=1,"
            f"${letra_cantidad_horas}{numero_fila}<>\"\"),"
            f"SUBTOTAL(109,${letra_total_cargo}{numero_fila}),0)"
        )

    expresion_por_acumulador = {
        "total": expresion_total_afectado_visible,
        "horas": expresion_horas_afectadas_visibles,
        "puntos_horas": expresion_puntos_horas_afectados_visibles,
    }
    columna_por_acumulador = {
        "total": total_visible_grupo_columna,
        "horas": horas_visible_grupo_columna,
        "puntos_horas": puntos_horas_visible_grupo_columna,
    }
    letra_por_acumulador = {
        "total": letra_total_visible_grupo,
        "horas": letra_horas_visible_grupo,
        "puntos_horas": letra_puntos_horas_visible_grupo,
    }

    anchos_columnas = [10] * total_columnas
    ultima_fila_por_grupo_total = {}
    cue_anterior = None
    indice_cue = 0
    fila_actual = fila_inicio

    for indice, fila_normalizada in enumerate(filas_normalizadas or []):
        fila_con_metadata = fila_normalizada.copy()
        fila_con_metadata["_grupo_visual_cueanexo"] = fila_normalizada.get(
            "cueanexo", ""
        )
        clave_grupo = obtener_clave_grupo_visual_reunida(
            fila_con_metadata,
            indice=indice,
        )
        cueanexo = normalizar_cueanexo_exportacion(
            fila_con_metadata.get("_grupo_visual_cueanexo", "")
        )
        clave_cue = _obtener_cue_grupo_visual_reunida(fila_con_metadata)
        tiene_anexo = int(bool(cueanexo))
        if clave_cue and clave_cue != cue_anterior:
            indice_cue += 1
        if fuentes_total_general_estandar:
            clave_total_general = obtener_clave_total_general_cueanexo_reunida(
                fila_normalizada,
                indice=indice,
            )
        elif fuentes_totales_tecnicos:
            clave_total_general = obtener_clave_total_general(
                fila_normalizada,
                schema,
            )
        else:
            clave_total_general = None
        fila_anterior_grupo_total = (
            ultima_fila_por_grupo_total.get(clave_total_general)
            if requiere_resumen_dinamico
            else None
        )

        for columna_indice, columna in enumerate(columnas_config, start=1):
            valor_excel = _valor_excel_exportacion(
                fila_normalizada.get(columna["source"], "")
            )
            celda = ws.cell(
                row=fila_actual,
                column=columna_indice,
                value=valor_excel,
            )
            celda.alignment = cuerpo_alignment
            valor_texto = "" if valor_excel is None else str(valor_excel)
            anchos_columnas[columna_indice - 1] = max(
                anchos_columnas[columna_indice - 1],
                min(len(valor_texto) + 4, 42),
            )
            if _es_columna_texto_excel(columna["titulo"]):
                if celda.value not in (None, ""):
                    celda.value = str(celda.value)
                celda.number_format = "@"
            elif _es_columna_numerica_excel(columna["titulo"]) and isinstance(
                celda.value, (int, float)
            ):
                celda.number_format = (
                    "#,##0"
                    if _es_columna_cantidad_entera_excel(columna["titulo"])
                    else "#,##0.00"
                )

            source = columna.get("source")
            if source in fuentes_resumen_dinamico:
                tipo_acumulador = acumulador_por_source[source]
                grupo_visto_anterior = (
                    "0"
                    if fila_anterior_grupo_total is None
                    else f"${letra_grupo_visto_visible}{fila_anterior_grupo_total}"
                )
                celda.value = (
                    f"=IF(AND(SUBTOTAL(103,${letra_fila_visible}{fila_actual})=1,"
                    f"{grupo_visto_anterior}=0),"
                    f"${letra_por_acumulador[tipo_acumulador]}{fila_actual},\"\")"
                )
                celda.number_format = (
                    "#,##0" if source == "total_horas_catedra" else "#,##0.00"
                )

        ws.cell(row=fila_actual, column=fila_visible_columna, value=1)
        ws.cell(row=fila_actual, column=clave_grupo_columna, value=clave_grupo)
        ws.cell(row=fila_actual, column=clave_cue_columna, value=clave_cue)
        ws.cell(row=fila_actual, column=tiene_anexo_columna, value=tiene_anexo)
        ws.cell(
            row=fila_actual,
            column=cue_con_separador_columna,
            value=int(bool(clave_cue and indice_cue > 1)),
        )
        ultimo_grupo_anterior = (
            '""'
            if fila_actual == fila_inicio
            else f"${letra_ultimo_grupo_visible}{fila_actual - 1}"
        )
        ultimo_cue_anterior = (
            '""'
            if fila_actual == fila_inicio
            else f"${letra_ultimo_cue_visible}{fila_actual - 1}"
        )
        ws.cell(
            row=fila_actual,
            column=ultimo_grupo_visible_columna,
            value=(
                f"=IF(SUBTOTAL(103,${letra_fila_visible}{fila_actual}),"
                f"${letra_clave_grupo}{fila_actual},{ultimo_grupo_anterior})"
            ),
        )
        ws.cell(
            row=fila_actual,
            column=ultimo_cue_visible_columna,
            value=(
                f"=IF(SUBTOTAL(103,${letra_fila_visible}{fila_actual}),"
                f"${letra_clave_cue}{fila_actual},{ultimo_cue_anterior})"
            ),
        )

        if requiere_resumen_dinamico:
            ws.cell(
                row=fila_actual,
                column=es_afectado_columna,
                value=int(
                    fila_normalizada.get("estado_pof_codigo")
                    == CargoPof.EstadoPof.AFECTADO
                ),
            )
            if requiere_total_cargo:
                ws.cell(
                    row=fila_actual,
                    column=total_cargo_columna,
                    value=(
                        _valor_excel_exportacion(fila_normalizada.get("total", 0))
                        or 0
                    ),
                )
            if requiere_cantidad_horas:
                ws.cell(
                    row=fila_actual,
                    column=cantidad_horas_columna,
                    value=_valor_excel_exportacion(
                        fila_normalizada.get("cantidad_horas", "")
                    ),
                )
            for tipo_acumulador, columna_acumulador in columna_por_acumulador.items():
                if columna_acumulador is None:
                    continue
                ws.cell(
                    row=fila_actual,
                    column=columna_acumulador,
                    value=(
                        f"={expresion_por_acumulador[tipo_acumulador](fila_actual)}"
                    ),
                )
            grupo_visto_anterior = (
                "0"
                if fila_anterior_grupo_total is None
                else f"${letra_grupo_visto_visible}{fila_anterior_grupo_total}"
            )
            ws.cell(
                row=fila_actual,
                column=grupo_visto_visible_columna,
                value=(
                    f"=IF(SUBTOTAL(103,${letra_fila_visible}{fila_actual})=1,1,"
                    f"{grupo_visto_anterior})"
                ),
            )
            if fila_anterior_grupo_total is not None:
                for (
                    tipo_acumulador,
                    columna_acumulador,
                ) in columna_por_acumulador.items():
                    if columna_acumulador is None:
                        continue
                    ws.cell(
                        row=fila_anterior_grupo_total,
                        column=columna_acumulador,
                        value=(
                            f"={expresion_por_acumulador[tipo_acumulador](fila_anterior_grupo_total)}+"
                            f"${letra_por_acumulador[tipo_acumulador]}{fila_actual}"
                        ),
                    )
            ultima_fila_por_grupo_total[clave_total_general] = fila_actual

        cue_anterior = clave_cue or None
        fila_actual += 1

    if mensaje and not filas_normalizadas:
        ws.append([mensaje])

    ultima_fila_datos = fila_actual - 1
    hay_filas_datos = ultima_fila_datos >= fila_inicio
    if hay_filas_datos:
        ocultar_repetido_dxf = DifferentialStyle(
            numFmt=NumberFormat(numFmtId=164, formatCode=";;;"),
        )
        for repeticion, clave_estado, indices in (
            (REPETIR_POR_CUEANEXO, letra_ultimo_grupo_visible, []),
            (REPETIR_POR_CUE, letra_ultimo_cue_visible, []),
        ):
            indices.extend(
                indice
                for indice, columna in enumerate(columnas_config, start=1)
                if (
                    columna.get("repetir") == repeticion
                    and columna.get("source") not in fuentes_resumen_dinamico
                )
            )
            if not indices:
                continue
            rangos = " ".join(
                f"{get_column_letter(indice)}{fila_inicio}:"
                f"{get_column_letter(indice)}{ultima_fila_datos}"
                for indice in indices
            )
            clave_actual = (
                letra_clave_grupo
                if repeticion == REPETIR_POR_CUEANEXO
                else letra_clave_cue
            )
            ws.conditional_formatting.add(
                rangos,
                Rule(
                    type="expression",
                    formula=[
                        f"AND(SUBTOTAL(103,${letra_fila_visible}{fila_inicio})=1,"
                        f"${clave_actual}{fila_inicio}=${clave_estado}{fila_encabezado})"
                    ],
                    dxf=ocultar_repetido_dxf,
                ),
            )

        rango_datos = (
            f"A{fila_inicio}:{get_column_letter(total_columnas)}{ultima_fila_datos}"
        )
        ws.conditional_formatting.add(
            rango_datos,
            FormulaRule(
                formula=[
                    f"AND(SUBTOTAL(103,${letra_fila_visible}{fila_inicio})=1,"
                    f"${letra_cue_con_separador}{fila_inicio}=1,"
                    f"${letra_clave_cue}{fila_inicio}<>"
                    f"${letra_ultimo_cue_visible}{fila_encabezado})"
                ],
                border=Border(top=borde_superior_cue),
                stopIfTrue=True,
            ),
        )
        ws.conditional_formatting.add(
            rango_datos,
            FormulaRule(
                formula=[
                    f"AND(SUBTOTAL(103,${letra_fila_visible}{fila_inicio})=1,"
                    f"${letra_tiene_anexo}{fila_inicio}=1,"
                    f"${letra_clave_grupo}{fila_inicio}<>"
                    f"${letra_ultimo_grupo_visible}{fila_encabezado},"
                    f"${letra_clave_cue}{fila_inicio}="
                    f"${letra_ultimo_cue_visible}{fila_encabezado})"
                ],
                border=Border(top=borde_superior_anexo),
            ),
        )

    for indice, ancho in enumerate(anchos_columnas, start=1):
        ws.column_dimensions[get_column_letter(indice)].width = ancho

    return indices_resumen_dinamico


def _normalizar_columna_excel(valor):
    return str(valor or "").strip().lower()


def _compactar_columna_excel(valor):
    return "".join(
        caracter
        for caracter in _normalizar_columna_excel(valor)
        if caracter.isalnum()
    )


def _es_columna_texto_excel(columna):
    nombre = _normalizar_columna_excel(columna)
    compacto = _compactar_columna_excel(columna)
    return (
        nombre in COLUMNAS_TEXTO_EXCEL
        or compacto in {
            "cue",
            "subcue",
            "cueanexo",
            "cui",
            "cuof",
            "subcuof",
            "cuecui",
            "cuebloquefinal",
            "cueanexo",
            "cuibloquefinal",
            "cuianexo",
            "anexo",
            "nanexo",
            "numeroestablecimiento",
            "n",
            "nº",
            "bp n°",
        }
    )


def _es_columna_numerica_excel(columna):
    nombre = _normalizar_columna_excel(columna)
    if nombre in COLUMNAS_NUMERICAS_EXCEL:
        return True
    return any(
        texto in nombre
        for texto in ("cantidad", "puntos", "total")
    )


def _es_columna_cantidad_entera_excel(columna):
    nombre = _normalizar_columna_excel(columna)
    compacto = _compactar_columna_excel(columna)
    return (
        nombre in COLUMNAS_CANTIDAD_ENTERA_EXCEL
        or compacto in {
            "cantidad",
            "cantidadcargos",
            "cantidadhoras",
            "totalhorascatedra",
        }
    )


def _aplicar_formato_columnas_excel(ws, columnas, fila_inicio):
    from openpyxl.styles import Alignment

    for indice, columna in enumerate(columnas, start=1):
        es_texto = _es_columna_texto_excel(columna)
        es_numerica = _es_columna_numerica_excel(columna)
        es_cantidad_entera = _es_columna_cantidad_entera_excel(columna)
        for fila in ws.iter_rows(
            min_row=fila_inicio,
            min_col=indice,
            max_col=indice,
        ):
            celda = fila[0]
            celda.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
            if es_texto:
                if celda.value not in (None, ""):
                    celda.value = str(celda.value)
                celda.number_format = "@"
            elif es_numerica and isinstance(celda.value, (int, float)):
                celda.number_format = "#,##0" if es_cantidad_entera else "#,##0.00"


def _aplicar_autofiltro_excel(ws, fila_encabezado, max_columna):
    if max_columna:
        ultima_fila = max(ws.max_row, fila_encabezado)
        ws.auto_filter.ref = (
            f"A{fila_encabezado}:"
            f"{ws.cell(row=ultima_fila, column=max_columna).coordinate}"
        )


def _aplicar_bordes_secciones_excel(ws, secciones, fila_inicio, max_columna):
    from openpyxl.styles import Border, Side

    borde_fuerte = Side(style="medium", color="111827")
    fila_actual = fila_inicio

    for seccion in secciones or []:
        cantidad_filas = len(seccion.get("filas", []))
        if not cantidad_filas:
            continue

        primera_fila = fila_actual
        ultima_fila = fila_actual + cantidad_filas - 1

        for columna in range(1, max_columna + 1):
            celda_inicio = ws.cell(row=primera_fila, column=columna)
            celda_inicio.border = Border(
                top=borde_fuerte,
                bottom=celda_inicio.border.bottom,
                left=celda_inicio.border.left,
                right=celda_inicio.border.right,
            )

            celda_fin = ws.cell(row=ultima_fila, column=columna)
            celda_fin.border = Border(
                top=celda_fin.border.top,
                bottom=borde_fuerte,
                left=celda_fin.border.left,
                right=celda_fin.border.right,
            )

        fila_actual = ultima_fila + 1


def _aplicar_bordes_grupos_visual_excel(
    ws,
    separadores_filas,
    fila_inicio,
    max_columna,
    color_anexo="111827",
):
    """
    Aplica los separadores visuales de CUE y CUEANEXO al workbook.

    - Usa medium azul para nuevos CUE.
    - Usa thin con el color recibido para nuevos CUEANEXO.
    - Permite conservar el color histórico de Proyecto Especial.
    """
    from openpyxl.styles import Border, Side

    borde_cue = Side(style="medium", color="2444D8")
    borde_anexo = Side(style="thin", color=color_anexo)

    for indice, separador in enumerate(separadores_filas or []):
        if separador.get("es_inicio_cue"):
            borde_superior = borde_cue
        elif separador.get("es_inicio_anexo"):
            borde_superior = borde_anexo
        else:
            continue

        numero_fila = fila_inicio + indice
        for columna in range(1, max_columna + 1):
            celda = ws.cell(row=numero_fila, column=columna)
            celda.border = Border(
                top=borde_superior,
                bottom=celda.border.bottom,
                left=celda.border.left,
                right=celda.border.right,
            )


def _texto_filtros_excel_exportacion(contexto):
    """Describe el alcance y el filtro aplicado a una exportación POF."""
    if contexto.get("es_proyecto_especial"):
        return contexto.get("filtros_excel") or "Sin filtros"

    if contexto.get("alcance_excel") == "todos":
        return "Ninguno (exportación completa)"

    busquedas = contexto.get("busquedas_columnas") or {}
    if not busquedas:
        return "Sin filtros"

    partes = []
    for columna_id, valor in busquedas.items():
        etiqueta = next(
            (
                columna["label"]
                for columna in COLUMNAS_BUSQUEDA_EXPORTACION
                if columna["id"] == columna_id
            ),
            columna_id,
        )
        partes.append(f"{etiqueta}: {valor}")
    return " · ".join(partes)


def _crear_respuesta_excel_exportacion(contexto):
    """
    Genera el workbook Excel respetando el pipeline de filas ya construido.

    - Optimiza solo Reunidas comunes combinando formato y anchos en una pasada.
    - Conserva el recorrido baseline completo para Proyecto Especial.
    - Agrega metadatos de generación sin cambiar filas ni columnas exportadas.
    - Mantiene bordes, filtros, freeze panes y estilos funcionales existentes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    reunida = contexto.get("reunida", {})
    es_proyecto_especial = contexto.get("es_proyecto_especial", False)
    columnas = contexto.get("columnas", [])
    filas = contexto.get("filas_exportacion", [])
    filas_normalizadas = contexto.get("filas_normalizadas_exportacion", [])
    columnas_config = contexto.get("columnas_exportacion_config", [])
    separadores_filas = contexto.get("separadores_filas_exportacion", [])
    secciones = contexto.get("secciones_exportacion", [])
    mensaje = contexto.get("mensaje_exportacion", "")
    nombre_archivo = contexto.get("nombre_archivo") or "POF.xlsx"
    titulo_hoja = contexto.get("titulo_hoja") or reunida.get("nivel") or "POF"
    titulo_excel = contexto.get("titulo_excel") or f"POF - {reunida.get('nivel') or ''} {reunida.get('anio') or ''}".strip()

    ws.title = _normalizar_titulo_hoja(titulo_hoja)
    ws.append([titulo_excel])
    ws.append([
        timezone.localtime(timezone.now()).strftime("Generado el %d/%m/%Y %H:%M")
    ])
    ws.append([f"Filtros aplicados: {_texto_filtros_excel_exportacion(contexto)}"])
    ws.append(columnas)

    for celda in ws[1]:
        celda.font = Font(bold=True, size=13)

    ws[2][0].font = Font(size=10)
    ws[3][0].font = Font(size=10)
    ws[3][0].alignment = Alignment(wrap_text=True)

    for celda in ws[4]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor="2444D8")
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    indices_resumen_dinamico = []
    if es_proyecto_especial:
        for fila in filas:
            ws.append([_valor_excel_exportacion(valor) for valor in fila])
        if mensaje and not filas:
            ws.append([mensaje])
    else:
        indices_resumen_dinamico = _escribir_excel_reunida_filtrable(
            ws,
            filas_normalizadas,
            columnas_config,
            contexto.get("schema_exportacion") or {},
            mensaje=mensaje,
            fila_encabezado=4,
        )

    max_columna = max(len(columnas), 1)
    for numero_fila in (1, 2, 3):
        ws.merge_cells(
            start_row=numero_fila,
            start_column=1,
            end_row=numero_fila,
            end_column=max_columna,
        )
    ws.freeze_panes = "A5"

    _aplicar_autofiltro_excel(ws, fila_encabezado=4, max_columna=max_columna)
    if indices_resumen_dinamico:
        from openpyxl.worksheet.filters import FilterColumn

        for indice in indices_resumen_dinamico:
            ws.auto_filter.filterColumn.append(
                FilterColumn(
                    colId=indice - 1,
                    hiddenButton=True,
                    showButton=False,
                )
            )
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    if es_proyecto_especial:
        _aplicar_formato_columnas_excel(ws, columnas, fila_inicio=5)
    if es_proyecto_especial:
        _aplicar_bordes_secciones_excel(
            ws,
            secciones,
            fila_inicio=5,
            max_columna=max_columna,
        )
    if es_proyecto_especial:
        _aplicar_bordes_grupos_visual_excel(
            ws,
            separadores_filas,
            fila_inicio=5,
            max_columna=max_columna,
            color_anexo="CBD5E1",
        )

    if es_proyecto_especial:
        _ajustar_ancho_columnas_excel(ws, fila_inicio=4)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


@pof_visualizacion_required
def inicio(request):
    contexto = {
        "pof_solo_visualizacion": not usuario_tiene_acceso_completo_pof(
            request.user
        )
    }
    return render(request, "reunidas_pof/inicio.html", contexto)


@pof_required
def cargar_cargos(request):
    """
    Renderiza la pantalla de Alta de Cargos POF para Reunida.

    - La ruta definitiva es `/reunidas_pof/reunida/cargar/`.
    - `/reunidas_pof/cargar/` queda como redirección legacy.
    - Mantiene la compatibilidad legacy de Proyecto Especial cuando entra por `/cargar/`.
    """
    cabecera_tipo = str(request.GET.get("cabecera_tipo", "") or "").strip().upper()
    nombre_ruta = getattr(getattr(request, "resolver_match", None), "url_name", "") or ""
    es_ruta_legacy = nombre_ruta == "cargar_cargos_legacy"

    if es_ruta_legacy and cabecera_tipo == "PROYECTO_ESPECIAL":
        proyecto_especial_id = str(request.GET.get("proyecto_especial_id", "") or "").strip()
        query = urlencode({"proyecto_especial_id": proyecto_especial_id}) if proyecto_especial_id else ""
        destino = reverse("reunidas_pof:cargar_cargos_proyecto_especial")
        if query:
            destino = f"{destino}?{query}"
        return redirect(destino)

    if es_ruta_legacy:
        query_params = {}
        anio = str(request.GET.get("anio", "") or "").strip()
        nivel = str(request.GET.get("nivel", "") or "").strip()
        if anio:
            query_params["anio"] = anio
        if nivel:
            query_params["nivel"] = nivel

        destino = reverse("reunidas_pof:cargar_cargos")
        if query_params:
            destino = f"{destino}?{urlencode(query_params)}"
        return redirect(destino)

    contexto = construir_contexto_carga(request)
    return render(request, "reunidas_pof/cargar_cargos.html", contexto)


@pof_required
def cargar_cargos_proyecto_especial(request):
    """
    Renderiza directamente la pantalla especializada de Alta de Proyecto Especial.

    - Permite abrir la pantalla sin `proyecto_especial_id` para elegir cabecera.
    - Resuelve el Proyecto Especial indicado cuando el parametro viene informado.
    - Entrega el listado de proyectos necesario para el selector especializado.
    - Muestra un error en la misma plantilla cuando el identificador informado no existe o no es valido.
    """
    proyecto_especial_id = str(request.GET.get("proyecto_especial_id", "") or "").strip()
    proyectos_especiales = ProyectosEspecialesPof.objects.all().order_by(
        "-anio", "nombre", "resolucion"
    )
    contexto = {
        "mensaje_error": "",
        "proyecto_especial": None,
        "proyecto_especial_resolucion": "",
        "proyectos_especiales": proyectos_especiales,
    }

    if not proyecto_especial_id:
        return render(request, "reunidas_pof/cargar_cargos_proyecto_especial.html", contexto)

    if not proyecto_especial_id.isdigit():
        contexto["mensaje_error"] = "El Proyecto Especial seleccionado no es valido."
        return render(request, "reunidas_pof/cargar_cargos_proyecto_especial.html", contexto)

    try:
        proyecto_especial = ProyectosEspecialesPof.objects.get(pk=proyecto_especial_id)
    except ProyectosEspecialesPof.DoesNotExist:
        contexto["mensaje_error"] = "El Proyecto Especial seleccionado no existe o no esta disponible."
        return render(request, "reunidas_pof/cargar_cargos_proyecto_especial.html", contexto)

    contexto.update(
        {
            "proyecto_especial": proyecto_especial,
            "proyecto_especial_resolucion": proyecto_especial.resolucion or "",
        }
    )
    return render(request, "reunidas_pof/cargar_cargos_proyecto_especial.html", contexto)


@pof_required
def reunidas_pof(request):
    contexto = construir_contexto_reunidas(request)
    return render(request, "reunidas_pof/reunidas.html", contexto)


@pof_required
def crear_reunida(request):
    """
    Crea una Reunida POF y hereda su estado inicial dentro de una transacción.

    - Mantiene al formulario como entrada exclusiva de año y nivel.
    - Delega la resolución de base al modelo y la clonación al servicio dedicado.
    - Revierte la cabecera si la herencia no puede completarse íntegramente.
    """
    if request.method == "POST":
        form = ReunidaPofForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    reunida = form.save()
                    heredar_estado_inicial_reunida(reunida, usuario=request.user)
            except IntegrityError:
                form.add_error(None, DUPLICADO_REUNIDA_POF)
            except ValidationError as error:
                form.add_error(None, error)
            else:
                messages.success(
                    request,
                    "Reunida POF creada correctamente.",
                    fail_silently=True,
                )
                return redirect("reunidas_pof:reunidas_pof")
    else:
        form = ReunidaPofForm()

    return render(request, "reunidas_pof/crear_reunida.html", {"form": form})


def _obtener_filtros_proyectos_con_errores(request):
    filtros = {
        "anio": "",
        "proyecto_especial_id": "",
    }
    errores = {}

    anio, error = validar_anio_filtro(request.GET.get("anio", ""))
    filtros["anio"] = anio
    if error:
        errores["anio"] = error

    proyecto_especial_id = str(request.GET.get("proyecto_especial_id", "") or "").strip()
    if proyecto_especial_id:
        if not proyecto_especial_id.isdigit():
            errores["proyecto_especial_id"] = "El Proyecto Especial seleccionado no es válido."
        elif not ProyectosEspecialesPof.objects.filter(pk=proyecto_especial_id).exists():
            errores["proyecto_especial_id"] = "El Proyecto Especial seleccionado no existe."
        else:
            filtros["proyecto_especial_id"] = proyecto_especial_id

    return filtros, errores

@pof_required
def proyectos_especiales_pof(request):
    filtros, errores_filtros = _obtener_filtros_proyectos_con_errores(request)
    filtro_anio = filtros["anio"]
    filtro_proyecto_especial_id = filtros["proyecto_especial_id"]

    proyectos = (
        ProyectosEspecialesPof.objects.select_related("proyecto_base_anterior")
        .annotate(
            ultima_modificacion_historial=Max(
                "localizaciones__cargos__movimientos__fecha"
            )
        )
    )
    proyectos_select = ProyectosEspecialesPof.objects.all()

    if errores_filtros:
        proyectos = proyectos.none()
    else:
        if filtro_anio:
            anio_entero = int(filtro_anio)
            proyectos = proyectos.filter(anio=anio_entero)
            proyectos_select = proyectos_select.filter(anio=anio_entero)

        if filtro_proyecto_especial_id:
            proyectos = proyectos.filter(pk=filtro_proyecto_especial_id)

    proyectos = proyectos.order_by("-anio", "nombre", "resolucion")
    proyectos_select = proyectos_select.order_by("-anio", "nombre", "resolucion")

    try:
        proyectos_select_lista = list(proyectos_select)
        paginator = Paginator(proyectos, 10)
        page_obj = paginator.get_page(request.GET.get("page"))
        total_registros = paginator.count
        proyectos_pagina = page_obj.object_list
        for proyecto in proyectos_pagina:
            ultima_modificacion_historial = proyecto.ultima_modificacion_historial
            proyecto.ultima_modificacion = proyecto.actualizado_en
            if ultima_modificacion_historial and (
                proyecto.ultima_modificacion is None
                or ultima_modificacion_historial > proyecto.ultima_modificacion
            ):
                proyecto.ultima_modificacion = ultima_modificacion_historial
        tabla_proyectos_no_migrada = False
    except (ProgrammingError, OperationalError):
        paginator = Paginator([], 10)
        page_obj = paginator.get_page(1)
        total_registros = 0
        proyectos_pagina = []
        proyectos_select_lista = []
        tabla_proyectos_no_migrada = True

    query_params = request.GET.copy()
    query_params.pop("page", None)
    query_params.pop("nombre", None)
    query_params.pop("resolucion", None)
    query_params.pop("page_size", None)

    for nombre, valor in (
        ("anio", filtro_anio),
        ("proyecto_especial_id", filtro_proyecto_especial_id),
    ):
        if valor:
            query_params[nombre] = valor
        else:
            query_params.pop(nombre, None)

    return render(
        request,
        "reunidas_pof/proyectos_especiales.html",
        {
            "proyectos": proyectos_pagina,
            "proyectos_select": proyectos_select_lista,
            "total_registros": total_registros,
            "showing_start": page_obj.start_index() if total_registros else 0,
            "showing_end": page_obj.end_index() if total_registros else 0,
            "page_obj": page_obj,
            "paginator": paginator,
            "page_range": list(paginator.get_elided_page_range(number=page_obj.number)),
            "query_params_base": query_params.urlencode(),
            "filtro_anio": filtro_anio,
            "filtro_proyecto_especial_id": filtro_proyecto_especial_id,
            "errores_filtros": errores_filtros,
            "mensaje_filtros": MENSAJE_FILTROS_INVALIDOS if errores_filtros else "",
            "tabla_proyectos_no_migrada": tabla_proyectos_no_migrada,
        },
    )


@pof_required
def crear_proyecto_especial(request):
    if request.method == "POST":
        form = ProyectoEspecialPofForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    proyecto = form.save()
                    heredar_estado_inicial_proyecto_especial(
                        proyecto,
                        usuario=request.user,
                    )
            except IntegrityError:
                form.add_error(None, "Ya existe un Proyecto Especial POF para ese año, nombre y resolución.")
            except ValidationError as error:
                form.add_error(None, error)
            else:
                messages.success(
                    request,
                    "Proyecto Especial POF creado correctamente.",
                    fail_silently=True,
                )
                return redirect("reunidas_pof:proyectos_especiales_pof")
    else:
        form = ProyectoEspecialPofForm()

    return render(
        request,
        "reunidas_pof/proyecto_especial_form.html",
        {"form": form, "titulo": "CREAR PROYECTO ESPECIAL"},
    )


@pof_required
def editar_proyecto_especial(request, pk):
    proyecto = get_object_or_404(ProyectosEspecialesPof, pk=pk)

    if request.method == "POST":
        form = ProyectoEspecialPofForm(request.POST, instance=proyecto)
        if form.is_valid():
            try:
                form.save()
            except IntegrityError:
                form.add_error(None, "Ya existe un Proyecto Especial para ese año, nombre y resolución.")
            except ValidationError as error:
                form.add_error(None, error)
            else:
                messages.success(
                    request,
                    "Proyecto Especial actualizado correctamente.",
                    fail_silently=True,
                )
                return redirect("reunidas_pof:proyectos_especiales_pof")
    else:
        form = ProyectoEspecialPofForm(instance=proyecto)

    return render(
        request,
        "reunidas_pof/proyecto_especial_form.html",
        {"form": form, "titulo": "Editar Proyecto Especial POF", "proyecto": proyecto},
    )


@pof_api_required
@require_POST
def eliminar_reunida(request, pk):
    """
    Solicita la eliminación segura de una Reunida desde su acción administrativa.

    - Acepta solo POST y delega la decisión de eliminabilidad al servicio transaccional.
    - Devuelve una respuesta JSON reutilizando el contrato API existente del módulo.
    - No recibe criterios de borrado ni estado de dependencias desde el navegador.
    """
    resultado = eliminar_reunida_pof(pk)
    return _respuesta_api_desde_resultado(resultado)


@pof_api_required
@require_POST
def eliminar_proyecto_especial(request, pk):
    """
    Solicita la eliminación segura de un Proyecto Especial desde su listado.

    - Acepta solo POST y mantiene la decisión de integridad dentro del servicio.
    - Devuelve una respuesta JSON reutilizando el contrato API existente del módulo.
    - No elimina ni valida dependencias a partir de datos enviados por la interfaz.
    """
    resultado = eliminar_proyecto_especial_pof(pk)
    return _respuesta_api_desde_resultado(resultado)


@pof_required
def cargos_pof(request):
    """Muestra la pantalla puente de Cargos POF."""
    return render(request, "reunidas_pof/cargos.html")


@pof_required
def consultar_cargos(request):
    contexto = construir_contexto_consulta(request)
    return render(request, "reunidas_pof/consultar_cargos.html", contexto)


@pof_visualizacion_required
def visualizacion_cargos_localizacion(request):
    contexto = construir_contexto_visualizacion_cargos_localizacion(request)
    contexto["pof_solo_visualizacion"] = not usuario_tiene_acceso_completo_pof(
        request.user
    )
    return render(request, "reunidas_pof/visualizacion_cargos_localizacion.html", contexto)


@pof_visualizacion_api_required
@require_GET
def visualizacion_cargos_localizacion_datos(request):
    payload = construir_payload_visualizacion_cargos_localizacion(request)
    return JsonResponse(payload)


@pof_visualizacion_api_required
@require_GET
def visualizacion_cargos_localizacion_filtros(request):
    try:
        opciones = construir_opciones_filtros_visualizacion_cargos_localizacion(request)
    except Exception:
        logger.exception("Error al construir opciones de filtros de Visualizacion de Cargos.")
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "No se pudieron cargar las opciones de filtros.",
                "opciones": {},
            },
            status=500,
        )
    return JsonResponse(opciones)


def _respuesta_excel_visualizacion_cargos_localizacion(payload):
    response = HttpResponse(
        payload["contenido"],
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{payload["nombre_archivo"]}"'
    return response


@pof_visualizacion_api_required
@require_GET
def visualizacion_cargos_localizacion_exportar_filtros(request):
    payload = construir_excel_visualizacion_cargos_localizacion(
        request,
        exportar_todo=False,
    )
    response = _respuesta_excel_visualizacion_cargos_localizacion(payload)
    return _marcar_descarga_excel_lista(request, response)


@pof_visualizacion_api_required
@require_GET
def visualizacion_cargos_localizacion_exportar_todo(request):
    payload = construir_excel_visualizacion_cargos_localizacion(
        request,
        exportar_todo=True,
    )
    response = _respuesta_excel_visualizacion_cargos_localizacion(payload)
    return _marcar_descarga_excel_lista(request, response)


@pof_required
def detalle_reunida(request):
    contexto = construir_contexto_detalle_reunida(request)
    return render(request, "reunidas_pof/detalle_reunida.html", contexto)


@pof_api_required
@require_GET
def detalle_reunida_opciones_filtro(request):
    """
    Devuelve por JSON un catálogo remoto del Detalle común.

    - Acepta únicamente campos validados por el servicio de Padrón.
    - Ejecuta una consulta de lectura por solicitud y no persiste resultados.
    - Responde errores genéricos sin exponer detalles de SQL o infraestructura.
    """
    campo = (request.GET.get("campo") or "").strip()
    try:
        opciones = obtener_opciones_filtro_detalle_padron(campo)
    except ValueError:
        return api_error_validacion(
            "El campo de filtro no es válido.",
            {"campo": ["El campo solicitado no está habilitado."]},
        )
    except DatabaseError:
        logger.exception("Error al cargar el catálogo remoto del Detalle: %s", campo)
        return api_error_interno(
            "No se pudieron cargar las opciones del filtro.",
        )

    return JsonResponse({"ok": True, "campo": campo, "opciones": opciones})


@pof_required
def historial_movimientos(request):
    contexto = construir_contexto_historial(request)
    return render(request, "reunidas_pof/historial_movimientos.html", contexto)


@pof_api_required
def detalle_movimiento_pof(request, movimiento_id):
    try:
        detalle = obtener_detalle_movimiento_pof(movimiento_id)
    except ObjectDoesNotExist:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "El movimiento indicado no existe.",
                "errores": {"movimiento_id": ["No se encontró el movimiento."]},
            },
            status=404,
        )
    except Exception:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "No se pudo obtener el detalle del movimiento.",
            },
            status=500,
        )

    return JsonResponse({"ok": True, "movimiento": detalle})


@pof_api_required
@require_GET
def historial_cantidad_cargos_pof(request):
    """
    Devuelve el historial real de modificaciones de cantidad de cargos POF.

    - Acepta uno o varios cargo_id de una misma fila consolidada.
    - Valida los identificadores en backend.
    - Reutiliza el servicio de historial de cantidad existente.
    - No modifica datos.
    """
    try:
        historial = obtener_historial_cantidad_cargos_pof(
            request.GET.getlist("cargo_id")
        )
    except ValidationError as error:
        return api_error_validacion(
            "Los cargos indicados no son validos para consultar este historial.",
            getattr(
                error,
                "message_dict",
                {"cargo_ids": error.messages},
            ),
        )
    except ObjectDoesNotExist:
        return api_error_no_encontrado(
            "No se encontro uno de los cargos indicados.",
            {
                "cargo_ids": [
                    "Uno o mas cargos no existen."
                ]
            },
        )
    except Exception:
        logger.exception(
            "Error interno al obtener el historial de cantidad de cargos POF"
        )
        return api_error_interno(
            "Ocurrio un error interno al obtener el historial de cantidad."
        )

    return api_ok(data=historial)


@pof_api_required
@require_GET
def historial_observacion_cargos_pof(request):
    """Devuelve cambios reales de observación para una fila exportada."""
    try:
        historial = obtener_historial_observacion_cargos_pof(
            request.GET.getlist("cargo_id")
        )
    except ValidationError as error:
        return api_error_validacion(
            "Los cargos indicados no son validos para consultar este historial.",
            getattr(
                error,
                "message_dict",
                {"cargo_ids": error.messages},
            ),
        )
    except ObjectDoesNotExist:
        return api_error_no_encontrado(
            "No se encontro uno de los cargos indicados.",
            {"cargo_ids": ["Uno o mas cargos no existen."]},
        )
    except Exception:
        logger.exception(
            "Error interno al obtener el historial de observacion de cargos POF"
        )
        return api_error_interno(
            "Ocurrio un error interno al obtener el historial de observacion."
        )

    return api_ok(data=historial)


@pof_api_required
@require_GET
def historial_estado_cargos_pof(request):
    """
    Devuelve el historial real de cambios entre Afectado y Desafectado.

    - Acepta uno o varios cargo_id de una misma fila consolidada.
    - Valida los identificadores en backend.
    - No expone errores internos ni modifica datos.
    """
    try:
        historial = obtener_historial_estado_cargos_pof(
            request.GET.getlist("cargo_id")
        )
    except ValidationError as error:
        return api_error_validacion(
            "Los cargos indicados no son validos para consultar este historial.",
            getattr(
                error,
                "message_dict",
                {"cargo_ids": error.messages},
            ),
        )
    except ObjectDoesNotExist:
        return api_error_no_encontrado(
            "No se encontro uno de los cargos indicados.",
            {
                "cargo_ids": [
                    "Uno o mas cargos no existen."
                ]
            },
        )
    except Exception:
        logger.exception(
            "Error interno al obtener el historial de estado de cargos POF"
        )
        return api_error_interno(
            "Ocurrio un error interno al obtener el historial de estado."
        )

    return api_ok(data=historial)


def _obtener_payload_json(request):
    try:
        return json.loads(request.body or "{}"), None
    except json.JSONDecodeError:
        return None, api_error_validacion(
            "El cuerpo de la solicitud no es un JSON válido.",
            {"payload": ["JSON inválido."]},
        )


def _respuesta_api_desde_resultado(resultado):
    if resultado.get("ok"):
        data = {
            clave: valor
            for clave, valor in resultado.items()
            if clave not in {"ok", "tipo", "mensaje", "errores"}
        }
        return api_ok(resultado.get("mensaje", ""), data=data)

    tipo = resultado.get("tipo") or "validacion"
    mensaje = resultado.get("mensaje", "No se pudo completar la acción.")
    errores = resultado.get("errores", {})

    if tipo == "sin_cambios":
        return api_error_sin_cambios(mensaje)
    if tipo == "no_encontrado":
        return api_error_no_encontrado(mensaje, errores)
    if tipo == "permiso":
        return api_error_permiso(mensaje)
    if tipo == "sesion":
        return api_error_sesion(mensaje)
    if tipo == "interno":
        return api_error_interno(mensaje)
    return api_error_validacion(mensaje, errores)


def _validar_parametros_grupo_detalle(cueanexo, cuof):
    """
    Valida los parametros canonicos del endpoint de cargos por Anexo.

    - Exige `cueanexo` de 9 digitos para preservar la clave tecnica real.
    - Conserva `cuof` solo como compatibilidad de enlaces viejos.
    - Devuelve un mapa de errores apto para las respuestas JSON del modulo.
    """
    errores = {}

    if not cueanexo:
        errores["cueanexo"] = ["Debe indicar un CUEANEXO."]
    elif not cueanexo.isdigit() or len(cueanexo) != 9:
        errores["cueanexo"] = ["El CUEANEXO debe contener exactamente 9 digitos."]

    return errores


@pof_api_required
def detalle_cargo_pof(request, cargo_id):
    try:
        detalle = obtener_detalle_cargo_pof(cargo_id)
    except ObjectDoesNotExist:
        return api_error_no_encontrado(
            "No se encontró el cargo solicitado.",
            {"cargo_id": ["No se encontró el cargo solicitado."]},
        )
    except Exception:
        logger.exception("Error interno al obtener detalle de cargo POF %s", cargo_id)
        return api_error_interno(
            "Ocurrió un error interno al obtener el cargo. Informe al administrador."
        )

    return api_ok(data={"cargo": detalle})


@pof_api_required
@require_GET
def detalle_reunida_grupo_cargos(request, reunida_id):
    """
    Devuelve por JSON los cargos normalizados de un grupo Anexo de la Reunida.

    - Restringe la consulta a una Reunida comun valida mediante `reunida_id`.
    - Filtra por la clave operativa `reunida + cueanexo`; `cuof` queda como compatibilidad.
    - Reutiliza la misma normalizacion y consolidacion actual del Detalle.
    - Marca en lote las filas con cambios reales de Estado POF para el historial bajo demanda.
    """
    cueanexo = request.GET.get("cueanexo", "").strip()
    cuof = request.GET.get("cuof", "").strip()
    errores = _validar_parametros_grupo_detalle(cueanexo, cuof)
    if errores:
        return api_error_validacion(
            "Hay errores de validacion en los parametros del grupo solicitado.",
            errores,
        )

    try:
        reunida = ReunidaPof.objects.get(pk=reunida_id)
    except ReunidaPof.DoesNotExist:
        return api_error_no_encontrado(
            "No se encontro la Reunida solicitada.",
            {"reunida_id": ["No se encontro la POF solicitada."]},
        )

    try:
        cargos_grupo = obtener_cargos_grilla_reunida(reunida=reunida).filter(
            localizacion__cueanexo=cueanexo,
        )
        grilla_detalle = construir_grilla_pof_desde_cargos(
            cargos=cargos_grupo,
            nivel_codigo=reunida.nivel,
            contexto="DETALLE_REUNIDA",
            espejo=False,
            incluir_historial_estado=True,
        )
        grupo_operativo = obtener_grupo_operativo_detalle(
            grilla_detalle.get("grupos_operativos_detalle", []),
            cueanexo=cueanexo,
            cuof=cuof,
        )
    except Exception:
        logger.exception(
            "Error interno al cargar el grupo operativo de detalle para reunida %s, cueanexo %s, cuof %s",
            reunida_id,
            cueanexo,
            cuof,
        )
        return api_error_interno(
            "Ocurrio un error interno al cargar los cargos del grupo solicitado."
        )

    if not grupo_operativo:
        return api_error_no_encontrado(
            "No existe un grupo con el CUEANEXO indicado dentro de la Reunida.",
            {
                "cueanexo": ["No existe un grupo con el CUEANEXO indicado en la POF."],
            },
        )

    return JsonResponse({
        "ok": True,
        **serializar_grupo_operativo_detalle(grupo_operativo),
    })


@pof_api_required
@require_GET
def detalle_proyecto_especial_localizacion_cargos(
    request,
    proyecto_especial_id,
    localizacion_id,
):
    try:
        proyecto = ProyectosEspecialesPof.objects.get(pk=proyecto_especial_id)
    except ProyectosEspecialesPof.DoesNotExist:
        return api_error_no_encontrado(
            "No se encontro el Proyecto Especial solicitado.",
            {"proyecto_especial_id": ["No se encontro el Proyecto Especial solicitado."]},
        )

    try:
        LocalizacionPof.objects.get(
            pk=localizacion_id,
            proyecto_especial_id=proyecto.id,
        )
    except LocalizacionPof.DoesNotExist:
        return api_error_no_encontrado(
            "No existe la localizacion indicada dentro del Proyecto Especial.",
            {
                "localizacion_id": [
                    "La localizacion no pertenece al Proyecto Especial solicitado."
                ],
            },
        )

    try:
        payload = construir_payload_cargos_detalle_proyecto_localizacion(
            proyecto.id,
            localizacion_id,
        )
    except Exception:
        logger.exception(
            "Error interno al cargar cargos de Proyecto Especial %s, localizacion %s",
            proyecto.id,
            localizacion_id,
        )
        return api_error_interno(
            "Ocurrio un error interno al cargar los cargos de la localizacion solicitada."
        )

    if not payload:
        return api_error_no_encontrado(
            "No hay cargos para la localizacion indicada.",
            {"localizacion_id": ["La localizacion no posee cargos disponibles."]},
        )

    return JsonResponse({"ok": True, **payload})


@pof_api_required
@require_POST
def modificar_cargo_pof_view(request, cargo_id):
    payload, respuesta_error = _obtener_payload_json(request)
    if respuesta_error:
        return respuesta_error

    usuario = request.user if request.user.is_authenticated else None
    try:
        resultado = modificar_cargo_pof(cargo_id, payload, usuario=usuario)
    except ValidationError as error:
        logger.warning(
            "Validacion no controlada al modificar cargo POF %s: %s",
            cargo_id,
            error,
        )
        return api_error_validacion("No se pudo modificar el cargo.", error)
    except IntegrityError as error:
        logger.warning(
            "Integridad no controlada al modificar cargo POF %s: %s",
            cargo_id,
            error,
        )
        return api_error_validacion(
            "No se pudo modificar el cargo.",
            {"__all__": ["Los datos enviados no cumplen las reglas de integridad."]},
        )
    except Exception:
        logger.exception("Error interno al modificar cargo POF %s", cargo_id)
        return api_error_interno(
            "Ocurrió un error interno al guardar. Informe al administrador."
        )
    return _respuesta_api_desde_resultado(resultado)


@pof_api_required
@require_POST
def cambiar_estado_cargo_pof_view(request, cargo_id):
    payload, respuesta_error = _obtener_payload_json(request)
    if respuesta_error:
        return respuesta_error

    usuario = request.user if request.user.is_authenticated else None
    estado_pof = payload.get("estado_pof") if isinstance(payload, dict) else ""
    try:
        resultado = cambiar_estado_cargo_pof(cargo_id, estado_pof, usuario=usuario)
    except Exception:
        logger.exception("Error interno al cambiar estado de cargo POF %s", cargo_id)
        return api_error_interno(
            "Ocurrió un error interno al guardar. Informe al administrador."
        )
    return _respuesta_api_desde_resultado(resultado)


@pof_api_required
@require_POST
def eliminar_cargo_pof_view(request, cargo_id):
    usuario = request.user if request.user.is_authenticated else None
    try:
        resultado = eliminar_cargo_pof(cargo_id, usuario=usuario)
    except Exception:
        logger.exception("Error interno al eliminar cargo POF %s", cargo_id)
        return api_error_interno(
            "Ocurrió un error interno al guardar. Informe al administrador."
        )
    return _respuesta_api_desde_resultado(resultado)


@pof_api_required
def validar_reunida(request):
    resultado = validar_cabecera_reunida(
        request.GET.get("anio", ""),
        request.GET.get("nivel", ""),
    )
    status = 200 if resultado.get("ok") else 400
    return JsonResponse(resultado, status=status)


@pof_api_required
@require_GET
def previsualizar_reunida_base(request):
    """
    Devuelve la previsualización mínima de una nueva Reunida y su base posible.

    - Valida año y nivel sin crear ni modificar registros.
    - Busca únicamente la combinación exacta de nivel y año inmediatamente anterior.
    - No expone IDs ni permite que el frontend decida la FK real de creación.
    """
    anio_texto = str(request.GET.get("anio", "")).strip()
    nivel = normalizar_nivel(request.GET.get("nivel", ""))

    errores = {}
    niveles_validos = {
        valor
        for valor, _ in ReunidaPof.Nivel.choices
    }

    if not anio_texto.isdigit() or len(anio_texto) != 4:
        errores["anio"] = "El año debe tener 4 dígitos numéricos."

    if not nivel or nivel not in niveles_validos:
        errores["nivel"] = "Debe seleccionar un nivel válido."

    if errores:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "Revisá los campos marcados antes de continuar.",
                "errores": errores,
            },
            status=400,
        )

    anio = int(anio_texto)
    anio_maximo = timezone.localdate().year + 1

    if anio > anio_maximo:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "Revisá los campos marcados antes de continuar.",
                "errores": {
                    "anio": "El año no puede superar el año próximo.",
                },
            },
            status=400,
        )

    if ReunidaPof.objects.filter(anio=anio, nivel=nivel).exists():
        return JsonResponse(
            {
                "ok": False,
                "mensaje": DUPLICADO_REUNIDA_POF,
                "errores": {},
            },
            status=400,
        )

    reunida_base = ReunidaPof.objects.filter(
        anio=anio - 1,
        nivel=nivel,
    ).only("anio", "nivel").first()

    return JsonResponse(
        {
            "ok": True,
            "destino": {
                "anio": anio,
                "nivel_nombre": ReunidaPof.Nivel(nivel).label,
            },
            "base": (
                {
                    "existe": True,
                    "anio": reunida_base.anio,
                    "nivel_nombre": reunida_base.get_nivel_display(),
                }
                if reunida_base
                else {
                    "existe": False,
                    "anio": anio - 1,
                    "nivel_nombre": ReunidaPof.Nivel(nivel).label,
                }
            ),
        }
    )


@pof_api_required
def buscar_padron(request):
    """
    Busca ofertas de padrón respetando la cabecera validada de la carga.

    - Para Reunidas exige año y nivel válidos antes de consultar padrón.
    - Propaga el nivel validado al servicio para marcar ofertas sugeridas.
    - Mantiene el comportamiento actual de Proyecto Especial cuando no aplica nivel.
    """
    cueanexo = request.GET.get("cueanexo", "").strip()
    cue = request.GET.get("cue", "").strip()
    cuof = request.GET.get("cuof", "").strip()
    cabecera_tipo = request.GET.get("cabecera_tipo", "REUNIDA").strip().upper()
    anio = request.GET.get("anio", "").strip()
    nivel = request.GET.get("nivel", "").strip()
    nivel_reunida = ""

    if cabecera_tipo != "PROYECTO_ESPECIAL":
        validacion_cabecera = validar_cabecera_reunida(anio, nivel)
        if not validacion_cabecera.get("ok"):
            return JsonResponse(
                {
                    "ok": False,
                    "mensaje": "Primero valide una Cabecera POF con año y nivel.",
                    "errores": validacion_cabecera.get("errores", {}),
                    "resultados": [],
                },
                status=400,
            )
        nivel_reunida = validacion_cabecera.get("reunida", {}).get("nivel", "") or normalizar_nivel(nivel)

    if not any([cueanexo, cue, cuof]):
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "Debe ingresar un CUEANEXO, CUE o CUOF.",
                "resultados": [],
            },
            status=400,
        )

    try:
        resultados = buscar_ofertas_padron(
            cueanexo=cueanexo,
            cue=cue,
            cuof=cuof,
            nivel_reunida=nivel_reunida,
        )
    except Exception as error:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "No se pudo consultar el padrón materializado.",
                "detalle": str(error),
                "resultados": [],
            },
            status=500,
        )

    if not resultados:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "No se encontraron ofertas para la búsqueda ingresada.",
                "resultados": [],
            },
            status=404,
        )

    return JsonResponse({"ok": True, "cantidad": len(resultados), "resultados": resultados})


@pof_api_required
@require_GET
def buscar_padron_proyecto_especial(request):
    cueanexo = request.GET.get("cueanexo", "").strip()
    cue = request.GET.get("cue", "").strip()
    anexo = request.GET.get("anexo", "").strip()

    if cueanexo and (not cueanexo.isdigit() or len(cueanexo) != 9):
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "El CUEANEXO debe tener 9 dígitos.",
                "resultados": [],
            },
            status=400,
        )
    if not cueanexo and cue and (not cue.isdigit() or len(cue) != 7):
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "El CUE debe tener 7 dígitos.",
                "resultados": [],
            },
            status=400,
        )
    if not cueanexo and anexo and (not anexo.isdigit() or len(anexo) != 2):
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "El anexo debe tener 2 dígitos.",
                "resultados": [],
            },
            status=400,
        )

    if not cueanexo and cue and anexo:
        cueanexo = construir_cueanexo_sin_guion(cue, anexo)

    if not cueanexo and not cue:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "Ingresá un CUE, Anexo o CUEANEXO para buscar en padrón.",
                "resultados": [],
            },
            status=400,
        )

    try:
        resultados = buscar_ofertas_padron(
            cueanexo=cueanexo,
            cue="" if cueanexo else cue,
            nivel_reunida="",
        )
    except Exception as error:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "No se pudo consultar el padrón materializado.",
                "detalle": str(error),
                "resultados": [],
            },
            status=500,
        )

    if not resultados:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "No se encontraron ofertas en padrón para la búsqueda indicada. Podés usar ingreso manual controlado si corresponde.",
                "resultados": [],
            },
            status=404,
        )

    return JsonResponse({"ok": True, "cantidad": len(resultados), "resultados": resultados})


@pof_api_required
@require_GET
def catalogos_ingreso_manual_proyecto_especial(request):
    try:
        catalogos = obtener_catalogos_padron_ingreso_manual_pof()
    except Exception:
        logger.exception("Error al cargar catálogos de ingreso manual de Proyecto Especial.")
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "No se pudieron cargar los catálogos de ingreso manual.",
                "catalogos": {},
            },
            status=500,
        )

    catalogos.update({
        "estado_localizacion_padron": ["Activo", "Baja"],
        "estado_oferta_padron": ["Activo", "Baja"],
        "estado_establecimiento_padron": ["Activo", "Baja"],
    })
    return JsonResponse({"ok": True, "catalogos": catalogos})


@pof_api_required
@require_GET
def buscar_cuof_manual_proyecto_especial(request):
    resultado = buscar_cuof_manual_proyecto_especial_service(
        proyecto_especial_id=request.GET.get("proyecto_especial_id", ""),
        cuof=request.GET.get("cuof", ""),
    )

    if not resultado.get("ok"):
        return JsonResponse(resultado, status=400)

    return JsonResponse(resultado)


@pof_api_required
@require_GET
def buscar_ceic_proyecto_especial(request):
    q = request.GET.get("q", "")
    if not q:
        return api_error_validacion(
            "Ingresá un CEIC para buscar el cargo.",
            {"ceic": ["Ingresá un CEIC para buscar el cargo."]},
        )

    error_ceic = _validar_busqueda_ceic_codigo(q)
    if error_ceic:
        return api_error_validacion(
            error_ceic,
            {"ceic": [error_ceic]},
        )

    try:
        filas = _consultar_ceic_puntos(niveles_ceic=None, q=q, limite=5)
    except Exception:
        logger.exception("Error interno al buscar CEIC de Proyecto Especial")
        return api_error_interno(
            "Ocurrió un error interno al buscar CEIC. Informe al administrador."
        )

    resultados = [
        {
            "ceic": fila[0],
            "cargo": fila[1],
            "puntos": str(fila[2]),
            "nivel": fila[3],
            "es_sugerido": True,
        }
        for fila in filas
    ]

    return api_ok(data={"cantidad": len(resultados), "resultados": resultados})


@pof_api_required
@require_GET
def catalogo_ceic_proyecto_especial(request):
    """
    Devuelve el catálogo completo de CEIC activos disponible para Proyecto Especial.

    - No aplica compatibilidad por nivel ni reglas de sugeridos de Reunida.
    - Solicita el catálogo completo para no truncar códigos activos válidos.
    - Conserva la validación definitiva de CEIC y puntos en el servicio de guardado.
    """
    try:
        filas = _consultar_ceic_puntos(niveles_ceic=None, limite=None)
    except Exception:
        logger.exception("Error interno al cargar catálogo CEIC de Proyecto Especial")
        return api_error_interno(
            "Ocurrió un error interno al cargar el catálogo CEIC. Informe al administrador."
        )

    resultados = [
        {
            "ceic": str(fila[0]),
            "cargo": fila[1],
            "puntos": str(fila[2]),
            "nivel": fila[3],
            "es_sugerido": True,
        }
        for fila in filas
    ]

    return api_ok(data={"cantidad": len(resultados), "resultados": resultados})


@pof_api_required
def buscar_ceic(request):
    q = request.GET.get("q", "")
    nivel = request.GET.get("nivel", "").strip()
    modo = request.GET.get("modo", "sugeridos").strip().lower()
    niveles_ceic = _obtener_niveles_ceic_validos(nivel)

    if not niveles_ceic:
        return api_error_validacion(
            "Debe indicar un nivel de Reunida válido para buscar CEIC.",
            {"nivel": ["Debe indicar un nivel de Reunida válido para buscar CEIC."]},
        )

    error_ceic = _validar_busqueda_ceic_codigo(q)
    if error_ceic:
        return api_error_validacion(
            error_ceic,
            {"ceic": [error_ceic]},
        )

    try:
        filas = _consultar_ceic_puntos(
            niveles_ceic=None if modo == "otros" else niveles_ceic,
            q=q,
            limite=5,
        )
    except Exception:
        logger.exception("Error interno al buscar CEIC en alta de cargos")
        return api_error_interno(
            "Ocurrió un error interno al buscar CEIC. Informe al administrador."
        )

    resultados = [
        {
            "ceic": fila[0],
            "cargo": fila[1],
            "puntos": str(fila[2]),
            "nivel": fila[3],
            "es_sugerido": fila[3] in niveles_ceic,
        }
        for fila in filas
    ]

    return api_ok(data={"cantidad": len(resultados), "resultados": resultados})


@pof_api_required
@require_GET
def catalogo_ceic(request):
    """
    Carga el catálogo CEIC activo permitido para la Reunida seleccionada.

    - Exige un nivel válido para no exponer un catálogo general.
    - Devuelve solo CEIC activos y compatibles con el nivel recibido.
    - Reutiliza la misma matriz de niveles que la búsqueda dinámica.
    """
    nivel = request.GET.get("nivel", "").strip()
    niveles_ceic = _obtener_niveles_ceic_validos(nivel)

    if not niveles_ceic:
        return api_error_validacion(
            "Debe indicar un nivel de Reunida válido para cargar el catálogo CEIC.",
            {"nivel": ["Debe indicar un nivel de Reunida válido para cargar el catálogo CEIC."]},
        )

    try:
        filas = _consultar_ceic_puntos(niveles_ceic, limite=500)
    except Exception:
        logger.exception("Error interno al cargar catálogo CEIC")
        return api_error_interno(
            "Ocurrió un error interno al cargar el catálogo CEIC. Informe al administrador."
        )

    resultados = [
        {
            "ceic": str(fila[0]),
            "cargo": fila[1],
            "puntos": str(fila[2]),
            "nivel": fila[3],
            "es_sugerido": True,
        }
        for fila in filas
    ]

    return api_ok(data={"cantidad": len(resultados), "resultados": resultados})


@pof_api_required
@require_POST
def guardar_carga_pof(request):
    try:
        payload = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "El cuerpo de la solicitud no es un JSON válido.",
                "errores": {"payload": ["JSON inválido."]},
            },
            status=400,
        )

    cabecera_tipo = str(payload.get("cabecera_tipo", "") or "").strip().upper()
    proyecto_especial_id = str(payload.get("proyecto_especial_id", "") or "").strip()
    if cabecera_tipo == "PROYECTO_ESPECIAL" or proyecto_especial_id:
        validacion = validar_payload_guardar_carga_proyecto_especial(payload)
    else:
        validacion = validar_payload_guardar_carga(payload)
    if not validacion["ok"]:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "Hay errores de validación.",
                "errores": validacion["errores"],
            },
            status=400,
        )

    try:
        usuario = request.user if request.user.is_authenticated else None
        resultado = guardar_carga_pof_service(validacion["datos"], usuario=usuario)
    except Exception as error:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "Ocurrió un error inesperado.",
                "detalle": str(error),
            },
            status=500,
        )

    if not resultado.get("ok"):
        return JsonResponse(resultado, status=400)

    return JsonResponse(resultado, status=201)


@pof_api_required
@require_POST
def guardar_carga_proyecto_especial(request):
    try:
        payload = json.loads(request.body or "{}")
    except json.JSONDecodeError:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "El cuerpo de la solicitud no es un JSON válido.",
                "errores": {"payload": ["JSON inválido."]},
            },
            status=400,
        )

    validacion = validar_payload_guardar_carga_proyecto_especial(payload)
    if not validacion["ok"]:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "Hay errores de validación.",
                "errores": validacion["errores"],
            },
            status=400,
        )

    try:
        usuario = request.user if request.user.is_authenticated else None
        resultado = guardar_carga_pof_service(validacion["datos"], usuario=usuario)
    except Exception as error:
        return JsonResponse(
            {
                "ok": False,
                "mensaje": "Ocurrió un error inesperado.",
                "detalle": str(error),
            },
            status=500,
        )

    if not resultado.get("ok"):
        return JsonResponse(resultado, status=400)

    return JsonResponse(resultado, status=201)


@pof_required
def exportar_reunida(request):
    contexto = construir_contexto_exportacion(request)
    if request.GET.get("accion") == "excel":
        response = _crear_respuesta_excel_exportacion(contexto)
        return _marcar_descarga_excel_lista(request, response)

    return render(request, "reunidas_pof/exportar_reunida.html", contexto)
