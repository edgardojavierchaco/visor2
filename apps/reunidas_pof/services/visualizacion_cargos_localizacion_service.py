import logging
from decimal import Decimal
from io import BytesIO

from django.core.paginator import Paginator
from django.db import DatabaseError, OperationalError, ProgrammingError
from django.db.models import (
    Case,
    CharField,
    IntegerField,
    OuterRef,
    Prefetch,
    Q,
    Subquery,
    Sum,
    Value,
    When,
)
from django.db.models.functions import Cast, Concat, Substr, Trim
from django.http import QueryDict
from django.utils import timezone

from ..models import (
    ROL_POF_DIRECTOR,
    ROL_POF_REGIONAL,
    ROLES_POF_ACCESO_COMPLETO,
    ROLES_POF_SOLO_VISUALIZACION_COMPLETA,
    CargoPof,
    ProyectosEspecialesPof,
    ReunidaPof,
    SnapshotPadronLocalizacionPof,
    obtener_cueanexos_director_pof,
    obtener_regiones_usuario_pof,
    obtener_rol_usuario_pof,
    usuario_tiene_alcance_restringido_pof,
)
from .padron_materializadas_service import (
    REGIONES_EDUCATIVAS_POF,
    normalizar_region_padron,
    obtener_opciones_filtros_visualizacion_padron,
    obtener_variantes_region_padron,
)


GUION = "—"
logger = logging.getLogger(__name__)
CABECERA_PROYECTO_ESPECIAL = "PROYECTO_ESPECIAL"
CUES_POR_PAGINA_VISUALIZACION = 5
MAX_TERMINOS_BUSQUEDA_OBSERVACION = 5

VISUALIZACION_CARGOS_COLUMNAS = [
    {"id": "cueanexo", "label": "CUEANEXO", "visible_default": True},
    {"id": "cue", "label": "CUE", "visible_default": True},
    {"id": "anexo", "label": "Anexo", "visible_default": True},
    {"id": "cuof", "label": "CUOF", "visible_default": True},
    {"id": "cui", "label": "CUI", "visible_default": True},
    {"id": "numero_establecimiento", "label": "Nº Estab.", "visible_default": True},
    {"id": "nombre_establecimiento", "label": "Establecimiento", "visible_default": True},
    {"id": "region", "label": "Región", "visible_default": True},
    {"id": "localidad", "label": "Localidad", "visible_default": True},
    {"id": "departamento", "label": "Departamento", "visible_default": True},
    {"id": "ambito", "label": "Ámbito", "visible_default": False},
    {"id": "categoria", "label": "Categoría", "visible_default": False},
    {"id": "jornada", "label": "Jornada", "visible_default": False},
    {"id": "oferta", "label": "Oferta", "visible_default": True},
    {"id": "acronimo", "label": "Acrónimo", "visible_default": True},
    {"id": "ubicacion", "label": "Ubicación", "visible_default": False},
    {"id": "estado_localizacion_padron", "label": "Estado localización", "visible_default": False},
    {"id": "estado_oferta_padron", "label": "Estado oferta", "visible_default": False},
    {"id": "estado_establecimiento_padron", "label": "Estado establecimiento", "visible_default": False},
    {"id": "ceic", "label": "CEIC", "visible_default": True},
    {"id": "cargo", "label": "Cargo", "visible_default": True},
    {"id": "cantidad", "label": "Cantidad", "visible_default": True},
    {"id": "unidad_cantidad", "label": "Unidad", "visible_default": True},
    {"id": "puntos_asignados", "label": "Puntos", "visible_default": True},
    {"id": "total", "label": "Total", "visible_default": True},
    {
        "id": "total_general",
        "label": "Total General",
        "visible_default": True,
        "ordenable": False,
        "buscable": False,
    },
    {"id": "estado_pof", "label": "Estado POF", "visible_default": True},
    {"id": "observacion", "label": "Observación", "visible_default": False},
    {"id": "actualizado_en", "label": "Actualizado", "visible_default": False},
]

COLUMNAS_POR_ID = {columna["id"]: columna for columna in VISUALIZACION_CARGOS_COLUMNAS}
COLUMNAS_DEFAULT_IDS = [
    columna["id"]
    for columna in VISUALIZACION_CARGOS_COLUMNAS
    if columna["visible_default"]
]
COLUMNAS_TODAS_IDS = [columna["id"] for columna in VISUALIZACION_CARGOS_COLUMNAS]
COLUMNAS_BUSCABLES_IDS = [
    columna["id"]
    for columna in VISUALIZACION_CARGOS_COLUMNAS
    if columna.get("buscable", True)
]

SNAPSHOT_COLUMNAS = {
    "numero_establecimiento": "numero_establecimiento",
    "nombre_establecimiento": "nombre_establecimiento",
    "region": "region",
    "localidad": "localidad",
    "departamento": "departamento",
    "ubicacion": "ubicacion",
    "estado_localizacion_padron": "estado_localizacion_padron",
    "estado_establecimiento_padron": "estado_establecimiento_padron",
}

OFERTA_CARGO_COLUMNAS = {
    "ambito": ("ambito", "oferta_ambito"),
    "categoria": ("categoria", "oferta_categoria"),
    "jornada": ("jornada", "jornada_ofertalocal"),
    "oferta": ("oferta_real", "oferta"),
    "acronimo": ("acronimo",),
    "estado_oferta_padron": ("est_oferta", "estado_oferta_padron"),
}

OFERTAS_FILTRO_VISUALIZACION = (
    {"value": "Común - Ciclos de Enseñanza Artística", "label": "Común - Ciclos de Enseñanza Artística"},
    {"value": "Común - Domiciliaria-hospitalaria. Nivel Primario/EGB", "label": "Común - Domiciliaria-hospitalaria. Nivel Primario/EGB"},
    {"value": "Común - Jardín maternal", "label": "Común - Jardín maternal"},
    {"value": "Común - SNU", "label": "Común - SNU"},
    {"value": "Común - Trayecto Artístico Profesional", "label": "Común - Trayecto Artístico Profesional"},
    {"value": "Común - Cursos de Capacitación de SNU", "label": "Común - Cursos de Capacitación de SNU"},
    {"value": "Común - Otras Titulaciones Técnicas", "label": "Común - Otras Titulaciones Técnicas"},
    {"value": "Común - Trayecto técnico profesional", "label": "Común - Trayecto tecnico profesional"},
    {"value": "Común - Cursos y Talleres de Artística", "label": "Común - Cursos y Talleres de Artística"},
    {"value": "Común - Itinerario formativo", "label": "Común - Itinerario formativo"},
    {"value": "Común - Primaria de 7 años", "label": "Común - Primaria de 7 años"},
    {"value": "Común - Secundaria Completa req. 7 años", "label": "Común - Secundaria Completa req. 7 años"},
    {"value": "Sin Información", "label": "Sin Información"},
    {"value": "Común - Domiciliaria-hospitalaria. Nivel Inicial", "label": "Común - Domiciliaria-hospitalaria. Nivel Inicial"},
    {"value": "Común - Jardín de infantes", "label": "Común - Jardín de infantes"},
    {"value": "Común - Servicios complementarios", "label": "Común - Servicios Complementarios"},
    {"value": "Especial - Cursos/Talleres de la Escuela Especial", "label": "Especial - Cursos/Talleres de la Escuela Especial"},
    {"value": "Especial - Educación Integral para adolescentes y jóvenes", "label": "Especial - Educacion integral para adolescentes y jovenes"},
    {"value": "Especial - Integración", "label": "Especial - Integración"},
    {"value": "Especial - Primaria de 7 años", "label": "Especial - Primaria de 7 años"},
    {"value": "Especial - Domiciliaria-hospitalaria. Nivel Inicial", "label": "Especial - Domiciliaria-hospitalaria. Nivel Inicial"},
    {"value": "Especial - Jardín de infantes", "label": "Especial - Jardín de infantes"},
    {"value": "Especial - Domiciliaria-hospitalaria. Nivel Primario/EGB", "label": "Especial - Domiciliaria-hospitalaria. Nivel Primario/EGB"},
    {"value": "Especial - Educación Integral para Adolescentes y Jóvenes/Secundaria Completa req. 7 años", "label": "Especial - Educación Integral para Adolescentes y Jóvenes/Secundaria Completa req. 7 años"},
    {"value": "Especial - Jardín maternal", "label": "Especial - Jardín maternal"},
    {"value": "Adultos - Primaria", "label": "Adultos - Primaria"},
    {"value": "Adultos - Formación Profesional", "label": "Adultos - Formación Profesional"},
    {"value": "Adultos - Secundaria Completa", "label": "Adultos - Secundaria Completa"},
)

COLUMNAS_CENTRADAS = {
    "cueanexo",
    "cue",
    "anexo",
    "cuof",
    "cui",
    "numero_establecimiento",
    "region",
    "ceic",
    "cantidad",
    "unidad_cantidad",
    "puntos_asignados",
    "total",
    "total_general",
    "estado_pof",
    "actualizado_en",
}

COLUMNAS_NO_REPETIR_POR_LOCALIZACION = {
    "cueanexo",
    "cue",
    "anexo",
    "cuof",
    "cui",
    "numero_establecimiento",
    "nombre_establecimiento",
    "region",
    "localidad",
    "departamento",
    "ubicacion",
    "estado_localizacion_padron",
    "estado_establecimiento_padron",
}

FILTROS_TEXTO = (
    {"id": "cueanexo", "label": "CUEANEXO", "placeholder": "9 dígitos", "col": "pof-col-2"},
    {"id": "cue", "label": "CUE base", "placeholder": "7 dígitos", "col": "pof-col-2"},
    {"id": "anexo", "label": "Anexo", "placeholder": "2 dígitos", "col": "pof-col-2"},
    {"id": "cuof", "label": "CUOF", "placeholder": "CUOF", "col": "pof-col-2"},
    {"id": "cui", "label": "CUI", "placeholder": "CUI", "col": "pof-col-2"},
    {"id": "ceic", "label": "CEIC", "placeholder": "CEIC", "col": "pof-col-2"},
    {"id": "cargo", "label": "Cargo/denominación", "placeholder": "Cargo", "col": "pof-col-4"},
    {
        "id": "nombre_establecimiento",
        "label": "Nombre establecimiento",
        "placeholder": "Establecimiento",
        "col": "pof-col-4",
    },
    {
        "id": "numero_establecimiento",
        "label": "Número establecimiento",
        "placeholder": "Nº establecimiento",
        "col": "pof-col-2",
    },
)

FILTROS_SELECT_SNAPSHOT = (
    "region",
    "localidad",
    "departamento",
    "estado_localizacion_padron",
    "estado_establecimiento_padron",
)

FILTROS_SELECT_DEFINICIONES = (
    {"id": "estado_pof", "label": "Estado POF", "col": "pof-col-2"},
    {"id": "region", "label": "Región", "col": "pof-col-2"},
    {"id": "localidad", "label": "Localidad", "col": "pof-col-3"},
    {"id": "departamento", "label": "Departamento", "col": "pof-col-3"},
    {"id": "ambito", "label": "Ámbito", "col": "pof-col-2"},
    {"id": "categoria", "label": "Categoría", "col": "pof-col-3"},
    {"id": "jornada", "label": "Jornada", "col": "pof-col-3"},
    {"id": "oferta", "label": "Oferta", "col": "pof-col-4"},
    {"id": "acronimo", "label": "Acrónimo", "col": "pof-col-2"},
    {"id": "estado_localizacion_padron", "label": "Estado localización", "col": "pof-col-3"},
    {"id": "estado_oferta_padron", "label": "Estado oferta", "col": "pof-col-3"},
    {"id": "estado_establecimiento_padron", "label": "Estado establecimiento", "col": "pof-col-3"},
)

FILTROS_IDS = [filtro["id"] for filtro in FILTROS_TEXTO] + [
    filtro["id"] for filtro in FILTROS_SELECT_DEFINICIONES
]

OPERADORES_FILTRO = {
    "0": "parecido a",
    "1": "no parecido a",
    "2": "igual a",
    "3": "mayor a",
    "4": "mayor o igual a",
    "5": "menor a",
    "6": "menor o igual a",
    "7": "distinto de",
}

OPERADORES_TEXTO = ("0", "1", "2", "7")
OPERADORES_EXACTOS = ("2", "7")
OPERADORES_NUMERICOS = ("2", "7", "3", "4", "5", "6")

FILTROS_AVANZADOS_CAMPOS = [
    {"id": "cueanexo", "label": "CUEANEXO", "tipo": "text", "operadores": "text"},
    {"id": "cue", "label": "CUE", "tipo": "text", "operadores": "text"},
    {"id": "anexo", "label": "Anexo", "tipo": "text", "operadores": "text"},
    {"id": "cuof", "label": "CUOF", "tipo": "text", "operadores": "text"},
    {"id": "cui", "label": "CUI", "tipo": "text", "operadores": "text"},
    {"id": "nombre_establecimiento", "label": "Establecimiento", "tipo": "text", "operadores": "text"},
    {"id": "numero_establecimiento", "label": "Número establecimiento", "tipo": "text", "operadores": "text"},
    {"id": "region", "label": "Región", "tipo": "checklist", "operadores": "exact"},
    {"id": "localidad", "label": "Localidad", "tipo": "checklist", "operadores": "exact"},
    {"id": "departamento", "label": "Departamento", "tipo": "checklist", "operadores": "exact"},
    {"id": "ambito", "label": "Ámbito", "tipo": "checklist", "operadores": "exact"},
    {"id": "categoria", "label": "Categoría", "tipo": "checklist", "operadores": "exact"},
    {"id": "jornada", "label": "Jornada", "tipo": "checklist", "operadores": "exact"},
    {"id": "oferta", "label": "Oferta", "tipo": "checklist", "operadores": "exact"},
    {"id": "acronimo", "label": "Acrónimo", "tipo": "checklist", "operadores": "exact"},
    {"id": "estado_localizacion_padron", "label": "Estado localización", "tipo": "checklist", "operadores": "exact"},
    {"id": "estado_oferta_padron", "label": "Estado oferta", "tipo": "checklist", "operadores": "exact"},
    {"id": "estado_establecimiento_padron", "label": "Estado establecimiento", "tipo": "checklist", "operadores": "exact"},
    {"id": "ceic", "label": "CEIC", "tipo": "number", "operadores": "numeric"},
    {"id": "cargo", "label": "Cargo", "tipo": "text", "operadores": "text"},
    {"id": "cantidad", "label": "Cantidad", "tipo": "number", "operadores": "numeric"},
    {"id": "unidad_cantidad", "label": "Unidad", "tipo": "checklist", "operadores": "exact"},
    {"id": "puntos_asignados", "label": "Puntos", "tipo": "number", "operadores": "numeric"},
    {"id": "total", "label": "Total", "tipo": "number", "operadores": "numeric"},
    {"id": "estado_pof", "label": "Estado POF", "tipo": "checklist", "operadores": "exact"},
    {"id": "observacion", "label": "Observación", "tipo": "text", "operadores": "text"},
    {"id": "actualizado_en", "label": "Actualizado", "tipo": "text", "operadores": "text"},
]

FILTROS_AVANZADOS_POR_ID = {
    filtro["id"]: filtro for filtro in FILTROS_AVANZADOS_CAMPOS
}

FILTROS_AVANZADOS_LABELS = {
    filtro["id"]: filtro["label"] for filtro in FILTROS_AVANZADOS_CAMPOS
}

FILTROS_AVANZADOS_LOCALIZACION = {
    "cueanexo": "localizacion__cueanexo",
    "cuof": "localizacion__cuof",
    "cui": "localizacion__cui",
}

FILTROS_AVANZADOS_CARGO_TEXTO = {
    "cargo": "cargo",
    "observacion": "observacion",
    "actualizado_en": "actualizado_busqueda",
}

FILTROS_AVANZADOS_NUMERICOS = {
    "ceic": "ceic",
    "cantidad": "cantidad",
    "puntos_asignados": "puntos_asignados",
    "total": "total",
}

ORDER_LOCALIZACION = {
    "cueanexo": "localizacion__cueanexo",
    "cue": "localizacion__cueanexo",
    "anexo": "localizacion__cueanexo",
    "cuof": "localizacion__cuof",
    "cui": "localizacion__cui",
}

ORDER_CARGO = {
    "ambito": "ofertas_seleccionadas__0__ambito",
    "categoria": "ofertas_seleccionadas__0__categoria",
    "jornada": "ofertas_seleccionadas__0__jornada",
    "oferta": "oferta",
    "acronimo": "ofertas_seleccionadas__0__acronimo",
    "estado_oferta_padron": "ofertas_seleccionadas__0__est_oferta",
    "ceic": "ceic",
    "cargo": "cargo",
    "cantidad": "cantidad",
    "unidad_cantidad": "unidad_cantidad",
    "puntos_asignados": "puntos_asignados",
    "total": "total",
    "estado_pof": "estado_pof",
    "observacion": "observacion",
    "actualizado_en": "actualizado_en",
}


def _limpiar_texto(valor, max_len=180):
    return str(valor or "").strip()[:max_len]


def _resolver_contexto_visualizacion(request):
    cabecera_tipo = _limpiar_texto(request.GET.get("cabecera_tipo", ""), 40).upper()
    proyecto_especial_id = _limpiar_texto(request.GET.get("proyecto_especial_id", ""), 20)
    es_proyecto_especial = cabecera_tipo == CABECERA_PROYECTO_ESPECIAL or bool(proyecto_especial_id)
    proyecto_especial = None
    mensaje = ""

    if es_proyecto_especial:
        cabecera_tipo = CABECERA_PROYECTO_ESPECIAL
        if not proyecto_especial_id.isdigit():
            mensaje = "Debe seleccionar un Proyecto Especial POF valido."
        else:
            try:
                proyecto_especial = ProyectosEspecialesPof.objects.get(pk=proyecto_especial_id)
            except ProyectosEspecialesPof.DoesNotExist:
                mensaje = "No existe el Proyecto Especial POF seleccionado."
            except (ProgrammingError, OperationalError):
                mensaje = "No se pudo consultar el Proyecto Especial POF seleccionado."

    return {
        "es_proyecto_especial": es_proyecto_especial,
        "cabecera_tipo_activa": cabecera_tipo,
        "proyecto_especial_id": proyecto_especial_id,
        "proyecto_especial": proyecto_especial,
        "mensaje": mensaje,
    }


def _params_contexto_visualizacion(contexto_cabecera):
    proyecto = contexto_cabecera.get("proyecto_especial")
    if not contexto_cabecera.get("es_proyecto_especial"):
        return {}
    proyecto_id = proyecto.id if proyecto else contexto_cabecera.get("proyecto_especial_id")
    if proyecto_id:
        return {
            "cabecera_tipo": CABECERA_PROYECTO_ESPECIAL,
            "proyecto_especial_id": proyecto_id,
        }
    return {}


def _query_params_contexto(contexto_cabecera):
    return _query_params_desde_dict(_params_contexto_visualizacion(contexto_cabecera))


def _query_params_desde_dict(params):
    query = QueryDict(mutable=True)
    for clave, valor in (params or {}).items():
        query[clave] = str(valor)
    return query.urlencode()


def _normalizar_query_filtros(request, query=None):
    """
    Serializa la URL con una única representación de los filtros de cargos.

    - Conserva parámetros de contexto, columnas, orden y paginación fuera del
      contrato de filtros.
    - Reemplaza tripletas avanzadas y parámetros `col_*` por los filtros
      válidos que reconoce el servicio, traduciendo legacy a operador 0.
    - Permite que payloads, exportaciones y enlaces compartan exactamente los
      mismos criterios sin ejecutar consultas adicionales.
    """
    filtros_avanzados = _obtener_filtros_avanzados(request)
    resultado = (query if query is not None else request.GET).copy()

    for columna_id in COLUMNAS_BUSCABLES_IDS:
        resultado.pop(f"col_{columna_id}", None)
    for clave in ("campo_filtro", "operador_filtro", "valor_filtro"):
        resultado.pop(clave, None)

    for filtro in filtros_avanzados:
        resultado.appendlist("campo_filtro", filtro["campo"])
        resultado.appendlist("operador_filtro", filtro["operador"])
        resultado.appendlist("valor_filtro", filtro["valor"])

    return resultado


def _obtener_anios_disponibles_visualizacion(request):
    try:
        if usuario_tiene_alcance_restringido_pof(request.user):
            filas_anios = _obtener_queryset_base(request).values_list(
                "localizacion__reunida__anio",
                "localizacion__proyecto_especial__anio",
            ).distinct()
            anios = {
                int(anio)
                for anio_reunida, anio_proyecto in filas_anios
                for anio in (anio_reunida, anio_proyecto)
                if anio is not None
            }
        else:
            anios_reunidas = ReunidaPof.objects.values_list("anio", flat=True)
            anios_proyectos = ProyectosEspecialesPof.objects.values_list("anio", flat=True)
            anios = {
                int(anio)
                for anio in list(anios_reunidas) + list(anios_proyectos)
                if anio is not None
            }
    except (DatabaseError, ProgrammingError, OperationalError):
        logger.exception("No se pudieron obtener anios disponibles de Visualizacion de Cargos.")
        return []

    return sorted(anios, reverse=True)


def _resolver_anio_visualizacion(request, anios_disponibles):
    if not anios_disponibles:
        return None

    anio_param = str(request.GET.get("anio", "") or "").strip()
    if anio_param.isdigit():
        anio_int = int(anio_param)
        if anio_int in anios_disponibles:
            return anio_int

    anio_actual = timezone.localdate().year
    if anio_actual in anios_disponibles:
        return anio_actual

    return anios_disponibles[0]


def _contexto_anio_visualizacion(request, contexto_cabecera):
    if contexto_cabecera.get("es_proyecto_especial"):
        return [], None

    anios_disponibles = _obtener_anios_disponibles_visualizacion(request)
    return anios_disponibles, _resolver_anio_visualizacion(request, anios_disponibles)


def _aplicar_anio_visualizacion_global(queryset, anio_visualizacion):
    if anio_visualizacion is None:
        return queryset.none()

    return queryset.filter(
        Q(localizacion__reunida__anio=anio_visualizacion)
        | Q(localizacion__proyecto_especial__anio=anio_visualizacion)
    )


def _obtener_filtros(request):
    filtros = {}
    for filtro_id in FILTROS_IDS:
        filtros[filtro_id] = _limpiar_texto(request.GET.get(filtro_id, ""))

    estados_validos = {opcion[0] for opcion in CargoPof.EstadoPof.choices}
    if filtros.get("estado_pof") not in estados_validos:
        filtros["estado_pof"] = ""

    return filtros


def _obtener_busqueda_general(request):
    return _limpiar_texto(request.GET.get("q", ""), 240)


def _obtener_busquedas_columna(request):
    busquedas = {}
    for columna_id in COLUMNAS_BUSCABLES_IDS:
        valor = _limpiar_texto(request.GET.get(f"col_{columna_id}", ""), 120)
        if valor:
            busquedas[columna_id] = valor
    return busquedas


def _operadores_validos_para_campo(campo_id):
    definicion = FILTROS_AVANZADOS_POR_ID.get(campo_id)
    if not definicion:
        if campo_id in SNAPSHOT_COLUMNAS:
            return OPERADORES_TEXTO
        return ()
    if definicion["operadores"] == "exact":
        operadores = OPERADORES_EXACTOS
    elif definicion["operadores"] == "numeric":
        operadores = OPERADORES_NUMERICOS
    else:
        operadores = OPERADORES_TEXTO
    if campo_id in COLUMNAS_BUSCABLES_IDS and "0" not in operadores:
        return ("0",) + operadores
    return operadores


def _obtener_filtros_avanzados(request):
    """
    Normaliza filtros avanzados y búsquedas rápidas en una única representación.

    - Conserva las tripletas avanzadas recibidas en la URL.
    - Traduce `col_<campo>` legacy a operador `0` solo cuando no existe un
      filtro avanzado válido del mismo campo.
    - Mantiene múltiples valores intencionales del mismo campo para que el
      agrupamiento OR existente siga resolviéndolos en el ORM.
    """
    filtros = []
    campos = request.GET.getlist("campo_filtro")
    operadores = request.GET.getlist("operador_filtro")
    valores = request.GET.getlist("valor_filtro")

    for indice, campo_id in enumerate(campos):
        campo_id = _limpiar_texto(campo_id, 80)
        valor = _limpiar_texto(valores[indice] if indice < len(valores) else "", 240)
        operador = _limpiar_texto(
            operadores[indice] if indice < len(operadores) else "0",
            4,
        )

        if (
            not campo_id
            or not valor
            or campo_id not in FILTROS_AVANZADOS_POR_ID
            and campo_id not in SNAPSHOT_COLUMNAS
        ):
            continue
        if operador not in _operadores_validos_para_campo(campo_id):
            continue

        filtros.append({
            "indice": indice,
            "campo": campo_id,
            "operador": operador,
            "valor": valor,
            "origen": "avanzado",
        })

    campos_avanzados = {filtro["campo"] for filtro in filtros}
    for columna_id in COLUMNAS_BUSCABLES_IDS:
        if columna_id in campos_avanzados:
            continue
        valor = _limpiar_texto(request.GET.get(f"col_{columna_id}", ""), 240)
        if not valor:
            continue
        if (
            columna_id not in FILTROS_AVANZADOS_POR_ID
            and columna_id not in SNAPSHOT_COLUMNAS
        ):
            continue
        filtros.append({
            "indice": None,
            "campo": columna_id,
            "operador": "0",
            "valor": valor,
            "origen": "legacy_columna",
        })

    return filtros


def _obtener_busqueda_columna_activa(filtros_avanzados):
    """
    Resuelve qué filtro canónico puede reflejarse en la barra superior.

    - Solo muestra campos buscables con exactamente un valor y operador 0.
    - Oculta criterios distintos, múltiples o contradictorios para no
      sugerir una semántica `parecido a` que el backend no está aplicando.
    - Recorre el orden de columnas visible para conservar la selección usual.
    """
    candidatas = []
    for columna_id in COLUMNAS_BUSCABLES_IDS:
        criterios = [
            filtro
            for filtro in filtros_avanzados
            if filtro["campo"] == columna_id
            and filtro.get("valor")
        ]
        if len(criterios) == 1 and criterios[0]["operador"] == "0":
            candidatas.append((columna_id, criterios[0]["valor"]))
    if len(candidatas) == 1:
        return candidatas[0]
    return "cueanexo", ""


def _resolver_columnas_visibles(request, exportar_todo=False):
    if exportar_todo:
        return list(COLUMNAS_TODAS_IDS), "all"

    columnas_estado = _limpiar_texto(request.GET.get("columnas_estado", ""))
    if columnas_estado == "none":
        return [], "none"

    visibles = [
        columna_id
        for columna_id in request.GET.getlist("visible_col")
        if columna_id in COLUMNAS_POR_ID
    ]
    if visibles:
        return visibles, "custom"

    return list(COLUMNAS_DEFAULT_IDS), "default"


def _query_sin_parametros(request, parametros):
    query = request.GET.copy()
    for parametro in parametros:
        query.pop(parametro, None)
    query.pop("page", None)
    query.pop("page_size", None)
    return query.urlencode()


def _query_limpia(columnas_visibles, base_params=None):
    query = QueryDict(mutable=True)
    for clave, valor in (base_params or {}).items():
        query[clave] = str(valor)
    if columnas_visibles:
        for columna_id in columnas_visibles:
            query.appendlist("visible_col", columna_id)
    else:
        query["columnas_estado"] = "none"
    return query.urlencode()


def _query_exportar_filtros(request, base_params=None):
    query = _normalizar_query_filtros(request)
    query.pop("page", None)
    query.pop("page_size", None)
    for clave, valor in (base_params or {}).items():
        query[clave] = str(valor)
    return query.urlencode()


def _estado_pof_options():
    return [
        {"value": "", "label": "Todos"},
        {"value": CargoPof.EstadoPof.AFECTADO, "label": "Afectado"},
        {"value": CargoPof.EstadoPof.DESAFECTADO, "label": "Desafectado / Baja"},
    ]


def _choices_options(choices):
    return [
        {"value": str(codigo), "label": str(etiqueta)}
        for codigo, etiqueta in choices
        if str(codigo)
    ]


def _opciones_oferta_filtro_visualizacion():
    return [dict(opcion) for opcion in OFERTAS_FILTRO_VISUALIZACION]


def _normalizar_opcion(valor):
    texto = _limpiar_texto(valor, 240)
    return texto if texto and texto != GUION else ""


def _opciones_distintas(queryset, campo):
    opciones = []
    vistos = set()
    for valor in queryset.order_by(campo).values_list(campo, flat=True).distinct():
        texto = _normalizar_opcion(valor)
        if not texto or texto in vistos:
            continue
        vistos.add(texto)
        opciones.append(texto)
    return opciones


def _opciones_localizacion(campo):
    return _opciones_distintas(CargoPof.objects.exclude(**{campo: ""}), campo)


def _opciones_cue_partes(parte):
    cueanexos = (
        CargoPof.objects.exclude(localizacion__cueanexo="")
        .order_by("localizacion__cueanexo")
        .values_list("localizacion__cueanexo", flat=True)
        .distinct()
    )
    opciones = []
    vistos = set()
    for cueanexo in cueanexos:
        texto = _limpiar_texto(cueanexo, 20)
        valor = texto[:7] if parte == "cue" else texto[7:]
        valor = _normalizar_opcion(valor)
        if not valor or valor in vistos:
            continue
        vistos.add(valor)
        opciones.append(valor)
    return opciones


def _opciones_snapshot_global(campo):
    localizaciones_con_cargos = CargoPof.objects.values_list("localizacion_id", flat=True)
    queryset = SnapshotPadronLocalizacionPof.objects.filter(
        vigente=True,
        localizacion_id__in=localizaciones_con_cargos,
    ).exclude(**{campo: ""})
    return _opciones_distintas(queryset, campo)


def _opciones_cargo(campo):
    return _opciones_distintas(CargoPof.objects.all(), campo)


def _opciones_cargo_decimal(campo):
    opciones = []
    vistos = set()
    for valor in (
        CargoPof.objects.order_by(campo)
        .values_list(campo, flat=True)
        .distinct()
    ):
        texto = _decimal_texto(valor)
        if not texto or texto in vistos:
            continue
        vistos.add(texto)
        opciones.append(texto)
    return opciones


def _opciones_desde_queryset(queryset, campo):
    return _opciones_distintas(queryset.exclude(**{campo: ""}), campo)


def _opciones_ofertas_cargo_desde_queryset(queryset, campo_id):
    aliases = OFERTA_CARGO_COLUMNAS[campo_id]
    opciones = set()

    for ofertas, oferta_cargo in queryset.values_list(
        "ofertas_seleccionadas",
        "oferta",
    ):
        for oferta in ofertas or []:
            if not isinstance(oferta, dict):
                continue
            for alias in aliases:
                texto = _normalizar_opcion(oferta.get(alias))
                if texto:
                    opciones.add(texto)
                    break

        if campo_id == "oferta" and not ofertas:
            for valor in str(oferta_cargo or "").split(","):
                texto = _normalizar_opcion(valor)
                if texto:
                    opciones.add(texto)

    return sorted(opciones, key=str.casefold)


def _opciones_choices_desde_queryset(queryset, campo, choices):
    etiquetas = {str(codigo): str(etiqueta) for codigo, etiqueta in choices}
    return [
        {"value": valor, "label": etiquetas.get(valor, valor)}
        for valor in _opciones_desde_queryset(queryset, campo)
    ]


def _opciones_filtros_visualizacion_restringida(request):
    opciones = {campo["id"]: [] for campo in FILTROS_AVANZADOS_CAMPOS}
    queryset = _queryset_visualizacion(request, ignorar_filtros=True)
    queryset_snapshot = queryset.filter(
        localizacion__snapshots_padron__vigente=True
    )

    for campo in FILTROS_SELECT_SNAPSHOT:
        campo_snapshot = f"localizacion__snapshots_padron__{SNAPSHOT_COLUMNAS[campo]}"
        valores = _opciones_desde_queryset(
            queryset_snapshot,
            campo_snapshot,
        )
        if campo == "region":
            regiones_presentes = {
                normalizar_region_padron(valor)
                for valor in valores
                if normalizar_region_padron(valor)
            }
            opciones[campo] = [
                region
                for region in REGIONES_EDUCATIVAS_POF
                if region in regiones_presentes
            ]
        else:
            opciones[campo] = valores

    for campo in OFERTA_CARGO_COLUMNAS:
        valores = set(_opciones_ofertas_cargo_desde_queryset(queryset, campo))
        campo_snapshot = f"localizacion__snapshots_padron__{campo}"
        valores.update(_opciones_desde_queryset(queryset_snapshot, campo_snapshot))
        opciones[campo] = sorted(valores, key=str.casefold)

    opciones["unidad_cantidad"] = _opciones_choices_desde_queryset(
        queryset,
        "unidad_cantidad",
        CargoPof.UnidadCantidad.choices,
    )
    opciones["estado_pof"] = _opciones_choices_desde_queryset(
        queryset,
        "estado_pof",
        CargoPof.EstadoPof.choices,
    )
    return opciones


def construir_opciones_filtros_visualizacion_cargos_localizacion(request=None):
    if request is not None:
        rol = obtener_rol_usuario_pof(request.user)
        if rol in {ROL_POF_REGIONAL, ROL_POF_DIRECTOR}:
            try:
                opciones = _opciones_filtros_visualizacion_restringida(request)
                opciones["oferta"] = _opciones_oferta_filtro_visualizacion()
                return opciones
            except (DatabaseError, ProgrammingError, OperationalError):
                logger.exception(
                    "No se pudieron obtener opciones de filtros para el alcance POF."
                )
                return {campo["id"]: [] for campo in FILTROS_AVANZADOS_CAMPOS}
        if (
            rol not in ROLES_POF_ACCESO_COMPLETO
            and rol not in ROLES_POF_SOLO_VISUALIZACION_COMPLETA
        ):
            return {campo["id"]: [] for campo in FILTROS_AVANZADOS_CAMPOS}

    opciones = {campo["id"]: [] for campo in FILTROS_AVANZADOS_CAMPOS}
    opciones.update({
        "unidad_cantidad": _choices_options(CargoPof.UnidadCantidad.choices),
        "estado_pof": _choices_options(CargoPof.EstadoPof.choices),
    })

    try:
        opciones.update(obtener_opciones_filtros_visualizacion_padron())
    except (DatabaseError, ProgrammingError, OperationalError):
        logger.exception("No se pudieron obtener opciones de filtros desde Padron Interno.")

    opciones["oferta"] = _opciones_oferta_filtro_visualizacion()
    return opciones


def _opciones_snapshot(campo, seleccionado=""):
    try:
        valores = list(
            SnapshotPadronLocalizacionPof.objects.filter(vigente=True)
            .exclude(**{campo: ""})
            .order_by(campo)
            .values_list(campo, flat=True)
            .distinct()[:500]
        )
    except (ProgrammingError, OperationalError):
        valores = []

    if campo == "region":
        regiones_presentes = {
            normalizar_region_padron(valor)
            for valor in valores
            if normalizar_region_padron(valor)
        }
        opciones = [{"value": "", "label": "Todos"}]
        opciones.extend(
            {"value": region, "label": region}
            for region in REGIONES_EDUCATIVAS_POF
            if region in regiones_presentes
        )
        if seleccionado and seleccionado not in regiones_presentes:
            region_seleccionada = normalizar_region_padron(seleccionado)
            if region_seleccionada:
                opciones.insert(1, {
                    "value": region_seleccionada,
                    "label": region_seleccionada,
                })
        return opciones

    opciones = [{"value": "", "label": "Todos"}]
    vistos = set()
    for valor in valores:
        texto = _limpiar_texto(valor)
        if not texto or texto in vistos:
            continue
        vistos.add(texto)
        opciones.append({"value": texto, "label": texto})

    if seleccionado and seleccionado not in vistos:
        opciones.insert(1, {"value": seleccionado, "label": seleccionado})

    return opciones


def _construir_filtros_config(filtros):
    filtros_config = []
    for definicion in FILTROS_TEXTO:
        filtros_config.append({
            **definicion,
            "tipo": "text",
            "valor": filtros.get(definicion["id"], ""),
        })

    for definicion in FILTROS_SELECT_DEFINICIONES:
        filtro_id = definicion["id"]
        if filtro_id == "estado_pof":
            opciones = _estado_pof_options()
        else:
            opciones = _opciones_snapshot(filtro_id, filtros.get(filtro_id, ""))
        filtros_config.append({
            **definicion,
            "tipo": "select",
            "valor": filtros.get(filtro_id, ""),
            "opciones": opciones,
        })

    return filtros_config


def _snapshots_vigentes_queryset():
    return SnapshotPadronLocalizacionPof.objects.filter(
        vigente=True
    ).order_by("-fecha_snapshot")


def _aplicar_alcance_visualizacion(queryset, user):
    rol = obtener_rol_usuario_pof(user)

    if (
        rol in ROLES_POF_ACCESO_COMPLETO
        or rol in ROLES_POF_SOLO_VISUALIZACION_COMPLETA
    ):
        return queryset

    if rol == ROL_POF_REGIONAL:
        regiones = obtener_regiones_usuario_pof(user)
        if not regiones:
            return queryset.none()
        return queryset.filter(
            localizacion__snapshots_padron__vigente=True,
            localizacion__snapshots_padron__region__in=regiones,
        )

    if rol == ROL_POF_DIRECTOR:
        cueanexos = obtener_cueanexos_director_pof(user)
        if not cueanexos:
            return queryset.none()
        return queryset.filter(localizacion__cueanexo__in=cueanexos)

    return queryset.none()


def _obtener_queryset_base(request):
    queryset = (
        CargoPof.objects.select_related(
            "localizacion",
            "localizacion__reunida",
            "localizacion__proyecto_especial",
            "lote_carga",
        )
        .prefetch_related(
            Prefetch(
                "localizacion__snapshots_padron",
                queryset=_snapshots_vigentes_queryset(),
                to_attr="snapshots_vigentes",
            )
        )
    )
    return _aplicar_alcance_visualizacion(queryset, request.user)


def _agregar_anotaciones_busqueda(queryset):
    return queryset.annotate(
        ceic_busqueda=Cast("ceic", CharField()),
        cantidad_busqueda=Cast("cantidad", CharField()),
        puntos_busqueda=Cast("puntos_asignados", CharField()),
        total_busqueda=Cast("total", CharField()),
        actualizado_busqueda=Cast("actualizado_en", CharField()),
    )


def _snapshot_filter_q(campo, valor, exacto=False):
    if campo == "region" and exacto:
        variantes = obtener_variantes_region_padron(valor)
        if variantes:
            regiones_q = Q()
            for variante in variantes:
                regiones_q |= Q(
                    localizacion__snapshots_padron__region__iexact=variante
                )
            return Q(localizacion__snapshots_padron__vigente=True) & regiones_q

    lookup = "iexact" if exacto else "icontains"
    return Q(
        localizacion__snapshots_padron__vigente=True,
        **{f"localizacion__snapshots_padron__{campo}__{lookup}": valor},
    )


def _texto_lookup_q(campo, operador, valor):
    lookup = "icontains" if operador in {"0", "1"} else "iexact"
    return Q(**{f"{campo}__{lookup}": valor})


def _snapshot_lookup_q(campo, operador, valor):
    if campo == "region" and operador in {"2", "7"}:
        variantes = obtener_variantes_region_padron(valor)
        if variantes:
            regiones_q = Q()
            for variante in variantes:
                regiones_q |= Q(
                    localizacion__snapshots_padron__region__iexact=variante
                )
            return Q(localizacion__snapshots_padron__vigente=True) & regiones_q

    lookup = "icontains" if operador in {"0", "1"} else "iexact"
    return Q(
        localizacion__snapshots_padron__vigente=True,
        **{f"localizacion__snapshots_padron__{campo}__{lookup}": valor},
    )


def _oferta_cargo_lookup_q(campo_id, operador, valor):
    exacto = operador in {"2", "7"}
    sin_ofertas = Q(ofertas_seleccionadas=[]) | Q(ofertas_seleccionadas__isnull=True)

    if campo_id == "oferta":
        consulta = Q(oferta__icontains=valor)
        return consulta | (
            Q(oferta="")
            & sin_ofertas
            & _snapshot_filter_q("oferta", valor, exacto=exacto)
        )

    if not exacto:
        consulta = Q(ofertas_seleccionadas__icontains=valor)
    else:
        consulta = Q()
        for alias in OFERTA_CARGO_COLUMNAS[campo_id]:
            consulta |= Q(ofertas_seleccionadas__contains=[{alias: valor}])

    return consulta | (
        sin_ofertas
        & _snapshot_filter_q(campo_id, valor, exacto=exacto)
    )


def _cue_lookup_q(operador, valor):
    if operador in {"2", "7"}:
        return Q(localizacion__cueanexo__startswith=valor)
    return Q(localizacion__cueanexo__icontains=valor)


def _anexo_lookup_q(operador, valor):
    if operador in {"2", "7"}:
        return Q(localizacion__cueanexo__endswith=valor)
    return Q(localizacion__cueanexo__icontains=valor)


def _decimal_valor(valor):
    try:
        return Decimal(str(valor).replace(",", "."))
    except Exception:
        return None


def _numero_lookup_q(campo, operador, valor):
    numero = _decimal_valor(valor)
    if numero is None:
        return None

    if campo == "ceic":
        if numero != numero.to_integral_value():
            return None
        numero = int(numero)

    lookup_por_operador = {
        "2": "exact",
        "7": "exact",
        "3": "gt",
        "4": "gte",
        "5": "lt",
        "6": "lte",
    }
    lookup = lookup_por_operador.get(operador)
    if not lookup:
        return None
    return Q(**{f"{campo}__{lookup}": numero})


def _filtro_avanzado_q(campo_id, operador, valor):
    if operador == "0" and campo_id in COLUMNAS_BUSCABLES_IDS:
        return _busqueda_columna_q(campo_id, valor)
    if campo_id == "cue":
        return _cue_lookup_q(operador, valor)
    if campo_id == "anexo":
        return _anexo_lookup_q(operador, valor)
    if campo_id in FILTROS_AVANZADOS_LOCALIZACION:
        return _texto_lookup_q(FILTROS_AVANZADOS_LOCALIZACION[campo_id], operador, valor)
    if campo_id in OFERTA_CARGO_COLUMNAS:
        return _oferta_cargo_lookup_q(campo_id, operador, valor)
    if campo_id in SNAPSHOT_COLUMNAS:
        return _snapshot_lookup_q(SNAPSHOT_COLUMNAS[campo_id], operador, valor)
    if campo_id in FILTROS_AVANZADOS_CARGO_TEXTO:
        return _texto_lookup_q(FILTROS_AVANZADOS_CARGO_TEXTO[campo_id], operador, valor)
    if campo_id in FILTROS_AVANZADOS_NUMERICOS:
        return _numero_lookup_q(FILTROS_AVANZADOS_NUMERICOS[campo_id], operador, valor)
    if campo_id == "estado_pof":
        return Q(estado_pof=valor)
    if campo_id == "unidad_cantidad":
        return Q(unidad_cantidad=valor)
    return None


def _aplicar_filtro_q(queryset, filtro_q, operador):
    if filtro_q is None:
        return queryset
    if operador in {"1", "7"}:
        return queryset.exclude(filtro_q)
    return queryset.filter(filtro_q)


def _aplicar_filtros_avanzados(queryset, filtros_avanzados):
    grupos = {}
    filtros_sueltos = []

    for filtro in filtros_avanzados:
        if filtro["operador"] in {"0", "2"}:
            grupos.setdefault((filtro["campo"], filtro["operador"]), []).append(filtro["valor"])
        else:
            filtros_sueltos.append(filtro)

    for (campo_id, operador), valores in grupos.items():
        filtro_grupo = Q()
        for valor in valores:
            filtro_q = _filtro_avanzado_q(campo_id, operador, valor)
            if filtro_q is not None:
                filtro_grupo |= filtro_q
        if filtro_grupo:
            queryset = queryset.filter(filtro_grupo)

    for filtro in filtros_sueltos:
        filtro_q = _filtro_avanzado_q(
            filtro["campo"],
            filtro["operador"],
            filtro["valor"],
        )
        queryset = _aplicar_filtro_q(queryset, filtro_q, filtro["operador"])

    return queryset


def _aplicar_filtros(queryset, filtros):
    if filtros.get("cueanexo"):
        queryset = queryset.filter(localizacion__cueanexo__icontains=filtros["cueanexo"])
    if filtros.get("cue"):
        queryset = queryset.filter(localizacion__cueanexo__startswith=filtros["cue"])
    if filtros.get("anexo"):
        queryset = queryset.filter(localizacion__cueanexo__endswith=filtros["anexo"])
    if filtros.get("cuof"):
        queryset = queryset.filter(localizacion__cuof__icontains=filtros["cuof"])
    if filtros.get("cui"):
        queryset = queryset.filter(localizacion__cui__icontains=filtros["cui"])
    if filtros.get("ceic"):
        if filtros["ceic"].isdigit():
            queryset = queryset.filter(ceic=int(filtros["ceic"]))
        else:
            queryset = queryset.filter(ceic_busqueda__icontains=filtros["ceic"])
    if filtros.get("cargo"):
        queryset = queryset.filter(cargo__icontains=filtros["cargo"])
    if filtros.get("estado_pof"):
        queryset = queryset.filter(estado_pof=filtros["estado_pof"])
    if filtros.get("nombre_establecimiento"):
        queryset = queryset.filter(
            _snapshot_filter_q("nombre_establecimiento", filtros["nombre_establecimiento"])
        )
    if filtros.get("numero_establecimiento"):
        queryset = queryset.filter(
            _snapshot_filter_q("numero_establecimiento", filtros["numero_establecimiento"])
        )

    for campo in FILTROS_SELECT_SNAPSHOT:
        if filtros.get(campo):
            queryset = queryset.filter(_snapshot_filter_q(campo, filtros[campo], exacto=True))

    for campo in OFERTA_CARGO_COLUMNAS:
        if filtros.get(campo):
            queryset = queryset.filter(
                _oferta_cargo_lookup_q(campo, "2", filtros[campo])
            )

    return queryset


def _estado_o_unidad_q(campo, valor, choices, extras=None):
    valor_normalizado = valor.lower()
    consulta = Q(**{f"{campo}__icontains": valor.upper()})
    for codigo, etiqueta in choices:
        etiqueta_normalizada = str(etiqueta).lower()
        extra = (extras or {}).get(codigo, "")
        if valor_normalizado in etiqueta_normalizada or (
            extra and valor_normalizado in extra.lower()
        ):
            consulta |= Q(**{campo: codigo})
    return consulta


def _busqueda_token_q(token):
    snapshot_q = Q()
    for campo in (
        "nombre_establecimiento",
        "numero_establecimiento",
        "region",
        "localidad",
        "departamento",
        "ubicacion",
    ):
        snapshot_q |= Q(**{f"localizacion__snapshots_padron__{campo}__icontains": token})

    consulta = (
        Q(localizacion__cueanexo__icontains=token)
        | Q(localizacion__cuof__icontains=token)
        | Q(localizacion__cui__icontains=token)
        | Q(ceic_busqueda__icontains=token)
        | Q(cargo__icontains=token)
        | Q(oferta__icontains=token)
        | Q(ofertas_seleccionadas__icontains=token)
        | Q(observacion__icontains=token)
        | _estado_o_unidad_q(
            "estado_pof",
            token,
            CargoPof.EstadoPof.choices,
            extras={CargoPof.EstadoPof.DESAFECTADO: "Baja"},
        )
        | _estado_o_unidad_q("unidad_cantidad", token, CargoPof.UnidadCantidad.choices)
        | (Q(localizacion__snapshots_padron__vigente=True) & snapshot_q)
    )
    return consulta


def _aplicar_busqueda_general(queryset, busqueda):
    if not busqueda:
        return queryset

    for token in busqueda.split():
        queryset = queryset.filter(_busqueda_token_q(token))

    return queryset


def _observacion_parecido_q(valor):
    """
    Construye la consulta por términos para Observación con operador 0.

    - Ignora espacios iniciales, finales y repetidos mediante `split()`.
    - Limita la consulta a cinco términos útiles.
    - Combina cada término con `AND` usando `observacion__icontains` en PostgreSQL.
    """
    consulta = Q()
    for termino in str(valor or "").split()[:MAX_TERMINOS_BUSQUEDA_OBSERVACION]:
        consulta &= Q(observacion__icontains=termino)
    return consulta


def _busqueda_columna_q(columna_id, valor):
    """
    Construye la consulta ORM de la búsqueda rápida de una columna.

    - Centraliza la semántica `parecido a` que también usa el filtro avanzado
      canónico con operador 0.
    - Conserva búsquedas parciales, sufijos y estados según cada columna.
    - Busca Observación por hasta cinco términos independientes solo con operador 0.
    - Devuelve solo expresiones Q; no consulta ni materializa registros.
    """
    if columna_id == "cueanexo":
        return Q(localizacion__cueanexo__icontains=valor)
    if columna_id == "cue":
        return Q(localizacion__cueanexo__icontains=valor)
    if columna_id == "anexo":
        return Q(localizacion__cueanexo__endswith=valor)
    if columna_id == "cuof":
        return Q(localizacion__cuof__icontains=valor)
    if columna_id == "cui":
        return Q(localizacion__cui__icontains=valor)
    if columna_id in OFERTA_CARGO_COLUMNAS:
        return _oferta_cargo_lookup_q(columna_id, "0", valor)
    if columna_id in SNAPSHOT_COLUMNAS:
        return _snapshot_filter_q(SNAPSHOT_COLUMNAS[columna_id], valor)
    if columna_id == "ceic":
        return Q(ceic_busqueda__icontains=valor)
    if columna_id == "cargo":
        return Q(cargo__icontains=valor)
    if columna_id == "cantidad":
        return Q(cantidad_busqueda__icontains=valor)
    if columna_id == "unidad_cantidad":
        return _estado_o_unidad_q("unidad_cantidad", valor, CargoPof.UnidadCantidad.choices)
    if columna_id == "puntos_asignados":
        return Q(puntos_busqueda__icontains=valor)
    if columna_id == "total":
        return Q(total_busqueda__icontains=valor)
    if columna_id == "estado_pof":
        return _estado_o_unidad_q(
            "estado_pof",
            valor,
            CargoPof.EstadoPof.choices,
            extras={CargoPof.EstadoPof.DESAFECTADO: "Baja"},
        )
    if columna_id == "observacion":
        return _observacion_parecido_q(valor)
    if columna_id == "actualizado_en":
        return Q(actualizado_busqueda__icontains=valor)

    return None


def _aplicar_busqueda_columna(queryset, columna_id, valor):
    """
    Aplica al queryset la expresión de búsqueda rápida de una columna.

    - Reutiliza la misma expresión Q que el filtro avanzado operador 0.
    - Mantiene la búsqueda como una operación ORM sin datasets intermedios.
    - Deja intacto el queryset si la columna no es reconocida.
    """
    filtro_q = _busqueda_columna_q(columna_id, valor)
    return queryset.filter(filtro_q) if filtro_q is not None else queryset


def _aplicar_busquedas_columna(queryset, busquedas_columna):
    for columna_id, valor in busquedas_columna.items():
        queryset = _aplicar_busqueda_columna(queryset, columna_id, valor)
    return queryset


def _snapshot_subquery(campo):
    return (
        SnapshotPadronLocalizacionPof.objects.filter(
            localizacion_id=OuterRef("localizacion_id"),
            vigente=True,
        )
        .order_by("-fecha_snapshot")
        .values(campo)[:1]
    )


def _aplicar_orden(queryset, request):
    orden = _limpiar_texto(request.GET.get("orden", ""))
    direccion = _limpiar_texto(request.GET.get("dir", "asc")).lower()
    prefijo = "-" if direccion == "desc" else ""

    if orden in SNAPSHOT_COLUMNAS:
        alias = f"orden_{orden}"
        queryset = queryset.annotate(**{alias: Subquery(_snapshot_subquery(SNAPSHOT_COLUMNAS[orden]))})
        campo_orden = alias
    elif orden in ORDER_LOCALIZACION:
        campo_orden = ORDER_LOCALIZACION[orden]
    elif orden in ORDER_CARGO:
        campo_orden = ORDER_CARGO[orden]
    else:
        return queryset.order_by("localizacion__cueanexo", "localizacion__cuof", "ceic", "id")

    orden_final = [f"{prefijo}{campo_orden}"]
    for campo_secundario in ("localizacion__cueanexo", "localizacion__cuof", "ceic", "id"):
        if campo_secundario != campo_orden:
            orden_final.append(campo_secundario)
    return queryset.order_by(*orden_final)


def _obtener_snapshot_vigente(localizacion):
    snapshots = getattr(localizacion, "snapshots_vigentes", [])
    return snapshots[0] if snapshots else None


def _valor_o_guion(valor):
    return valor if valor not in (None, "") else GUION


def _decimal_texto(valor):
    numero = Decimal(valor or 0)
    if numero == numero.to_integral_value():
        return str(int(numero))
    return str(numero.quantize(Decimal("0.01")))


def _fecha_texto(valor):
    if not valor:
        return GUION
    return timezone.localtime(valor).strftime("%d/%m/%Y %H:%M")


def _obtener_clave_agrupacion_visualizacion(localizacion):
    """Agrupa por CUEANEXO o, cuando está vacío, por CUOF."""
    cueanexo = _limpiar_texto(localizacion.cueanexo)
    if cueanexo:
        return f"CUEANEXO:{cueanexo}"

    cuof = _limpiar_texto(localizacion.cuof)
    if cuof:
        return f"CUOF:{cuof}"

    return f"LOCALIZACION:{localizacion.pk}"


def _anotar_clave_total_general(queryset):
    """
    Anota la clave de agrupacion usada por Total General sin cargar cargos en Python.

    - Prioriza CUEANEXO y usa CUOF cuando el CUEANEXO esta vacio.
    - Conserva la localizacion como ultimo respaldo para casos sin ambos valores.
    - Replica la limpieza de espacios aplicada por la agrupacion historica.
    """
    queryset = queryset.annotate(
        _cueanexo_total_general=Trim("localizacion__cueanexo"),
        _cuof_total_general=Trim("localizacion__cuof"),
    )
    return queryset.annotate(
        _clave_total_general=Case(
            When(
                Q(_cueanexo_total_general__isnull=False)
                & ~Q(_cueanexo_total_general=""),
                then=Concat(
                    Value("CUEANEXO:"),
                    "_cueanexo_total_general",
                ),
            ),
            When(
                Q(_cuof_total_general__isnull=False)
                & ~Q(_cuof_total_general=""),
                then=Concat(
                    Value("CUOF:"),
                    "_cuof_total_general",
                ),
            ),
            default=Concat(
                Value("LOCALIZACION:"),
                Cast("localizacion_id", CharField()),
            ),
            output_field=CharField(),
        )
    )


def _armar_contexto_totales_desde_agregados(cargos, totales_agrupados):
    """
    Adapta los totales agregados al contrato usado por la serializacion de cargos.

    - Conserva una clave por cada cargo materializado por el consumidor.
    - Convierte agregados nulos, sin cargos AFECTADO, en cero.
    - Mantiene las mismas claves de contexto que el calculo historico en Python.
    """
    claves_por_cargo = {
        cargo.id: _obtener_clave_agrupacion_visualizacion(cargo.localizacion)
        for cargo in cargos
    }
    totales_por_clave = {
        clave: total or Decimal("0")
        for clave, total in totales_agrupados
    }
    for clave in claves_por_cargo.values():
        totales_por_clave.setdefault(clave, Decimal("0"))

    return {
        "claves_por_cargo": claves_por_cargo,
        "totales_por_clave": totales_por_clave,
    }


def _construir_contexto_totales_generales_orm(queryset, cargos):
    """
    Calcula Total General en SQL desde el queryset autorizado recibido.

    - Deduplica primero los cargos autorizados para evitar sumas por joins repetidos.
    - Agrupa por CUEANEXO o por CUOF/localizacion con la semantica vigente.
    - Suma exclusivamente cargo.total de registros con estado AFECTADO.
    """
    ids_autorizados = queryset.order_by().values("pk").distinct()
    queryset_agregacion = CargoPof.objects.filter(
        pk__in=Subquery(ids_autorizados)
    )
    totales_agrupados = (
        _anotar_clave_total_general(queryset_agregacion)
        .order_by()
        .values("_clave_total_general")
        .annotate(
            _total_general=Sum(
                "total",
                filter=Q(estado_pof=CargoPof.EstadoPof.AFECTADO),
            )
        )
        .values_list("_clave_total_general", "_total_general")
    )
    return _armar_contexto_totales_desde_agregados(cargos, totales_agrupados)


def _serializar_cargo(cargo, contexto_totales_generales=None):
    localizacion = cargo.localizacion
    snapshot = _obtener_snapshot_vigente(localizacion)
    contexto_totales_generales = contexto_totales_generales or {}
    clave_total_general = contexto_totales_generales.get("claves_por_cargo", {}).get(
        cargo.id
    )
    totales_por_clave = contexto_totales_generales.get("totales_por_clave", {})
    total_general = (
        _decimal_texto(totales_por_clave[clave_total_general])
        if clave_total_general in totales_por_clave
        else GUION
    )

    fila = {
        "id": cargo.id,
        "_clave_localizacion_grupo": _obtener_clave_agrupacion_visualizacion(localizacion),
        "_clave_total_general": clave_total_general,
        "_cueanexo_visual": _limpiar_texto(localizacion.cueanexo),
        "cueanexo": _valor_o_guion(localizacion.cueanexo),
        "cue": _valor_o_guion(localizacion.cue_base),
        "anexo": _valor_o_guion(localizacion.anexo_localizacion),
        "cuof": _valor_o_guion(localizacion.cuof),
        "cui": _valor_o_guion(localizacion.cui),
        "ceic": str(cargo.ceic or ""),
        "cargo": _valor_o_guion(cargo.cargo),
        "cantidad": _decimal_texto(cargo.cantidad),
        "unidad_cantidad": cargo.get_unidad_cantidad_display(),
        "puntos_asignados": _decimal_texto(cargo.puntos_asignados),
        "total": _decimal_texto(cargo.total),
        "total_general": total_general,
        "estado_pof": cargo.get_estado_pof_display(),
        "estado_pof_codigo": cargo.estado_pof,
        "observacion": _valor_o_guion(cargo.observacion),
        "actualizado_en": _fecha_texto(cargo.actualizado_en),
    }

    for columna_id, campo in SNAPSHOT_COLUMNAS.items():
        valor = getattr(snapshot, campo, "") if snapshot else ""
        if columna_id == "region":
            valor = normalizar_region_padron(valor)
        fila[columna_id] = _valor_o_guion(valor)

    ofertas = cargo.ofertas_seleccionadas or []
    for columna_id, aliases in OFERTA_CARGO_COLUMNAS.items():
        valores = []
        vistos = set()
        for oferta in ofertas:
            if not isinstance(oferta, dict):
                continue
            valor = ""
            for alias in aliases:
                valor = _limpiar_texto(oferta.get(alias), 240)
                if valor:
                    break
            if valor and valor not in vistos:
                vistos.add(valor)
                valores.append(valor)

        valor = ", ".join(valores)
        if columna_id == "oferta":
            valor = _limpiar_texto(cargo.oferta, 1000) or valor
        if not valor and snapshot:
            valor = getattr(snapshot, columna_id, "")
        fila[columna_id] = _valor_o_guion(valor)

    return fila


def _aplicar_no_repeticion_visual(fila_raw, clave_localizacion_anterior):
    clave_localizacion_actual = fila_raw.get("_clave_localizacion_grupo", "")
    fila_display = fila_raw.copy()

    if clave_localizacion_actual == clave_localizacion_anterior:
        for columna_id in COLUMNAS_NO_REPETIR_POR_LOCALIZACION:
            fila_display[columna_id] = ""

    return fila_display, clave_localizacion_actual or None


def _obtener_claves_jerarquicas_visualizacion(fila_raw):
    """
    Obtiene las claves de CUE y anexo usadas para separar grupos visuales.

    - Usa los primeros siete caracteres de CUEANEXO como CUE base.
    - Usa el resto de CUEANEXO como anexo dentro del CUE.
    - Conserva la misma semantica de agrupacion usada por la tabla.
    """
    cueanexo_actual = fila_raw.get("_cueanexo_visual", "")
    cue_actual = cueanexo_actual[:7] if len(cueanexo_actual) >= 7 else ""
    anexo_actual = cueanexo_actual[7:] if len(cueanexo_actual) > 7 else ""
    return cueanexo_actual, cue_actual, anexo_actual


def _aplicar_no_repeticion_total_general(fila_display, claves_totales_vistas):
    clave_total_general = fila_display.get("_clave_total_general")
    if clave_total_general in claves_totales_vistas:
        fila_display["total_general"] = ""
    elif clave_total_general:
        claves_totales_vistas.add(clave_total_general)


def _armar_filas_tabla(cargos, columnas, contexto_totales_generales):
    """
    Construye las filas visibles del visualizador conservando sus agrupaciones.

    - Marca los cambios de CUE y CUEANEXO con la misma jerarquia del exportador.
    - Mantiene la no repeticion visual por localizacion.
    - Aplica Total General una sola vez por clave de agrupacion.
    """
    filas = []
    clave_localizacion_anterior = None
    claves_totales_vistas = set()
    cue_anterior = None
    anexo_anterior = None

    for cargo in cargos:
        fila_raw = _serializar_cargo(cargo, contexto_totales_generales)
        cueanexo_actual, cue_actual, anexo_actual = (
            _obtener_claves_jerarquicas_visualizacion(fila_raw)
        )
        clave_localizacion_actual = fila_raw.get("_clave_localizacion_grupo", "")
        es_inicio_cue = bool(
            filas and cue_actual and cue_actual != cue_anterior
        )
        es_inicio_anexo = bool(
            filas and not es_inicio_cue and cue_actual == cue_anterior
            and anexo_actual and anexo_actual != anexo_anterior
        )
        es_inicio_localizacion = bool(
            filas and not cueanexo_actual
            and clave_localizacion_actual != clave_localizacion_anterior
        )
        fila_display, clave_localizacion_anterior = _aplicar_no_repeticion_visual(
            fila_raw,
            clave_localizacion_anterior,
        )
        _aplicar_no_repeticion_total_general(fila_display, claves_totales_vistas)
        filas.append({
            "id": fila_raw["id"],
            "estado_pof_codigo": str(fila_raw.get("estado_pof_codigo", "")).lower(),
            "es_inicio_cue": es_inicio_cue,
            "es_inicio_anexo": es_inicio_anexo,
            "es_inicio_localizacion": es_inicio_localizacion,
            "celdas": [
                {
                    "id": columna["id"],
                    "valor": fila_display.get(columna["id"], GUION),
                    "visible": columna.get("visible", False),
                    "centrada": columna["id"] in COLUMNAS_CENTRADAS,
                    "badge": columna["id"] == "estado_pof",
                }
                for columna in columnas
            ],
        })
        cue_anterior = cue_actual or None
        anexo_anterior = anexo_actual or None
    return filas


def _armar_columnas(request, columnas_visibles_ids):
    orden_actual = _limpiar_texto(request.GET.get("orden", ""))
    dir_actual = _limpiar_texto(request.GET.get("dir", "asc")).lower()
    columnas = []

    for columna in VISUALIZACION_CARGOS_COLUMNAS:
        columna_id = columna["id"]
        ordenable = columna.get("ordenable", True)
        buscable = columna.get("buscable", True)
        querystring_orden = ""
        if ordenable:
            query = request.GET.copy()
            query.pop("page", None)
            query.pop("page_size", None)
            query["orden"] = columna_id
            query["dir"] = "desc" if orden_actual == columna_id and dir_actual != "desc" else "asc"
            querystring_orden = query.urlencode()

        columnas.append({
            **columna,
            "visible": columna_id in columnas_visibles_ids,
            "ordenable": ordenable,
            "buscable": buscable,
            "order_querystring": querystring_orden,
            "order_active": ordenable and orden_actual == columna_id,
            "order_dir": dir_actual if ordenable and orden_actual == columna_id else "",
            "busqueda": (
                _limpiar_texto(request.GET.get(f"col_{columna_id}", ""))
                if buscable else ""
            ),
        })

    return columnas


def _valor_filtro_label(campo_id, valor):
    choices = ()
    if campo_id == "estado_pof":
        choices = CargoPof.EstadoPof.choices
    elif campo_id == "unidad_cantidad":
        choices = CargoPof.UnidadCantidad.choices

    for codigo, etiqueta in choices:
        if str(codigo) == str(valor):
            return str(etiqueta)
    return valor


def _armar_chips(filtros_avanzados):
    """
    Construye chips desde la única representación canónica de filtros.

    - Representa búsquedas rápidas como filtros avanzados con operador 0.
    - Conserva la agrupación OR de múltiples valores intencionales.
    - Resume el operador de igualdad con el formato `Campo: Valor`.
    """
    chips = []

    for filtro in filtros_avanzados:
        campo_id = filtro["campo"]
        etiqueta = FILTROS_AVANZADOS_LABELS.get(
            campo_id,
            COLUMNAS_POR_ID.get(campo_id, {}).get("label", campo_id),
        )
        operador = filtro["operador"]
        operador_label = OPERADORES_FILTRO.get(operador, OPERADORES_FILTRO["0"])
        valor_label = _valor_filtro_label(campo_id, filtro["valor"])
        texto = (
            f"{etiqueta}: {valor_label}"
            if operador == "2"
            else f"{etiqueta} {operador_label}: {valor_label}"
        )
        chips.append({
            "tipo": "avanzado",
            "indice": filtro.get("indice"),
            "origen": filtro.get("origen", "avanzado"),
            "campo": campo_id,
            "etiqueta": etiqueta,
            "operador": operador,
            "operador_label": operador_label,
            "valor": filtro["valor"],
            "valor_label": valor_label,
            "texto": texto,
        })

    return chips


def _queryset_visualizacion(request, ignorar_filtros=False):
    queryset = _obtener_queryset_base(request)
    contexto_cabecera = _resolver_contexto_visualizacion(request)
    if contexto_cabecera["es_proyecto_especial"]:
        proyecto = contexto_cabecera["proyecto_especial"]
        queryset = (
            queryset.filter(localizacion__proyecto_especial_id=proyecto.id)
            if proyecto
            else queryset.none()
        )
    else:
        _anios_visualizacion, anio_visualizacion = _contexto_anio_visualizacion(
            request,
            contexto_cabecera,
        )
        queryset = _aplicar_anio_visualizacion_global(queryset, anio_visualizacion)

    filtros = {} if ignorar_filtros else _obtener_filtros(request)
    filtros_avanzados = [] if ignorar_filtros else _obtener_filtros_avanzados(request)
    busqueda = "" if ignorar_filtros else _obtener_busqueda_general(request)

    if busqueda or filtros.get("ceic") or filtros_avanzados:
        queryset = _agregar_anotaciones_busqueda(queryset)

    if not ignorar_filtros:
        queryset = _aplicar_filtros(queryset, filtros)
        queryset = _aplicar_filtros_avanzados(queryset, filtros_avanzados)
        queryset = _aplicar_busqueda_general(queryset, busqueda)
    else:
        return queryset.distinct().order_by("localizacion__cueanexo", "localizacion__cuof", "ceic", "id")

    return _aplicar_orden(queryset.distinct(), request)


def _anotar_unidad_paginacion_cue(queryset):
    """
    Anota la unidad estable usada para paginar el visualizador POF.

    - Agrupa los CUEANEXO por sus siete digitos de CUE base.
    - Conserva el agrupador CUOF vigente para Proyecto Especial sin CUE.
    - Opera sobre el queryset ya autorizado y filtrado que recibe.
    """
    sin_cue = Q(localizacion__cueanexo="") | Q(localizacion__cueanexo__isnull=True)
    return queryset.annotate(
        _unidad_paginacion_cue=Case(
            When(
                sin_cue,
                then=Concat(Value("0:CUOF:"), "localizacion__cuof"),
            ),
            default=Concat(
                Value("1:CUE:"),
                Substr("localizacion__cueanexo", 1, 7),
            ),
            output_field=CharField(),
        )
    )


def _orden_unidades_paginacion(request):
    """
    Resuelve el orden de las unidades CUE para la pagina solicitada.

    - Mantiene orden descendente cuando el usuario ordena CUE o CUEANEXO asi.
    - Usa orden ascendente estable para las demas columnas de cargo.
    - No acepta campos libres como expresiones ORM.
    """
    orden = _limpiar_texto(request.GET.get("orden", ""))
    direccion = _limpiar_texto(request.GET.get("dir", "asc")).lower()
    if orden in {"cue", "cueanexo"} and direccion == "desc":
        return "-_unidad_paginacion_cue"
    return "_unidad_paginacion_cue"


def _paginar_unidades_cue(queryset, request):
    """
    Obtiene desde PostgreSQL las cinco unidades CUE de la pagina pedida.

    - Parte del queryset con alcance y filtros ya aplicados.
    - Cuenta y pagina claves distintas sin materializar los cargos completos.
    - Normaliza paginas invalidas o fuera de rango mediante `Paginator.get_page`.
    """
    unidades_queryset = (
        _anotar_unidad_paginacion_cue(queryset)
        .order_by(_orden_unidades_paginacion(request))
        .values_list("_unidad_paginacion_cue", flat=True)
        .distinct()
    )
    paginator = Paginator(unidades_queryset, CUES_POR_PAGINA_VISUALIZACION)
    page_obj = paginator.get_page(request.GET.get("page", 1))
    unidades_pagina = list(page_obj.object_list)
    return unidades_pagina, page_obj


def _restringir_queryset_a_unidades_cue(queryset, unidades_pagina):
    """
    Restringe cargos ya autorizados a las unidades CUE de una pagina.

    - No reconstruye el alcance ni consulta anexos fuera del queryset recibido.
    - Mantiene juntos los cargos de cada CUE y el orden interno solicitado.
    - Devuelve un queryset vacio cuando la pagina no contiene unidades.
    """
    if not unidades_pagina:
        return queryset.none()

    orden_actual = tuple(queryset.query.order_by)
    orden_unidad = Case(
        *[
            When(_unidad_paginacion_cue=unidad, then=Value(indice))
            for indice, unidad in enumerate(unidades_pagina)
        ],
        default=Value(len(unidades_pagina)),
        output_field=IntegerField(),
    )
    return (
        _anotar_unidad_paginacion_cue(queryset)
        .filter(_unidad_paginacion_cue__in=unidades_pagina)
        .annotate(_orden_unidad_paginacion=orden_unidad)
        .order_by("_orden_unidad_paginacion", *orden_actual)
    )


def _construir_metadatos_paginacion(page_obj, total_registros_pagina):
    """
    Serializa metadatos minimos para la futura UI de paginacion por CUE.

    - Conserva fija la pagina en cinco unidades completas.
    - Informa navegacion anterior y siguiente sin alterar claves existentes.
    - Distingue cargos visibles en la pagina del total de cargos filtrados.
    """
    return {
        "pagina_actual": page_obj.number,
        "total_paginas": page_obj.paginator.num_pages,
        "cues_por_pagina": CUES_POR_PAGINA_VISUALIZACION,
        "total_cues": page_obj.paginator.count,
        "total_registros_pagina": total_registros_pagina,
        "tiene_anterior": page_obj.has_previous(),
        "pagina_anterior": (
            page_obj.previous_page_number() if page_obj.has_previous() else None
        ),
        "tiene_siguiente": page_obj.has_next(),
        "pagina_siguiente": (
            page_obj.next_page_number() if page_obj.has_next() else None
        ),
    }


def construir_contexto_visualizacion_cargos_localizacion(request, incluir_opciones=True):
    """
    Construye la vista paginada del visualizador sin ampliar el alcance POF.

    - Aplica permisos, contexto y filtros antes de seleccionar los CUE paginables.
    - Materializa solo cargos de las cinco unidades CUE de la pagina.
    - Conserva los calculos vigentes de cantidad, puntos, total y Total General.
    """
    contexto_cabecera = _resolver_contexto_visualizacion(request)
    base_params_contexto = _params_contexto_visualizacion(contexto_cabecera)
    anios_visualizacion, anio_visualizacion = _contexto_anio_visualizacion(
        request,
        contexto_cabecera,
    )
    if anio_visualizacion is not None:
        base_params_contexto["anio"] = str(anio_visualizacion)
    filtros = _obtener_filtros(request)
    filtros_avanzados = _obtener_filtros_avanzados(request)
    busqueda = _obtener_busqueda_general(request)
    busquedas_columna = _obtener_busquedas_columna(request)
    busqueda_columna_id, busqueda_columna_valor = _obtener_busqueda_columna_activa(
        filtros_avanzados
    )
    columnas_visibles_ids, columnas_estado = _resolver_columnas_visibles(request)
    columnas = _armar_columnas(request, columnas_visibles_ids)
    columnas_visibles = [columna for columna in columnas if columna["visible"]]
    queryset = _queryset_visualizacion(request)

    try:
        total_registros = queryset.count()
        unidades_pagina, page_obj = _paginar_unidades_cue(queryset, request)
        queryset_pagina = _restringir_queryset_a_unidades_cue(
            queryset,
            unidades_pagina,
        )
        cargos = list(queryset_pagina)
        queryset_totales_pagina = _restringir_queryset_a_unidades_cue(
            _queryset_visualizacion(request, ignorar_filtros=True),
            unidades_pagina,
        )
        contexto_totales_generales = _construir_contexto_totales_generales_orm(
            queryset_totales_pagina,
            cargos,
        )
        filas = _armar_filas_tabla(cargos, columnas, contexto_totales_generales)
        paginacion = _construir_metadatos_paginacion(page_obj, len(cargos))
        tabla_no_migrada = False
    except (ProgrammingError, OperationalError):
        total_registros = 0
        filas = []
        paginator = Paginator([], CUES_POR_PAGINA_VISUALIZACION)
        page_obj = paginator.get_page(1)
        paginacion = _construir_metadatos_paginacion(page_obj, 0)
        tabla_no_migrada = True

    contexto = {
        "filtros": filtros,
        "filtros_avanzados": filtros_avanzados,
        "filtros_campos": FILTROS_AVANZADOS_CAMPOS,
        "filtros_opciones": (
            construir_opciones_filtros_visualizacion_cargos_localizacion(request)
            if incluir_opciones else {}
        ),
        "operadores_filtro": OPERADORES_FILTRO,
        "busqueda_general": busqueda,
        "busquedas_columna": busquedas_columna,
        "busqueda_columna_id": busqueda_columna_id,
        "busqueda_columna_valor": busqueda_columna_valor,
        "filtros_activos": _armar_chips(
            filtros_avanzados,
        ),
        "limpiar_filtros_querystring": _query_limpia(columnas_visibles_ids, base_params_contexto),
        "columnas": columnas,
        "columnas_visibles": columnas_visibles,
        "columnas_visibles_count": len(columnas_visibles),
        "columnas_total_count": len(VISUALIZACION_CARGOS_COLUMNAS),
        "columnas_estado": columnas_estado,
        "columnas_colspan": max(len(columnas), 1),
        "filas": filas,
        "querystring_exportar_filtros": _query_exportar_filtros(request, base_params_contexto),
        "querystring_exportar_todo": _query_params_desde_dict(base_params_contexto),
        "total_registros": total_registros,
        "paginacion": paginacion,
        "tabla_visualizacion_no_migrada": tabla_no_migrada,
        "es_proyecto_especial": contexto_cabecera["es_proyecto_especial"],
        "anios_visualizacion": anios_visualizacion,
        "anio_visualizacion": anio_visualizacion,
        "proyecto_especial": contexto_cabecera["proyecto_especial"],
        "cabecera_tipo_activa": contexto_cabecera["cabecera_tipo_activa"],
        "proyecto_especial_id": contexto_cabecera["proyecto_especial_id"],
        "proyecto_especial_id_activo": contexto_cabecera["proyecto_especial_id"],
        "mensaje_contexto_visualizacion": contexto_cabecera["mensaje"],
    }
    return contexto


def construir_payload_visualizacion_cargos_localizacion(request):
    """
    Construye el contrato JSON del visualizador paginado por CUE.

    - Mantiene todas las claves historicas del endpoint de datos.
    - Agrega solo metadatos de paginacion para la Etapa 2.
    - Reutiliza el mismo alcance, filtros y calculos de la vista HTML.
    """
    contexto = construir_contexto_visualizacion_cargos_localizacion(
        request,
        incluir_opciones=False,
    )
    return {
        "ok": True,
        "columnas": contexto["columnas"],
        "columnas_visibles_count": contexto["columnas_visibles_count"],
        "columnas_total_count": contexto["columnas_total_count"],
        "columnas_colspan": contexto["columnas_colspan"],
        "filas": contexto["filas"],
        "filtros_activos": contexto["filtros_activos"],
        "querystring": _normalizar_query_filtros(request).urlencode(),
        "querystring_exportar_filtros": contexto["querystring_exportar_filtros"],
        "querystring_exportar_todo": contexto["querystring_exportar_todo"],
        "total_registros": contexto["total_registros"],
        "paginacion": contexto["paginacion"],
        "tabla_visualizacion_no_migrada": contexto["tabla_visualizacion_no_migrada"],
        "es_proyecto_especial": contexto["es_proyecto_especial"],
        "proyecto_especial_id": contexto["proyecto_especial_id"],
    }


def _texto_filtros_excel(request, exportar_todo=False):
    contexto_cabecera = _resolver_contexto_visualizacion(request)
    partes_contexto = []
    proyecto = contexto_cabecera.get("proyecto_especial")
    if contexto_cabecera.get("es_proyecto_especial") and proyecto:
        partes_contexto.append(
            f"Proyecto Especial: {proyecto.anio} - {proyecto.nombre}"
        )
    elif not contexto_cabecera.get("es_proyecto_especial"):
        _anios_visualizacion, anio_visualizacion = _contexto_anio_visualizacion(
            request,
            contexto_cabecera,
        )
        if anio_visualizacion is not None:
            partes_contexto.append(f"Año: {anio_visualizacion}")

    if exportar_todo:
        return " | ".join(partes_contexto) if partes_contexto else "Sin filtros aplicados"

    filtros = _obtener_filtros(request)
    busqueda = _obtener_busqueda_general(request)
    partes = list(partes_contexto)

    if busqueda:
        partes.append(f"Búsqueda: {busqueda}")

    labels_filtros = {
        filtro["id"]: filtro["label"]
        for filtro in list(FILTROS_TEXTO) + list(FILTROS_SELECT_DEFINICIONES)
    }
    for filtro_id, valor in filtros.items():
        if valor:
            partes.append(f"{labels_filtros.get(filtro_id, filtro_id)}: {valor}")

    for filtro in _obtener_filtros_avanzados(request):
        etiqueta = FILTROS_AVANZADOS_LABELS.get(filtro["campo"], filtro["campo"])
        operador = OPERADORES_FILTRO.get(filtro["operador"], OPERADORES_FILTRO["0"])
        valor = _valor_filtro_label(filtro["campo"], filtro["valor"])
        partes.append(f"{etiqueta} {operador}: {valor}")

    return " | ".join(partes) if partes else "Sin filtros aplicados"


def _valor_excel(valor):
    if isinstance(valor, Decimal):
        return float(valor)
    return "" if valor == GUION else valor


def _valor_excel_columna(columna_id, valor):
    if columna_id == "total_general" and valor not in (None, "", GUION):
        return float(Decimal(valor))
    return _valor_excel(valor)


def _iterar_cargos_exportacion_por_lotes(queryset, chunk_size=1000):
    """
    Recorre el queryset de exportacion en lotes sin conservar su cache global.

    - Usa `iterator` con un `chunk_size` explicito para mantener compatible el prefetch vigente.
    - Conserva exactamente el orden del queryset recibido.
    - Materializa solo el lote actual, que luego puede asociarse al mapa de claves por cargo.
    """
    lote = []
    for cargo in queryset.iterator(chunk_size=chunk_size):
        lote.append(cargo)
        if len(lote) == chunk_size:
            yield lote
            lote = []
    if lote:
        yield lote


def construir_excel_visualizacion_cargos_localizacion(request, exportar_todo=False):
    """
    Construye el Excel del visualizador con el alcance y formato vigentes.

    - Calcula Total General en SQL desde el queryset autorizado sin filtros de pantalla.
    - Recorre secuencialmente el queryset en lotes de 1.000 sin cachear todos los cargos.
    - Calcula anchos durante la escritura y reutiliza estilos identicos por tipo de celda.
    - Mantiene columnas, estilos y valores de la exportacion sin cambios funcionales.
    """
    from copy import copy as copy_style

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    contexto_cabecera = _resolver_contexto_visualizacion(request)
    queryset = _queryset_visualizacion(request, ignorar_filtros=exportar_todo)
    contexto_totales_generales = _construir_contexto_totales_generales_orm(
        _queryset_visualizacion(request, ignorar_filtros=True),
        (),
    )
    columnas_ids, _ = _resolver_columnas_visibles(request, exportar_todo=exportar_todo)
    if not columnas_ids:
        columnas_ids = list(COLUMNAS_DEFAULT_IDS)
    columnas = [COLUMNAS_POR_ID[columna_id] for columna_id in columnas_ids]

    wb = Workbook()
    ws = wb.active
    ws.title = "Cargos POF"

    titulo_font = Font(bold=True, size=13)
    subtitulo_font = Font(size=10)
    filtros_font = Font(size=10)
    filtros_alignment = Alignment(wrap_text=True)
    encabezado_font = Font(bold=True, color="FFFFFF")
    encabezado_fill = PatternFill("solid", fgColor="2444D8")
    encabezado_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    cuerpo_alignment = Alignment(
        vertical="top",
        wrap_text=True,
    )
    borde_superior_cue = Side(style="medium", color="2444D8")
    borde_superior_anexo = Side(style="thin", color="111827")

    total_columnas = len(columnas)
    ultima_columna = get_column_letter(total_columnas)
    proyecto = contexto_cabecera.get("proyecto_especial")
    titulo = (
        f"Proyecto Especial POF - {proyecto.anio} - {proyecto.nombre}"
        if contexto_cabecera.get("es_proyecto_especial") and proyecto
        else "Cargos por Localización POF"
    )
    subtitulo = timezone.localtime(timezone.now()).strftime("Generado el %d/%m/%Y %H:%M")

    ws.merge_cells(f"A1:{ultima_columna}1")
    ws["A1"] = titulo
    ws["A1"].font = titulo_font

    ws.merge_cells(f"A2:{ultima_columna}2")
    ws["A2"] = subtitulo
    ws["A2"].font = subtitulo_font

    ws.merge_cells(f"A3:{ultima_columna}3")
    ws["A3"] = f"Filtros aplicados: {_texto_filtros_excel(request, exportar_todo)}"
    ws["A3"].font = filtros_font
    ws["A3"].alignment = filtros_alignment

    fila_encabezado = 4
    anchos_columnas = [
        max(12, min(len(columna["label"]) + 4, 42))
        for columna in columnas
    ]
    for indice, columna in enumerate(columnas, start=1):
        celda = ws.cell(row=fila_encabezado, column=indice, value=columna["label"])
        celda.font = encabezado_font
        celda.fill = encabezado_fill
        celda.alignment = encabezado_alignment

    fila_actual = fila_encabezado + 1
    clave_localizacion_anterior = None
    claves_totales_vistas = set()
    cue_anterior = None
    anexo_anterior = None
    for cargos_lote in _iterar_cargos_exportacion_por_lotes(queryset):
        contexto_totales_generales["claves_por_cargo"] = {
            cargo.id: _obtener_clave_agrupacion_visualizacion(cargo.localizacion)
            for cargo in cargos_lote
        }
        for cargo in cargos_lote:
            fila_raw = _serializar_cargo(cargo, contexto_totales_generales)
            _, cue_actual, anexo_actual = _obtener_claves_jerarquicas_visualizacion(
                fila_raw
            )
            es_inicio_cue = bool(
                fila_actual > fila_encabezado + 1
                and cue_actual
                and cue_actual != cue_anterior
            )
            es_inicio_anexo = bool(
                fila_actual > fila_encabezado + 1
                and not es_inicio_cue
                and cue_actual == cue_anterior
                and anexo_actual
                and anexo_actual != anexo_anterior
            )
            borde_superior = (
                borde_superior_cue
                if es_inicio_cue
                else borde_superior_anexo
                if es_inicio_anexo
                else None
            )
            fila_display, clave_localizacion_anterior = _aplicar_no_repeticion_visual(
                fila_raw,
                clave_localizacion_anterior,
            )
            _aplicar_no_repeticion_total_general(
                fila_display,
                claves_totales_vistas,
            )
            for indice, columna in enumerate(columnas, start=1):
                valor_excel = _valor_excel_columna(
                    columna["id"],
                    fila_display.get(columna["id"], ""),
                )
                celda = ws.cell(
                    row=fila_actual,
                    column=indice,
                    value=valor_excel,
                )
                valor_texto = "" if valor_excel is None else str(valor_excel)
                anchos_columnas[indice - 1] = max(
                    anchos_columnas[indice - 1],
                    min(len(valor_texto) + 2, 42),
                )
                celda.alignment = cuerpo_alignment
                if borde_superior is not None:
                    borde = copy_style(celda.border)
                    borde.top = borde_superior
                    celda.border = borde
            fila_actual += 1
            cue_anterior = cue_actual or None
            anexo_anterior = anexo_actual or None

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{fila_encabezado}:{ultima_columna}{max(fila_actual - 1, fila_encabezado)}"

    for indice, ancho in enumerate(anchos_columnas, start=1):
        letra = get_column_letter(indice)
        ws.column_dimensions[letra].width = ancho

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return {
        "contenido": buffer.getvalue(),
        "nombre_archivo": (
            f"Proyecto_Especial_POF_{proyecto.anio}_{proyecto.id}.xlsx"
            if contexto_cabecera.get("es_proyecto_especial") and proyecto and exportar_todo
            else f"Proyecto_Especial_POF_{proyecto.anio}_{proyecto.id}_Filtros.xlsx"
            if contexto_cabecera.get("es_proyecto_especial") and proyecto
            else "Cargos_por_Localizacion_POF.xlsx"
            if exportar_todo
            else "Cargos_por_Localizacion_POF_Filtros.xlsx"
        ),
    }
