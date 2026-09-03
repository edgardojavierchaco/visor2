from inspect import unwrap
from decimal import Decimal
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import MagicMock, Mock, patch

from django.core.exceptions import PermissionDenied
from django.http import JsonResponse
from django.test import RequestFactory, SimpleTestCase
from openpyxl import load_workbook

from . import forms as pof_forms
from . import models, permisos, views
from .models import (
    ROL_POF_DIRECTOR,
    ROL_POF_REGIONAL,
    ROLES_POF_ACCESO_COMPLETO,
    ROLES_POF_SOLO_VISUALIZACION_COMPLETA,
    VCapaUnicaOfertasAnt,
    obtener_cueanexos_director_pof,
    obtener_regiones_usuario_pof,
)
from .services import carga_service
from .services import exportacion_reunida as exportacion_service
from .services import guardado_pof_service
from .services import visualizacion_cargos_localizacion_service as visualizacion_service


class RolesPofTests(SimpleTestCase):
    def test_roles_centralizados_coinciden_con_la_matriz(self):
        self.assertEqual(ROLES_POF_ACCESO_COMPLETO, {"Pof", "Administrador"})
        self.assertIn("Director de Nivel Inicial", ROLES_POF_SOLO_VISUALIZACION_COMPLETA)
        self.assertNotIn(ROL_POF_REGIONAL, ROLES_POF_SOLO_VISUALIZACION_COMPLETA)
        self.assertNotIn(ROL_POF_DIRECTOR, ROLES_POF_SOLO_VISUALIZACION_COMPLETA)


class ExportacionReunidaExcelTests(SimpleTestCase):
    def _abrir_excel_comun(self, columnas_config, filas, schema):
        contexto = {
            "es_proyecto_especial": False,
            "columnas": [columna["titulo"] for columna in columnas_config],
            "filas_exportacion": [],
            "filas_normalizadas_exportacion": filas,
            "columnas_exportacion_config": columnas_config,
            "schema_exportacion": schema,
            "separadores_filas_exportacion": [],
            "secciones_exportacion": [],
            "mensaje_exportacion": "",
            "nombre_archivo": "reunida.xlsx",
            "titulo_hoja": "Reunida",
            "titulo_excel": "Reunida POF",
            "reunida": {},
        }
        respuesta = views._crear_respuesta_excel_exportacion(contexto)
        return load_workbook(BytesIO(respuesta.content), data_only=False).active

    def _columnas_auxiliares(self, ws, cantidad_columnas_visibles):
        return {
            ws.cell(row=4, column=indice).value: indice
            for indice in range(cantidad_columnas_visibles + 1, ws.max_column + 1)
        }

    def test_clave_total_general_estandar_prioriza_cueanexo_y_aplica_fallbacks(self):
        helper = exportacion_service.obtener_clave_total_general_cueanexo_reunida

        self.assertEqual(
            helper({"cueanexo": " 111111100 ", "cuof": "10", "cue": "1111111"}),
            "CUEANEXO:111111100",
        )
        self.assertNotEqual(
            helper({"cueanexo": "111111100", "cue": "1111111"}),
            helper({"cueanexo": "111111101", "cue": "1111111"}),
        )
        self.assertEqual(helper({"cueanexo": "", "cuof": " 20 "}), "CUOF:20")
        self.assertEqual(
            helper({"cueanexo": "", "cuof": "", "localizacion_id": 30}),
            "LOCALIZACION:30",
        )
        self.assertEqual(helper({}, indice=4), "FILA:4")

    def test_preview_total_general_separa_cueanexos_del_mismo_cue(self):
        filas = [
            {"cueanexo": "111111100", "cue": "1111111", "cargo_id": 1},
            {"cueanexo": "111111100", "cue": "1111111", "cargo_id": 2},
            {"cueanexo": "111111101", "cue": "1111111", "cargo_id": 3},
        ]
        totales = {
            "CUEANEXO:111111100": Decimal("300"),
            "CUEANEXO:111111101": Decimal("50"),
        }
        columnas = [
            {"source": "cueanexo", "repetir": "por_cueanexo"},
            {"source": "total_general", "repetir": "por_cue"},
        ]

        exportacion_service._aplicar_totales_generales_preview_reunida(
            filas,
            "PRIMARIA",
            totales,
        )
        proyectadas = exportacion_service._proyectar_filas_exportacion(
            "PRIMARIA",
            filas,
            columnas,
        )

        self.assertEqual(
            proyectadas,
            [
                ["111111100", Decimal("300")],
                ["", ""],
                ["111111101", Decimal("50")],
            ],
        )
        self.assertEqual(filas[0]["total_general_exportacion"], Decimal("300"))
        self.assertEqual(filas[2]["total_general_exportacion"], Decimal("50"))

    def test_excel_total_general_no_encadena_cueanexos_del_mismo_cue(self):
        columnas_config = [
            {"source": "cueanexo", "titulo": "CUEANEXO", "repetir": "por_cueanexo"},
            {
                "source": "total_general",
                "titulo": "Total General",
                "repetir": "por_cue",
            },
        ]
        filas = [
            {
                "cueanexo": "111111100",
                "cue": "1111111",
                "estado_pof_codigo": views.CargoPof.EstadoPof.AFECTADO,
                "total": Decimal("100"),
            },
            {
                "cueanexo": "111111100",
                "cue": "1111111",
                "estado_pof_codigo": views.CargoPof.EstadoPof.AFECTADO,
                "total": Decimal("200"),
            },
            {
                "cueanexo": "111111100",
                "cue": "1111111",
                "estado_pof_codigo": views.CargoPof.EstadoPof.DESAFECTADO,
                "total": Decimal("0"),
            },
            {
                "cueanexo": "111111101",
                "cue": "1111111",
                "estado_pof_codigo": views.CargoPof.EstadoPof.AFECTADO,
                "total": Decimal("50"),
            },
        ]

        ws = self._abrir_excel_comun(
            columnas_config,
            filas,
            {"columnas": [], "grupo_total_general": ("cue",)},
        )
        auxiliares = self._columnas_auxiliares(ws, len(columnas_config))
        letra_total = ws.cell(
            4,
            auxiliares["_total_visible_grupo"],
        ).column_letter
        letra_grupo_visto = ws.cell(
            4,
            auxiliares["_grupo_visto_visible"],
        ).column_letter
        letra_es_afectado = ws.cell(
            4,
            auxiliares["_es_afectado"],
        ).column_letter

        self.assertIn(f"${letra_total}6", ws[f"{letra_total}5"].value)
        self.assertIn(f"${letra_total}7", ws[f"{letra_total}6"].value)
        self.assertNotIn(f"${letra_total}8", ws[f"{letra_total}7"].value)
        self.assertEqual(ws[f"{letra_es_afectado}7"].value, 0)
        self.assertNotIn(f"${letra_grupo_visto}7", ws[f"{letra_grupo_visto}8"].value)
        self.assertIn("0=0", ws["B8"].value)
        self.assertEqual(
            [columna.colId for columna in ws.auto_filter.filterColumn],
            [1],
        )

    def test_totales_tecnicos_visibles_usan_acumuladores_dinamicos_por_grupo(self):
        columnas_config = [
            {"source": "cueanexo", "titulo": "CUEANEXO", "repetir": "por_cueanexo"},
            {
                "source": "total_horas_catedra",
                "titulo": "Total Horas Cátedra",
                "repetir": "siempre",
            },
            {
                "source": "puntos_horas_catedra",
                "titulo": "Puntos Horas Cátedra",
                "repetir": "siempre",
            },
            {"source": "total_puntos", "titulo": "Total Puntos", "repetir": "siempre"},
        ]
        schema = {
            "columnas": [
                {"key": "total_horas_catedra"},
                {"key": "puntos_horas_catedra"},
                {"key": "total_puntos"},
            ],
            "grupo_total_general": ("cue",),
        }
        filas = [
            {
                "cueanexo": "111111100",
                "cue": "1111111",
                "estado_pof_codigo": views.CargoPof.EstadoPof.AFECTADO,
                "cantidad_horas": Decimal("10"),
                "total": Decimal("100"),
                "total_horas_catedra": Decimal("10"),
                "puntos_horas_catedra": Decimal("100"),
                "total_puntos": Decimal("100"),
            },
            {
                "cueanexo": "111111100",
                "cue": "1111111",
                "estado_pof_codigo": views.CargoPof.EstadoPof.AFECTADO,
                "cantidad_horas": "",
                "total": Decimal("200"),
                "total_horas_catedra": "",
                "puntos_horas_catedra": "",
                "total_puntos": Decimal("200"),
            },
            {
                "cueanexo": "111111100",
                "cue": "1111111",
                "estado_pof_codigo": views.CargoPof.EstadoPof.DESAFECTADO,
                "cantidad_horas": Decimal("5"),
                "total": Decimal("50"),
                "total_horas_catedra": Decimal("5"),
                "puntos_horas_catedra": Decimal("50"),
                "total_puntos": Decimal("50"),
            },
            {
                "cueanexo": "222222200",
                "cue": "2222222",
                "estado_pof_codigo": views.CargoPof.EstadoPof.AFECTADO,
                "cantidad_horas": "",
                "total": Decimal("25"),
            },
            {
                "cueanexo": "111111101",
                "cue": "1111111",
                "estado_pof_codigo": views.CargoPof.EstadoPof.AFECTADO,
                "cantidad_horas": "",
                "total": Decimal("300"),
            },
        ]

        ws = self._abrir_excel_comun(columnas_config, filas, schema)
        auxiliares = self._columnas_auxiliares(ws, len(columnas_config))

        self.assertEqual(len(auxiliares), 14)
        self.assertIn("_cantidad_horas", auxiliares)
        self.assertIn("_horas_visible_grupo", auxiliares)
        self.assertIn("_puntos_horas_visible_grupo", auxiliares)
        self.assertTrue(
            all(
                ws.column_dimensions[ws.cell(4, indice).column_letter].hidden
                for indice in auxiliares.values()
            )
        )
        self.assertNotIn(ws.cell(4, auxiliares["_fila_visible"]).column_letter, ws.auto_filter.ref)

        letras = {
            nombre: ws.cell(4, indice).column_letter
            for nombre, indice in auxiliares.items()
        }
        self.assertIn(f"${letras['_horas_visible_grupo']}5", ws["B5"].value)
        self.assertIn(f"${letras['_puntos_horas_visible_grupo']}5", ws["C5"].value)
        self.assertIn(f"${letras['_total_visible_grupo']}5", ws["D5"].value)
        self.assertEqual(ws["B5"].number_format, "#,##0")
        self.assertEqual(ws["C5"].number_format, "#,##0.00")
        self.assertEqual(ws["D5"].number_format, "#,##0.00")

        formula_horas = ws.cell(5, auxiliares["_horas_visible_grupo"]).value
        formula_puntos_horas = ws.cell(
            5, auxiliares["_puntos_horas_visible_grupo"]
        ).value
        self.assertIn(
            f"SUBTOTAL(109,${letras['_cantidad_horas']}5)",
            formula_horas,
        )
        self.assertIn(
            f"SUBTOTAL(109,${letras['_total_cargo']}5)",
            formula_puntos_horas,
        )
        self.assertNotIn("puntos", formula_puntos_horas.lower())
        for acumulador in (
            "_total_visible_grupo",
            "_horas_visible_grupo",
            "_puntos_horas_visible_grupo",
        ):
            formula_intercalada = ws.cell(7, auxiliares[acumulador]).value
            self.assertIn(f"${letras[acumulador]}9", formula_intercalada)

        self.assertEqual(
            {columna.colId for columna in ws.auto_filter.filterColumn},
            {1, 2, 3},
        )
        self.assertTrue(
            all(
                columna.hiddenButton and columna.showButton is False
                for columna in ws.auto_filter.filterColumn
            )
        )
        reglas_condicionales = [
            regla
            for reglas in ws.conditional_formatting._cf_rules.values()
            for regla in reglas
        ]
        self.assertTrue(
            any(
                regla.dxf
                and regla.dxf.numFmt
                and regla.dxf.numFmt.formatCode == ";;;"
                and regla.dxf.font is None
                for regla in reglas_condicionales
            )
        )
        self.assertTrue(
            any(
                regla.dxf
                and regla.dxf.border
                and regla.dxf.border.top.style == "medium"
                and regla.dxf.border.top.color.rgb == "003B5CFF"
                for regla in reglas_condicionales
            )
        )
        self.assertTrue(
            any(
                regla.dxf
                and regla.dxf.border
                and regla.dxf.border.top.style == "medium"
                and regla.dxf.border.top.color.rgb == "009CA3AF"
                for regla in reglas_condicionales
            )
        )

    def test_total_puntos_no_es_dinamico_fuera_del_schema_tecnico(self):
        columnas_config = [
            {"source": "cueanexo", "titulo": "CUEANEXO", "repetir": "por_cueanexo"},
            {"source": "total_puntos", "titulo": "Total Puntos", "repetir": "siempre"},
        ]
        fila = {
            "cueanexo": "111111100",
            "cue": "1111111",
            "total_puntos": Decimal("125"),
        }

        ws = self._abrir_excel_comun(
            columnas_config,
            [fila],
            {"columnas": [{"key": "total_puntos"}]},
        )

        self.assertEqual(ws["B5"].value, 125)
        self.assertEqual(ws.max_column, len(columnas_config) + 7)
        self.assertFalse(ws.auto_filter.filterColumn)
        self.assertIsNone(ws.parent.calculation.forceFullCalc)

    def test_total_general_normal_conserva_once_helpers(self):
        columnas_config = [
            {"source": "cueanexo", "titulo": "CUEANEXO", "repetir": "por_cueanexo"},
            {
                "source": "total_general",
                "titulo": "Total General",
                "repetir": "por_cue",
            },
        ]
        fila = {
            "cueanexo": "111111100",
            "cue": "1111111",
            "estado_pof_codigo": views.CargoPof.EstadoPof.AFECTADO,
            "total": Decimal("125"),
        }

        ws = self._abrir_excel_comun(
            columnas_config,
            [fila],
            {"columnas": [], "grupo_total_general": ("cue",)},
        )

        self.assertEqual(ws.max_column, len(columnas_config) + 11)
        self.assertTrue(ws["B5"].value.startswith("=IF(AND(SUBTOTAL(103,"))
        self.assertEqual(
            [columna.colId for columna in ws.auto_filter.filterColumn],
            [1],
        )

    def test_proyecto_especial_conserva_writer_historico_sin_auxiliares_comunes(self):
        """Proyecto Especial no debe entrar al writer filtrable de Reunidas."""
        contexto = {
            "es_proyecto_especial": True,
            "columnas": ["CUEANEXO", "Cargo"],
            "filas_exportacion": [["123456700", "Cargo A"]],
            "filas_normalizadas_exportacion": [
                {"cueanexo": "123456700", "cargo": "Cargo A"}
            ],
            "columnas_exportacion_config": [
                {"source": "cueanexo", "titulo": "CUEANEXO"},
                {"source": "cargo", "titulo": "Cargo"},
            ],
            "separadores_filas_exportacion": [],
            "secciones_exportacion": [],
            "mensaje_exportacion": "",
            "nombre_archivo": "proyecto.xlsx",
            "titulo_hoja": "Proyecto Especial",
            "titulo_excel": "Proyecto Especial POF",
            "reunida": {},
        }

        respuesta = views._crear_respuesta_excel_exportacion(contexto)
        ws = load_workbook(BytesIO(respuesta.content), data_only=False).active

        self.assertEqual(ws.max_column, 2)
        self.assertEqual(ws.auto_filter.ref, "A4:B5")
        self.assertEqual(ws.freeze_panes, "A5")
        self.assertFalse(ws.column_dimensions["C"].hidden)
        self.assertFalse(
            any(
                isinstance(celda.value, str) and celda.value.startswith("=")
                for fila in ws.iter_rows()
                for celda in fila
            )
        )


class AniosDisponiblesCargaPofTests(SimpleTestCase):
    def test_validacion_cabecera_admite_anio_posterior_si_la_pof_existe(self):
        reunida = SimpleNamespace(
            id=10,
            anio=2099,
            nivel="ADULTOS",
            get_nivel_display=lambda: "Adultos",
        )
        manager = MagicMock()
        manager.filter.return_value.first.return_value = reunida

        with patch.object(carga_service.ReunidaPof, "objects", manager):
            resultado = carga_service.validar_cabecera_reunida(2099, "ADULTOS")

        self.assertTrue(resultado["ok"])
        manager.filter.assert_called_once_with(anio=2099, nivel="ADULTOS")

    def test_formulario_guardado_admite_anio_posterior_si_la_pof_existe(self):
        manager = MagicMock()
        manager.filter.return_value.exists.return_value = True

        with patch.object(pof_forms.ReunidaPof, "objects", manager):
            formulario = pof_forms.GuardarCargaPofForm({
                "cabecera_tipo": "REUNIDA",
                "anio": 2099,
                "nivel": "ADULTOS",
                "tipo_operacion": "AFECTADO",
            })

            self.assertTrue(formulario.is_valid(), formulario.errors)

    def test_servicio_guardado_admite_anio_posterior_si_la_pof_existe(self):
        manager = MagicMock()
        manager.filter.return_value.exists.return_value = True
        datos = {
            "cabecera_tipo": "REUNIDA",
            "anio": 2099,
            "nivel": "ADULTOS",
            "tipo_operacion": "AFECTADO",
            "padron": {
                "padron_cueanexo": "123456700",
                "cuof_loc": "123",
            },
            "cargos": [{
                "ceic": "1",
                "cantidad": 1,
                "unidad_cantidad": "CARGO",
                "observacion": "",
            }],
        }

        with patch.object(guardado_pof_service.ReunidaPof, "objects", manager):
            errores = guardado_pof_service._validar_datos_guardado_minimos(datos)

        self.assertNotIn("anio", errores)


class AsociacionesPofTests(SimpleTestCase):
    def test_regiones_elimina_nulos_espacios_y_duplicados(self):
        cursor = MagicMock()
        cursor.fetchall.return_value = [
            (" Región I ",),
            (None,),
            ("",),
            ("Región I",),
            ("Región II",),
        ]
        cursor_contexto = MagicMock()
        cursor_contexto.__enter__.return_value = cursor

        user = SimpleNamespace(username="20123456789")
        with patch.object(
            models.connection,
            "cursor",
            return_value=cursor_contexto,
        ):
            regiones = obtener_regiones_usuario_pof(user)

        self.assertEqual(regiones, {"Región I", "Región II"})
        self.assertEqual(cursor.execute.call_args.args[1], ["20123456789"])

    def test_director_prioriza_padron_cueanexo_y_conserva_el_valor_completo(self):
        manager = MagicMock()
        manager.using.return_value.filter.return_value.values_list.return_value = [
            (" 123456700 ", "999999999"),
            (None, "123456701"),
            ("123456700", None),
            (None, ""),
        ]
        user = SimpleNamespace(username="20-12345678-9")

        with patch.object(VCapaUnicaOfertasAnt, "objects", manager):
            cueanexos = obtener_cueanexos_director_pof(user)

        self.assertEqual(cueanexos, {"123456700", "123456701"})
        filtro = manager.using.return_value.filter.call_args.kwargs
        self.assertIn("20-12345678-9", filtro["resploc_cuitcuil__in"])
        self.assertIn("20123456789", filtro["resploc_cuitcuil__in"])


class DecoradoresPofTests(SimpleTestCase):
    def setUp(self):
        self.request_factory = RequestFactory()
        self.user = SimpleNamespace(is_authenticated=True)

    def test_vista_administrativa_rechaza_usuario_sin_acceso_completo(self):
        request = self.request_factory.get("/administracion/")
        request.user = self.user
        vista = permisos.pof_required(lambda request: JsonResponse({"ok": True}))

        with patch.object(permisos, "usuario_tiene_acceso_completo_pof", return_value=False):
            with self.assertRaises(PermissionDenied):
                vista(request)

    def test_api_administrativa_devuelve_403(self):
        request = self.request_factory.post("/administracion/")
        request.user = self.user
        vista = permisos.pof_api_required(lambda request: JsonResponse({"ok": True}))

        with patch.object(permisos, "usuario_tiene_acceso_completo_pof", return_value=False):
            response = vista(request)

        self.assertEqual(response.status_code, 403)

    def test_api_visualizacion_admite_capacidad_de_consulta(self):
        request = self.request_factory.get("/visualizacion/")
        request.user = self.user
        vista = permisos.pof_visualizacion_api_required(
            lambda request: JsonResponse({"ok": True})
        )

        with patch.object(permisos, "usuario_puede_ver_visualizacion_pof", return_value=True):
            response = vista(request)

        self.assertEqual(response.status_code, 200)

    def test_inicio_muestra_acceso_rapido_limitado_a_usuario_solo_visualizacion(self):
        request = self.request_factory.get("/")
        request.user = self.user
        response_esperada = Mock()

        with patch.object(views, "usuario_tiene_acceso_completo_pof", return_value=False), patch.object(
            views,
            "render",
            return_value=response_esperada,
        ) as render_mock:
            response = unwrap(views.inicio)(request)

        self.assertIs(response, response_esperada)
        render_mock.assert_called_once_with(
            request,
            "reunidas_pof/inicio.html",
            {"pof_solo_visualizacion": True},
        )


class VisualizacionFiltrosCanonicosTests(SimpleTestCase):
    """
    Verifica la normalizacion compartida entre busqueda rapida y filtros avanzados.

    - Mantiene el CUE como busqueda parcial en la ruta canonica.
    - Traduce URLs legacy sin duplicar criterios del mismo campo.
    - Conserva la acumulacion y el OR de multiples valores avanzados.
    """

    def setUp(self):
        self.request_factory = RequestFactory()

    def test_filtro_cue_parecido_usa_icontains(self):
        consulta = visualizacion_service._filtro_avanzado_q("cue", "0", "313")

        self.assertEqual(
            consulta.children,
            [("localizacion__cueanexo__icontains", "313")],
        )

    def test_busqueda_columna_cue_usa_icontains(self):
        queryset = Mock()
        visualizacion_service._aplicar_busqueda_columna(queryset, "cue", "313")

        filtro_q = queryset.filter.call_args.args[0]
        self.assertEqual(
            filtro_q.children,
            [("localizacion__cueanexo__icontains", "313")],
        )

    def test_busqueda_rapida_numerica_usa_la_anotacion_textual_canonica(self):
        filtro_q = visualizacion_service._filtro_avanzado_q("cantidad", "0", "2")

        self.assertEqual(
            filtro_q.children,
            [("cantidad_busqueda__icontains", "2")],
        )

    def test_columna_legacy_se_convierte_en_un_chip_avanzado(self):
        request = self.request_factory.get("/visualizacion/?col_cue=313")

        filtros = visualizacion_service._obtener_filtros_avanzados(request)
        chips = visualizacion_service._armar_chips(filtros)

        self.assertEqual(
            filtros,
            [{
                "indice": None,
                "campo": "cue",
                "operador": "0",
                "valor": "313",
                "origen": "legacy_columna",
            }],
        )
        self.assertEqual(len(chips), 1)
        self.assertEqual(chips[0]["texto"], "CUE parecido a: 313")
        self.assertEqual(chips[0]["tipo"], "avanzado")
        self.assertEqual(chips[0]["origen"], "legacy_columna")

    def test_filtro_avanzado_gana_sobre_columna_legacy_del_mismo_campo(self):
        request = self.request_factory.get(
            "/visualizacion/?col_cue=313&campo_filtro=cue&operador_filtro=2&valor_filtro=522522"
        )

        filtros = visualizacion_service._obtener_filtros_avanzados(request)

        self.assertEqual(
            [(filtro["campo"], filtro["operador"], filtro["valor"]) for filtro in filtros],
            [("cue", "2", "522522")],
        )

    def test_columna_legacy_se_conserva_si_la_tripleta_avanzada_es_invalida(self):
        request = self.request_factory.get(
            "/visualizacion/?col_cue=313&campo_filtro=cue&operador_filtro=99&valor_filtro=522522"
        )

        filtros = visualizacion_service._obtener_filtros_avanzados(request)

        self.assertEqual(
            [(filtro["campo"], filtro["operador"], filtro["valor"]) for filtro in filtros],
            [("cue", "0", "313")],
        )

    def test_columnas_legacy_de_campos_distintos_se_acumulan(self):
        request = self.request_factory.get(
            "/visualizacion/?col_cue=313&col_cuof=ABC"
        )

        filtros = visualizacion_service._obtener_filtros_avanzados(request)

        self.assertEqual(
            [(filtro["campo"], filtro["valor"]) for filtro in filtros],
            [("cue", "313"), ("cuof", "ABC")],
        )

    def test_multiples_valores_avanzados_del_mismo_campo_se_conservan(self):
        request = self.request_factory.get(
            "/visualizacion/?campo_filtro=oferta&campo_filtro=oferta"
            "&operador_filtro=0&operador_filtro=0"
            "&valor_filtro=Primaria&valor_filtro=Secundaria"
        )

        filtros = visualizacion_service._obtener_filtros_avanzados(request)

        self.assertEqual(
            [(filtro["campo"], filtro["valor"]) for filtro in filtros],
            [("oferta", "Primaria"), ("oferta", "Secundaria")],
        )

    def test_chips_conservan_multiples_valores_del_mismo_campo(self):
        filtros = [
            {"indice": 0, "campo": "cue", "operador": "0", "valor": "313"},
            {"indice": 1, "campo": "cue", "operador": "0", "valor": "522"},
        ]

        chips = visualizacion_service._armar_chips(filtros)

        self.assertEqual(
            [chip["valor"] for chip in chips],
            ["313", "522"],
        )

    def test_barra_no_muestra_un_valor_si_hay_dos_campos_canonicos(self):
        filtros = [
            {"campo": "cue", "operador": "0", "valor": "313"},
            {"campo": "cuof", "operador": "0", "valor": "ABC"},
        ]

        self.assertEqual(
            visualizacion_service._obtener_busqueda_columna_activa(filtros),
            ("cueanexo", ""),
        )

    def test_barra_no_muestra_un_operador_distinto_de_parecido(self):
        filtros = [{"campo": "cue", "operador": "2", "valor": "313"}]

        self.assertEqual(
            visualizacion_service._obtener_busqueda_columna_activa(filtros),
            ("cueanexo", ""),
        )

    def test_querystring_y_exportacion_reemplazan_columna_legacy(self):
        request = self.request_factory.get(
            "/visualizacion/?anio=2025&col_cue=313&page=2&page_size=50"
        )

        query = visualizacion_service._normalizar_query_filtros(request)
        exportacion = visualizacion_service._query_exportar_filtros(request)

        self.assertNotIn("col_cue", query)
        self.assertIn("campo_filtro=cue", query.urlencode())
        self.assertNotIn("page=", exportacion)
        self.assertNotIn("page_size=", exportacion)
        self.assertIn("operador_filtro=0", exportacion)
        self.assertIn("valor_filtro=313", exportacion)

    def test_querystring_no_duplica_legacy_cuando_gana_el_filtro_avanzado(self):
        request = self.request_factory.get(
            "/visualizacion/?col_cue=313&campo_filtro=cue&operador_filtro=2&valor_filtro=522522"
        )

        query = visualizacion_service._normalizar_query_filtros(request)

        self.assertEqual(query.getlist("campo_filtro"), ["cue"])
        self.assertEqual(query.getlist("operador_filtro"), ["2"])
        self.assertEqual(query.getlist("valor_filtro"), ["522522"])


class VisualizacionOrdenamientoTests(SimpleTestCase):
    def setUp(self):
        self.request_factory = RequestFactory()

    def _armar_columnas(self, querystring=""):
        request = self.request_factory.get(f"/visualizacion/?{querystring}")
        return visualizacion_service._armar_columnas(
            request,
            visualizacion_service.COLUMNAS_DEFAULT_IDS,
        )

    def _obtener_columna(self, columnas, columna_id):
        return next(columna for columna in columnas if columna["id"] == columna_id)

    def test_sin_orden_activo_el_siguiente_enlace_genera_asc(self):
        columnas = self._armar_columnas("anio=2025")
        cueanexo = self._obtener_columna(columnas, "cueanexo")
        query = self.request_factory.get(f"/?{cueanexo['order_querystring']}").GET

        self.assertEqual(query["orden"], "cueanexo")
        self.assertEqual(query["dir"], "asc")
        self.assertFalse(any(columna["order_active"] for columna in columnas))

    def test_orden_asc_activo_el_siguiente_enlace_genera_desc(self):
        columnas = self._armar_columnas("orden=cueanexo&dir=asc")
        cueanexo = self._obtener_columna(columnas, "cueanexo")
        query = self.request_factory.get(f"/?{cueanexo['order_querystring']}").GET

        self.assertEqual(query["orden"], "cueanexo")
        self.assertEqual(query["dir"], "desc")
        self.assertTrue(cueanexo["order_active"])
        self.assertEqual(cueanexo["order_dir"], "asc")

    def test_orden_desc_activo_el_siguiente_enlace_vuelve_al_predeterminado(self):
        columnas = self._armar_columnas("orden=cueanexo&dir=desc")
        cueanexo = self._obtener_columna(columnas, "cueanexo")
        query = self.request_factory.get(f"/?{cueanexo['order_querystring']}").GET

        self.assertNotIn("orden", query)
        self.assertNotIn("dir", query)
        columnas_predeterminadas = self._armar_columnas(query.urlencode())
        self.assertFalse(
            any(columna["order_active"] for columna in columnas_predeterminadas)
        )

    def test_vuelta_al_predeterminado_conserva_contexto_y_elimina_paginacion(self):
        columnas = self._armar_columnas(
            "anio=2025&cabecera_tipo=PROYECTO_ESPECIAL&proyecto_especial_id=34"
            "&campo_filtro=cue&operador_filtro=0&valor_filtro=313&q=maestra"
            "&visible_col=cueanexo&visible_col=cargo&page=3&page_size=50"
            "&orden=cueanexo&dir=desc"
        )
        cueanexo = self._obtener_columna(columnas, "cueanexo")
        query = self.request_factory.get(f"/?{cueanexo['order_querystring']}").GET

        self.assertEqual(query["anio"], "2025")
        self.assertEqual(query["cabecera_tipo"], "PROYECTO_ESPECIAL")
        self.assertEqual(query["proyecto_especial_id"], "34")
        self.assertEqual(query.getlist("campo_filtro"), ["cue"])
        self.assertEqual(query.getlist("operador_filtro"), ["0"])
        self.assertEqual(query.getlist("valor_filtro"), ["313"])
        self.assertEqual(query["q"], "maestra")
        self.assertEqual(query.getlist("visible_col"), ["cueanexo", "cargo"])
        self.assertNotIn("orden", query)
        self.assertNotIn("dir", query)
        self.assertNotIn("page", query)
        self.assertNotIn("page_size", query)

    def test_total_general_continua_sin_ser_ordenable(self):
        columnas = self._armar_columnas()
        total_general = self._obtener_columna(columnas, "total_general")

        self.assertFalse(total_general["ordenable"])
        self.assertEqual(total_general["order_querystring"], "")
        self.assertFalse(total_general["order_active"])

    def test_sin_orden_aplica_el_orden_predeterminado_existente(self):
        queryset = Mock()
        request = self.request_factory.get("/visualizacion/")

        resultado = visualizacion_service._aplicar_orden(queryset, request)

        self.assertIs(resultado, queryset.order_by.return_value)
        queryset.order_by.assert_called_once_with(
            "localizacion__cueanexo",
            "localizacion__cuof",
            "ceic",
            "id",
        )


class AlcanceVisualizacionPofTests(SimpleTestCase):
    def test_acceso_completo_y_consulta_general_no_reciben_filtro_de_datos(self):
        for rol in ("Pof", "Administrador", "Director de Nivel Inicial"):
            with self.subTest(rol=rol):
                queryset = Mock()
                with patch.object(
                    visualizacion_service,
                    "obtener_rol_usuario_pof",
                    return_value=rol,
                ):
                    resultado = visualizacion_service._aplicar_alcance_visualizacion(
                        queryset,
                        SimpleNamespace(),
                    )

                self.assertIs(resultado, queryset)
                queryset.filter.assert_not_called()
                queryset.none.assert_not_called()

    def test_regional_filtra_por_snapshot_vigente_y_regiones_asociadas(self):
        queryset = Mock()
        queryset.filter.return_value = Mock()
        user = SimpleNamespace()

        with patch.object(
            visualizacion_service,
            "obtener_rol_usuario_pof",
            return_value=ROL_POF_REGIONAL,
        ), patch.object(
            visualizacion_service,
            "obtener_regiones_usuario_pof",
            return_value={"Región I", "Región II"},
        ):
            resultado = visualizacion_service._aplicar_alcance_visualizacion(
                queryset,
                user,
            )

        self.assertIs(resultado, queryset.filter.return_value)
        queryset.filter.assert_called_once_with(
            localizacion__snapshots_padron__vigente=True,
            localizacion__snapshots_padron__region__in={"Región I", "Región II"},
        )

    def test_regional_sin_asociaciones_no_recibe_acceso_general(self):
        queryset = Mock()
        queryset.none.return_value = Mock()

        with patch.object(
            visualizacion_service,
            "obtener_rol_usuario_pof",
            return_value=ROL_POF_REGIONAL,
        ), patch.object(
            visualizacion_service,
            "obtener_regiones_usuario_pof",
            return_value=set(),
        ):
            resultado = visualizacion_service._aplicar_alcance_visualizacion(
                queryset,
                SimpleNamespace(),
            )

        self.assertIs(resultado, queryset.none.return_value)

    def test_director_filtra_por_cueanexo_completo_exacto(self):
        queryset = Mock()
        queryset.filter.return_value = Mock()
        cueanexos = {"123456700", "123456701"}

        with patch.object(
            visualizacion_service,
            "obtener_rol_usuario_pof",
            return_value=ROL_POF_DIRECTOR,
        ), patch.object(
            visualizacion_service,
            "obtener_cueanexos_director_pof",
            return_value=cueanexos,
        ):
            resultado = visualizacion_service._aplicar_alcance_visualizacion(
                queryset,
                SimpleNamespace(),
            )

        self.assertIs(resultado, queryset.filter.return_value)
        queryset.filter.assert_called_once_with(localizacion__cueanexo__in=cueanexos)

    def test_director_sin_asociaciones_no_recibe_acceso_general(self):
        queryset = Mock()
        queryset.none.return_value = Mock()

        with patch.object(
            visualizacion_service,
            "obtener_rol_usuario_pof",
            return_value=ROL_POF_DIRECTOR,
        ), patch.object(
            visualizacion_service,
            "obtener_cueanexos_director_pof",
            return_value=set(),
        ):
            resultado = visualizacion_service._aplicar_alcance_visualizacion(
                queryset,
                SimpleNamespace(),
            )

        self.assertIs(resultado, queryset.none.return_value)
