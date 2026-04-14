from apps.evaluaciones_educativas.models import *
from apps.evaluaciones_educativas.forms.forms import *
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.db import transaction
from django.contrib.auth.decorators import login_required
from django.templatetags.static import static
from django.db.models import Count,Q,Avg
from datetime import date, datetime
import psycopg2
from psycopg2 import extras
import os
from openpyxl import Workbook


@login_required
def carga_alumno(request,grado_public_id):
	grado=get_object_or_404(Grado,public_id= grado_public_id)
	alumno_form = AlumnoForm()
	grado_form = GradoForm(instance=grado)
	seccion_form = SeccionForm()
	if request.method == 'POST':
		alumno_form = AlumnoForm(request.POST)
		grado_form = GradoForm(request.POST)
		seccion_form = SeccionForm(request.POST)
		if alumno_form.is_valid() and grado_form.is_valid() and seccion_form.is_valid():
		   #una instancia a la vez
			with transaction.atomic():
				turno_seccion=seccion_form.cleaned_data["turno"]
				nombre_seccion=seccion_form.cleaned_data["seccion"]
				instancia_seccion, creado_seccion = Seccion.objects.get_or_create(
				seccion=nombre_seccion,
				turno=turno_seccion,
				grado=grado
				)
				
				alumno = alumno_form.save(commit=False)
				alumno.seccion = instancia_seccion
				alumno.save()
				instancia_evaluacion, creando_evaluacion=EvaluacionFluidezLectora.objects.get_or_create(
				alumno_id=alumno.id,cantidad_palabras_leidas=None, pregunta_1=None, pregunta_2=None, pregunta_3=None, pregunta_4=None, pregunta_5=None, pregunta_6=None, asistencia='AUSENTE',encargado_carga='DIRECTOR')
			return redirect("evaluaciones_educativas:asistencia", alumno_public_id=alumno.public_id)
			
	context = {
		'alumno_form': alumno_form,
		'grado_form': grado_form,
		'seccion_form': seccion_form,
		'grado_public':grado_public_id,
			   }
	return render(request, "alumno.html", context)

@login_required
def editar_alumno(request,alumno_public_id):
	instancia_alumno=get_object_or_404(Alumno,public_id=alumno_public_id)
	instancia_seccion=get_object_or_404(Seccion,id=instancia_alumno.seccion_id)
	instancia_grado=get_object_or_404(Grado,id=instancia_seccion.grado_id)
	alumno_form = AlumnoForm(instance=instancia_alumno)
	seccion_form = SeccionForm(instance=instancia_seccion)
	grado_form = GradoForm(instance=instancia_grado)
	if request.method == 'POST':
		alumno_form = AlumnoForm(request.POST, instance=instancia_alumno)
		grado_form = GradoForm(request.POST, instance=instancia_grado)
		seccion_form = SeccionForm(request.POST, instance=instancia_seccion)
		if alumno_form.is_valid() and grado_form.is_valid() and seccion_form.is_valid():
			with transaction.atomic():
				nombre_grado=grado_form.cleaned_data["nombre_grado"]
				cueanexo_grado=grado_form.cleaned_data["cueanexo"]
				instancia_grado, creado_grado=Grado.objects.get_or_create(
					nombre_grado=nombre_grado,
					cueanexo=cueanexo_grado
					)
				turno_seccion=seccion_form.cleaned_data["turno"]
				nombre_seccion=seccion_form.cleaned_data["seccion"]
				instancia_seccion, creado_seccion = Seccion.objects.get_or_create(
				seccion=nombre_seccion,
				turno=turno_seccion,
				grado=instancia_grado
				)
				alumno = alumno_form.save(commit=False)
				alumno.seccion = instancia_seccion
				alumno.save()
			return redirect("evaluaciones_educativas:editar_asistencia", alumno_public_id=alumno.public_id)
	context = {
		'alumno_form': alumno_form,
		 'grado_form': grado_form,
		'seccion_form': seccion_form,
		'grado_public':instancia_grado.public_id,
			   }
	return render(request, "alumno.html", context)

@login_required
def lista(request,grado_public_id): 
	instancia_grado=get_object_or_404(Grado,public_id=grado_public_id)
	instancia_seccion=Seccion.objects.filter(grado_id=instancia_grado)
	alumnos = Alumno.objects.filter(seccion_id__in=instancia_seccion).order_by('nombre')
	# print(instancia_grado)
	# print('-'*50)
	# print(instancia_seccion)
	# print('-'*50)
	# print(alumnos)
	evaluacion = EvaluacionFluidezLectora.objects.filter(alumno__in=alumnos)
	# print('-'*50)
	# print(evaluacion)
	contexto = {
		'lista_alumnos': alumnos,
		'evaluciones': evaluacion,
		'nombre_grado':instancia_grado.nombre_grado,
		'grado_public_id':instancia_grado.public_id
	}
	return render(request,"lista.html", contexto)
#-----------------lista para grados------------------
@login_required
def lista_grado(request,grado): 
	#---------logica para obtener cueanexo por medio de username--------------
	usuario= request.user
	if usuario.is_authenticated:
		name=usuario.username
	#-----logica para DNI+CUEANEXO---------
	if len(name)>9  and len(name)<=17:
		#DNI+CUEANEXO
		nombre_usuario_cueanexo=name[8:]
	else:
		nombre_usuario_cueanexo=name
	#cueanexo=int(nombre_usuario_cueanexo)   
#-----logica para DNI+CUEANEXO---------
	if grado =='SEGUNDO':
		grado='2do Año/Grado'
	elif grado =='TERCERO':
			grado='3er Año/Grado'
	try:
		instancia_grado=Grado.objects.get(cueanexo=nombre_usuario_cueanexo, nombre_grado=grado)
		return redirect("evaluaciones_educativas:lista", grado_public_id=instancia_grado.public_id)
	except Grado.DoesNotExist:
		return render(request,"lista.html")
@login_required
def grado(request):
	usuario= request.user
	if usuario.is_authenticated:
		name=usuario.username
		#-----logica para DNI+CUEANEXO---------
		if len(name)>9  and len(name)<=17:
			#DNI+CUEANEXO
			nombre_usuario_cueanexo=name[8:]
		else:
			nombre_usuario_cueanexo=name
		#cueanexo=int(nombre_usuario_cueanexo)
		#-----logica para DNI+CUEANEXO---------
		
	opciones_grado = [
					 ('2do Año/Grado', '2do Año/Grado'),
					  ('3er Año/Grado','3er Año/Grado')
						]
	grado_form_data = GradoViewForm()
	if request.method == 'POST':
		grado_form_data = GradoViewForm(request.POST)
		grado_form_data.fields['grado'].choices = opciones_grado
		if grado_form_data.is_valid():
			with transaction.atomic():
				grado=grado_form_data.cleaned_data["grado"]
				#print(grado)
				instancia_grado, creado_grado=Grado.objects.get_or_create(
					nombre_grado=grado,
					cueanexo=nombre_usuario_cueanexo
					)
				grado_public=instancia_grado.public_id
			return redirect("evaluaciones_educativas:carga_alumno", grado_public_id=grado_public)
	else:
		grado_form_data.fields['grado'].choices = opciones_grado
	contexto = {
		'grado_form_data': grado_form_data
	}
	return render(request,"grados.html", contexto)

	
@login_required
def carga_evaluacion(request, alumno_public_id):
	alumno_id=get_object_or_404(Alumno, public_id=alumno_public_id)
	instancia_seccion=get_object_or_404(Seccion,id=alumno_id.seccion_id)
	instancia_grado=get_object_or_404(Grado,id=instancia_seccion.grado_id)
	grado_public=instancia_grado.public_id
	# print(instancia_grado.nombre_grado)
	if instancia_grado.nombre_grado =='SEGUNDO':
		cantidad_palabra_maxima=170
	else:
		cantidad_palabra_maxima=211
		
	evaluacion_existente = None
	try:
		# Buscamos el examen que se creó previamente con get_or_create
		# ASUMO que tu modelo se llama EvaluacionFluidezLectora
		evaluacion_existente = EvaluacionFluidezLectora.objects.get(alumno=alumno_id.id)
	except EvaluacionFluidezLectora.DoesNotExist:
		# Si no existe, el formulario será de CREACIÓN (INSERT)
		pass

	if request.method == 'POST':
		# 3. Rama POST: Se usa la instancia para forzar la ACTUALIZACIÓN (Sobrescritura)
		
		# Al pasar 'instance=evaluacion_existente' el form sabe qué registro modificar
		# Aunque el usuario lo haya visto vacío, los datos que se guardarán son los nuevos.
		form = EvaluacionFluidezForm(request.POST, 
									 max_cantidad_palabra=cantidad_palabra_maxima,
									 instance=evaluacion_existente)
		if form.is_valid():
			with transaction.atomic():
				evaluacion = form.save(commit=False)
				evaluacion.alumno = alumno_id
				evaluacion.asistencia ='PRESENTE'
				evaluacion.encargado_carga='DIRECTOR'
				evaluacion.save()
			return redirect("evaluaciones_educativas:lista", grado_public_id=grado_public)
	else:
		#Instancia vacia para metodo get
		form = EvaluacionFluidezForm(max_cantidad_palabra=cantidad_palabra_maxima)
		#Creacion de diccionario para el Post
	context = {'form': form,
			   'alumno':alumno_id}
	return render(request, "evaluacion.html", context)

@login_required
def editar_evaluacion(request, alumno_public_id):
	alumno_id=get_object_or_404(Alumno,public_id=alumno_public_id)
	instancia_seccion=get_object_or_404(Seccion,id=alumno_id.seccion_id)
	instancia_grado=get_object_or_404(Grado,id=instancia_seccion.grado_id)
	grado_public=instancia_grado.public_id
	instancia_evaluacion=get_object_or_404(EvaluacionFluidezLectora,alumno_id=alumno_id.id)
	if instancia_grado.nombre_grado =='SEGUNDO':
		cantidad_palabra_maxima=170
	else:
		cantidad_palabra_maxima=211
	form=EvaluacionFluidezForm(instance=instancia_evaluacion, max_cantidad_palabra=cantidad_palabra_maxima)
	if request.method == 'POST':
		form=EvaluacionFluidezForm(request.POST,instance=instancia_evaluacion,max_cantidad_palabra=cantidad_palabra_maxima)
		if form.is_valid():
			with transaction.atomic():
				evaluacion=form.save(commit=False)
				evaluacion.alumno_id = alumno_id
				evaluacion.asistencia='PRESENTE'
				evaluacion.encargado_carga='DIRECTOR'
				evaluacion.save()
			return redirect("evaluaciones_educativas:lista", grado_public_id=grado_public)
	context = {
		'form': form,
		'alumno':alumno_id
		}
	return render(request, "evaluacion.html", context)

@login_required
def asistencia(request,alumno_public_id):
	alumno_id=get_object_or_404(Alumno, public_id=alumno_public_id)
	#SI instanciamos aca recibe la evalaucion que se creo en carga...
	instancia_evaluacion=get_object_or_404(EvaluacionFluidezLectora, alumno_id=alumno_id.id)
	instancia_seccion=get_object_or_404(Seccion,id=alumno_id.seccion_id)
	instancia_grado=get_object_or_404(Grado,id=instancia_seccion.grado_id)
	grado_public=instancia_grado.public_id
	if request.method == 'POST':
			form = AsistenciaForm(request.POST)
			if form.is_valid():
				with transaction.atomic():
					asistencia= form.cleaned_data["asistencia"]
					if asistencia: 
						return redirect("evaluaciones_educativas:carga_evaluacion", alumno_public_id=alumno_id.public_id)
					else:
						#llamamos a funcion ausentismo
						evaluacion=ausentismo_evaluacion(instancia_evaluacion)
						evaluacion.save()
						return redirect("evaluaciones_educativas:lista", grado_public_id=grado_public)
	else:
		form = AsistenciaForm()
	context = {'form': form,
			   'alumno':alumno_id
			   }
	return render(request,"asistencia.html",context)

@login_required
def editar_asistencia(request,alumno_public_id):
	#INSTANCIAS
	alumno_id=get_object_or_404(Alumno, public_id=alumno_public_id)
	instancia_seccion=get_object_or_404(Seccion,id=alumno_id.seccion_id)
	instancia_grado=get_object_or_404(Grado,id=instancia_seccion.grado_id)
	grado_public=instancia_grado.public_id
	if request.method == 'POST':
		form = AsistenciaForm(request.POST)
		#Campos de tabla evaluacion
		if form.is_valid():
			with transaction.atomic():
				asistencia = form.cleaned_data["asistencia"]
				#verificamos alumno presente
				if asistencia: 
					return redirect("evaluaciones_educativas:editar_evaluacion", alumno_public_id=alumno_id.public_id)
				else:
					#Recien instanciamos en el ELSE 
					instancia_evaluacion=get_object_or_404(EvaluacionFluidezLectora, alumno_id=alumno_id.id)
					evaluacion=ausentismo_evaluacion(instancia_evaluacion)
					evaluacion.save()
					return redirect("evaluaciones_educativas:lista", grado_public_id=grado_public)
	else:
		asistencia_form = AsistenciaForm()
	context = {'form': asistencia_form}
	return render(request,"asistencia.html",context)

@login_required
def borrar_registro_alumno(request,alumno_public_id):
	alumno_id=get_object_or_404(Alumno, public_id=alumno_public_id)
	instancia_seccion=get_object_or_404(Seccion,id=alumno_id.seccion_id)
	instancia_grado=get_object_or_404(Grado,id=instancia_seccion.grado_id)
	grado_public=instancia_grado.public_id
	if request.method == 'POST':
		form = BorrarRegistroAlumnoForm(request.POST)
		if form.is_valid():
			with transaction.atomic():
				eleccion= form.cleaned_data["borrar"]
				if eleccion:
					alumno_id.delete()
					return redirect("evaluaciones_educativas:lista", grado_public_id=grado_public)
				else:
					return redirect("evaluaciones_educativas:lista", grado_public_id=grado_public)
	else:
		form = BorrarRegistroAlumnoForm()
	context = {'form': form,
			   'alumno':alumno_id
			   }
	return render(request,"borrar_registro_alumno.html",context)
#DESCARGAR EXCEL
@login_required
def descargar_excel(request,grado_public_id):
	instancia_grado=get_object_or_404(Grado,public_id=grado_public_id)
	instancia_seccion=Seccion.objects.filter(grado_id=instancia_grado)
	alumnos = Alumno.objects.filter(seccion_id__in=instancia_seccion).order_by('nombre')
	evaluacion = EvaluacionFluidezLectora.objects.filter(alumno__in=alumnos)
	# 1. Configurar la respuesta HTTP para un archivo Excel
	# El 'mimetype' (o Content-Type) es crucial para que el navegador sepa que es un archivo .xlsx
	response = HttpResponse(
		content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
	)
	
	# 2. Configurar el encabezado Content-Disposition
	# Esto le dice al navegador que DEBE descargar el contenido y le asigna un nombre de archivo.
	if instancia_grado.nombre_grado=='2do Año/Grado':
		nombre_grado='2do_Grado_'
	elif instancia_grado.nombre_grado=='3er Año/Grado':
		nombre_grado='3er_Grado_'
	else:
		nombre_grado='_'
	response['Content-Disposition'] = f'attachment; filename="reporte_fluidez_{nombre_grado}noviembre_2025.xlsx"'

	# 3. Generar el contenido del Excel (lo mismo que tenías)
	wb = Workbook()
	ws = wb.active
	fecha_hora_actual = datetime.now()
	ws['A1'] = f'CUEANEXO: {instancia_grado.cueanexo}'
	ws['G1'] = f'FECHA Y HORA:  {fecha_hora_actual.strftime("%d/%m/%Y %I:%M:%S %p")}'
	lista=['NOMBRE','APELLIDO','DNI','COMUNIDAD INDíGENA','DISCAPACIDAD','GRADO','SECCIÓN','TURNO','ASISTENCIA','FLUIDEZ','P1','P2','P3','P4','P5','P6']
	#print(alumnos)
	ws.append(lista)
	for i,v in enumerate(evaluacion):
		ws[f'A{i + 3}']=v.alumno.nombre
		ws[f'B{i + 3}']=v.alumno.apellido
		ws[f'C{i + 3}']=v.alumno.dni
		ws[f'D{i + 3}']=v.alumno.comunidad_indigena
		ws[f'E{i + 3}']=v.alumno.discapacidad
		ws[f'F{i + 3}']=v.alumno.seccion.grado.nombre_grado
		ws[f'G{i + 3}']=v.alumno.seccion.seccion
		ws[f'H{i + 3}']=v.alumno.seccion.turno
		ws[f'I{i + 3}']=v.asistencia
		ws[f'J{i + 3}']=v.cantidad_palabras_leidas
		ws[f'K{i + 3}']=v.pregunta_1
		ws[f'L{i + 3}']=v.pregunta_2
		ws[f'M{i + 3}']=v.pregunta_3
		ws[f'N{i + 3}']=v.pregunta_4
		ws[f'O{i + 3}']=v.pregunta_5
		ws[f'P{i + 3}']=v.pregunta_6

	wb.save(response)
	# 5. Retornar la respuesta al navegador
	return response

# @login_required
# def monitoreo(request):
#     instancia_grado_cueanexo=Grado.objects.all()
#     # instancia_grado=Grado.objects.filter(cueanexo__in=instancia_grado_cueanexo).values_list('nombre_grado',flat=True)
#     # # for i in instancia_grado:
#     # #     print(i)
#     contexto={'grados':instancia_grado_cueanexo}
#     return render(request,"monitoreo.html", contexto)

#-----------------------LOGICA PARA VISUALIZAR DATOS ---------------------------

@login_required
def analisis_evaluaciones_noviembre_2025(request):
	contexto = {
		'alumnos_evaluados_segundo': [],
		'alumnos_evaluados_tercero': [],
		'grado': None,
		'seccion': None,
		'form_grado': None,
		'form_seccion': None,
		'rol':None,
		'listar':True,
	}
	
	usuario = request.user
	cuil = usuario.username
	# Formateo de CUIL
	#cuil_con_caracter = f"{cuil[:2]}-{cuil[2:10]}-{cuil[10:]}"
	nivel_acceso=request.user.nivelacceso_id
	contexto['rol']=nivel_acceso
	# Inicializamos el primer formulario siempre
	form_cueanexo = CueanexoForm(request.POST or None, cuil=cuil,nivel_acceso=nivel_acceso)
	contexto["form_cueanexo"] = form_cueanexo
	lista_cueanexos=[]
	if request.method == 'POST':
		if form_cueanexo.is_valid():
			cueanexo = form_cueanexo.cleaned_data['cueanexo_seleccionado']
			#print(f'cueanexo intento{cueanexo}')
			# Inicializamos el form de Grado (aquí podrías pasarle el cueanexo si el form fuera dinámico)
			form_grado = Grado_select_Form(request.POST or None)
			contexto["form_grado"] = form_grado
			
			if form_grado.is_valid():
				nombre_grado = form_grado.cleaned_data['grado_seleccion']
				
				if nombre_grado:
					# transformar varibale cueanexo a lista 
					lista_cueanexos.append(cueanexo)
					# Usamos filter().first() para evitar errores si no existe
					grado_obj = Grado.objects.filter(nombre_grado=nombre_grado, cueanexo__in=lista_cueanexos)
					#print(len(grado_obj))
					#grado_obj = Grado.objects.filter(nombre_grado=nombre_grado, cueanexo__in=lista_cueanexos).first()
					if grado_obj:
						#print('entre')
						# Inicializamos sección con el ID del grado
						form_seccion = SeccionTurnoForm(request.POST or None, grados=grado_obj)
						#form_seccion = SeccionTurnoForm(request.POST or None, grados=grado_obj.id)
						contexto["form_seccion"] = form_seccion
						
						secciones = None
						turno = None
						
						if form_seccion.is_valid():
							seleccionado = form_seccion.cleaned_data['seleccion']
							if seleccionado != 'TODOS':
								# 'seleccionado' aquí suele ser el objeto Seccion si es ModelChoiceField
								secciones = seleccionado.seccion
								turno = seleccionado.turno

						# CARGA DE DATOS: Se ejecuta si hay grado_obj, independientemente de la sección
						if nombre_grado == '2do Año/Grado':
						#if grado_obj.nombre_grado == '2do Año/Grado':
							datos = analisis_segundo_grado_grafico(nombre_grado, lista_cueanexos, secciones, turno)
							contexto["grado"] = "segundo"
						else:
							datos = analisis_tercer_grado_grafico(nombre_grado, lista_cueanexos, secciones, turno)
							contexto["grado"] = "tercero"
						
						contexto.update(datos)
						contexto["seccion"] = secciones
						contexto["turno"] = turno
				else:
					contexto['grado'] = None
	nombre_archivo = 'evaluaciones_educativas/pdf/INFORME-FLUIDEZ-LECTORA-NOVIEMBRE2025.pdf'
# Generamos la URL y le pegamos los parámetros del visor
	contexto["material_pdf"] = static(nombre_archivo)
	return render(request, "analisis_evaluaciones_noviembre_2025.html", contexto)
#------------------logica para regional---------------------------
@login_required
def analisis_evaluaciones_regional_noviembre_2025(request):
	contexto = {
		'alumnos_evaluados_segundo': [],
		'alumnos_evaluados_tercero': [],
		'grado': None,
		'seccion': None,
		'form_grado': None,
		'form_seccion': None,
		'rol':None,
		'listar':None,
	}
	
	usuario = request.user
	cuil = usuario.username
	# Formateo de CUIL
	#cuil_con_caracter = f"{cuil[:2]}-{cuil[2:10]}-{cuil[10:]}"
	nivel_acceso=request.user.nivelacceso_id
	#print(f'REGIONALLLLL{nivel_acceso}')
	contexto['rol']=nivel_acceso
	form_director_regional=DirectorForm(request.POST or None)
	# Inicializamos el primer formulario siempre
	lista_cueanexos=[]
	if request.method == 'POST':
		if form_director_regional.is_valid():
			sector=form_director_regional.cleaned_data['sector']
			ambito=form_director_regional.cleaned_data['ambito']
			form_cueanexo = CueanexoForm(request.POST or None, cuil=cuil,nivel_acceso=nivel_acceso,sector=sector,ambito=ambito)
			contexto["form_cueanexo"] = form_cueanexo
			if form_cueanexo.is_valid():
				cueanexo = form_cueanexo.cleaned_data['cueanexo_seleccionado']
				#print(f'cueanexo intento{cueanexo}')
				if cueanexo == 'TODOS' and nivel_acceso !='Director/a':
					# Esto te devuelve la lista de tuplas (valor, etiqueta)
					lista_cueanexos = [valor for valor, etiqueta in form_cueanexo.fields['cueanexo_seleccionado'].choices]
					#print(f'lista{len(lista_cueanexos)}')
				else:
					lista_cueanexos.append(cueanexo)
					contexto['listar']=True
				# Inicializamos el form de Grado (aquí podrías pasarle el cueanexo si el form fuera dinámico)
				form_grado = Grado_select_Form(request.POST or None)
				contexto["form_grado"] = form_grado
				
				if form_grado.is_valid():
					nombre_grado = form_grado.cleaned_data['grado_seleccion']
					
					if nombre_grado:
						# transformar varibale cueanexo a lista 
						# lista_cueanexos=[cueanexo]
						# Usamos filter().first() para evitar errores si no existe
						grado_obj = Grado.objects.filter(nombre_grado=nombre_grado, cueanexo__in=lista_cueanexos)
						#print(len(grado_obj))
						#grado_obj = Grado.objects.filter(nombre_grado=nombre_grado, cueanexo__in=lista_cueanexos).first()
						if grado_obj:
							# Inicializamos sección con el ID del grado
							form_seccion = SeccionTurnoForm(request.POST or None, grados=grado_obj)
							#form_seccion = SeccionTurnoForm(request.POST or None, grados=grado_obj.id)
							contexto["form_seccion"] = form_seccion
							
							secciones = None
							turno = None
							
							if form_seccion.is_valid():
								seleccionado = form_seccion.cleaned_data['seleccion']
								if seleccionado != 'TODOS':
									# 'seleccionado' aquí suele ser el objeto Seccion si es ModelChoiceField
									secciones = seleccionado.seccion
									turno = seleccionado.turno

							# CARGA DE DATOS: Se ejecuta si hay grado_obj, independientemente de la sección
							if nombre_grado == '2do Año/Grado':
							#if grado_obj.nombre_grado == '2do Año/Grado':
								datos = analisis_segundo_grado_grafico(nombre_grado, lista_cueanexos, secciones, turno)
								contexto["grado"] = "segundo"
							else:
								datos = analisis_tercer_grado_grafico(nombre_grado, lista_cueanexos, secciones, turno)
								contexto["grado"] = "tercero"
							
							contexto.update(datos)
							contexto["seccion"] = secciones
							contexto["turno"] = turno
					else:
						contexto['grado'] = None

	contexto["form_director_regional"] = form_director_regional
	nombre_archivo = 'evaluaciones_educativas/pdf/INFORME-FLUIDEZ-LECTORA-NOVIEMBRE2025.pdf'
# Generamos la URL y le pegamos los parámetros del visor
	contexto["material_pdf"] = static(nombre_archivo)
	return render(request, "analisis_evaluaciones_noviembre_2025.html", contexto)
#--------------------fin de logica para regional----------------
#---------------------logica para SUBSE Y MINISTRO---------------------------
@login_required
def analisis_evaluaciones_ministros_noviembre_2025(request):
	contexto = {
		'alumnos_evaluados_segundo': [],
		'alumnos_evaluados_tercero': [],
		'grado': None,
		'seccion': None,
		'form_grado': None,
		'form_seccion': None,
		'rol':None,
		'listar':None,
	}
	
	usuario = request.user
	cuil = usuario.username
	# Formateo de CUIL
	#cuil_con_caracter = f"{cuil[:2]}-{cuil[2:10]}-{cuil[10:]}"
	nivel_acceso=request.user.nivelacceso_id
	#print(f'REGIONALLLLL{nivel_acceso}')
	contexto['rol']=nivel_acceso
	# Inicializamos el primer formulario siempre
	#print(nivel_acceso)
	form_director_nivel=DirectorNivelForm(request.POST or None)
	#form_cueanexo = CueanexoForm(request.POST or None, cuil=cuil,nivel_acceso=nivel_acceso)
	lista_cueanexos=[]
	if request.method == 'POST':
		if form_director_nivel.is_valid():
				sector=form_director_nivel.cleaned_data['sector']
				ambito=form_director_nivel.cleaned_data['ambito']
				region = form_director_nivel.cleaned_data['region']
				form_cueanexo =  CueanexoForm(request.POST or None, cuil=cuil,nivel_acceso=nivel_acceso,sector=sector,ambito=ambito,region=region)
				contexto["form_cueanexo"] = form_cueanexo    
				if form_cueanexo.is_valid():
					cueanexo = form_cueanexo.cleaned_data['cueanexo_seleccionado']
					#print(f'cueanexo intento{cueanexo}')
					if cueanexo == 'TODOS' and nivel_acceso !='Director/a':
						# Esto te devuelve la lista de tuplas (valor, etiqueta)
						lista_cueanexos = [valor for valor, etiqueta in form_cueanexo.fields['cueanexo_seleccionado'].choices]
						#print(f'lista{len(lista_cueanexos)}')
					else:
						lista_cueanexos.append(cueanexo)
						contexto['listar']=True
					# Inicializamos el form de Grado (aquí podrías pasarle el cueanexo si el form fuera dinámico)
					form_grado = Grado_select_Form(request.POST or None)
					contexto["form_grado"] = form_grado
					
					if form_grado.is_valid():
						nombre_grado = form_grado.cleaned_data['grado_seleccion']
						
						if nombre_grado:
							# transformar varibale cueanexo a lista 
							# lista_cueanexos=[cueanexo]
							# Usamos filter().first() para evitar errores si no existe
							grado_obj = Grado.objects.filter(nombre_grado=nombre_grado, cueanexo__in=lista_cueanexos)
							#print(len(grado_obj))
							#grado_obj = Grado.objects.filter(nombre_grado=nombre_grado, cueanexo__in=lista_cueanexos).first()
							if grado_obj:
								# Inicializamos sección con el ID del grado
								form_seccion = SeccionTurnoForm(request.POST or None, grados=grado_obj)
								#form_seccion = SeccionTurnoForm(request.POST or None, grados=grado_obj.id)
								contexto["form_seccion"] = form_seccion
								
								secciones = None
								turno = None
								
								if form_seccion.is_valid():
									seleccionado = form_seccion.cleaned_data['seleccion']
									if seleccionado != 'TODOS':
										# 'seleccionado' aquí suele ser el objeto Seccion si es ModelChoiceField
										secciones = seleccionado.seccion
										turno = seleccionado.turno

								# CARGA DE DATOS: Se ejecuta si hay grado_obj, independientemente de la sección
								if nombre_grado == '2do Año/Grado':
								#if grado_obj.nombre_grado == '2do Año/Grado':
									datos = analisis_segundo_grado_grafico(nombre_grado, lista_cueanexos, secciones, turno)
									contexto["grado"] = "segundo"
								else:
									datos = analisis_tercer_grado_grafico(nombre_grado, lista_cueanexos, secciones, turno)
									contexto["grado"] = "tercero"
								
								contexto.update(datos)
								contexto["seccion"] = secciones
								contexto["turno"] = turno
						else:
							contexto['grado'] = None
							
	contexto['form_director_nivel']=form_director_nivel
	#contexto["form_cueanexo"] = form_cueanexo
	nombre_archivo = 'evaluaciones_educativas/pdf/INFORME-FLUIDEZ-LECTORA-NOVIEMBRE2025.pdf'
# Generamos la URL y le pegamos los parámetros del visor
	contexto["material_pdf"] = static(nombre_archivo)
	return render(request, "analisis_evaluaciones_noviembre_2025.html", contexto)
#----------------------FIN SUBSE Y MINISTRO--------------------------------
#-----------------------FIN LOGICA PARA VISUALIZAR DATOS ---------------------------

def ausentismo_evaluacion(instancia_evaluacion):
	evaluacion_campos=instancia_evaluacion._meta.fields
	for i in evaluacion_campos:
		if i.name == 'asistencia':
			#FUNCION DE PYTHON para estabalecer valores a campos de un objeto
			setattr(instancia_evaluacion, i.name, 'AUSENTE')
			#verificamos campos PK y NOT NULL
		if not i.primary_key and i.null:
			setattr(instancia_evaluacion, i.name, None)
	return instancia_evaluacion

#---------LOGICA PARA ANALISIS DE LAS EVALUACIONES EDUCATIVAS--------------------------

def analisis_segundo_grado_grafico(grado_seleccionado,cueanexo,secciones ,turno):
#-----------------------LOGICA PARA SEGUNDO GRADO------------------------
	#print('entre a analisis 2do')
	if type(cueanexo) is str:
		grado=Grado.objects.filter(nombre_grado__icontains=grado_seleccionado, cueanexo=cueanexo)
	else:
		#print('ENTREEE')
		grado=Grado.objects.filter(nombre_grado__icontains=grado_seleccionado, cueanexo__in=cueanexo)
	#print(grado)
	if secciones and turno:
		secciones=Seccion.objects.filter(grado__in=grado, seccion=secciones,turno=turno)
	#print(seccion)
	else:
		secciones=Seccion.objects.filter(grado__in=grado)
	alumnos=Alumno.objects.filter(seccion_id__in=secciones)
	evaluaciones = EvaluacionFluidezLectora.objects.filter(alumno__in=alumnos)
	#comprension lectora
	#presentes=evaluaciones.filter(asistencia='PRESENTE')
	#POSIBLE OLUCION
	evaluaciones_con_comprension,conteos=comprension_lectora(alumnos,grado_seleccionado)
	datos = evaluaciones.aggregate(
	# Asistencia
	presentes=Count('alumno_id', filter=Q(asistencia='PRESENTE')),
	ausentes=Count('alumno_id', filter=Q(asistencia='AUSENTE')),
	
	# Desempeño
	promedio_general=Avg('cantidad_palabras_leidas',filter=Q(asistencia='PRESENTE')),
	debajo_del_basico=Count('alumno_id', filter=Q(asistencia='PRESENTE',cantidad_palabras_leidas__lt=21)),
	basico=Count('alumno_id', filter=Q(asistencia='PRESENTE',cantidad_palabras_leidas__gte=21, cantidad_palabras_leidas__lte=46)),
	satisfactorio=Count('alumno_id', filter=Q(asistencia='PRESENTE',cantidad_palabras_leidas__gt=46, cantidad_palabras_leidas__lte=70)),
	avanzado=Count('alumno_id', filter=Q(asistencia='PRESENTE',cantidad_palabras_leidas__gt=70))
	)
	#print(datos['debajo_del_basico'])
	alumnos_presentes=datos['presentes']
	alumnos_ausentes=datos['ausentes']
	#SOLUCIONAR LAS DIVISIONES POR 0
	nivel_debajo_del_basico = round((datos['debajo_del_basico']/alumnos_presentes)*100,1)if alumnos_presentes else 0
	nivel_basico = round((datos['basico']/alumnos_presentes)*100, 1)if alumnos_presentes else 0
	nivel_satisfactorio=round((datos['satisfactorio']/alumnos_presentes)*100, 1)if alumnos_presentes else 0
	nivel_avanzado=round((datos['avanzado']/alumnos_presentes)*100, 1)if alumnos_presentes else 0
	#INDICE DE PROMEDIO (ENTERO)
	nivel_debajo_del_basico_comprension = round((conteos['debajo_del_basico']/alumnos_presentes)*100,1)if alumnos_presentes else 0
	nivel_basico_comprension = round((conteos['basico']/alumnos_presentes)*100, 1)if alumnos_presentes else 0
	nivel_satisfactorio_comprension=round((conteos['satisfactorio']/alumnos_presentes)*100, 1)if alumnos_presentes else 0
	nivel_avanzado_comprension=round((conteos['avanzado']/alumnos_presentes)*100, 1)if alumnos_presentes else 0
	promedio = datos['promedio_general']#Modificar para trabajar mejor los decimales
	#print(promedio)
	#print(evaluacion)
	contexto={
	'CUEANEXO':cueanexo,
	'nivel_educativo':'Primario',
	'alumnos_evaluados_segundo': evaluaciones_con_comprension,
	'asistencia_segundo': [f'Presentes {alumnos_presentes}', f'Ausentes {alumnos_ausentes}'],
	'valores_asistencia_segundo': [alumnos_presentes, alumnos_ausentes],
	'etiquetas_nivel_desempeno_segundo': [f'Debajo del Basico {nivel_debajo_del_basico}%', f'Básico {nivel_basico}%',f'Satisfactorio {nivel_satisfactorio}%',f'Avanzado {nivel_avanzado}%'],
	'valores_desempeno_segundo': [nivel_debajo_del_basico, nivel_basico, nivel_satisfactorio, nivel_avanzado],
	'etiquetas_nivel_desempeno_comprension_segundo': [f'Debajo del Basico {nivel_debajo_del_basico_comprension}%', f'Básico {nivel_basico_comprension}%',f'Satisfactorio {nivel_satisfactorio_comprension}%',f'Avanzado {nivel_avanzado_comprension}%'],
	'valores_desempeno_comprension_segundo': [nivel_debajo_del_basico_comprension, nivel_basico_comprension, nivel_satisfactorio_comprension, nivel_avanzado_comprension],
	'etiqueta_promedios_segundo': 'Promedio de Palabras Leídas Correctamente por Minuto',
	'valor_promedio_segundo': [promedio],
	}
	return contexto

		#----------------FIN LOGICA 2 GRADO ---------

def analisis_tercer_grado_grafico(grado_seleccionado,cueanexo,secciones,turno):
#-----------------------LOGICA PARA TERCER GRADO------------------------
	#PARTICIPACION
	#print('entre a analisis 3ero')
	
		#cueanexo=['220000200','220009000']
	if type(cueanexo) is str:
		grado=Grado.objects.filter(nombre_grado__icontains=grado_seleccionado, cueanexo=cueanexo)
	else:
		#print('ENTREEE')
		grado=Grado.objects.filter(nombre_grado__icontains=grado_seleccionado, cueanexo__in=cueanexo)
	#print(grado)
	if secciones and turno:
		secciones=Seccion.objects.filter(grado__in=grado, seccion=secciones,turno=turno)
	#print(seccion)
	else:
		secciones=Seccion.objects.filter(grado__in=grado)
	alumnos=Alumno.objects.filter(seccion_id__in=secciones)

	evaluaciones = EvaluacionFluidezLectora.objects.filter(alumno__in=alumnos)
	#presentes=evaluaciones.filter(asistencia='PRESENTE')
	#POSIBLE OLUCION
	evaluaciones_con_comprension,conteos=comprension_lectora(alumnos,grado_seleccionado)
	#evaluaciones_con_comprension,conteos=comprension_lectora(evaluaciones,grado_seleccionado)
	datos = evaluaciones.aggregate(
	# Asistencia
	presentes=Count('alumno_id', filter=Q(asistencia='PRESENTE')),
	ausentes=Count('alumno_id', filter=Q(asistencia='AUSENTE')),
	
	# Desempeño
	promedio_general=Avg('cantidad_palabras_leidas',filter=Q(asistencia='PRESENTE')),
	debajo_del_basico=Count('alumno_id', filter=Q(asistencia='PRESENTE',cantidad_palabras_leidas__lt=30)),
	basico=Count('alumno_id', filter=Q(asistencia='PRESENTE',cantidad_palabras_leidas__gte=30, cantidad_palabras_leidas__lte=60)),
	satisfactorio=Count('alumno_id', filter=Q(asistencia='PRESENTE',cantidad_palabras_leidas__gt=61, cantidad_palabras_leidas__lte=90)),
	avanzado=Count('alumno_id', filter=Q(asistencia='PRESENTE',cantidad_palabras_leidas__gt=90))
	)
	alumnos_presentes=datos['presentes']
	alumnos_ausentes=datos['ausentes']
	nivel_debajo_del_basico = round((datos['debajo_del_basico']/alumnos_presentes)*100, 1)if alumnos_presentes else 0
	nivel_basico = round((datos['basico']/alumnos_presentes)*100, 1)if alumnos_presentes else 0
	nivel_satisfactorio=round((datos['satisfactorio']/alumnos_presentes)*100, 1)if alumnos_presentes else 0
	nivel_avanzado=round((datos['avanzado']/alumnos_presentes)*100, 1)if alumnos_presentes else 0
	#grafico
	nivel_debajo_del_basico_comprension = round((conteos['debajo_del_basico']/alumnos_presentes)*100,1)if alumnos_presentes else 0
	nivel_basico_comprension = round((conteos['basico']/alumnos_presentes)*100, 1)if alumnos_presentes else 0
	nivel_satisfactorio_comprension=round((conteos['satisfactorio']/alumnos_presentes)*100, 1)if alumnos_presentes else 0
	nivel_avanzado_comprension=round((conteos['avanzado']/alumnos_presentes)*100, 1)if alumnos_presentes else 0
	#INDICE DE PROMEDIO (ENTERO)
	promedio = datos['promedio_general']#Modificar para trabajar mejor los decimales
	#print(promedio)
	#print(evaluacion)
	contexto={
	'CUEANEXO':cueanexo,
	'nivel_educativo':'Primario',
	'alumnos_evaluados_tercero': evaluaciones_con_comprension,
	'asistencia_tercero': [f'Presentes {alumnos_presentes}', f'Ausentes {alumnos_ausentes}'],
	'valores_asistencia_tercero': [alumnos_presentes, alumnos_ausentes],
	'etiquetas_nivel_desempeno_tercero': [f'Debajo del Basico {nivel_debajo_del_basico}%', f'Básico {nivel_basico}%',f'Satisfactorio {nivel_satisfactorio}%',f'Avanzado {nivel_avanzado}%'],
	'valores_desempeno_tercero': [nivel_debajo_del_basico, nivel_basico, nivel_satisfactorio, nivel_avanzado],
	'etiquetas_nivel_desempeno_comprension_tercero': [f'Debajo del Basico {nivel_debajo_del_basico_comprension}%', f'Básico {nivel_basico_comprension}%',f'Satisfactorio {nivel_satisfactorio_comprension}%',f'Avanzado {nivel_avanzado_comprension}%'],
	'valores_desempeno_comprension_tercero': [nivel_debajo_del_basico_comprension, nivel_basico_comprension, nivel_satisfactorio_comprension, nivel_avanzado_comprension],
	'etiqueta_promedios_tercero': 'Promedio de Palabras Leídas Correctamente por Minuto',
	'valor_promedio_tercero': [promedio],}
	return contexto


#Crear funcion para logica de COMPRENSION LECTORA, 
def comprension_lectora(alumnos,grado_seleccionado):
	#print(type(grado_seleccionado))
	#print('en comprension')
	#resultados = [] # Usamos una lista para guardar a todos
	evaluaciones=EvaluacionFluidezLectora.objects.filter(alumno__in=alumnos, asistencia='PRESENTE')
	conteos = {'debajo_del_basico': 0, 'basico': 0, 'satisfactorio': 0, 'avanzado': 0}
	if grado_seleccionado == '2do Año/Grado':
		#print('entre en 2 comprensio')
		for i in evaluaciones:
			# Reiniciamos los puntos para CADA alumno dentro del bucle
			p1 = 1 if i.pregunta_1 == 'B' else 0
			p2 = 1 if i.pregunta_2 == 'C' else 0
			p3 = 1 if i.pregunta_3 == 'B' else 0
			
			# IMPORTANTE: Usamos punto para el decimal (1.5)
			p4 = 1.50 if i.pregunta_4 == 'A' else 0
			p5 = 1.50 if i.pregunta_5 == 'B' else 0
			p6 = 1.50 if i.pregunta_6 == 'C' else 0
			#print(f'{i.pregunta_1}{i.pregunta_2}{i.pregunta_3}{i.pregunta_4}{i.pregunta_5}{i.pregunta_6}')
			puntaje_total = p1 + p2 + p3 + p4 + p5 + p6
			i.puntaje_comprension = puntaje_total
			if puntaje_total < 3.40:
				conteos['debajo_del_basico'] += 1
			elif puntaje_total <= 5.20:
				conteos['basico'] += 1
			elif puntaje_total <= 6.75:
				conteos['satisfactorio'] += 1
			else:
				conteos['avanzado'] += 1

			#print(f'{i.puntaje_comprension} alumno:{i.alumno.nombre}' )
	else:
		for i in evaluaciones:
		# Reiniciamos los puntos para CADA alumno dentro del bucle
			p1 = 1 if i.pregunta_1 == 'B' else 0
			p2 = 1 if i.pregunta_2 == 'C' else 0
			p3 = 1 if i.pregunta_3 == 'C' else 0
			
			# IMPORTANTE: Usamos punto para el decimal (1.5)
			p4 = 1.50 if i.pregunta_4 == 'A' else 0
			p5 = 1.50 if i.pregunta_5 == 'C' else 0
			p6 = 1.50 if i.pregunta_6 == 'C' else 0
			#print(f'{i.pregunta_1}{i.pregunta_2}{i.pregunta_3}{i.pregunta_4}{i.pregunta_5}{i.pregunta_6}')
			puntaje_total = p1 + p2 + p3 + p4 + p5 + p6
			if puntaje_total < 3.40:
				conteos['debajo_del_basico'] += 1
			elif puntaje_total <= 5.20:
				conteos['basico'] += 1
			elif puntaje_total <= 6.75:
				conteos['satisfactorio'] += 1
			else:
				conteos['avanzado'] += 1
			
			# Guardamos un diccionario por cada alumno
			i.puntaje_comprension = puntaje_total
			#print(f'{i.puntaje_comprension} alumno:{i.alumno.nombre}' )
	
	return evaluaciones, conteos # Devolvemos la lista completa


#-----------------------LOGICA PARA VISUALIZAR DATOS ---------------------------

@login_required
def analisis_evaluaciones_mayo_2025(request):
	contexto = {
		'alumnos_evaluados_segundo': [],
		'alumnos_evaluados_tercero': [],
		'grado': None,
		'seccion': None,
		'form_cueanexo': None,
		'form_grado': None,
		'form_seccion': None,
		'datos':None,
	}
	usuario = request.user
	cuil = usuario.username
	# Formateo de CUIL
	cuil_con_caracter = f"{cuil[:2]}-{cuil[2:10]}-{cuil[10:]}"
	#------------------------abri psycopg2-----------------
	db_params = {
		"host": os.getenv('POSTGRES_HOST'),
		"database": os.getenv('POSTGRES_DB'),
		"user": os.getenv('POSTGRES_USER'),
		"password": os.getenv('POSTGRES_PASSWORD'),
		"port": os.getenv('POSTGRES_PORT'),
		"options":"-c search_path=operativoschaco",
	}
	conn = None # La definimos fuera para poder cerrarla en el 'finally'
	qs = CapaUnicaOfertas.objects.filter(resploc_cuitcuil=cuil_con_caracter, oferta='Común - Primaria de 7 años ').only('cueanexo')
	lista_cueanexos_validos=[]
	for i in qs:
		lista_cueanexos_validos.append(f'{i.cueanexo}')
	cueanexo_tuplas=tuple(lista_cueanexos_validos)
	try:
		# 1. Establecer la conexión
		conn = psycopg2.connect(**db_params)
		
		# 2. Crear el cursor (usamos RealDictCursor para traer nombres de columnas)
		with conn.cursor(cursor_factory=extras.RealDictCursor) as cur:
			if lista_cueanexos_validos:
				query_secciones=f"""
						SELECT DISTINCT cueanexo FROM operativoschaco."Examen_Fluidez_Segundo" WHERE cueanexo in %s
						UNION
						SELECT DISTINCT cueanexo FROM operativoschaco."Examen_Fluidez_Tercero" WHERE cueanexo in %s
						ORDER BY cueanexo
						"""
				cur.execute(query_secciones,(cueanexo_tuplas, cueanexo_tuplas))
				cueanexos = cur.fetchall()
		# Inicializamos el primer formulario siempre
				form_cueanexo = CueanexoManualForm(request.POST or None, opciones=cueanexos)
				contexto["form_cueanexo"] = form_cueanexo
				if request.method == 'POST':
					if form_cueanexo.is_valid():
						cueanexo = form_cueanexo.cleaned_data['cueanexo_seleccionado']
						cuenexo_comilla=f"{cueanexo}"
						# Inicializamos el form de Grado (aquí podrías pasarle el cueanexo si el form fuera dinámico)
						form_grado = GradoManualForm(request.POST or None)
						contexto["form_grado"] = form_grado
						
						if form_grado.is_valid():
							nombre_grado = form_grado.cleaned_data['grado_seleccion']
							#print(nombre_grado)
							if nombre_grado:# ACA DEBERIAMOS DECIRLE SI ES 2 O 3
								query_secciones=f"""
								SELECT distinct(division)
								FROM operativoschaco."Examen_Fluidez_{nombre_grado}"
								WHERE cueanexo=%s
								ORDER BY division
								"""
								cur.execute(query_secciones,[cuenexo_comilla])
								secciones = cur.fetchall()
								form_seccion = SeccionTurnoManualForm(request.POST or None, opciones=secciones)
								contexto["form_seccion"] = form_seccion
								
								secciones = None
								#turno = None
									
								if form_seccion.is_valid():
									seleccionado = form_seccion.cleaned_data['seleccion']
									#print(seleccionado)
									if seleccionado != 'TODOS':
										# 'seleccionado' aquí suele ser el objeto Seccion si es ModelChoiceField
										secciones = seleccionado
										# turno = seleccionado.turno
								#print(secciones)
								# CARGA DE DATOS: Se ejecuta si hay grado_obj, independientemente de la sección
								if nombre_grado == 'Segundo':
									obtener_consulta = analisis_segundo_grado_mayo_2025_grafico(cueanexo, secciones)
									cur.execute(obtener_consulta['consulta'])
									datos = cur.fetchall()
									contexto["grado"] = "Segundo Grado"
								else:
									obtener_consulta = analisis_tercer_grado_mayo_2025_grafico(cueanexo, secciones)
									cur.execute(obtener_consulta['consulta'])
									datos = cur.fetchall()
									contexto["grado"] = "Tercer Grado"
							
								contexto["datos"]= datos
								contexto["seccion"] = secciones
						else:
							contexto['grado'] = None
	except (Exception, psycopg2.DatabaseError) as error:
		# 5. Manejo de errores
		print(f"Error al conectar o consultar: {error}")

	finally:
		# 6. Cerrar la conexión pase lo que pase
		if conn is not None:
			conn.close()
			print("Conexión cerrada.")            
	#------------------------CERRAR psycopg2-----------------
	#print(contexto)
	return render(request, "analisis_evaluaciones_mayo_2025.html", contexto)


def analisis_segundo_grado_mayo_2025_grafico(cueanexo,secciones):
	# print('entreee')
	contexto={}
	query = f"""
	SELECT *,
			CASE 
				WHEN velocidad >= 0 and velocidad <= 20  THEN 'Por debajo del básico'
				WHEN velocidad > 20 and velocidad <= 42  THEN 'Básico'
				WHEN velocidad > 42 and velocidad <= 51  THEN 'Satisfactorio'
				WHEN velocidad > 51 and velocidad <= 53  THEN 'Avanzado'
				ELSE 'Calificacion Incorrecta'
			END AS "Desempeño_Velocidad",
				CASE 
				
				WHEN "precision" >= 0 and "precision" <= 30  THEN 'Por debajo del básico'
				WHEN "precision" > 30 and "precision" <= 59  THEN 'Básico'
				WHEN "precision" > 59 and "precision" <= 73  THEN 'Satisfactorio'
				WHEN "precision" > 73 and "precision" <= 76  THEN 'Avanzado'
				ELSE 'Calificacion Incorrecta'
			END AS "Desempeño_Precisión",
			CASE
				WHEN prosodia = '1'  THEN 'Por debajo del básico'
				WHEN prosodia = '2' THEN 'Básico'
				WHEN prosodia = '3' THEN 'Satisfactorio'
				ELSE 'Avanzado'
			END AS "Desempeño_prosodia"
	FROM "Examen_Fluidez_Segundo"
	WHERE CUEANEXO='{cueanexo}'
	"""
	#print(secciones)
	if secciones:
		seccion_seleccionada=f"AND division='{secciones}'"
		query=query + seccion_seleccionada
	contexto['consulta']=query
	return contexto



def analisis_tercer_grado_mayo_2025_grafico(cueanexo,secciones):
	contexto={}

	query = f"""
	SELECT *,
			CASE 
				WHEN velocidad >= 0 and velocidad <= 34  THEN 'Por debajo del básico'
				WHEN velocidad > 34 and velocidad <= 69  THEN 'Básico'
				WHEN velocidad > 69 and velocidad <= 86  THEN 'Satisfactorio'
				WHEN velocidad > 86 and velocidad <= 88  THEN 'Avanzado'
				ELSE 'Calificacion Incorrecta'
			END AS "Desempeño_Velocidad",
				CASE 
				
				WHEN "precision" >= 0 and "precision" <= 36  THEN 'Por debajo del básico'
				WHEN "precision" > 36 and "precision" <= 73  THEN 'Básico'
				WHEN "precision" > 73 and "precision" <= 91  THEN 'Satisfactorio'
				WHEN "precision" > 91 and "precision" <= 93  THEN 'Avanzado'
				ELSE 'Calificacion Incorrecta'
			END AS "Desempeño_Precisión",
			CASE
				WHEN prosodia = '1'  THEN 'Por debajo del básico'
				WHEN prosodia = '2' THEN 'Básico'
				WHEN prosodia = '3' THEN 'Satisfactorio'
				ELSE 'Avanzado'
			END AS "Desempeño_prosodia"
	FROM "Examen_Fluidez_Tercero"
	WHERE CUEANEXO='{cueanexo}'
	"""
	if secciones:
		seccion_seleccionada=f"AND division='{secciones}'"
		query=query + seccion_seleccionada
	contexto['consulta']=query
	return contexto
#-----------------------FIN LOGICA PARA VISUALIZAR DATOS ---------------------------
#-----------------------Inicio LOGICA DIRECTORES REGIONALES PARA VISUALIZAR DATOS ---------------------------
# def obtener_regional(cuil):
#      #consutla a select * from public.usuarios_regionalusuarios where usuario=cuil_con_caracter
#     try:
#         # 1. Establecer la conexión
#         conn = psycopg2.connect(**db_params)
		
#         # 2. Crear el cursor (usamos RealDictCursor para traer nombres de columnas)
#         with conn.cursor(cursor_factory=extras.RealDictCursor) as cur:
#             if cuil:
#                 cuil_string=f'{cuil}'
#                 query_regiones=f"""
#                         SELECT region_loc
#                         FROM public.usuarios_regionalusuarios
#                         WHERE usuario= %s
#                         """
#                 cur.execute(query_regiones,(cuil_string))
#                 regional = cur.fetchall()
#     except (Exception, psycopg2.DatabaseError) as error:
#     # 5. Manejo de errores
#         print(f"Error al conectar o consultar: {error}")

#     finally:
#         # 6. Cerrar la conexión pase lo que pase
#         if conn is not None:
#             conn.close()
#             print("Conexión cerrada.") 

#     return regional
#-----------------------FIN LOGICA DIRECTORES REGIONALES PARA VISUALIZAR DATOS ---------------------------  


