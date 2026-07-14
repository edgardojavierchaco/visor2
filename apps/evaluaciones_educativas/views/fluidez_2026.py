from apps.evaluaciones_educativas.models.fluidez_2026 import *
from apps.evaluaciones_educativas.forms.fluidez_2026 import *
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse,JsonResponse
from django.core.paginator import Paginator
from django.db import transaction
from django.contrib.auth.decorators import login_required
from django.templatetags.static import static
from django.db.models import Count,Q,Avg
from datetime import date, datetime
import psycopg2
from psycopg2 import extras
import os
from openpyxl import Workbook
import json
#from apps.consultasge.models import CapaUnicaOfertas
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.urls import reverse
from urllib.parse import urlencode # Necesario para armar los parámetros de la URL


# @login_required
# def inicio(request):
# 	#-------------OBTENER USERNAME------------
# 	# print(request.user)
# 	# print(request.name)
# 		#-------------OBTENER USERNAME------------
# 	cuil=27308542489
# 	grado=None
# 	cueanexo_form=CueanexoFluidez2026ViewForm(request.POST or None,cuil=cuil)
# 	grado_form=GradoFluidez2026ViewForm()
# 	if request.method == 'POST':
# 		with transaction.atomic():
# 			if cueanexo_form.is_valid():
# 				cueanexo=cueanexo_form.cleaned_data["cueanexo"]
# 				grado_form=GradoFluidez2026ViewForm(request.POST, cueanexo=cueanexo)
# 				if grado_form.is_valid():
# 					grado=grado_form.cleaned_data["grado"]
# 					print(grado)
# 					return redirect("evaluaciones_educativas:fluidez_2026:lista",grado_public_id=grado)
# 	contexto={'cueanexo_form':cueanexo_form,
# 		   'grado_form':grado_form,
# 		   'grado':grado
# 		   }
# 	return render(request, "fluidez_2026/inicio.html",contexto)
@login_required
def lista(request,fid_actual=None):
	
	usuario= request.user
	cuil = usuario.username
	#print(cuil)
	if not fid_actual:
	# 1. Variables por defecto
		fid_actual = request.GET.get('fid')
	cueanexo_id = None
	grado_public_id = None

	alumnos_qs = None
	qs_secciones = None
	grado = None
	
	# Instanciamos los formularios vacíos
	cueanexo_form = CueanexoFluidez2026ViewForm(cuil=cuil)
	grado_form = GradoFluidez2026ViewForm()

	# ==========================================
	# LÓGICA DE CAPTURA DE DATOS (POST)
	# ==========================================
	if request.method == 'POST':
		# 1. Tomamos los datos tal cual los envió el HTML
		cueanexo_nuevo = request.POST.get('cueanexo')
		grado_nuevo = request.POST.get('grado')
		
		# 2. Leemos qué teníamos guardado ANTES de este cambio
		datos_anteriores = request.session.get(f"filtro_{fid_actual}", {})
		cueanexo_anterior = datos_anteriores.get('cueanexo')
		
		# 3. LA MAGIA: Si el cueanexo cambió, el grado viejo ya no sirve
		if cueanexo_nuevo != cueanexo_anterior:
			grado_nuevo = None  # Lo reseteamos a la fuerza
			
		# 4. Guardamos o sobreescribimos los datos en ESA misma carpeta
		request.session[f"filtro_{fid_actual}"] = {
			'cueanexo': cueanexo_nuevo,
			'grado': grado_nuevo
		}
		
		request.session.modified = True
		
		# Redirigimos
		return redirect(f"{request.path}?fid={fid_actual}")

	# ==========================================
	# LÓGICA DE RECUPERACIÓN (GET)
	# ==========================================
	else:
		if fid_actual:
			datos_guardados = request.session.get(f"filtro_{fid_actual}")
			
			if datos_guardados:
				# 1. Extraemos los datos CRUDOS directamente de la memoria
				cueanexo_id = datos_guardados.get('cueanexo')
				grado_public_id = datos_guardados.get('grado')
				#print(f'cue:{cueanexo_id}, grado{grado_public_id}')
				# 2. Pasamos los datos a los formularios SOLO para que el HTML 
				# mantenga visualmente seleccionadas las opciones correctas.
				cueanexo_form = CueanexoFluidez2026ViewForm(datos_guardados, cuil=cuil)
				if cueanexo_id:
					grado_form = GradoFluidez2026ViewForm(datos_guardados, cueanexo=cueanexo_id)


	# ==========================================
	# LÓGICA DE BÚSQUEDA A LA BASE DE DATOS
	# ==========================================
	# Esto se ejecutará siempre que tengamos los dos datos, sin depender del .is_valid()
	if cueanexo_id and grado_public_id:
		
		grado = GradoFluidez2026.objects.get(public_id=grado_public_id)
		
		if grado.nombre_grado == '2do Año/Grado':
			nombre_grado = '2do Grado/Año'
		else:
			nombre_grado = '3er Grado/Año'
			
		lista_dnis = list(
			TablaTemporalAlumnoFluidez2026.objects
			.filter(cueanexo=cueanexo_id, anio=nombre_grado)
			.values_list('numero_de_documento', flat=True)
		)
		# print(lista_dnis)
		# print(cueanexo_id)
		# print(nombre_grado)
		lista = list(
			AlumnoFluidez2026.objects
			.filter(~Q(dni__in=lista_dnis), seccion__grado__cueanexo=int(cueanexo_id),seccion__grado__nombre_grado=grado.nombre_grado)
			.values_list('dni', flat=True)
		)
		# print('-'*50)
		# print(lista)
		
		lista_dnis.extend(lista)
		alumnos_qs = AlumnoFluidez2026.objects.filter(dni__in=lista_dnis)
		# print(alumnos_qs)
		
		qs_secciones = SeccionFluidez2026.objects.filter(
			grado__public_id=grado.public_id
		)
	# ==========================================
	# RENDERIZADO
	# ==========================================
	contexto = {
		'cueanexo_form': cueanexo_form,
		'grado_form': grado_form,
		'grado': grado,
		'alumnos': alumnos_qs,
		'secciones_turnos_disponibles': qs_secciones,
		'opciones_comunidad_indigena': AlumnoFluidez2026.OPCIONES_COMUNIDAD_INDIGENA,
		'opciones_discapacidad': AlumnoFluidez2026.OPCIONES_DISCAPACIDAD,
		'fid_actual': fid_actual,
	}
	
	return render(request, "fluidez_2026/lista.html", contexto)

#-----------logica de actualizar seccion-------------------
@login_required
def actualizar_seccion(request, alumno_public_id):
	"""Actualiza la sección, turno, comunidad_indigena y discapacidad de un alumno."""
	if request.method == 'POST':
		fid = request.POST.get('fid')
		
		# 1. Obtenemos al alumno de forma segura
		alumno = get_object_or_404(AlumnoFluidez2026, public_id=alumno_public_id)

		nueva_seccion_turno_public_id = request.POST.get('seccion_turno')
		comunidad = request.POST.get('comunidad_indigena')
		discapacidad = request.POST.get('discapacidad')

		# 2. Actualizar sección y turno
		if nueva_seccion_turno_public_id:
			# Usamos get_object_or_404 para evitar crasheos si el ID es inválido
			seccion_obj = get_object_or_404(SeccionFluidez2026, public_id=nueva_seccion_turno_public_id)
			alumno.seccion = seccion_obj
		else:
			# Si permite desasignar la sección al seleccionar la opción vacía
			alumno.seccion = None 

		# 3. Actualizar comunidad indígena
		valores_comunidad = [str(v) for v, _ in AlumnoFluidez2026.OPCIONES_COMUNIDAD_INDIGENA]
		if comunidad in valores_comunidad:
			alumno.comunidad_indigena = comunidad
		elif not comunidad: # Si manda vacío, limpiamos el campo (si tu modelo lo permite)
			alumno.comunidad_indigena = None

		# 4. Actualizar discapacidad
		valores_discapacidad = [str(v) for v, _ in AlumnoFluidez2026.OPCIONES_DISCAPACIDAD]
		if discapacidad in valores_discapacidad:
			alumno.discapacidad = discapacidad
		elif not discapacidad: # Limpiamos si viene vacío
			alumno.discapacidad = None

		# 5. Guardamos cambios
		alumno.save()

		# 6. Redirección con el token de pestaña
		base_url = reverse('evaluaciones_educativas:fluidez_2026:lista')
		
		if fid:
			return redirect(f"{base_url}?fid={fid}")
		
		return redirect(base_url)
	
	# Si por alguna razón entra por GET (alguien teclea la URL directo), 
	# simplemente lo devolvemos a la lista.
	return redirect('evaluaciones_educativas:fluidez_2026:lista')

#------------------fin logica de actualizar seccion-------------------









@login_required
def lista_examen(request,fid_actual=None):
	
	usuario= request.user
	cuil = usuario.username
	if not fid_actual:
	# 1. Variables por defecto
		fid_actual = request.GET.get('fid')

	cueanexo_id = None
	grado_public_id = None

	alumnos_qs = None
	qs_secciones = None
	grado = None
	
	# Instanciamos los formularios vacíos
	cueanexo_form = CueanexoFluidez2026ViewForm(cuil=cuil)
	grado_form = GradoFluidez2026ViewForm()

	# ==========================================
	# LÓGICA DE CAPTURA DE DATOS (POST)
	# ==========================================
	if request.method == 'POST':
		# 1. Tomamos los datos tal cual los envió el HTML
		cueanexo_nuevo = request.POST.get('cueanexo')
		grado_nuevo = request.POST.get('grado')
		
		# 2. Leemos qué teníamos guardado ANTES de este cambio
		datos_anteriores = request.session.get(f"filtro_{fid_actual}", {})
		cueanexo_anterior = datos_anteriores.get('cueanexo')
		
		# 3. LA MAGIA: Si el cueanexo cambió, el grado viejo ya no sirve
		if cueanexo_nuevo != cueanexo_anterior:
			grado_nuevo = None  # Lo reseteamos a la fuerza
			
		# 4. Guardamos o sobreescribimos los datos en ESA misma carpeta
		request.session[f"filtro_{fid_actual}"] = {
			'cueanexo': cueanexo_nuevo,
			'grado': grado_nuevo
		}
		
		request.session.modified = True
		
		# Redirigimos
		return redirect(f"{request.path}?fid={fid_actual}")

	# ==========================================
	# LÓGICA DE RECUPERACIÓN (GET)
	# ==========================================
	else:
		#print(fid_actual)
		if fid_actual:
			datos_guardados = request.session.get(f"filtro_{fid_actual}")
			
			if datos_guardados:
				# 1. Extraemos los datos CRUDOS directamente de la memoria
				cueanexo_id = datos_guardados.get('cueanexo')
				grado_public_id = datos_guardados.get('grado')
				
				# 2. Pasamos los datos a los formularios SOLO para que el HTML 
				# mantenga visualmente seleccionadas las opciones correctas.
				cueanexo_form = CueanexoFluidez2026ViewForm(datos_guardados, cuil=cuil)
				if cueanexo_id:
					grado_form = GradoFluidez2026ViewForm(datos_guardados, cueanexo=cueanexo_id)


	# ==========================================
	# LÓGICA DE BÚSQUEDA A LA BASE DE DATOS
	# ==========================================
	# Esto se ejecutará siempre que tengamos los dos datos, sin depender del .is_valid()
	if cueanexo_id and grado_public_id:
		
		grado = GradoFluidez2026.objects.get(public_id=grado_public_id)
		
		if grado.nombre_grado == '2do Año/Grado':
			nombre_grado = '2do Grado/Año'
		else:
			nombre_grado = '3er Grado/Año'
			
		lista_dnis = list(
			TablaTemporalAlumnoFluidez2026.objects
			.filter(cueanexo=cueanexo_id, anio=nombre_grado)
			.values_list('numero_de_documento', flat=True)
		)
		
		lista = list(
			AlumnoFluidez2026.objects
			.filter(~Q(dni__in=lista_dnis), seccion__grado__cueanexo=int(cueanexo_id),seccion__grado__nombre_grado=grado.nombre_grado)
			.values_list('dni', flat=True)
		)
		
		lista_dnis.extend(lista)
		alumnos_qs = AlumnoFluidez2026.objects.filter(dni__in=lista_dnis).select_related('evaluacionfluidezlectorafluidez2026')
		
		# qs_secciones = SeccionFluidez2026.objects.filter(
		#     grado__cueanexo=cueanexo_id
		# )

	# ==========================================
	# RENDERIZADO
	# ==========================================
	contexto = {
		'cueanexo_form': cueanexo_form,
		'grado_form': grado_form,
		'grado': grado,
		'alumnos': alumnos_qs,
		'fid_actual': fid_actual,
	}
	
	return render(request, "fluidez_2026/lista_examen.html", contexto)

@login_required
def carga_alumno(request,fid_actual,grado_public_id):
	usuario= request.user
	cuil = usuario.username
	#print(grado_public_id)
	grado=get_object_or_404(GradoFluidez2026, public_id=grado_public_id)
	alumno_form = AlumnoFluidez2026Form()
	grado_form = GradoFluidez2026Form(instance=grado)
	#grado_form = None
	seccion_form = SeccionFluidez2026Form(cueanexo=grado.cueanexo,nombre_grado=grado.nombre_grado)
	if request.method == 'POST':
		alumno_form = AlumnoFluidez2026Form(request.POST)
		#grado_form = GradoFluidez2026Form(request.POST)
		seccion_form = SeccionFluidez2026Form(request.POST, cueanexo=grado.cueanexo,nombre_grado=grado.nombre_grado)
		# if alumno_form.is_valid() and grado_form.is_valid() and seccion_form.is_valid():
		if alumno_form.is_valid() and seccion_form.is_valid():
		   #una instancia a la vez
			with transaction.atomic():
				seccion_turno_id=seccion_form.cleaned_data["seccion_turno"]
				#print(seccion_turno_id)
				instancia_seccion, creado_seccion = SeccionFluidez2026.objects.get_or_create(
				id=seccion_turno_id,
				grado=grado
				)
				
				alumno = alumno_form.save(commit=False)
				alumno.seccion = instancia_seccion
				alumno.save()
				# instancia_evaluacion, creando_evaluacion=EvaluacionFluidezLectoraFluidez2026.objects.get_or_create(
				# alumno_id=alumno.id,cantidad_palabras_leidas=None, pregunta_1=None, pregunta_2=None, pregunta_3=None, pregunta_4=None, pregunta_5=None, pregunta_6=None, asistencia='AUSENTE',encargado_carga='DIRECTOR')
			return redirect("evaluaciones_educativas:fluidez_2026:lista",  fid_actual=fid_actual)
			
	context = {
		'alumno_form': alumno_form,
		'grado_form': grado_form,
		'seccion_form': seccion_form,
		'fid_actual':fid_actual,
			   }
	return render(request, "fluidez_2026/alumno.html", context)

@login_required
def editar_alumno(request,alumno_public_id, fid_actual):
	#print(fid_actual)
	instancia_alumno=get_object_or_404(AlumnoFluidez2026,public_id=alumno_public_id)
	#print(instancia_alumno)
	if not instancia_alumno.seccion.grado.cueanexo:
		alumno_datos=get_object_or_404(TablaTemporalAlumnoFluidez2026,numero_de_documento=instancia_alumno.dni)
		cueanexo=alumno_datos.cueanexo
		anio=alumno_datos.anio
	cueanexo=instancia_alumno.seccion.grado.cueanexo
	anio=instancia_alumno.seccion.grado.nombre_grado
	seccion=instancia_alumno.seccion.id
	#print(seccion)
	#print(anio)
	# print(alumno_datos.cueanexo)
	# print(alumno_datos.anio)
	# if anio == '2do Grado/Año':
	# 	nombre_grado='2do Año/Grado'
	# else:
	# 	nombre_grado='3er Año/Grado'
	#print(nombre_grado)
	#secciones=SeccionFluidez2026.objects.filter(grado__nombre_grado=alumno_datos.anio,grado__cueanexo=alumno_datos.cueanexo)
	instancia_grado=get_object_or_404(GradoFluidez2026,cueanexo=cueanexo,nombre_grado=anio)
	# instancia_grado=get_object_or_404(GradoFluidez2026,id=instancia_seccion.grado_id)
	alumno_form = AlumnoFluidez2026Form(instance=instancia_alumno)
	#print('hola')
	seccion_form = SeccionFluidez2026Form(cueanexo=instancia_grado.cueanexo,nombre_grado=instancia_grado.nombre_grado,initial={'seccion_turno':seccion})
	grado_form = GradoFluidez2026Form(instance=instancia_grado)
	if request.method == 'POST':
		alumno_form = AlumnoFluidez2026Form(request.POST, instance=instancia_alumno)
		grado_form = GradoFluidez2026Form(request.POST, instance=instancia_grado)
		seccion_form = SeccionFluidez2026Form(request.POST, cueanexo=cueanexo,nombre_grado=instancia_grado.nombre_grado)
		if alumno_form.is_valid() and grado_form.is_valid() and seccion_form.is_valid():
			with transaction.atomic():
				#nombre_grado=grado_form.cleaned_data["nombre_grado"]
				#cueanexo_grado=grado_form.cleaned_data["cueanexo"]
				instancia_grado, creado_grado=GradoFluidez2026.objects.get_or_create(
					nombre_grado=instancia_grado.nombre_grado,
					cueanexo=cueanexo
					)
				seccion_turno_id=seccion_form.cleaned_data["seccion_turno"]
				#print(seccion_turno_id)
				instancia_seccion, creado_seccion = SeccionFluidez2026.objects.get_or_create(
				id=seccion_turno_id,
				grado=instancia_grado
				)
				alumno = alumno_form.save(commit=False)
				alumno.seccion = instancia_seccion
				alumno.save()
			return redirect("evaluaciones_educativas:fluidez_2026:lista", fid_actual=fid_actual)
	context = {
		'alumno_form': alumno_form,
		 'grado_form': grado_form,
		'seccion_form': seccion_form,
		'grado_public':instancia_grado.public_id,
		'fid_actual':fid_actual,
			   }
	return render(request, "fluidez_2026/alumno.html", context)

# @login_required
# def lista(request,grado_public_id):
# 	contexto = {
# 	'cueanexo':None,
# 	'grado_form_data': None,
# 	'lista_alumnos': None,
# 	'evaluciones': None,
# 	'nombre_grado':None,
# 	'grado_public_id':None
# 	}
# 	usuario= request.user
# 	name=usuario.username
# 	opciones_grado = [
# 					 ('2do Año/Grado', '2do Año/Grado'),
# 					  ('3er Año/Grado','3er Año/Grado')
# 						]
# 	grado_form_data = GradoViewForm(cuil=name)
# 	grado_form_data.fields['grado'].choices = opciones_grado
# 	contexto["grado_form_data"] = grado_form_data
# 	if request.method == 'POST':
# 		grado_form_data = GradoViewForm(request.POST,cuil=name)
# 		grado_form_data.fields['grado'].choices = opciones_grado
# 		contexto["grado_form_data"] = grado_form_data
# 		if grado_form_data.is_valid():
# 			with transaction.atomic():
# 				grado=grado_form_data.cleaned_data["grado"]
# 				cueanexo=grado_form_data.cleaned_data["cueanexo"]
# 				if int(cueanexo) not in listaCueanexoPermitidos:
# 					raise PermissionDenied("No tienes permiso para acceder a esta sección.")
# 				else:
# 					try:
# 						instancia_grado=Grado.objects.get(cueanexo=cueanexo, nombre_grado=grado)
# 						return redirect("evaluaciones_educativas:fluidez_2026:lista", grado_public_id=instancia_grado.public_id)
# 					except Grado.DoesNotExist:
# 						messages.error(request, "No tiene un grado disponible para los datos ingresados.")
	
# 						# 2. Obtienes la URL de donde venía el usuario
# 						return_url = request.META.get('HTTP_REFERER', '/') 
	
# 						# 3. Rediriges a esa misma URL
# 						return redirect(return_url)

# 	else:
# 		try:
# 			instancia_grado=Grado.objects.get(public_id=grado_public_id)
# 			instancia_seccion=Seccion.objects.filter(grado_id=instancia_grado)
# 			alumnos = Alumno.objects.filter(seccion_id__in=instancia_seccion).order_by('nombre')
# 			evaluacion = EvaluacionFluidezLectora.objects.filter(alumno__in=alumnos)
# 			contexto["nombre_grado"] = instancia_grado.nombre_grado
# 			contexto["cueanexo"] = instancia_grado.cueanexo
# 			contexto["lista_alumnos"] = alumnos
# 			contexto["evaluciones"] = evaluacion
# 			contexto["grado_public_id"] = instancia_grado.public_id
# 		except Grado.DoesNotExist:
# 			raise PermissionDenied("No tiene grado disponible.")
# 	return render(request,"fluidez_2026/lista.html", contexto)
# #-----------------lista para grados------------------
# @login_required
# def lista_grado(request,grado): 
# 	#---------logica para obtener cueanexo por medio de username--------------
# 	#contexto={}
# 	usuario= request.user
# 	name=usuario.username
# 	opciones_grado = [
# 					 ('2do Año/Grado', '2do Año/Grado'),
# 					  ('3er Año/Grado','3er Año/Grado')
# 						]
# 	grado_form_data = GradoViewForm(cuil=name)
# 	if request.method == 'POST':
# 		grado_form_data = GradoViewForm(request.POST,cuil=name)
# 		grado_form_data.fields['grado'].choices = opciones_grado
# 		if grado_form_data.is_valid():
# 			with transaction.atomic():
# 				grado=grado_form_data.cleaned_data["grado"]
# 				cueanexo=grado_form_data.cleaned_data["cueanexo"]
# 				#print(f'{cueanexo}{grado}')
# 				if int(cueanexo) not in listaCueanexoPermitidos:
# 					raise PermissionDenied("No tienes permiso para acceder a esta sección.")
# 				else:
# 					try:
# 						#print(f'{cueanexo}{grado}')
# 						instancia_grado=Grado.objects.get(cueanexo=cueanexo, nombre_grado=grado)
# 						return redirect("evaluaciones_educativas:fluidez_2026:lista", grado_public_id=instancia_grado.public_id)
# 					except Grado.DoesNotExist:
# 						return render(request,"fluidez_2026/lista.html",contexto)
# 	else:
# 		grado_form_data.fields['grado'].choices = opciones_grado
# 	contexto = {
# 		'grado_form_data': grado_form_data
# 	}
# 	return render(request,"fluidez_2026/lista.html",contexto)
	
# @login_required
# def grado(request):
# 	usuario= request.user
# 	name=usuario.username
# 	#numeroCueanexo = obtener_cueanexo(name)
# 	# if usuario.is_authenticated:
# 		# name=usuario.username
# 		#-----logica para DNI+CUEANEXO---------
# 		# if len(name)>9  and len(name)<=17:
# 		# 	#DNI+CUEANEXO
# 		# 	nombre_usuario_cueanexo=name[8:]
# 		# else:
# 		# 	nombre_usuario_cueanexo=name
# 		# #cueanexo=int(nombre_usuario_cueanexo)
# 		#-----logica para DNI+CUEANEXO---------
# 	#if numeroCueanexo not in listaCueanexoPermitidos:
# 	#	raise PermissionDenied("No tienes permiso para acceder a esta sección.")
# 	opciones_grado = [
# 					 ('2do Año/Grado', '2do Año/Grado'),
# 					  ('3er Año/Grado','3er Año/Grado')
# 						]
# 	grado_form_data = GradoViewForm(cuil=name)
# 	if request.method == 'POST':
# 		grado_form_data = GradoViewForm(request.POST,cuil=name)
# 		grado_form_data.fields['grado'].choices = opciones_grado
# 		if grado_form_data.is_valid():
# 			with transaction.atomic():
# 				grado=grado_form_data.cleaned_data["grado"]
# 				cueanexo=grado_form_data.cleaned_data["cueanexo"]
# 				if int(cueanexo) not in listaCueanexoPermitidos:
# 					raise PermissionDenied("No tienes permiso para acceder a esta sección.")
# 				else:
# 					instancia_grado, creado_grado=Grado.objects.get_or_create(
# 					nombre_grado=grado,
# 					cueanexo=cueanexo
# 					)
# 				grado_public=instancia_grado.public_id
# 			return redirect("evaluaciones_educativas:fluidez_2026:carga_alumno", grado_public_id=grado_public)
# 	else:
# 		grado_form_data.fields['grado'].choices = opciones_grado
# 	contexto = {
# 		'grado_form_data': grado_form_data
# 	}
# 	return render(request,"fluidez_2026/grados.html", contexto)

	
@login_required
def carga_evaluacion(request, alumno_public_id, fid_actual):
	#fid_actual = request.POST.get('fid')
	#print(fid_actual)
	alumno_id=get_object_or_404(AlumnoFluidez2026, public_id=alumno_public_id)
	instancia_seccion=get_object_or_404(SeccionFluidez2026,id=alumno_id.seccion_id)
	instancia_grado=get_object_or_404(GradoFluidez2026,id=instancia_seccion.grado_id)
	grado_public=instancia_grado.public_id
	if instancia_grado.nombre_grado =='2do Año/Grado':
		cantidad_palabra_maxima=196
	else:
		cantidad_palabra_maxima=265
		
	evaluacion_existente = None
	try:

		evaluacion_existente = EvaluacionFluidezLectoraFluidez2026.objects.get(alumno=alumno_id.id)
	except EvaluacionFluidezLectoraFluidez2026.DoesNotExist:
		pass
	if request.method == 'POST':
		# 3. Rama POST: Se usa la instancia para forzar la ACTUALIZACIÓN (Sobrescritura)
		
		# Al pasar 'instance=evaluacion_existente' el form sabe qué registro modificar
		# Aunque el usuario lo haya visto vacío, los datos que se guardarán son los nuevos.
		form = EvaluacionFluidez2026Form(request.POST, 
									 max_cantidad_palabra=cantidad_palabra_maxima,
									 instance=evaluacion_existente)
		if form.is_valid():
			with transaction.atomic():
				evaluacion = form.save(commit=False)
				evaluacion.alumno = alumno_id
				evaluacion.encargado_carga='DIRECTOR'
				if form.cleaned_data["asistencia"] == 'AUSENTE':
					evaluacion=ausentismo_evaluacion(evaluacion)
				else:
					evaluacion.asistencia='PRESENTE'

				evaluacion.save()
			return redirect("evaluaciones_educativas:fluidez_2026:lista_examen", fid_actual=fid_actual)
	else:
		#Instancia vacia para metodo get
		form = EvaluacionFluidez2026Form(max_cantidad_palabra=cantidad_palabra_maxima)
		#Creacion de diccionario para el Post
	context = {'form': form,
			   'alumno':alumno_id,
			   'fid_actual':fid_actual
			   }
	return render(request, "fluidez_2026/evaluacion.html", context)

@login_required
def editar_evaluacion(request, alumno_public_id, fid_actual):
	alumno_id=get_object_or_404(AlumnoFluidez2026,public_id=alumno_public_id)
	instancia_seccion=get_object_or_404(SeccionFluidez2026,id=alumno_id.seccion_id)
	instancia_grado=get_object_or_404(GradoFluidez2026,id=instancia_seccion.grado_id)
	grado_public=instancia_grado.public_id
	instancia_evaluacion=get_object_or_404(EvaluacionFluidezLectoraFluidez2026,alumno_id=alumno_id.id)
	if instancia_grado.nombre_grado =='2do Año/Grado':
		cantidad_palabra_maxima=196
	else:
		cantidad_palabra_maxima=265
	form=EvaluacionFluidez2026Form(instance=instancia_evaluacion, max_cantidad_palabra=cantidad_palabra_maxima)
	if request.method == 'POST':
		form=EvaluacionFluidez2026Form(request.POST,instance=instancia_evaluacion,max_cantidad_palabra=cantidad_palabra_maxima)
		if form.is_valid():
			with transaction.atomic():
				evaluacion=form.save(commit=False)
				evaluacion.alumno_id = alumno_id
				evaluacion.encargado_carga='DIRECTOR'
				if form.cleaned_data["asistencia"] == 'AUSENTE':
					evaluacion=ausentismo_evaluacion(instancia_evaluacion)
				else:
					evaluacion.asistencia='PRESENTE'
				evaluacion.save()
			return redirect("evaluaciones_educativas:fluidez_2026:lista_examen", fid_actual=fid_actual)
	context = {
		'form': form,
		'alumno':alumno_id,
		'fid_actual':fid_actual,
		}
	return render(request, "fluidez_2026/evaluacion.html", context)

# @login_required
# def asistencia(request,alumno_public_id):
# 	alumno_id=get_object_or_404(Alumno, public_id=alumno_public_id)
# 	#SI instanciamos aca recibe la evalaucion que se creo en carga...
# 	instancia_evaluacion=get_object_or_404(EvaluacionFluidezLectora, alumno_id=alumno_id.id)
# 	instancia_seccion=get_object_or_404(Seccion,id=alumno_id.seccion_id)
# 	instancia_grado=get_object_or_404(Grado,id=instancia_seccion.grado_id)
# 	grado_public=instancia_grado.public_id
# 	if request.method == 'POST':
# 			form = AsistenciaForm(request.POST)
# 			if form.is_valid():
# 				with transaction.atomic():
# 					asistencia= form.cleaned_data["asistencia"]
# 					if asistencia: 
# 						return redirect("evaluaciones_educativas:fluidez_2026:carga_evaluacion", alumno_public_id=alumno_id.public_id)
# 					else:
# 						#llamamos a funcion ausentismo
# 						evaluacion=ausentismo_evaluacion(instancia_evaluacion)
# 						evaluacion.save()
# 						return redirect("evaluaciones_educativas:fluidez_2026:lista", grado_public_id=grado_public)
# 	else:
# 		form = AsistenciaForm()
# 	context = {'form': form,
# 			   'alumno':alumno_id
# 			   }
# 	return render(request,"fluidez_2026/asistencia.html",context)

@login_required
def editar_asistencia(request,alumno_public_id):
	#INSTANCIAS
	alumno_id=get_object_or_404(AlumnoFluidez2026, public_id=alumno_public_id)
	instancia_seccion=get_object_or_404(SeccionFluidez2026,id=alumno_id.seccion_id)
	instancia_grado=get_object_or_404(GradoFluidez2026,id=instancia_seccion.grado_id)
	grado_public=instancia_grado.public_id
	if request.method == 'POST':
		form = AsistenciaFluidez2026Form(request.POST)
		#Campos de tabla evaluacion
		if form.is_valid():
			with transaction.atomic():
				asistencia = form.cleaned_data["asistencia"]
				#verificamos alumno presente
				if asistencia: 
					return redirect("evaluaciones_educativas:fluidez_2026:editar_evaluacion", alumno_public_id=alumno_id.public_id)
				else:
					#Recien instanciamos en el ELSE 
					instancia_evaluacion=get_object_or_404(EvaluacionFluidezLectoraFluidez2026, alumno_id=alumno_id.id)
					evaluacion=ausentismo_evaluacion(instancia_evaluacion)
					evaluacion.save()
					return redirect("evaluaciones_educativas:fluidez_2026:lista", grado_public_id=grado_public)
	else:
		asistencia_form = AsistenciaFluidez2026Form()
	context = {'form': asistencia_form}
	return render(request,"fluidez_2026/asistencia.html",context)

@login_required
def borrar_registro_alumno(request,alumno_public_id):
	alumno_id=get_object_or_404(AlumnoFluidez2026, public_id=alumno_public_id)
	instancia_seccion=get_object_or_404(SeccionFluidez2026,id=alumno_id.seccion_id)
	instancia_grado=get_object_or_404(GradoFluidez2026,id=instancia_seccion.grado_id)
	grado_public=instancia_grado.public_id
	if request.method == 'POST':
		form = BorrarRegistroAlumnoForm(request.POST)
		if form.is_valid():
			with transaction.atomic():
				eleccion= form.cleaned_data["borrar"]
				if eleccion:
					alumno_id.delete()
					return redirect("evaluaciones_educativas:fluidez_2026:lista")
					#return redirect("evaluaciones_educativas:fluidez_2026:lista", grado_public_id=grado_public)
				else:
					return redirect("evaluaciones_educativas:fluidez_2026:lista")
					#return redirect("evaluaciones_educativas:fluidez_2026:lista", grado_public_id=grado_public)
	else:
		form = BorrarRegistroAlumnoForm()
	context = {'form': form,
			   'alumno':alumno_id
			   }
	return render(request,"fluidez_2026/borrar_registro_alumno.html",context)
#DESCARGAR EXCEL
@login_required
def descargar_excel(request,grado_public_id):
	instancia_grado=get_object_or_404(GradoFluidez2026,public_id=grado_public_id)
	instancia_seccion=SeccionFluidez2026.objects.filter(grado_id=instancia_grado)
	alumnos = AlumnoFluidez2026.objects.filter(seccion_id__in=instancia_seccion).order_by('nombre')
	evaluacion = EvaluacionFluidezLectoraFluidez2026.objects.filter(alumno__in=alumnos)
	# 1. Configurar la respuesta HTTP para un archivo Excel
	# El 'mimetype' (o Content-Type) es crucial para que el navegador sepa que es un archivo .xlsx
	response = HttpResponse(
		content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
	)
	
	# 2. Configurar el encabezado Content-Disposition
	# Esto le dice al navegador que DEBE descargar el contenido y le asigna un nombre de archivo.
	if instancia_grado.nombre_grado=='2do Año/Grado':
		nombre_grado='2do_Grado_'
	elif instancia_grado.nombre_grado=='3er Año/Grado':
		nombre_grado='3er_Grado_'
	else:
		nombre_grado='_'
	response['Content-Disposition'] = f'attachment; filename="reporte_fluidez_{nombre_grado}junio_2026.xlsx"'

	# 3. Generar el contenido del Excel (lo mismo que tenías)
	wb = Workbook()
	ws = wb.active
	fecha_hora_actual = datetime.now()
	ws['A1'] = f'CUEANEXO: {instancia_grado.cueanexo}'
	ws['G1'] = f'FECHA Y HORA:  {fecha_hora_actual.strftime("%d/%m/%Y %I:%M:%S %p")}'
	lista=['NOMBRE','APELLIDO','DNI','COMUNIDAD INDíGENA','DISCAPACIDAD','GRADO','SECCIÓN','TURNO','ASISTENCIA','FLUIDEZ','P1','P2','P3','P4','P5','P6']
	#print(alumnos)
	ws.append(lista)
	for i,v in enumerate(evaluacion):
		ws[f'A{i + 3}']=v.alumno.nombre
		ws[f'B{i + 3}']=v.alumno.apellido
		ws[f'C{i + 3}']=v.alumno.dni
		ws[f'D{i + 3}']=v.alumno.comunidad_indigena
		ws[f'E{i + 3}']=v.alumno.discapacidad
		ws[f'F{i + 3}']=v.alumno.seccion.grado.nombre_grado
		ws[f'G{i + 3}']=v.alumno.seccion.seccion
		ws[f'H{i + 3}']=v.alumno.seccion.turno
		ws[f'I{i + 3}']=v.asistencia
		ws[f'J{i + 3}']=v.cantidad_palabras_leidas
		ws[f'K{i + 3}']=v.pregunta_1
		ws[f'L{i + 3}']=v.pregunta_2
		ws[f'M{i + 3}']=v.pregunta_3
		ws[f'N{i + 3}']=v.pregunta_4
		ws[f'O{i + 3}']=v.pregunta_5
		ws[f'P{i + 3}']=v.pregunta_6

	wb.save(response)
	# 5. Retornar la respuesta al navegador
	return response

@login_required
def completar_carga(request,grado_public_id):
	#instancia_grado=get_object_or_404(GradoFluidez2026,public_id=grado_public_id)
	estado_carga=True
	lista_inicial_conteo=None
	lista_final_conteo=None
	numero=None
	instancia_grado = GradoFluidez2026.objects.get(public_id=grado_public_id)
		
	if instancia_grado.nombre_grado == '2do Año/Grado':
		nombre_grado = '2do Grado/Año'
	else:
		nombre_grado = '3er Grado/Año'
		
	lista_dnis = list(
		TablaTemporalAlumnoFluidez2026.objects
		.filter(cueanexo=instancia_grado.cueanexo, anio=nombre_grado)
		.values_list('numero_de_documento', flat=True)
	)
	
	lista = list(
		AlumnoFluidez2026.objects
		.filter(~Q(dni__in=lista_dnis), seccion__grado__cueanexo=int(instancia_grado.cueanexo),seccion__grado__nombre_grado=instancia_grado.nombre_grado)
		.values_list('dni', flat=True)
	)
	lista_dnis.extend(lista)
	alumnos_qs = AlumnoFluidez2026.objects.filter(dni__in=lista_dnis)
	lista_inicial_conteo=alumnos_qs.count()
	evaluaciones=EvaluacionFluidezLectoraFluidez2026.objects.filter(alumno__in=alumnos_qs).count()
	lista_final_conteo=evaluaciones
	if instancia_grado.estado_carga == True:
		# Si estaba CERRADO y hacen clic, quieren ABRIRLO.
		# Lo abrimos siempre, sin importar si los conteos coinciden o no.
		instancia_grado.estado_carga = False
		instancia_grado.save()
	else:
		# Si estaba ABIERTO y hacen clic, quieren CERRARLO.
		# ACÁ SÍ validamos que hayan evaluado a todos.
		if lista_inicial_conteo == lista_final_conteo:
			instancia_grado.estado_carga = True
			instancia_grado.save()
		else:
			# Faltan alumnos, cancelamos la acción.
			estado_carga = False
			numero = lista_inicial_conteo - lista_final_conteo
	#print(f'es numero{numero}')
	return JsonResponse({
		'es_valido': estado_carga,
		'nuevo_estado': instancia_grado.estado_carga,
		'faltantes': numero
	})
#---------------------------------------------------------------------------------

@login_required
def analisis_evaluaciones_junio_2026(request):
	contexto = {
		"alumnos_evaluados_segundo": [],
		"alumnos_evaluados_tercero": [],
		"grado": None,
		"seccion": None,
		"form_grado": None,
		"form_seccion": None,
		"rol": None,
		"listar": True,
	}

	usuario = request.user
	cuil = usuario.username
	nivel_acceso = request.user.nivelacceso_id
	# Formateo de CUIL
	# cuil_con_caracter = f"{cuil[:2]}-{cuil[2:10]}-{cuil[10:]}"
	#nivel_acceso = "Director/a"
	contexto["rol"] = nivel_acceso
	# Inicializamos el primer formulario siempre
	form_cueanexo = CueanexoForm(
		request.POST or None, cuil=cuil, nivel_acceso=nivel_acceso
	)
	contexto["form_cueanexo"] = form_cueanexo
	lista_cueanexos = []
	if request.method == "POST":
		if form_cueanexo.is_valid():
			cueanexo = form_cueanexo.cleaned_data["cueanexo_seleccionado"]
			# print(f'cueanexo intento{cueanexo}')
			# Inicializamos el form de Grado (aquí podrías pasarle el cueanexo si el form fuera dinámico)
			form_grado = Grado_selec_2026_Form(request.POST or None)
			contexto["form_grado"] = form_grado

			if form_grado.is_valid():
				nombre_grado = form_grado.cleaned_data["grado_seleccion"]

				if nombre_grado:
					# transformar varibale cueanexo a lista
					lista_cueanexos.append(cueanexo)
					# Usamos filter().first() para evitar errores si no existe
					grado_obj = GradoFluidez2026.objects.filter(
						nombre_grado=nombre_grado, cueanexo__in=lista_cueanexos
					)
					# print(len(grado_obj))
					# grado_obj = Grado.objects.filter(nombre_grado=nombre_grado, cueanexo__in=lista_cueanexos).first()
					if grado_obj:
						# print('entre')
						# Inicializamos sección con el ID del grado
						form_seccion = SeccionTurnoForm(
							request.POST or None, grados=grado_obj
						)
						# form_seccion = SeccionTurnoForm(request.POST or None, grados=grado_obj.id)
						contexto["form_seccion"] = form_seccion

						secciones = None
						turno = None

						if form_seccion.is_valid():
							seleccionado = form_seccion.cleaned_data["seleccion"]
							if seleccionado != "TODOS":
								# 'seleccionado' aquí suele ser el objeto Seccion si es ModelChoiceField
								secciones = seleccionado.seccion
								turno = seleccionado.turno

						# CARGA DE DATOS: Se ejecuta si hay grado_obj, independientemente de la sección
						if nombre_grado == "2do Año/Grado":
							# if grado_obj.nombre_grado == '2do Año/Grado':
							datos = analisis_segundo_grado_grafico(
								nombre_grado, lista_cueanexos, secciones, turno
							)
							contexto["grado"] = "segundo"
						else:
							datos = analisis_tercer_grado_grafico(
								nombre_grado, lista_cueanexos, secciones, turno
							)
							contexto["grado"] = "tercero"

						contexto.update(datos)
						contexto["seccion"] = secciones
						contexto["turno"] = turno
				else:
					contexto["grado"] = None
	# nombre_archivo = (
	# 	"evaluaciones_educativas/pdf/INFORME-FLUIDEZ-LECTORA-NOVIEMBRE2026.pdf"
	# )
	# # Generamos la URL y le pegamos los parámetros del visor
	# contexto["material_pdf"] = static(nombre_archivo)
	return render(
		request, "fluidez_2026/analisis_evaluaciones_junio_2026.html", contexto
	)


# #------------------logica para regional---------------------------
@login_required
def analisis_evaluaciones_regional_junio_2026(request):
	contexto = {
		"alumnos_evaluados_segundo": [],
		"alumnos_evaluados_tercero": [],
		"grado": None,
		"seccion": None,
		"form_grado": None,
		"form_seccion": None,
		"rol": None,
		"listar": None,
	}

	usuario = request.user
	cuil = usuario.username
	# Formateo de CUIL
	# cuil_con_caracter = f"{cuil[:2]}-{cuil[2:10]}-{cuil[10:]}"
	nivel_acceso = request.user.nivelacceso_id
	# print(f'REGIONALLLLL{nivel_acceso}')
	contexto["rol"] = nivel_acceso
	form_director_regional = DirectorForm(request.POST or None)
	# Inicializamos el primer formulario siempre
	lista_cueanexos = []
	if request.method == "POST":
		if form_director_regional.is_valid():
			sector = form_director_regional.cleaned_data["sector"]
			ambito = form_director_regional.cleaned_data["ambito"]
			form_cueanexo = CueanexoForm(
				request.POST or None,
				cuil=cuil,
				nivel_acceso=nivel_acceso,
				sector=sector,
				ambito=ambito,
			)
			contexto["form_cueanexo"] = form_cueanexo
			if form_cueanexo.is_valid():
				cueanexo = form_cueanexo.cleaned_data["cueanexo_seleccionado"]
				# print(f'cueanexo intento{cueanexo}')
				if cueanexo == "TODOS" and nivel_acceso != "Director/a":
					# Esto te devuelve la lista de tuplas (valor, etiqueta)
					lista_cueanexos = [
						valor
						for valor, etiqueta in form_cueanexo.fields[
							"cueanexo_seleccionado"
						].choices
					]
					# print(f'lista{len(lista_cueanexos)}')
				else:
					lista_cueanexos.append(cueanexo)
					contexto["listar"] = True
				# Inicializamos el form de Grado (aquí podrías pasarle el cueanexo si el form fuera dinámico)
				form_grado = Grado_selec_2026_Form(request.POST or None)
				contexto["form_grado"] = form_grado

				if form_grado.is_valid():
					nombre_grado = form_grado.cleaned_data["grado_seleccion"]

					if nombre_grado:
						# transformar varibale cueanexo a lista
						# lista_cueanexos=[cueanexo]
						# Usamos filter().first() para evitar errores si no existe
						grado_obj = GradoFluidez2026.objects.filter(
							nombre_grado=nombre_grado, cueanexo__in=lista_cueanexos
						)
						# print(len(grado_obj))
						# grado_obj = Grado.objects.filter(nombre_grado=nombre_grado, cueanexo__in=lista_cueanexos).first()
						if grado_obj:
							# Inicializamos sección con el ID del grado
							form_seccion = SeccionTurnoForm(
								request.POST or None, grados=grado_obj
							)
							# form_seccion = SeccionTurnoForm(request.POST or None, grados=grado_obj.id)
							contexto["form_seccion"] = form_seccion

							secciones = None
							turno = None

							if form_seccion.is_valid():
								seleccionado = form_seccion.cleaned_data["seleccion"]
								if seleccionado != "TODOS":
									# 'seleccionado' aquí suele ser el objeto Seccion si es ModelChoiceField
									secciones = seleccionado.seccion
									turno = seleccionado.turno

							# CARGA DE DATOS: Se ejecuta si hay grado_obj, independientemente de la sección
							if nombre_grado == "2do Año/Grado":
								# if grado_obj.nombre_grado == '2do Año/Grado':
								datos = analisis_segundo_grado_grafico(
									nombre_grado, lista_cueanexos, secciones, turno
								)
								contexto["grado"] = "segundo"
							else:
								datos = analisis_tercer_grado_grafico(
									nombre_grado, lista_cueanexos, secciones, turno
								)
								contexto["grado"] = "tercero"

							contexto.update(datos)
							contexto["seccion"] = secciones
							contexto["turno"] = turno
					else:
						contexto["grado"] = None

	contexto["form_director_regional"] = form_director_regional
	# nombre_archivo = (
	# 	"evaluaciones_educativas/pdf/INFORME-FLUIDEZ-LECTORA-NOVIEMBRE2026.pdf"
	# )
	# # Generamos la URL y le pegamos los parámetros del visor
	# contexto["material_pdf"] = static(nombre_archivo)
	return render(
		request, "fluidez_2026/analisis_evaluaciones_junio_2026.html", contexto
	)


# #--------------------fin de logica para regional----------------
# #---------------------logica para SUBSE Y MINISTRO---------------------------
@login_required
def analisis_evaluaciones_ministros_junio_2026(request):
	contexto = {
		"alumnos_evaluados_segundo": [],
		"alumnos_evaluados_tercero": [],
		"grado": None,
		"seccion": None,
		"form_grado": None,
		"form_seccion": None,
		"rol": None,
		"listar": None,
	}

	usuario = request.user
	cuil = usuario.username
	# Formateo de CUIL
	# cuil_con_caracter = f"{cuil[:2]}-{cuil[2:10]}-{cuil[10:]}"
	nivel_acceso = request.user.nivelacceso_id
	# print(f'REGIONALLLLL{nivel_acceso}')
	contexto["rol"] = nivel_acceso
	print(nivel_acceso)
	# Inicializamos el primer formulario siempre
	# print(nivel_acceso)
	form_director_nivel = DirectorNivelForm(request.POST or None)
	# form_cueanexo = CueanexoForm(request.POST or None, cuil=cuil,nivel_acceso=nivel_acceso)
	lista_cueanexos = []
	if request.method == "POST":
		if form_director_nivel.is_valid():
			sector = form_director_nivel.cleaned_data["sector"]
			ambito = form_director_nivel.cleaned_data["ambito"]
			region = form_director_nivel.cleaned_data["region"]
			form_cueanexo = CueanexoForm(
				request.POST or None,
				cuil=cuil,
				nivel_acceso=nivel_acceso,
				sector=sector,
				ambito=ambito,
				region=region,
			)
			contexto["form_cueanexo"] = form_cueanexo
			if form_cueanexo.is_valid():
				cueanexo = form_cueanexo.cleaned_data["cueanexo_seleccionado"]
				# print(f'cueanexo intento{cueanexo}')
				if cueanexo == "TODOS" and nivel_acceso != "Director/a":
					# Esto te devuelve la lista de tuplas (valor, etiqueta)
					lista_cueanexos = [
						valor
						for valor, etiqueta in form_cueanexo.fields[
							"cueanexo_seleccionado"
						].choices
					]
					# print(f'lista{len(lista_cueanexos)}')
				else:
					lista_cueanexos.append(cueanexo)
					contexto["listar"] = True
				# Inicializamos el form de Grado (aquí podrías pasarle el cueanexo si el form fuera dinámico)
				form_grado = Grado_selec_2026_Form(request.POST or None)
				contexto["form_grado"] = form_grado

				if form_grado.is_valid():
					nombre_grado = form_grado.cleaned_data["grado_seleccion"]

					if nombre_grado:
						# transformar varibale cueanexo a lista
						# lista_cueanexos=[cueanexo]
						# Usamos filter().first() para evitar errores si no existe
						grado_obj = GradoFluidez2026.objects.filter(
							nombre_grado=nombre_grado, cueanexo__in=lista_cueanexos
						)
						# print(len(grado_obj))
						# grado_obj = Grado.objects.filter(nombre_grado=nombre_grado, cueanexo__in=lista_cueanexos).first()
						if grado_obj:
							# Inicializamos sección con el ID del grado
							form_seccion = SeccionTurnoForm(
								request.POST or None, grados=grado_obj
							)
							# form_seccion = SeccionTurnoForm(request.POST or None, grados=grado_obj.id)
							contexto["form_seccion"] = form_seccion

							secciones = None
							turno = None

							if form_seccion.is_valid():
								seleccionado = form_seccion.cleaned_data["seleccion"]
								if seleccionado != "TODOS":
									# 'seleccionado' aquí suele ser el objeto Seccion si es ModelChoiceField
									secciones = seleccionado.seccion
									turno = seleccionado.turno

							# CARGA DE DATOS: Se ejecuta si hay grado_obj, independientemente de la sección
							if nombre_grado == "2do Año/Grado":
								# if grado_obj.nombre_grado == '2do Año/Grado':
								datos = analisis_segundo_grado_grafico(
									nombre_grado, lista_cueanexos, secciones, turno
								)
								contexto["grado"] = "segundo"
							else:
								datos = analisis_tercer_grado_grafico(
									nombre_grado, lista_cueanexos, secciones, turno
								)
								contexto["grado"] = "tercero"

							contexto.update(datos)
							contexto["seccion"] = secciones
							contexto["turno"] = turno
					else:
						contexto["grado"] = None

	contexto["form_director_nivel"] = form_director_nivel
	#contexto["form_cueanexo"] = form_cueanexo
	# nombre_archivo = (
	# 	"evaluaciones_educativas/pdf/INFORME-FLUIDEZ-LECTORA-NOVIEMBRE2026.pdf"
	# )
	# # Generamos la URL y le pegamos los parámetros del visor
	# contexto["material_pdf"] = static(nombre_archivo)
	return render(
		request, "fluidez_2026/analisis_evaluaciones_junio_2026.html", contexto
	)

# #----------------------FIN SUBSE Y MINISTRO--------------------------------
# #-----------------------FIN LOGICA PARA VISUALIZAR DATOS ---------------------------

def ausentismo_evaluacion(instancia_evaluacion):
	evaluacion_campos=instancia_evaluacion._meta.fields
	for i in evaluacion_campos:
		if i.name == 'asistencia':
			#FUNCION DE PYTHON para estabalecer valores a campos de un objeto
			setattr(instancia_evaluacion, i.name, 'AUSENTE')
			#verificamos campos PK y NOT NULL
		if not i.primary_key and i.null:
			setattr(instancia_evaluacion, i.name, None)
	return instancia_evaluacion

# #---------LOGICA PARA ANALISIS DE LAS EVALUACIONES EDUCATIVAS--------------------------

##---------LOGICA PARA ANALISIS DE LAS EVALUACIONES EDUCATIVAS--------------------------


def analisis_segundo_grado_grafico(grado_seleccionado, cueanexo, secciones, turno):
	# #-----------------------LOGICA PARA SEGUNDO GRADO------------------------
	# print('entre a analisis 2do')
	if type(cueanexo) is str:
		grado = GradoFluidez2026.objects.filter(
			nombre_grado__icontains=grado_seleccionado, cueanexo=cueanexo
		)
	else:
		# print('ENTREEE')
		grado = GradoFluidez2026.objects.filter(
			nombre_grado__icontains=grado_seleccionado, cueanexo__in=cueanexo
		)
	# print(grado)
	if secciones and turno:
		secciones = SeccionFluidez2026.objects.filter(
			grado__in=grado, seccion=secciones, turno=turno
		)
	# print(seccion)
	else:
		secciones = SeccionFluidez2026.objects.filter(grado__in=grado)
	alumnos = AlumnoFluidez2026.objects.filter(seccion_id__in=secciones)
	evaluaciones = EvaluacionFluidezLectoraFluidez2026.objects.filter(
		alumno__in=alumnos
	)
	# comprension lectora
	# presentes=evaluaciones.filter(asistencia='PRESENTE')
	# POSIBLE OLUCION
	evaluaciones_con_comprension, conteos = comprension_lectora(
		alumnos, grado_seleccionado
	)
	datos = evaluaciones.aggregate(
		# Asistencia
		presentes=Count("alumno_id", filter=Q(asistencia="PRESENTE")),
		ausentes=Count("alumno_id", filter=Q(asistencia="AUSENTE")),
		# Desempeño
		promedio_general=Avg(
			"cantidad_palabras_leidas", filter=Q(asistencia="PRESENTE")
		),
		debajo_del_basico=Count(
			"alumno_id",
			filter=Q(asistencia="PRESENTE", cantidad_palabras_leidas__lt=21),
		),
		basico=Count(
			"alumno_id",
			filter=Q(
				asistencia="PRESENTE",
				cantidad_palabras_leidas__gte=21,
				cantidad_palabras_leidas__lte=46,
			),
		),
		satisfactorio=Count(
			"alumno_id",
			filter=Q(
				asistencia="PRESENTE",
				cantidad_palabras_leidas__gt=46,
				cantidad_palabras_leidas__lte=70,
			),
		),
		avanzado=Count(
			"alumno_id",
			filter=Q(asistencia="PRESENTE", cantidad_palabras_leidas__gt=70),
		),
	)
	# print(datos['debajo_del_basico'])
	alumnos_presentes = datos["presentes"]
	alumnos_ausentes = datos["ausentes"]
	# SOLUCIONAR LAS DIVISIONES POR 0
	nivel_debajo_del_basico = (
		round((datos["debajo_del_basico"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	nivel_basico = (
		round((datos["basico"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	nivel_satisfactorio = (
		round((datos["satisfactorio"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	nivel_avanzado = (
		round((datos["avanzado"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	# INDICE DE PROMEDIO (ENTERO)
	nivel_debajo_del_basico_comprension = (
		round((conteos["debajo_del_basico"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	nivel_basico_comprension = (
		round((conteos["basico"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	nivel_satisfactorio_comprension = (
		round((conteos["satisfactorio"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	nivel_avanzado_comprension = (
		round((conteos["avanzado"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	promedio = datos["promedio_general"]  # Modificar para trabajar mejor los decimales
	# print(promedio)
	# print(evaluacion)
	contexto = {
		"CUEANEXO": cueanexo,
		"nivel_educativo": "Primario",
		"alumnos_evaluados_segundo": evaluaciones_con_comprension,
		"asistencia_segundo": [
			f"Presentes {alumnos_presentes}",
			f"Ausentes {alumnos_ausentes}",
		],
		"valores_asistencia_segundo": [alumnos_presentes, alumnos_ausentes],
		"etiquetas_nivel_desempeno_segundo": [
			f"Debajo del Basico {nivel_debajo_del_basico}%",
			f"Básico {nivel_basico}%",
			f"Satisfactorio {nivel_satisfactorio}%",
			f"Avanzado {nivel_avanzado}%",
		],
		"valores_desempeno_segundo": [
			nivel_debajo_del_basico,
			nivel_basico,
			nivel_satisfactorio,
			nivel_avanzado,
		],
		"etiquetas_nivel_desempeno_comprension_segundo": [
			f"Debajo del Basico {nivel_debajo_del_basico_comprension}%",
			f"Básico {nivel_basico_comprension}%",
			f"Satisfactorio {nivel_satisfactorio_comprension}%",
			f"Avanzado {nivel_avanzado_comprension}%",
		],
		"valores_desempeno_comprension_segundo": [
			nivel_debajo_del_basico_comprension,
			nivel_basico_comprension,
			nivel_satisfactorio_comprension,
			nivel_avanzado_comprension,
		],
		"etiqueta_promedios_segundo": "Promedio de Palabras Leídas Correctamente por Minuto",
		"valor_promedio_segundo": [promedio],
		"dni_alumnos_segundo": json.dumps(
			[
				str(ev.alumno.dni) if ev.alumno.dni else "Sin DNI"
				for ev in evaluaciones_con_comprension
				if ev.asistencia == "PRESENTE"
				and ev.cantidad_palabras_leidas is not None
			]
		),
		"valores_fluidez_segundo": json.dumps(
			[
				float(ev.cantidad_palabras_leidas)
				for ev in evaluaciones_con_comprension
				if ev.asistencia == "PRESENTE"
				and ev.cantidad_palabras_leidas is not None
			]
		),
	}
	return contexto

	# ----------------FIN LOGICA 2 GRADO ---------


def analisis_tercer_grado_grafico(grado_seleccionado, cueanexo, secciones, turno):
	# -----------------------LOGICA PARA TERCER GRADO------------------------
	# PARTICIPACION
	# print('entre a analisis 3ero')

	# cueanexo=['220000200','220009000']
	if type(cueanexo) is str:
		grado = GradoFluidez2026.objects.filter(
			nombre_grado__icontains=grado_seleccionado, cueanexo=cueanexo
		)
	else:
		# print('ENTREEE')
		grado = GradoFluidez2026.objects.filter(
			nombre_grado__icontains=grado_seleccionado, cueanexo__in=cueanexo
		)
	# print(grado)
	if secciones and turno:
		secciones = SeccionFluidez2026.objects.filter(
			grado__in=grado, seccion=secciones, turno=turno
		)
	# print(seccion)
	else:
		secciones = SeccionFluidez2026.objects.filter(grado__in=grado)
	alumnos = AlumnoFluidez2026.objects.filter(seccion_id__in=secciones)

	evaluaciones = EvaluacionFluidezLectoraFluidez2026.objects.filter(
		alumno__in=alumnos
	)
	# presentes=evaluaciones.filter(asistencia='PRESENTE')
	# POSIBLE OLUCION
	evaluaciones_con_comprension, conteos = comprension_lectora(
		alumnos, grado_seleccionado
	)
	# evaluaciones_con_comprension,conteos=comprension_lectora(evaluaciones,grado_seleccionado)
	datos = evaluaciones.aggregate(
		# Asistencia
		presentes=Count("alumno_id", filter=Q(asistencia="PRESENTE")),
		ausentes=Count("alumno_id", filter=Q(asistencia="AUSENTE")),
		# Desempeño
		promedio_general=Avg(
			"cantidad_palabras_leidas", filter=Q(asistencia="PRESENTE")
		),
		debajo_del_basico=Count(
			"alumno_id",
			filter=Q(asistencia="PRESENTE", cantidad_palabras_leidas__lt=30),
		),
		basico=Count(
			"alumno_id",
			filter=Q(
				asistencia="PRESENTE",
				cantidad_palabras_leidas__gte=30,
				cantidad_palabras_leidas__lte=60,
			),
		),
		satisfactorio=Count(
			"alumno_id",
			filter=Q(
				asistencia="PRESENTE",
				cantidad_palabras_leidas__gt=61,
				cantidad_palabras_leidas__lte=90,
			),
		),
		avanzado=Count(
			"alumno_id",
			filter=Q(asistencia="PRESENTE", cantidad_palabras_leidas__gt=90),
		),
	)
	alumnos_presentes = datos["presentes"]
	alumnos_ausentes = datos["ausentes"]
	nivel_debajo_del_basico = (
		round((datos["debajo_del_basico"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	nivel_basico = (
		round((datos["basico"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	nivel_satisfactorio = (
		round((datos["satisfactorio"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	nivel_avanzado = (
		round((datos["avanzado"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	# grafico
	nivel_debajo_del_basico_comprension = (
		round((conteos["debajo_del_basico"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	nivel_basico_comprension = (
		round((conteos["basico"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	nivel_satisfactorio_comprension = (
		round((conteos["satisfactorio"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	nivel_avanzado_comprension = (
		round((conteos["avanzado"] / alumnos_presentes) * 100, 1)
		if alumnos_presentes
		else 0
	)
	# INDICE DE PROMEDIO (ENTERO)
	promedio = datos["promedio_general"]  # Modificar para trabajar mejor los decimales
	# print(promedio)
	# print(evaluacion)
	contexto = {
		"CUEANEXO": cueanexo,
		"nivel_educativo": "Primario",
		"alumnos_evaluados_tercero": evaluaciones_con_comprension,
		"asistencia_tercero": [
			f"Presentes {alumnos_presentes}",
			f"Ausentes {alumnos_ausentes}",
		],
		"valores_asistencia_tercero": [alumnos_presentes, alumnos_ausentes],
		"etiquetas_nivel_desempeno_tercero": [
			f"Debajo del Basico {nivel_debajo_del_basico}%",
			f"Básico {nivel_basico}%",
			f"Satisfactorio {nivel_satisfactorio}%",
			f"Avanzado {nivel_avanzado}%",
		],
		"valores_desempeno_tercero": [
			nivel_debajo_del_basico,
			nivel_basico,
			nivel_satisfactorio,
			nivel_avanzado,
		],
		"etiquetas_nivel_desempeno_comprension_tercero": [
			f"Debajo del Basico {nivel_debajo_del_basico_comprension}%",
			f"Básico {nivel_basico_comprension}%",
			f"Satisfactorio {nivel_satisfactorio_comprension}%",
			f"Avanzado {nivel_avanzado_comprension}%",
		],
		"valores_desempeno_comprension_tercero": [
			nivel_debajo_del_basico_comprension,
			nivel_basico_comprension,
			nivel_satisfactorio_comprension,
			nivel_avanzado_comprension,
		],
		"etiqueta_promedios_tercero": "Promedio de Palabras Leídas Correctamente por Minuto",
		"valor_promedio_tercero": [promedio],
		"dni_alumnos_tercero": json.dumps(
			[
				str(ev.alumno.dni) if ev.alumno.dni else "Sin DNI"
				for ev in evaluaciones_con_comprension
				if ev.asistencia == "PRESENTE"
				and ev.cantidad_palabras_leidas is not None
			]
		),
		"valores_fluidez_tercero": json.dumps(
			[
				float(ev.cantidad_palabras_leidas)
				for ev in evaluaciones_con_comprension
				if ev.asistencia == "PRESENTE"
				and ev.cantidad_palabras_leidas is not None
			]
		),
	}
	return contexto


# Crear funcion para logica de COMPRENSION LECTORA,
def comprension_lectora(alumnos, grado_seleccionado):
	# print(type(grado_seleccionado))
	# print('en comprension')
	# resultados = [] # Usamos una lista para guardar a todos
	evaluaciones = EvaluacionFluidezLectoraFluidez2026.objects.filter(
		alumno__in=alumnos, asistencia="PRESENTE"
	).select_related("alumno")
	conteos = {"debajo_del_basico": 0, "basico": 0, "satisfactorio": 0, "avanzado": 0}
	if grado_seleccionado == "2do Año/Grado":
		# print('entre en 2 comprensio')
		for i in evaluaciones:
			# Reiniciamos los puntos para CADA alumno dentro del bucle
			p1 = 1 if i.pregunta_1 == "B" else 0
			p2 = 1 if i.pregunta_2 == "C" else 0
			p3 = 1 if i.pregunta_3 == "B" else 0

			# IMPORTANTE: Usamos punto para el decimal (1.5)
			p4 = 1.50 if i.pregunta_4 == "A" else 0
			p5 = 1.50 if i.pregunta_5 == "B" else 0
			p6 = 1.50 if i.pregunta_6 == "C" else 0
			# print(f'{i.pregunta_1}{i.pregunta_2}{i.pregunta_3}{i.pregunta_4}{i.pregunta_5}{i.pregunta_6}')
			puntaje_total = p1 + p2 + p3 + p4 + p5 + p6
			i.puntaje_comprension = puntaje_total
			if puntaje_total < 3.40:
				conteos["debajo_del_basico"] += 1
			elif puntaje_total <= 5.20:
				conteos["basico"] += 1
			elif puntaje_total <= 6.75:
				conteos["satisfactorio"] += 1
			else:
				conteos["avanzado"] += 1

			# print(f'{i.puntaje_comprension} alumno:{i.alumno.nombre}' )
	else:
		for i in evaluaciones:
			# Reiniciamos los puntos para CADA alumno dentro del bucle
			p1 = 1 if i.pregunta_1 == "B" else 0
			p2 = 1 if i.pregunta_2 == "C" else 0
			p3 = 1 if i.pregunta_3 == "C" else 0

			# IMPORTANTE: Usamos punto para el decimal (1.5)
			p4 = 1.50 if i.pregunta_4 == "A" else 0
			p5 = 1.50 if i.pregunta_5 == "C" else 0
			p6 = 1.50 if i.pregunta_6 == "C" else 0
			# print(f'{i.pregunta_1}{i.pregunta_2}{i.pregunta_3}{i.pregunta_4}{i.pregunta_5}{i.pregunta_6}')
			puntaje_total = p1 + p2 + p3 + p4 + p5 + p6
			if puntaje_total < 3.40:
				conteos["debajo_del_basico"] += 1
			elif puntaje_total <= 5.20:
				conteos["basico"] += 1
			elif puntaje_total <= 6.75:
				conteos["satisfactorio"] += 1
			else:
				conteos["avanzado"] += 1

			# Guardamos un diccionario por cada alumno
			i.puntaje_comprension = puntaje_total
			# print(f'{i.puntaje_comprension} alumno:{i.alumno.nombre}' )

	return evaluaciones, conteos  # Devolvemos la lista completa
