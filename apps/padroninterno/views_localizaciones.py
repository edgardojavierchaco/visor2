import json
import re
from datetime import date, datetime
from functools import lru_cache
from io import BytesIO
from .permisos import padron_interno_admin_o_gestor_required
from .views_fecha import get_contexto_fecha_padron
import openpyxl
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import connections
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import ensure_csrf_cookie
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

PADRON_DB = 'Padron'
PAGE_SIZE = 10
OFERTA_ESTADO_FILTER_OPTIONS = ('Activo', 'Inactivo', 'Baja', 'Inactivo sin Docentes')

PERIODO_FUNCIONAMIENTO_OPTIONS = [
    'Común',
    'Especial o de temporada',
    'Ambos Periodos',
    'Especial2',
    'Sin Información',
    'No corresponde',
    'Especial sept-junio',
]

MODALIDADES_COMPLEMENTARIAS_OPTIONS = [
    'Técnico Profesional',
    'Artística',
    'Rural',
    'Intercultural Bilingüe',
    'Contextos de privación de libertad',
    'Común',
    'Especial',
    'Adultos',
    'No Corresponde',
    'Domiciliaria y Hospitalaria (año anterior)',
]

UDT_FILTER_OPTIONS = [(str(code), f'U.D.T. {code}') for code in [1, 10, 11, 12, 13, 14, 15, 2, 3, 4, 5, 6, 7, 8, 9]]

ANTERIOR_REGIONAL_OPTIONS = [
    (str(code), f'REGION EDUCATIVA {code}')
    for code in [1, 10, 11, 12, 13, 2, 3, 4, 5, 6, 7, 8, 9]
]

PATRIMONIO_EDILICIO_OPTIONS = [
    ('1', 'Público'),
    ('2', 'Privado'),
    ('3', 'Alquilado'),
    ('4', 'Comodato'),
    ('5', 'Prestado'),
    ('6', 'Otros'),
]

REGIONAL_HASTA_2015_OPTIONS = [
    ('1', 'REGION EDUCATIVA 1'),
    ('10', 'REGION EDUCATIVA 10'),
    ('11', 'SUB REGIONAL 4'),
    ('2', 'REGION EDUCATIVA 2'),
    ('3', 'REGION EDUCATIVA 3'),
    ('4', 'REGION EDUCATIVA 4'),
    ('5', 'REGION EDUCATIVA 5'),
    ('6', 'REGION EDUCATIVA 6'),
    ('7', 'REGION EDUCATIVA 7'),
    ('8', 'REGION EDUCATIVA 8'),
    ('9', 'REGION EDUCATIVA 9'),
]

REGIONAL_HASTA_2020_OPTIONS = [
    ('1', 'R.E. 1'),
    ('2', 'R.E. 2'),
    ('3', 'R.E. 3'),
    ('4', 'R.E. 4'),
    ('5', 'R.E. 5'),
    ('6', 'R.E. 6'),
    ('7', 'R.E. 7'),
    ('8', 'R.E. 8'),
    ('9', 'R.E. 9'),
    ('10', 'R.E. 10'),
    ('11', 'R.E. 11'),
    ('SubReg. 6', 'SUB. R.E. 6'),
    ('12', 'R.E. 12'),
    ('Subs. 9', 'SUB. R.E. 9'),
    ('Subs. 5', 'SUB. R.E. 5'),
    ('Subs. 10', 'SUB. R.E. 10'),
]


_REPEATED_PAIR_RE = re.compile(r'^(.+)-\1$')


def _normalize_text(value, keep_linebreaks=False):
    if value is None:
        return ''

    text = str(value)
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    if text.strip().upper() == 'NULL':
        return ''

    if keep_linebreaks:
        lines = [re.sub(r'\s+', ' ', line).strip() for line in text.split('\n')]
        return '\n'.join([line for line in lines if line]).strip()

    return re.sub(r'\s+', ' ', text).strip()


def _normalize_optional_text(value, placeholders=None, keep_linebreaks=False):
    normalized = _normalize_text(value, keep_linebreaks=keep_linebreaks)
    if not normalized:
        return ''

    normalized_lower = normalized.lower()
    invalid_values = {
        '-',
        '.',
        '0',
        'null',
        's/inf.',
        's/inf',
        's/info',
        's/info.',
        's/datos',
        's/i',
        's/i.',
        'sin dato',
        'sin dato-',
        'no declara',
        'sin información',
        'sin informacion',
    }

    if placeholders:
        invalid_values.update({str(item).strip().lower() for item in placeholders if str(item).strip()})

    if normalized_lower in invalid_values:
        return ''

    return normalized


def _format_date(value):
    normalized = _normalize_optional_text(value)
    if not normalized:
        return ''

    if isinstance(value, datetime):
        dt_value = value
    elif isinstance(value, date):
        dt_value = datetime.combine(value, datetime.min.time())
    else:
        raw = normalized.split(' ')[0]
        dt_value = None
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%d-%m-%Y', '%d-%m-%y', '%d/%m/%y'):
            try:
                dt_value = datetime.strptime(raw, fmt)
                break
            except ValueError:
                continue
        if dt_value is None:
            return normalized

    meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    return f"{meses[dt_value.month - 1]} {dt_value.day:02d} {dt_value.year}"


def _bool_to_si_no(value):
    if value is None:
        return ''

    if isinstance(value, bool):
        return 'Si' if value else 'No'

    normalized = _normalize_text(value).lower()
    if normalized in {'true', 't', '1', 'si', 'sí'}:
        return 'Si'
    if normalized in {'false', 'f', '0', 'no'}:
        return 'No'
    return _normalize_text(value)


def _build_persona(apellido, nombre, documento='', force_empty_parentheses=False):
    apellido_txt = _normalize_text(apellido)
    nombre_txt = _normalize_text(nombre)
    documento_txt = _normalize_text(documento)

    if apellido_txt and nombre_txt:
        persona = f'{apellido_txt}, {nombre_txt}'
    else:
        persona = apellido_txt or nombre_txt

    if documento_txt:
        persona = f'{persona}({documento_txt})' if persona else documento_txt
    elif force_empty_parentheses and persona:
        persona = f'{persona}()'

    return persona


def _build_domicilio(calle, nro, cod_postal, localidad, departamento):
    calle_txt = _normalize_text(calle)
    nro_txt = _normalize_text(nro)
    cod_postal_txt = _normalize_text(cod_postal)
    localidad_txt = _normalize_text(localidad)
    departamento_txt = _normalize_text(departamento)
    localidad_con_cp = localidad_txt

    if cod_postal_txt and localidad_txt:
        localidad_con_cp = f'({cod_postal_txt}) {localidad_txt}'
    elif cod_postal_txt:
        localidad_con_cp = f'({cod_postal_txt})'

    primera_parte = ' '.join([value for value in [calle_txt, nro_txt] if value]).strip()
    return ' - '.join([value for value in [primera_parte, localidad_con_cp, departamento_txt] if value])


def _build_localidad_departamento(localidad, departamento):
    localidad_txt = _normalize_text(localidad)
    departamento_txt = _normalize_text(departamento)
    return ' - '.join([value for value in [localidad_txt, departamento_txt] if value])


def _build_phone_with_area(cod_area, telefono):
    cod_area_txt = _normalize_optional_text(cod_area, placeholders={'000'})
    telefono_txt = _normalize_optional_text(telefono, placeholders={'00000'})

    if cod_area_txt and telefono_txt:
        return f'{cod_area_txt} - {telefono_txt}'
    return telefono_txt or cod_area_txt


def _build_cue_anexo(cue, anexo):
    cue_txt = _normalize_text(cue)
    anexo_txt = _normalize_text(anexo)

    if not cue_txt:
        return ''
    if not anexo_txt:
        return cue_txt

    if anexo_txt.isdigit():
        anexo_txt = anexo_txt.zfill(2)

    return f'{cue_txt}-{anexo_txt}'


@lru_cache(maxsize=None)
def _get_campo_codigo_pairs(id_campo_prov):
    with connections[PADRON_DB].cursor() as cursor:
        cursor.execute(
            """
            SELECT codigo::text, COALESCE(descripcion, '')
            FROM campo_prov_codigo
            WHERE id_campo_prov = %s
            ORDER BY LENGTH(codigo::text) DESC, codigo::text
            """,
            [id_campo_prov],
        )
        return tuple((_normalize_text(code), _normalize_text(desc)) for code, desc in cursor.fetchall())


def _split_campo_codigo(value, id_campo_prov=None):
    normalized = _normalize_text(value)
    if not normalized:
        return '', ''

    if id_campo_prov is not None:
        for code, description in _get_campo_codigo_pairs(id_campo_prov):
            if not code:
                continue
            if normalized == code:
                return code, description or code
            prefix = f'{code}-'
            if normalized.startswith(prefix):
                return code, _normalize_text(normalized[len(prefix):]) or description or code

    repeated_match = _REPEATED_PAIR_RE.match(normalized)
    if repeated_match:
        half = _normalize_text(repeated_match.group(1))
        return half, half

    if '-' in normalized:
        code, description = normalized.split('-', 1)
        return _normalize_text(code), _normalize_text(description)

    return normalized, ''


def _format_desc_only(value, id_campo_prov=None):
    code, description = _split_campo_codigo(value, id_campo_prov)
    return description or code


def _format_desc_with_code(value, id_campo_prov=None):
    normalized = _normalize_text(value)
    code, description = _split_campo_codigo(value, id_campo_prov)
    if normalized == '-' and not code and not description:
        return '()'
    if description and code:
        return f'{description} ({code})'
    return description or code


def _format_code_then_desc(value, id_campo_prov=None):
    code, description = _split_campo_codigo(value, id_campo_prov)
    if code and description:
        return f'{code}, {description}'
    return code or description


def _format_desc_then_code(value, id_campo_prov=None):
    code, description = _split_campo_codigo(value, id_campo_prov)
    if code and description:
        return f'{description}, {code}'
    return description or code


RESPONSABLE_SQL = """
    BTRIM(
        CONCAT(
            COALESCE(vl.responsable_apellido, ''),
            CASE
                WHEN COALESCE(vl.responsable_nombre, '') <> '' AND COALESCE(vl.responsable_apellido, '') <> ''
                    THEN ', ' || vl.responsable_nombre
                ELSE COALESCE(vl.responsable_nombre, '')
            END,
            CASE
                WHEN vl.documento_responsable IS NOT NULL
                    THEN '(' || vl.documento_responsable::text || ')'
                ELSE ''
            END
        )
    )
"""

DOMICILIO_SEARCH_SQL = """
    COALESCE(
        NULLIF(
            TRIM(
                CONCAT(
                    COALESCE(NULLIF(BTRIM(dom.calle), ''), ''),
                    CASE
                        WHEN COALESCE(NULLIF(BTRIM(dom.nro), ''), '') <> ''
                            THEN ' ' || BTRIM(dom.nro)
                        ELSE ''
                    END,
                    CASE
                        WHEN COALESCE(NULLIF(BTRIM(dom.barrio), ''), '') <> ''
                            THEN ', ' || BTRIM(dom.barrio)
                        ELSE ''
                    END,
                    CASE
                        WHEN COALESCE(NULLIF(BTRIM(dom.cod_postal), ''), '') <> ''
                            THEN ' (' || BTRIM(dom.cod_postal) || ')'
                        ELSE ''
                    END
                )
            ),
            ''
        ),
        CONCAT_WS(
            ' ',
            COALESCE(vl.calle, ''),
            COALESCE(vl.nro, ''),
            COALESCE(vl.cod_postal, ''),
            COALESCE(vl.localidad_nombre, ''),
            COALESCE(vl.departamento_nombre, '')
        )
    )
"""

CAMPO_SQL = {
    'cue': 'vl.cue::text',
    'anexo': "COALESCE(BTRIM(vl.anexo), '')",
    'codigo_jurisdiccional': "COALESCE(BTRIM(vl.codigo_jurisdiccional), '')",
    'ambito': "COALESCE(BTRIM(vl.ambito), '')",
    'sede': "CASE WHEN vl.sede THEN 'Si' ELSE 'No' END",
    'sede_adm': "CASE WHEN vl.sede_administrativa THEN 'Si' ELSE 'No' END",
    'periodo_funcionamiento': "COALESCE(BTRIM(vl.periodo_funcionamiento), '')",
    'sector': "COALESCE(BTRIM(vl.sector), '')",
    'dependencia': "COALESCE(BTRIM(vl.dependencia), '')",
    'nombre': "COALESCE(BTRIM(vl.nombre), '')",
    'establecimiento': "COALESCE(BTRIM(ve.nombre), '')",
    'estado': "COALESCE(BTRIM(vl.estado_localizacion), '')",
    'responsable': RESPONSABLE_SQL,
    'tipo_oferta': "COALESCE(BTRIM(otipos.tipo_ofertas), '')",
    'localidad': "COALESCE(BTRIM(vl.localidad_nombre), '')",
    'departamento': "COALESCE(BTRIM(vl.departamento_nombre), '')",
    'domicilio_ppal': DOMICILIO_SEARCH_SQL,
    'cod_area_tel': "COALESCE(BTRIM(vl.telefono_cod_area), '')",
    'telefono': "COALESCE(BTRIM(vl.telefono), '')",
    'email': "COALESCE(BTRIM(vl.email), '')",
    'modalidades_complementarias': "COALESCE(BTRIM(mods.modalidades), '')",
    'tel_director_regional': "COALESCE(BTRIM(vl.cp_tedirregional), '')",
    'email_dir_regional': "COALESCE(BTRIM(vl.cp_emaildirregional), '')",
    'supervisor_tecnico': "COALESCE(BTRIM(vl.cp_supervisortecnico), '')",
    'microregion': "COALESCE(BTRIM(vl.cp_esvat4), '')",
    'udt': "COALESCE(BTRIM(vl.cp_esvat6), '')",
    'regional_actual': "COALESCE(BTRIM(vl.cp_esvat5), '')",
    'zona_provincial': "COALESCE(BTRIM(vl.cp_zonaprovincial), '')",
    'anterior_regional': "COALESCE(BTRIM(vl.cp_esvat3), '')",
    'director_regional': "COALESCE(BTRIM(vl.cp_directorregional), '')",
    'plan_de_obra': "COALESCE(BTRIM(vl.cp_plandeobra), '')",
    'patrimonio_edilicio': "COALESCE(BTRIM(vl.cp_patrimonioedilicio), '')",
    'fecha_creacion_edificio': "COALESCE(BTRIM(vl.cp_estfechacreacionedificio), '')",
    'nro_biblioteca': "COALESCE(BTRIM(vl.cp_esvar1), '')",
    'cuof': "COALESCE(BTRIM(vl.cp_esvar4), '')",
    'cui': "COALESCE(BTRIM(vl.cp_esvar2), '')",
    'cua': "COALESCE(BTRIM(vl.cp_esvar3), '')",
    'tipo_albergue': "COALESCE(BTRIM(vl.cp_esvat1), '')",
    'regional_hasta_2015': "COALESCE(BTRIM(vl.cp_p8104_localizacion_1019638033), '')",
    'inst_legal_edificio': "COALESCE(BTRIM(vl.cp_edif_instlegal_creaciondeestablecimiento), '')",
    'tel_supervisor': "COALESCE(BTRIM(loc_tel_sup.valor), '')",
    'email_supervisor': "COALESCE(BTRIM(loc_email_sup.valor), '')",
    'regional_hasta_2020': "COALESCE(BTRIM(vl.cp_reg_hasta_2020), '')",
}

VISIBLE_NAME_MAP = {
    'cue': 'Cue',
    'anexo': 'Anexo',
    'codigo_jurisdiccional': 'Codigo Jurisdiccional',
    'ambito': 'Ambito',
    'sede': 'Sede',
    'sede_adm': 'Sede Adm.',
    'periodo_funcionamiento': 'Período de Func.',
    'sector': 'Sector',
    'dependencia': 'Dependencia',
    'nombre': 'Nombre',
    'establecimiento': 'Establecimiento',
    'estado': 'Estado',
    'responsable': 'Responsable',
    'tipo_oferta': 'Tipo de Oferta',
    'localidad': 'Localidad',
    'departamento': 'Departamento',
    'domicilio_ppal': 'Domicilio Ppal.',
    'cod_area_tel': 'Código de Área Teléfonico',
    'telefono': 'Teléfono',
    'email': 'Email',
    'modalidades_complementarias': 'Modalidades Complementarias',
    'tel_director_regional': 'Teléfono Director Regional',
    'email_dir_regional': 'Email Director Regional',
    'supervisor_tecnico': 'Supervisor Técnico',
    'microregion': 'Microregión',
    'udt': 'Unidad de Desarrollo Territorial UDT',
    'regional_actual': 'Regional Educativa Actual',
    'zona_provincial': 'Zona Provincial',
    'anterior_regional': 'Anterior Regional Educativa',
    'director_regional': 'Director Regional',
    'plan_de_obra': 'Plan de Obra',
    'patrimonio_edilicio': 'Patrimonio Edilicio',
    'fecha_creacion_edificio': 'Fecha de creación del Edificio',
    'nro_biblioteca': 'Nro de Biblioteca asociado a la Localización',
    'cuof': 'CUOF Código Único de Oficina de la Localización',
    'cui': 'CUI Código Único de Infrestructura de la Localización',
    'cua': 'CUA Código Único de Agrupamiento de la Localización',
    'tipo_albergue': 'Tipo de Albergue',
    'regional_hasta_2015': 'Regional Educativa Hasta 2015',
    'inst_legal_edificio': 'Instrumento Legal de creación del Edificio',
    'tel_supervisor': 'Teléfono Supervisor',
    'email_supervisor': 'Email Supervisor',
    'regional_hasta_2020': 'Regional Ed. HASTA 2020',
}

COLUMNAS_EXPORTACION = [
    ('Cue', 'cue'),
    ('Anexo', 'anexo'),
    ('Codigo Jurisdiccional', 'codigo_jurisdiccional'),
    ('Ambito', 'ambito'),
    ('Sede', 'sede'),
    ('Sede Adm.', 'sede_adm'),
    ('Período de Func.', 'periodo_funcionamiento'),
    ('Nombre', 'nombre'),
    ('Establecimiento', 'establecimiento'),
    ('Responsable', 'responsable'),
    ('Estado', 'estado'),
    ('Sector', 'sector'),
    ('Dependencia', 'dependencia'),
    ('Tipo de Oferta', 'tipo_oferta'),
    ('Localidad', 'localidad'),
    ('Departamento', 'departamento'),
    ('Domicilio Ppal.', 'domicilio_ppal'),
    ('Código de Área Teléfonico', 'cod_area_tel'),
    ('Teléfono', 'telefono'),
    ('Email', 'email'),
    ('Modalidades Complementarias', 'modalidades_complementarias'),
    ('Teléfono Director Regional', 'tel_director_regional'),
    ('Email Director Regional', 'email_dir_regional'),
    ('Supervisor Técnico', 'supervisor_tecnico'),
    ('Microregión', 'microregion'),
    ('Unidad de Desarrollo Territorial UDT', 'udt'),
    ('Regional Educativa Actual', 'regional_actual'),
    ('Zona Provincial', 'zona_provincial'),
    ('Anterior Regional Educativa', 'anterior_regional'),
    ('Director Regional', 'director_regional'),
    ('Plan de Obra', 'plan_de_obra'),
    ('Patrimonio Edilicio', 'patrimonio_edilicio'),
    ('Fecha de creación del Edificio', 'fecha_creacion_edificio'),
    ('Nro de Biblioteca asociado a la Localización', 'nro_biblioteca'),
    ('CUOF Código Único de Oficina de la Localización', 'cuof'),
    ('CUI Código Único de Infrestructura de la Localización', 'cui'),
    ('CUA Código Único de Agrupamiento de la Localización', 'cua'),
    ('Tipo de Albergue', 'tipo_albergue'),
    ('Regional Educativa Hasta 2015', 'regional_hasta_2015'),
    ('Instrumento Legal de creación del Edificio', 'inst_legal_edificio'),
    ('Teléfono Supervisor', 'tel_supervisor'),
    ('Email Supervisor', 'email_supervisor'),
    ('Regional Ed. HASTA 2020', 'regional_hasta_2020'),
]

_SELECT_FIELDS = f"""
    SELECT
        vl.id_localizacion AS id,
        COALESCE(BTRIM(vl.cue::text), '') AS cue,
        COALESCE(BTRIM(vl.anexo), '') AS anexo,
        COALESCE(BTRIM(vl.codigo_jurisdiccional), '') AS codigo_jurisdiccional,
        COALESCE(BTRIM(vl.ambito), '') AS ambito,
        vl.sede,
        vl.sede_administrativa AS sede_adm,
        COALESCE(BTRIM(vl.periodo_funcionamiento), '') AS periodo_funcionamiento,
        COALESCE(BTRIM(vl.nombre), '') AS nombre,
        COALESCE(BTRIM(ve.nombre), '') AS establecimiento,
        {RESPONSABLE_SQL} AS responsable,
        COALESCE(BTRIM(vl.estado_localizacion), '') AS estado,
        COALESCE(BTRIM(vl.sector), '') AS sector,
        COALESCE(BTRIM(vl.dependencia), '') AS dependencia,
        COALESCE(BTRIM(otipos.tipo_ofertas), '') AS tipo_oferta,
        COALESCE(BTRIM(vl.localidad_nombre), '') AS localidad,
        COALESCE(BTRIM(vl.departamento_nombre), '') AS departamento,
        {DOMICILIO_SEARCH_SQL} AS domicilio_ppal,
        COALESCE(BTRIM(vl.telefono_cod_area), '') AS cod_area_tel,
        COALESCE(BTRIM(vl.telefono), '') AS telefono,
        COALESCE(BTRIM(vl.email), '') AS email,
        COALESCE(BTRIM(mods.modalidades), '') AS modalidades_complementarias,
        COALESCE(BTRIM(vl.cp_tedirregional), '') AS tel_director_regional,
        COALESCE(BTRIM(vl.cp_emaildirregional), '') AS email_dir_regional,
        COALESCE(BTRIM(vl.cp_supervisortecnico), '') AS supervisor_tecnico,
        COALESCE(BTRIM(vl.cp_esvat4), '') AS microregion,
        COALESCE(BTRIM(vl.cp_esvat6), '') AS udt,
        COALESCE(BTRIM(vl.cp_esvat5), '') AS regional_actual,
        COALESCE(BTRIM(vl.cp_zonaprovincial), '') AS zona_provincial,
        COALESCE(BTRIM(vl.cp_esvat3), '') AS anterior_regional,
        COALESCE(BTRIM(vl.cp_directorregional), '') AS director_regional,
        COALESCE(BTRIM(vl.cp_plandeobra), '') AS plan_de_obra,
        COALESCE(BTRIM(vl.cp_patrimonioedilicio), '') AS patrimonio_edilicio,
        COALESCE(BTRIM(vl.cp_estfechacreacionedificio), '') AS fecha_creacion_edificio,
        COALESCE(BTRIM(vl.cp_esvar1), '') AS nro_biblioteca,
        COALESCE(BTRIM(vl.cp_esvar4), '') AS cuof,
        COALESCE(BTRIM(vl.cp_esvar2), '') AS cui,
        COALESCE(BTRIM(vl.cp_esvar3), '') AS cua,
        COALESCE(BTRIM(vl.cp_esvat1), '') AS tipo_albergue,
        COALESCE(BTRIM(vl.cp_p8104_localizacion_1019638033), '') AS regional_hasta_2015,
        COALESCE(BTRIM(vl.cp_edif_instlegal_creaciondeestablecimiento), '') AS inst_legal_edificio,
        COALESCE(BTRIM(loc_tel_sup.valor), '') AS tel_supervisor,
        COALESCE(BTRIM(loc_email_sup.valor), '') AS email_supervisor,
        COALESCE(BTRIM(vl.cp_reg_hasta_2020), '') AS regional_hasta_2020
"""

_BASE_SQL = """
    FROM vp_localizaciones vl
    LEFT JOIN vp_establecimientos ve
      ON ve.id_establecimiento = vl.id_establecimiento
    LEFT JOIN loc_campo_prov_valor loc_tel_sup
      ON loc_tel_sup.id_localizacion = vl.id_localizacion
     AND loc_tel_sup.id_campo_prov = 1019638042
    LEFT JOIN loc_campo_prov_valor loc_email_sup
      ON loc_email_sup.id_localizacion = vl.id_localizacion
     AND loc_email_sup.id_campo_prov = 1019638043
    LEFT JOIN LATERAL (
        SELECT STRING_AGG(BTRIM(ot.descripcion), ', ' ORDER BY ot.c_oferta, ol.id_oferta_local) AS tipo_ofertas
        FROM oferta_local ol
        JOIN oferta_tipo ot ON ot.c_oferta = ol.c_oferta
        WHERE ol.id_localizacion = vl.id_localizacion
    ) otipos ON TRUE
    LEFT JOIN LATERAL (
        SELECT STRING_AGG(BTRIM(m2.descripcion), ', ' ORDER BY m2.orden, m2.descripcion) AS modalidades
        FROM localizacion_modalidad2_assn lm2
        JOIN modalidad2_tipo m2 ON m2.c_modalidad2 = lm2.c_modalidad2
        WHERE lm2.id_localizacion = vl.id_localizacion
    ) mods ON TRUE
    LEFT JOIN LATERAL (
        SELECT
            d.calle,
            d.nro,
            d.barrio,
            d.cod_postal
        FROM localizacion_domicilio ld
        JOIN domicilio d ON d.id_domicilio = ld.id_domicilio
        WHERE ld.id_localizacion = vl.id_localizacion
        ORDER BY CASE WHEN ld.c_tipo_dom = 1 THEN 0 ELSE 1 END, ld.c_tipo_dom, ld.id_domicilio
        LIMIT 1
    ) dom ON TRUE
"""


class _RawPage:
    def __init__(self, data_sql, count_sql, params, db_alias):
        self._data_sql = data_sql
        self._count_sql = count_sql
        self._params = params
        self._db_alias = db_alias
        self._length = None

    def __len__(self):
        if self._length is None:
            with connections[self._db_alias].cursor() as cursor:
                cursor.execute(self._count_sql, self._params)
                self._length = cursor.fetchone()[0]
        return self._length

    def __getitem__(self, item):
        if not isinstance(item, slice):
            raise TypeError('Los elementos deben pedirse como slice.')

        start = item.start or 0
        stop = item.stop or start
        limit = max(stop - start, 0)
        sql = f"{self._data_sql} LIMIT {limit} OFFSET {start}"

        with connections[self._db_alias].cursor() as cursor:
            cursor.execute(sql, self._params)
            cols = [d[0] for d in cursor.description]
            return [dict(zip(cols, row)) for row in cursor.fetchall()]


OPERADOR_SQL = {
    '0': ('ILIKE', lambda v: f'%{v}%'),
    '1': ('NOT ILIKE', lambda v: f'%{v}%'),
    '2': ('=', lambda v: v),
    '3': ('>', lambda v: v),
    '4': ('>=', lambda v: v),
    '5': ('<', lambda v: v),
    '6': ('<=', lambda v: v),
    '7': ('!=', lambda v: v),
}

_ACCENTED_CHARS = '\u00c1\u00c9\u00cd\u00d3\u00da\u00dc\u00d1\u00e1\u00e9\u00ed\u00f3\u00fa\u00fc\u00f1'
_UNACCENTED_CHARS = 'AEIOUUNaeiouun'
_ACCENT_TRANSLATION = str.maketrans(_ACCENTED_CHARS, _UNACCENTED_CHARS)


def _fold_filter_text(value):
    return _normalize_text(value).translate(_ACCENT_TRANSLATION).lower()


def _folded_sql(expr):
    return f"LOWER(TRANSLATE(BTRIM(COALESCE(({expr})::text, '')), '{_ACCENTED_CHARS}', '{_UNACCENTED_CHARS}'))"


def _filter_tokens(value):
    folded = _fold_filter_text(value)
    tokens = [token for token in re.split(r'[^a-z0-9]+', folded) if token]
    return list(dict.fromkeys(tokens))


def _like_token_values(value):
    return [f'%{token}%' for token in _filter_tokens(value)]


def _tipo_oferta_filter_values(oper, value):
    folded = _fold_filter_text(value)
    return _like_token_values(value) if oper in {'0', '1'} else [folded]


def _is_oferta_estado_filter_value(value):
    folded = _fold_filter_text(value)
    return any(folded == _fold_filter_text(option) for option in OFERTA_ESTADO_FILTER_OPTIONS)


def _split_tipo_oferta_filter_params(oper, values):
    tipo_param_groups = []
    estado_params = []

    for value in values:
        filter_values = _tipo_oferta_filter_values(oper, value)
        if _is_oferta_estado_filter_value(value):
            estado_params.extend(filter_values)
        else:
            tipo_param_groups.append(filter_values)

    return tipo_param_groups, estado_params


def _append_tipo_oferta_filter(clauses, params, oper, values):
    tipo_param_groups, estado_params = _split_tipo_oferta_filter_params(oper, values)
    if not tipo_param_groups and not estado_params:
        return

    clauses.append(_build_tipo_oferta_clause(
        oper,
        [len(group) for group in tipo_param_groups],
        len(estado_params),
    ))
    for group in tipo_param_groups:
        params.extend(group)
    params.extend(estado_params)


def _build_text_search_clause(expr, value):
    tokens = _filter_tokens(value)
    if not tokens:
        return '', []

    folded_expr = _folded_sql(expr)
    return ' AND '.join([f"{folded_expr} LIKE %s" for _ in tokens]), [f'%{token}%' for token in tokens]


def _tipo_oferta_clause_operator(oper):
    if oper in {'0', '1'}:
        return 'LIKE'
    if oper in {'2', '7'}:
        return '='
    return OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])[0]


def _build_tipo_oferta_clause(oper, tipo_value_counts=None, estado_count=0):
    tipo_value_counts = tipo_value_counts or []
    descripcion = _folded_sql('otf.descripcion')
    estado = _folded_sql('etf.descripcion')
    op_str = _tipo_oferta_clause_operator(oper)
    conditions = []

    if tipo_value_counts:
        tipo_checks = []
        for token_count in tipo_value_counts:
            token_checks = ' AND '.join([f"{descripcion} {op_str} %s" for _ in range(max(token_count, 1))])
            tipo_checks.append(f"({token_checks})")
        conditions.append(f"({' OR '.join(tipo_checks)})")

    if estado_count:
        estado_checks = ' OR '.join([f"{estado} {op_str} %s" for _ in range(max(estado_count, 1))])
        conditions.append(f"({estado_checks})")

    exists_operator = 'NOT EXISTS' if oper in {'1', '7'} else 'EXISTS'
    return (
        f"{exists_operator} ("
        "SELECT 1 "
        "FROM oferta_local olf "
        "JOIN oferta_tipo otf ON otf.c_oferta = olf.c_oferta "
        "LEFT JOIN estado_tipo etf ON etf.c_estado = olf.c_estado "
        "WHERE olf.id_localizacion = vl.id_localizacion "
        f"AND {' AND '.join(conditions)}"
        ")"
    )


def _build_modalidades_clause(oper, token_count=1):
    descripcion = _folded_sql('m2f.descripcion')

    if oper == '1':
        token_checks = ' AND '.join([f"{descripcion} LIKE %s" for _ in range(max(token_count, 1))])
        return (
            "NOT EXISTS ("
            "SELECT 1 "
            "FROM localizacion_modalidad2_assn lm2f "
            "JOIN modalidad2_tipo m2f ON m2f.c_modalidad2 = lm2f.c_modalidad2 "
            "WHERE lm2f.id_localizacion = vl.id_localizacion "
            f"AND {token_checks}"
            ")"
        )

    if oper == '7':
        return (
            "NOT EXISTS ("
            "SELECT 1 "
            "FROM localizacion_modalidad2_assn lm2f "
            "JOIN modalidad2_tipo m2f ON m2f.c_modalidad2 = lm2f.c_modalidad2 "
            "WHERE lm2f.id_localizacion = vl.id_localizacion "
            f"AND {descripcion} = %s"
            ")"
        )

    op_str, _ = OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])
    if oper in {'0', '2'}:
        op_str = 'LIKE' if oper == '0' else '='
    token_checks = ' AND '.join([f"{descripcion} {op_str} %s" for _ in range(max(token_count, 1))])
    return (
        "EXISTS ("
        "SELECT 1 "
        "FROM localizacion_modalidad2_assn lm2f "
        "JOIN modalidad2_tipo m2f ON m2f.c_modalidad2 = lm2f.c_modalidad2 "
        "WHERE lm2f.id_localizacion = vl.id_localizacion "
        f"AND {token_checks}"
        ")"
    )


def _build_where(request):
    clauses = []
    params = []
    grouped_positive_filters = {}
    geo_positive_filters = []

    campos = request.GET.getlist('campo_filtro')
    opers = request.GET.getlist('operador_filtro')
    valores = request.GET.getlist('valor_filtro')

    for index, campo in enumerate(campos):
        campo = campo.strip()
        oper = opers[index].strip() if index < len(opers) else '0'
        valor = valores[index].strip() if index < len(valores) else ''

        if not campo or not valor or campo not in CAMPO_SQL:
            continue

        if oper in {'0', '2'}:
            grouped_positive_filters.setdefault((campo, oper), [])
            if valor not in grouped_positive_filters[(campo, oper)]:
                grouped_positive_filters[(campo, oper)].append(valor)
            continue

        if campo == 'tipo_oferta':
            _append_tipo_oferta_filter(clauses, params, oper, [valor])
            continue

        if campo == 'modalidades_complementarias':
            mod_values = _like_token_values(valor) if oper in {'0', '1'} else [_fold_filter_text(valor)]
            clauses.append(_build_modalidades_clause(oper, len(mod_values)))
            params.extend(mod_values)
            continue

        col = CAMPO_SQL[campo]
        op_str, val_fn = OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])
        if oper == '0':
            token_clause, token_params = _build_text_search_clause(col, valor)
            if token_clause:
                clauses.append(token_clause)
                params.extend(token_params)
            continue

        clauses.append(f"{col}::text {op_str} %s")
        params.append(val_fn(valor))

    for (campo, oper), values in grouped_positive_filters.items():
        if campo in {'localidad', 'departamento'}:
            geo_positive_filters.append((campo, oper, values))
            continue

        if campo == 'tipo_oferta':
            _append_tipo_oferta_filter(clauses, params, oper, values)
            continue

        if campo == 'modalidades_complementarias':
            if len(values) == 1:
                mod_values = _like_token_values(values[0]) if oper in {'0', '1'} else [_fold_filter_text(values[0])]
                clauses.append(_build_modalidades_clause(oper, len(mod_values)))
                params.extend(mod_values)
                continue

            subclauses = []
            for value in values:
                mod_values = _like_token_values(value) if oper in {'0', '1'} else [_fold_filter_text(value)]
                subclauses.append(_build_modalidades_clause(oper, len(mod_values)))
                params.extend(mod_values)
            clauses.append(f"({' OR '.join(subclauses)})")
            continue

        col = CAMPO_SQL[campo]
        op_str, val_fn = OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])

        if len(values) == 1:
            if oper == '0':
                token_clause, token_params = _build_text_search_clause(col, values[0])
                if token_clause:
                    clauses.append(token_clause)
                    params.extend(token_params)
                continue

            clauses.append(f"{col}::text {op_str} %s")
            params.append(val_fn(values[0]))
            continue

        subclauses = []
        for value in values:
            if oper == '0':
                token_clause, token_params = _build_text_search_clause(col, value)
                if token_clause:
                    subclauses.append(token_clause)
                    params.extend(token_params)
                continue

            subclauses.append(f"{col}::text {op_str} %s")
            params.append(val_fn(value))
        clauses.append(f"({' OR '.join(subclauses)})")

    if geo_positive_filters:
        geo_subclauses = []

        for campo, oper, values in geo_positive_filters:
            col = CAMPO_SQL[campo]
            op_str, val_fn = OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])

            for value in values:
                if oper == '0':
                    token_clause, token_params = _build_text_search_clause(col, value)
                    if token_clause:
                        geo_subclauses.append(token_clause)
                        params.extend(token_params)
                    continue

                geo_subclauses.append(f"{col}::text {op_str} %s")
                params.append(val_fn(value))

        if geo_subclauses:
            clauses.append(f"({' OR '.join(geo_subclauses)})")

    q = request.GET.get('q', '').strip()
    if q:
        token_groups = []
        for token in _filter_tokens(q):
            global_search_clauses = [f"{_folded_sql(sql_expr)} LIKE %s" for sql_expr in CAMPO_SQL.values()]
            token_groups.append(f"({' OR '.join(global_search_clauses)})")
            params.extend([f'%{token}%'] * len(global_search_clauses))
        if token_groups:
            clauses.append(f"({' AND '.join(token_groups)})")

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ''
    return where, params


def _armar_texto_filtros(request):
    partes = []
    operadores_txt = {
        '0': 'parecido a',
        '1': 'no parecido a',
        '2': 'igual a',
        '3': 'mayor a',
        '4': 'mayor o igual a',
        '5': 'menor a',
        '6': 'menor o igual a',
        '7': 'distinto de',
    }

    if request.GET.get('formato') == 'excel_todo':
       return 'Sin filtros aplicados'

    campos = request.GET.getlist('campo_filtro')
    opers = request.GET.getlist('operador_filtro')
    valores = request.GET.getlist('valor_filtro')

    q = request.GET.get('q', '').strip()
    if q:
        partes.append(f'Búsqueda: {q}')

    for index, campo in enumerate(campos):
        campo = campo.strip()
        valor = valores[index].strip() if index < len(valores) else ''
        oper = opers[index].strip() if index < len(opers) else '0'

        if not campo or not valor:
            continue

        partes.append(
            f"{VISIBLE_NAME_MAP.get(campo, campo)} {operadores_txt.get(oper, 'parecido a')}: {valor}"
        )

    return ' | '.join(partes) if partes else 'Sin filtros aplicados'


def _formatear_fecha_estilo_seguimiento():
    ahora = datetime.now()
    dia = ahora.day
    mes = ahora.month
    anio = ahora.year
    hora_12 = ahora.strftime('%I').lstrip('0') or '0'
    minuto = ahora.strftime('%M')
    ampm = 'p. m.' if ahora.hour >= 12 else 'a. m.'
    return f'{dia}/{mes}/{anio} a las {hora_12}:{minuto} {ampm}'


def _resolver_columnas_exportar(request, formato):
    if formato == 'excel_todo':
        return COLUMNAS_EXPORTACION

    visibles = {value.strip() for value in request.GET.getlist('visible_col') if value.strip()}
    if not visibles:
        return COLUMNAS_EXPORTACION

    columnas = [item for item in COLUMNAS_EXPORTACION if item[1] in visibles]
    return columnas or COLUMNAS_EXPORTACION


def _exportar_excel(datos_exportar, formato, request):
    columnas_mapeo = _resolver_columnas_exportar(request, formato)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Localizaciones'

    total_columnas = len(columnas_mapeo)
    ultima_columna = get_column_letter(total_columnas)

    ws.merge_cells(f'A1:{ultima_columna}1')
    ws['A1'] = 'Informe Localizaciones'
    ws['A1'].font = Font(bold=True, size=10)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'A2:{ultima_columna}2')
    ws['A2'] = f'Informe generado el: {_formatear_fecha_estilo_seguimiento()}'
    ws['A2'].font = Font(size=9)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'A3:{ultima_columna}3')
    ws['A3'] = f'Filtros aplicados: {_armar_texto_filtros(request)}'
    ws['A3'].font = Font(size=9)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20

    header_row = 4
    for col_idx, (titulo_col, _) in enumerate(columnas_mapeo, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=titulo_col)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for item in datos_exportar:
        fila = []
        for _, clave in columnas_mapeo:
            fila.append(_normalize_text(item.get(clave), keep_linebreaks=True))
        ws.append(fila)

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{ultima_columna}{ws.max_row}'

    if formato == 'excel_todo':
        for col_num in range(1, total_columnas + 1):
            titulo_col = columnas_mapeo[col_num - 1][0]
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = min(max(len(str(titulo_col)) + 4, 14), 28)
    else:
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.font = Font(size=9)

        for col_num in range(1, total_columnas + 1):
            col_letter = get_column_letter(col_num)
            max_length = 0
            for row_num in range(1, ws.max_row + 1):
                value = ws[f'{col_letter}{row_num}'].value
                max_length = max(max_length, len(str(value)) if value is not None else 0)
            ws.column_dimensions[col_letter].width = min(max_length + 2, 42)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_archivo = datetime.now().strftime('%Y%m%d_%H%M')
    sufijo = 'Filtros' if formato == 'excel_pagina' else 'Todo'
    nombre_archivo = f'Localizaciones_{sufijo}_{fecha_archivo}.xlsx'

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response

def _dictfetchall(cursor):
    columns = [column[0] for column in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _fetch_one(sql, params):
    with connections[PADRON_DB].cursor() as cursor:
        cursor.execute(sql, params)
        rows = _dictfetchall(cursor)
    return rows[0] if rows else None


def _fetch_all(sql, params):
    with connections[PADRON_DB].cursor() as cursor:
        cursor.execute(sql, params)
        return _dictfetchall(cursor)


def _sanitize_json_payload(value):
    if isinstance(value, dict):
        return {key: _sanitize_json_payload(inner_value) for key, inner_value in value.items()}
    if isinstance(value, list):
        return [_sanitize_json_payload(item) for item in value]
    if value is None:
        return ''
    return value


def _make_option(value, label=None):
    value_txt = _normalize_optional_text(value)
    label_txt = _normalize_optional_text(label if label is not None else value)
    if not value_txt or not label_txt:
        return None
    return {'value': value_txt, 'label': label_txt}


def _fetch_catalog_options(cursor, sql, value_builder=None, label_builder=None):
    cursor.execute(sql)
    options = []
    seen = set()

    for row in cursor.fetchall():
        raw_value = value_builder(*row) if value_builder else row[0]
        raw_label = label_builder(*row) if label_builder else raw_value
        option = _make_option(raw_value, raw_label)
        if not option:
            continue
        key = (option['value'], option['label'])
        if key in seen:
            continue
        seen.add(key)
        options.append(option)

    return options


def _fetch_distinct_formatted_options(cursor, column_name, formatter):
    cursor.execute(
        f"""
        SELECT DISTINCT {column_name}
        FROM vp_localizaciones
        WHERE COALESCE(BTRIM({column_name}), '') <> ''
        """
    )
    options = []
    seen = set()

    for (raw_value,) in cursor.fetchall():
        raw_txt = _normalize_optional_text(raw_value, placeholders={'0'})
        if not raw_txt:
            continue
        label_txt = _normalize_optional_text(formatter(raw_txt), placeholders={'0'})
        if not label_txt:
            continue
        key = (raw_txt, label_txt)
        if key in seen:
            continue
        seen.add(key)
        options.append({'value': raw_txt, 'label': label_txt})

    options.sort(key=lambda item: item['label'])
    return options


def _fetch_distinct_label_options(cursor, column_name, label_builder):
    cursor.execute(
        f"""
        SELECT DISTINCT {column_name}
        FROM vp_localizaciones
        WHERE COALESCE(BTRIM({column_name}), '') <> ''
        """
    )
    options = []
    seen_labels = set()

    for (raw_value,) in cursor.fetchall():
        raw_txt = _normalize_text(raw_value)
        if not raw_txt or raw_txt.lower() == 'null' or raw_txt == '0':
            continue

        label_txt = _normalize_text(label_builder(raw_txt))
        if not label_txt:
            continue

        label_key = label_txt.casefold()
        if label_key in seen_labels:
            continue

        seen_labels.add(label_key)
        options.append({'value': raw_txt, 'label': label_txt})

    options.sort(key=lambda item: item['label'])
    return options


def _fetch_column_value_by_label(cursor, column_name, label_builder):
    cursor.execute(
        f"""
        SELECT DISTINCT {column_name}
        FROM vp_localizaciones
        WHERE COALESCE(BTRIM({column_name}), '') <> ''
        """
    )
    value_by_label = {}

    for (raw_value,) in cursor.fetchall():
        raw_txt = _normalize_text(raw_value)
        if not raw_txt or raw_txt.lower() == 'null' or raw_txt == '0':
            continue

        label_txt = _normalize_text(label_builder(raw_txt))
        if not label_txt:
            continue

        value_by_label.setdefault(label_txt.casefold(), raw_txt)

    return value_by_label


def _build_preferred_column_options(cursor, column_name, preferred_options, label_builder, fallback_builder):
    value_by_label = _fetch_column_value_by_label(cursor, column_name, label_builder)
    options = []

    for code, label in preferred_options:
        label_txt = _normalize_text(label)
        if not label_txt:
            continue

        value = value_by_label.get(label_txt.casefold(), fallback_builder(code, label_txt))
        options.append({'value': _normalize_text(value), 'label': label_txt})

    return options


def _build_preferred_value_options(cursor, sql, preferred_labels):
    cursor.execute(sql)
    value_by_label = {}

    for row in cursor.fetchall():
        value_txt = _normalize_text(row[0])
        if value_txt:
            value_by_label.setdefault(value_txt.casefold(), value_txt)

    return [
        {'value': value_by_label.get(_normalize_text(label).casefold(), _normalize_text(label)), 'label': _normalize_text(label)}
        for label in preferred_labels
        if _normalize_text(label)
    ]


def _code_dash_label_fallback(code, label):
    return f'{code}-{label}' if code else label


@lru_cache(maxsize=1)
def _get_filter_options():
    with connections[PADRON_DB].cursor() as cursor:
        cursor.execute("SELECT descripcion FROM ambito_tipo ORDER BY c_ambito")
        ambitos = [_normalize_text(row[0]) for row in cursor.fetchall() if _normalize_text(row[0])]

        cursor.execute("SELECT descripcion FROM sector_tipo ORDER BY c_sector")
        sectores = [_normalize_text(row[0]) for row in cursor.fetchall() if _normalize_text(row[0])]

        cursor.execute("SELECT descripcion FROM dependencia_tipo ORDER BY c_dependencia")
        dependencias = [_normalize_text(row[0]) for row in cursor.fetchall() if _normalize_text(row[0])]

        cursor.execute("SELECT descripcion FROM estado_tipo ORDER BY c_estado")
        estados = [_normalize_text(row[0]) for row in cursor.fetchall() if _normalize_text(row[0])]

        cursor.execute("SELECT BTRIM(descripcion) FROM oferta_tipo ORDER BY c_oferta")
        tipo_ofertas = [_normalize_text(row[0]) for row in cursor.fetchall() if _normalize_text(row[0])]

        cursor.execute(
            """
            SELECT DISTINCT BTRIM(localidad_nombre) AS valor
            FROM vp_localizaciones
            WHERE COALESCE(BTRIM(localidad_nombre), '') <> ''
            ORDER BY valor
            """
        )
        localidades = [_normalize_text(row[0]) for row in cursor.fetchall() if _normalize_text(row[0])]

        cursor.execute(
            """
            SELECT DISTINCT BTRIM(departamento_nombre) AS valor
            FROM vp_localizaciones
            WHERE COALESCE(BTRIM(departamento_nombre), '') <> ''
            ORDER BY valor
            """
        )
        departamentos = [_normalize_text(row[0]) for row in cursor.fetchall() if _normalize_text(row[0])]

        periodos = _build_preferred_value_options(
            cursor,
            """
            SELECT DISTINCT BTRIM(periodo_funcionamiento) AS valor
            FROM vp_localizaciones
            WHERE COALESCE(BTRIM(periodo_funcionamiento), '') <> ''
            ORDER BY valor
            """,
            PERIODO_FUNCIONAMIENTO_OPTIONS,
        )

        modalidades = _build_preferred_value_options(
            cursor,
            "SELECT descripcion FROM modalidad2_tipo ORDER BY c_modalidad2",
            MODALIDADES_COMPLEMENTARIAS_OPTIONS,
        )

        return {
            'ambito': ambitos,
            'sede': ['Si', 'No'],
            'sede_adm': ['Si', 'No'],
            'periodo_funcionamiento': periodos,
            'sector': sectores,
            'dependencia': dependencias,
            'estado': estados,
            'tipo_oferta': tipo_ofertas,
            'localidad': localidades,
            'departamento': departamentos,
            'modalidades_complementarias': modalidades,
            'microregion': _fetch_distinct_formatted_options(cursor, 'cp_esvat4', lambda value: _format_desc_then_code(value, 1019638010)),
            'udt': _build_preferred_column_options(cursor, 'cp_esvat6', UDT_FILTER_OPTIONS, lambda value: _format_desc_only(value, 1019638012), lambda code, label: code),
            'regional_actual': _fetch_distinct_formatted_options(cursor, 'cp_esvat5', lambda value: _format_code_then_desc(value, 1019638011)),
            'zona_provincial': _fetch_distinct_label_options(cursor, 'cp_zonaprovincial', _format_desc_only),
            'anterior_regional': _build_preferred_column_options(cursor, 'cp_esvat3', ANTERIOR_REGIONAL_OPTIONS, _format_desc_only, _code_dash_label_fallback),
            'patrimonio_edilicio': _build_preferred_column_options(cursor, 'cp_patrimonioedilicio', PATRIMONIO_EDILICIO_OPTIONS, _format_desc_only, _code_dash_label_fallback),
            'tipo_albergue': _fetch_distinct_formatted_options(cursor, 'cp_esvat1', lambda value: _format_desc_with_code(value, 1019638007)),
            'regional_hasta_2015': _build_preferred_column_options(cursor, 'cp_p8104_localizacion_1019638033', REGIONAL_HASTA_2015_OPTIONS, _format_desc_only, _code_dash_label_fallback),
            'regional_hasta_2020': _build_preferred_column_options(cursor, 'cp_reg_hasta_2020', REGIONAL_HASTA_2020_OPTIONS, lambda value: _format_desc_only(value, 1019638079), _code_dash_label_fallback),
        }


def _serialize_list_item(row):
    return {
        'id': row.get('id'),
        'cue': _normalize_text(row.get('cue')),
        'anexo': _normalize_text(row.get('anexo')),
        'codigo_jurisdiccional': _normalize_text(row.get('codigo_jurisdiccional')),
        'ambito': _normalize_text(row.get('ambito')),
        'sede': '1' if _bool_to_si_no(row.get('sede')) == 'Si' else '',
        'sede_adm': '1' if _bool_to_si_no(row.get('sede_adm')) == 'Si' else '',
        'periodo_funcionamiento': _normalize_text(row.get('periodo_funcionamiento')),
        'nombre': _normalize_text(row.get('nombre')),
        'establecimiento': _normalize_text(row.get('establecimiento')),
        'responsable': _normalize_text(row.get('responsable')),
        'estado': _normalize_text(row.get('estado')),
        'sector': _normalize_text(row.get('sector')),
        'dependencia': _normalize_text(row.get('dependencia')),
        'tipo_oferta': _normalize_text(row.get('tipo_oferta')),
        'localidad': _normalize_text(row.get('localidad')),
        'departamento': _normalize_text(row.get('departamento')),
        'domicilio_ppal': _normalize_text(row.get('domicilio_ppal')),
        'cod_area_tel': _normalize_optional_text(row.get('cod_area_tel'), placeholders={'000'}),
        'telefono': _normalize_optional_text(row.get('telefono'), placeholders={'00000'}),
        'email': _normalize_optional_text(row.get('email')),
        'modalidades_complementarias': _normalize_text(row.get('modalidades_complementarias')),
        'tel_director_regional': _normalize_optional_text(row.get('tel_director_regional')),
        'email_dir_regional': _normalize_optional_text(row.get('email_dir_regional')),
        'supervisor_tecnico': _format_desc_only(row.get('supervisor_tecnico'), 1019638046),
        'microregion': _format_desc_only(row.get('microregion'), 1019638010),
        'udt': _format_desc_only(row.get('udt'), 1019638012),
        'regional_actual': _format_desc_only(row.get('regional_actual'), 1019638011),
        'zona_provincial': _format_desc_only(_normalize_optional_text(row.get('zona_provincial'))),
        'anterior_regional': _format_desc_only(row.get('anterior_regional')),
        'director_regional': _format_desc_only(row.get('director_regional'), 1019638045),
        'plan_de_obra': _normalize_optional_text(row.get('plan_de_obra')),
        'patrimonio_edilicio': _format_desc_only(_normalize_optional_text(row.get('patrimonio_edilicio'))),
        'fecha_creacion_edificio': _normalize_optional_text(row.get('fecha_creacion_edificio')),
        'nro_biblioteca': _normalize_optional_text(row.get('nro_biblioteca')),
        'cuof': _normalize_optional_text(row.get('cuof')),
        'cui': _normalize_optional_text(row.get('cui')),
        'cua': _normalize_optional_text(row.get('cua')),
        'tipo_albergue': _format_desc_only(row.get('tipo_albergue'), 1019638007),
        'regional_hasta_2015': _format_desc_only(row.get('regional_hasta_2015')),
        'inst_legal_edificio': _normalize_optional_text(row.get('inst_legal_edificio')),
        'tel_supervisor': _normalize_optional_text(row.get('tel_supervisor'), placeholders={'00000'}),
        'email_supervisor': _normalize_optional_text(row.get('email_supervisor')),
        'regional_hasta_2020': _format_desc_only(row.get('regional_hasta_2020'), 1019638079),
    }


def _serialize_localizacion(row):
    responsable_fallback = (
        str(row.get('id_responsable')).strip() == '-2'
        or _normalize_text(row.get('responsable_apellido')).lower() in {'sin información', 'sin informacion'}
    )

    return {
        'id_localizacion': row.get('id_localizacion'),
        'id_responsable': row.get('id_responsable'),
        'cue': _normalize_text(row.get('cue')),
        'anexo': _normalize_text(row.get('anexo')),
        'cue_anexo': _build_cue_anexo(row.get('cue'), row.get('anexo')),
        'estab_nombre': _normalize_text(row.get('establecimiento_nombre')),
        'estab_sector': _normalize_text(row.get('sector')),
        'estab_dependencia': _normalize_text(row.get('dependencia')),
        'nombre': _normalize_text(row.get('nombre')),
        'codigo_jurisdiccional': _normalize_text(row.get('codigo_jurisdiccional')),
        'sede': _bool_to_si_no(row.get('sede')),
        'sede_administrativa': _bool_to_si_no(row.get('sede_administrativa')),
        'ambito': _normalize_text(row.get('ambito')),
        'estado': _normalize_text(row.get('estado_localizacion')),
        'telefono_solo': _normalize_optional_text(row.get('telefono'), placeholders={'00000'}),
        'telefono': _build_phone_with_area(row.get('telefono_cod_area'), row.get('telefono')),
        'telefono_cod_area': _normalize_optional_text(row.get('telefono_cod_area'), placeholders={'000'}),
        'domicilio_principal': _build_domicilio(
            row.get('calle'),
            row.get('nro'),
            row.get('cod_postal'),
            row.get('localidad_nombre'),
            row.get('departamento_nombre'),
        ),
        'alternancia': _normalize_text(row.get('alternancia')),
        'periodo_funcionamiento': _normalize_text(row.get('periodo_funcionamiento')),
        'email': _normalize_optional_text(row.get('email')),
        'sitio_web': _normalize_optional_text(row.get('sitio_web')),
        'cooperadora': _normalize_text(row.get('cooperadora')),
        'responsable': _build_persona(
            row.get('responsable_apellido'),
            row.get('responsable_nombre'),
            row.get('documento_responsable'),
            force_empty_parentheses=responsable_fallback and not _normalize_text(row.get('documento_responsable')),
        ),
        'permanencia': _normalize_text(row.get('permanencia')),
        'observaciones': _normalize_text(row.get('observaciones'), keep_linebreaks=True),
        'fecha_creacion': _format_date(row.get('fecha_creacion')),
        'fecha_alta': _format_date(row.get('fecha_alta')),
        'fecha_baja': _format_date(row.get('fecha_baja')),
        'fecha_actualizacion': _format_date(row.get('fecha_actualizacion')),
        'modalidades_complementarias': _normalize_text(row.get('modalidades_complementarias')),
        'tipo_oferta': _normalize_text(row.get('tipo_oferta')),
        'localidad': _normalize_text(row.get('localidad_nombre')),
        'departamento': _normalize_text(row.get('departamento_nombre')),
        'cod_area_tel': _normalize_optional_text(row.get('telefono_cod_area'), placeholders={'000'}),
        'tel_director_regional': _normalize_optional_text(row.get('cp_tedirregional')),
        'email_dir_regional': _normalize_optional_text(row.get('cp_emaildirregional')),
        'supervisor_tecnico': _format_desc_then_code(row.get('cp_supervisortecnico'), 1019638046),
        'microregion': _format_desc_then_code(row.get('cp_esvat4'), 1019638010),
        'udt': _format_desc_then_code(row.get('cp_esvat6'), 1019638012),
        'regional_actual': _format_code_then_desc(row.get('cp_esvat5'), 1019638011),
        'zona_provincial': _format_desc_then_code(_normalize_optional_text(row.get('cp_zonaprovincial'))),
        'anterior_regional': _format_desc_then_code(row.get('cp_esvat3')),
        'director_regional': _format_desc_then_code(row.get('cp_directorregional'), 1019638045),
        'plan_de_obra': _normalize_optional_text(row.get('cp_plandeobra')),
        'patrimonio_edilicio': _format_desc_then_code(_normalize_optional_text(row.get('cp_patrimonioedilicio'))),
        'fecha_creacion_edificio': _normalize_optional_text(row.get('cp_estfechacreacionedificio')),
        'nro_biblioteca': _normalize_optional_text(row.get('cp_esvar1')),
        'cuof': _normalize_optional_text(row.get('cp_esvar4')),
        'cui': _normalize_optional_text(row.get('cp_esvar2')),
        'cua': _normalize_optional_text(row.get('cp_esvar3')),
        'tipo_albergue': _format_desc_with_code(row.get('cp_esvat1'), 1019638007),
        'regional_hasta_2015': _format_desc_then_code(row.get('cp_p8104_localizacion_1019638033')),
        'inst_legal_edificio': _normalize_optional_text(row.get('cp_edif_instlegal_creaciondeestablecimiento')),
        'tel_supervisor': _normalize_optional_text(row.get('tel_supervisor'), placeholders={'00000'}),
        'email_supervisor': _normalize_optional_text(row.get('email_supervisor')),
        'regional_hasta_2020': _format_desc_then_code(row.get('cp_reg_hasta_2020'), 1019638079),
    }


def _serialize_domicilio(row):
    return {
        'id_domicilio': row.get('id_domicilio'),
        'tipo_domicilio': _normalize_text(row.get('tipo_domicilio')),
        'calle': _normalize_text(row.get('calle')),
        'nro': _normalize_text(row.get('nro')),
        'barrio': _normalize_text(row.get('barrio')),
        'referencia': _normalize_text(row.get('referencia'), keep_linebreaks=True),
        'cod_postal': _normalize_text(row.get('cod_postal')),
        'cui': _normalize_text(row.get('cui')),
        'localidad': _build_localidad_departamento(row.get('localidad_nombre'), row.get('departamento_nombre')),
        'calle_fondo': _normalize_text(row.get('calle_fondo')),
        'calle_derecha': _normalize_text(row.get('calle_derecha')),
        'calle_izquierda': _normalize_text(row.get('calle_izquierda')),
        'fecha_actualizacion': _format_date(row.get('fecha_actualizacion')),
        'caption': _build_domicilio(
            row.get('calle'),
            row.get('nro'),
            row.get('cod_postal'),
            row.get('localidad_nombre'),
            row.get('departamento_nombre'),
        ),
    }


def _serialize_responsable(row):
    apellido = _normalize_text(row.get('apellido')) or _normalize_text(row.get('responsable_apellido')) or 'Sin información'
    nombre = _normalize_text(row.get('nombre')) or _normalize_text(row.get('responsable_nombre'))
    documento = _normalize_text(row.get('nro_documento')) or _normalize_text(row.get('documento_responsable'))

    return {
        'apellido': apellido,
        'nombre': nombre,
        'tipo_documento': _normalize_text(row.get('tipo_documento_desc')),
        'nro_documento': documento,
        'telefono': _normalize_optional_text(row.get('telefono')),
        'email': _normalize_optional_text(row.get('email')),
        'fecha_nacimiento': _format_date(row.get('fecha_nacimiento')),
        'nacionalidad': _normalize_text(row.get('nacionalidad')),
        'cuil_cuit': _normalize_optional_text(row.get('cuil_cuit')),
        'sexo': _normalize_text(row.get('sexo_desc')),
    }


def _serialize_oferta(row):
    return {
        'id_oferta_local': row.get('id_oferta_local'),
        'localizacion': _build_cue_anexo(row.get('cue'), row.get('anexo')),
        'oferta': _normalize_text(row.get('oferta')),
        'estado': _normalize_text(row.get('estado_ofertalocal')),
        'codigo_jurisdiccional': _normalize_text(row.get('codigo_jurisdiccional_oferta_local')),
        'subvencion': _normalize_text(row.get('subvencion_oferta_local')),
        'jornada': _normalize_text(row.get('jornada_ofertalocal')),
        'cuof': _normalize_optional_text(row.get('cp_efvar4')),
        'cuof_ryc': _normalize_optional_text(row.get('cp_cuof_ryc')),
        'cui': _normalize_optional_text(row.get('cp_efvar2')),
        'cua': _normalize_optional_text(row.get('cp_of_cua')),
        'udt': _format_desc_only(row.get('cp_efvar6')),
        'regional': _format_desc_only(row.get('cp_efvar5')),
        'acronimo': _format_desc_only(row.get('cp_acronimo')),
        'supervisor_tecnico': _format_desc_only(row.get('cp_supervisortecnico_oferta')),
        'tel_supervisor': _normalize_optional_text(row.get('cp_tesupervisor_oferta'), placeholders={'0'}),
        'email_supervisor': _normalize_optional_text(row.get('cp_mailsupervisor_oferta')),
        'fecha_inst_legal': _format_date(row.get('cp_of_fecha_inst_legal')),
        'nro_inst_legal': _normalize_optional_text(row.get('cp_of_nro_inst_legal')),
        'anio_creacion': _normalize_optional_text(row.get('cp_of_anio_inst_legal')),
        'descrip_inst_legal': _format_desc_only(row.get('cp_of_descrip_inst_legal')),
        'cabecera_anexo': _format_desc_only(row.get('cp_of_cab_anexo')),
        'ambito': _format_desc_only(row.get('cp_of_ambito')),
        'tipo_ed': _format_desc_only(row.get('cp_of_tipo_ed')),
        'nivel': _format_desc_only(row.get('cp_of_nivel')),
        'sector': _format_desc_only(row.get('cp_of_sector')),
        'categoria': _format_desc_only(row.get('cp_of_categoria')),
    }


def _format_historial_instr_legal(numero, descripcion):
    numero_txt = _normalize_text(numero)
    descripcion_txt = _normalize_text(descripcion)

    if numero_txt and descripcion_txt:
        return f'({numero_txt}) {descripcion_txt}'
    return descripcion_txt or numero_txt


def _serialize_historial(row):
    cue_anexo = _build_cue_anexo(row.get('cue'), row.get('anexo'))
    fecha_vigencia = _format_date(row.get('fecha_vigencia'))

    return {
        'caption': f'{cue_anexo} ({fecha_vigencia})' if cue_anexo and fecha_vigencia else cue_anexo or fecha_vigencia,
        'localizacion': cue_anexo,
        'movimiento': ' - '.join(
            [
                value
                for value in [
                    _normalize_text(row.get('id_movimiento')),
                    _normalize_text(row.get('cod_tipo_mov')),
                    _normalize_text(row.get('tipo_movimiento')),
                ]
                if value
            ]
        ),
        'estado': _normalize_text(row.get('estado')),
        'fecha_instr_legal': _format_date(row.get('fecha_inst_legal')),
        'fecha_vigencia': fecha_vigencia,
        'observacion': _normalize_text(row.get('observacion'), keep_linebreaks=True),
        'instr_legal': _format_historial_instr_legal(row.get('nro_instr_legal'), row.get('instr_legal')),
        'motivo': _normalize_text(row.get('motivo')),
        'usuario': _normalize_text(row.get('usuario')),
    }

@padron_interno_admin_o_gestor_required
def detalle_localizacion_json(request, id_localizacion):
    localizacion_sql = """
        SELECT
            vl.*,
            l.fecha_creacion,
            l.fecha_alta,
            l.fecha_baja,
            l.fecha_actualizacion,
            COALESCE(BTRIM(ve.nombre), '') AS establecimiento_nombre,
            COALESCE(BTRIM(otipos.tipo_ofertas), '') AS tipo_oferta,
            COALESCE(BTRIM(mods.modalidades), '') AS modalidades_complementarias,
            COALESCE(BTRIM(loc_tel_sup.valor), '') AS tel_supervisor,
            COALESCE(BTRIM(loc_email_sup.valor), '') AS email_supervisor
        FROM vp_localizaciones vl
        JOIN localizacion l
          ON l.id_localizacion = vl.id_localizacion
        LEFT JOIN vp_establecimientos ve
          ON ve.id_establecimiento = vl.id_establecimiento
        LEFT JOIN loc_campo_prov_valor loc_tel_sup
          ON loc_tel_sup.id_localizacion = vl.id_localizacion
         AND loc_tel_sup.id_campo_prov = 1019638042
        LEFT JOIN loc_campo_prov_valor loc_email_sup
          ON loc_email_sup.id_localizacion = vl.id_localizacion
         AND loc_email_sup.id_campo_prov = 1019638043
        LEFT JOIN LATERAL (
            SELECT STRING_AGG(BTRIM(ot.descripcion), ', ' ORDER BY ot.c_oferta, ol.id_oferta_local) AS tipo_ofertas
            FROM oferta_local ol
            JOIN oferta_tipo ot ON ot.c_oferta = ol.c_oferta
            WHERE ol.id_localizacion = vl.id_localizacion
        ) otipos ON TRUE
        LEFT JOIN LATERAL (
            SELECT STRING_AGG(BTRIM(m2.descripcion), ', ' ORDER BY m2.orden, m2.descripcion) AS modalidades
            FROM localizacion_modalidad2_assn lm2
            JOIN modalidad2_tipo m2 ON m2.c_modalidad2 = lm2.c_modalidad2
            WHERE lm2.id_localizacion = vl.id_localizacion
        ) mods ON TRUE
        WHERE vl.id_localizacion = %s
    """

    domicilios_sql = """
        SELECT
            ld.id_domicilio,
            ld.c_tipo_dom,
            dt.descripcion AS tipo_domicilio,
            d.calle,
            d.nro,
            d.barrio,
            d.referencia,
            d.cod_postal,
            d.cui,
            d.calle_fondo,
            d.calle_derecha,
            d.calle_izquierda,
            d.fecha_actualizacion,
            lt.nombre AS localidad_nombre,
            dept.nombre AS departamento_nombre
        FROM localizacion_domicilio ld
        JOIN domicilio d ON d.id_domicilio = ld.id_domicilio
        LEFT JOIN domicilio_tipo dt ON dt.c_tipo_dom = ld.c_tipo_dom
        LEFT JOIN localidad_tipo lt ON lt.c_localidad = d.c_localidad
        LEFT JOIN departamento_tipo dept ON dept.c_departamento = lt.c_departamento
        WHERE ld.id_localizacion = %s
        ORDER BY CASE WHEN ld.c_tipo_dom = 1 THEN 0 ELSE 1 END, ld.c_tipo_dom, ld.id_domicilio
    """

    domicilio_localizaciones_sql = """
        SELECT
            vl.*,
            COALESCE(BTRIM(otipos.tipo_ofertas), '') AS tipo_oferta,
            COALESCE(BTRIM(mods.modalidades), '') AS modalidades_complementarias,
            COALESCE(BTRIM(loc_tel_sup.valor), '') AS tel_supervisor,
            COALESCE(BTRIM(loc_email_sup.valor), '') AS email_supervisor
        FROM localizacion_domicilio ld
        JOIN vp_localizaciones vl
          ON vl.id_localizacion = ld.id_localizacion
        LEFT JOIN loc_campo_prov_valor loc_tel_sup
          ON loc_tel_sup.id_localizacion = vl.id_localizacion
         AND loc_tel_sup.id_campo_prov = 1019638042
        LEFT JOIN loc_campo_prov_valor loc_email_sup
          ON loc_email_sup.id_localizacion = vl.id_localizacion
         AND loc_email_sup.id_campo_prov = 1019638043
        LEFT JOIN LATERAL (
            SELECT STRING_AGG(BTRIM(ot.descripcion), ', ' ORDER BY ot.c_oferta, ol.id_oferta_local) AS tipo_ofertas
            FROM oferta_local ol
            JOIN oferta_tipo ot ON ot.c_oferta = ol.c_oferta
            WHERE ol.id_localizacion = vl.id_localizacion
        ) otipos ON TRUE
        LEFT JOIN LATERAL (
            SELECT STRING_AGG(BTRIM(m2.descripcion), ', ' ORDER BY m2.orden, m2.descripcion) AS modalidades
            FROM localizacion_modalidad2_assn lm2
            JOIN modalidad2_tipo m2 ON m2.c_modalidad2 = lm2.c_modalidad2
            WHERE lm2.id_localizacion = vl.id_localizacion
        ) mods ON TRUE
        WHERE ld.id_domicilio = %s
        ORDER BY vl.cue::text, COALESCE(vl.anexo::text, '') DESC, vl.id_localizacion
    """

    domicilio_ofertas_sql = """
        SELECT vol.*
        FROM localizacion_domicilio ld
        JOIN vp_oferta_local vol
          ON vol.id_localizacion = ld.id_localizacion
        WHERE ld.id_domicilio = %s
        ORDER BY vol.c_oferta, vol.id_oferta_local
    """

    responsable_sql = """
        SELECT
            vl.id_localizacion,
            vl.id_responsable,
            vl.responsable_apellido,
            vl.responsable_nombre,
            vl.documento_responsable,
            r.apellido,
            r.nombre,
            r.telefono,
            r.nro_documento,
            r.email,
            r.fecha_nacimiento,
            r.cuil_cuit,
            tdt.descripcion AS tipo_documento_desc,
            st.descripcion AS sexo_desc,
            ot.descripcion AS nacionalidad
        FROM vp_localizaciones vl
        LEFT JOIN responsable r ON r.id_responsable = vl.id_responsable
        LEFT JOIN tipo_documento_tipo tdt ON tdt.c_tipo_documento = r.c_tipo_documento
        LEFT JOIN sexo_tipo st ON st.c_sexo = r.c_sexo
        LEFT JOIN origen_tipo ot ON ot.c_origen = r.c_nacionalidad
        WHERE vl.id_localizacion = %s
    """

    ofertas_sql = """
        SELECT *
        FROM vp_oferta_local
        WHERE id_localizacion = %s
        ORDER BY c_oferta, id_oferta_local
    """

    historial_sql = """
        SELECT
            cl.id_localizacion,
            l.anexo,
            e.cue,
            m.id_movimiento,
            tm.cod_tipo_mov,
            tm.descripcion AS tipo_movimiento,
            est.descripcion AS estado,
            m.nro_instr_legal,
            il.descripcion AS instr_legal,
            mot.descripcion AS motivo,
            m.fecha_inst_legal,
            m.fecha_vigencia,
            m.observacion,
            u.nombre AS usuario
        FROM cambio_estado_localizacion cl
        JOIN localizacion l ON l.id_localizacion = cl.id_localizacion
        JOIN establecimiento e ON e.id_establecimiento = l.id_establecimiento
        JOIN movimiento m ON m.id_movimiento = cl.id_movimiento
        LEFT JOIN tipo_mov_tipo tm ON tm.c_tipo_mov = m.c_tipo_mov
        LEFT JOIN estado_tipo est ON est.c_estado = cl.c_estado
        LEFT JOIN instr_legal_tipo il ON il.c_instr_legal = m.c_instr_legal
        LEFT JOIN motivo_tipo mot ON mot.c_motivo = m.c_motivo
        LEFT JOIN usuario u ON u.id_usuario = m.id_usuario
        WHERE cl.id_localizacion = %s
        ORDER BY m.fecha_vigencia DESC, m.id_movimiento DESC
    """

    try:
        localizacion_row = _fetch_one(localizacion_sql, [id_localizacion])
        if not localizacion_row:
            return JsonResponse({'error': 'Localización no encontrada.'}, status=404)

        domicilios_rows = _fetch_all(domicilios_sql, [id_localizacion])
        responsable_row = _fetch_one(responsable_sql, [id_localizacion])
        ofertas_rows = _fetch_all(ofertas_sql, [id_localizacion])
        historial_rows = _fetch_all(historial_sql, [id_localizacion])

        domicilios_payload = []
        for domicilio_row in domicilios_rows:
            domicilio_id = domicilio_row.get('id_domicilio')
            linked_localizaciones = _fetch_all(domicilio_localizaciones_sql, [domicilio_id]) if domicilio_id else []
            linked_ofertas = _fetch_all(domicilio_ofertas_sql, [domicilio_id]) if domicilio_id else []

            domicilio_payload = _serialize_domicilio(domicilio_row)
            domicilio_payload['localizaciones'] = [_serialize_localizacion(row) for row in linked_localizaciones]
            domicilio_payload['ofertas'] = [_serialize_oferta(row) for row in linked_ofertas]
            domicilios_payload.append(domicilio_payload)

        payload = {
            'localizacion': _serialize_localizacion(localizacion_row),
            'domicilios': domicilios_payload,
            'responsable': _serialize_responsable(responsable_row or localizacion_row),
            'ofertas': [_serialize_oferta(row) for row in ofertas_rows],
            'historial': [_serialize_historial(row) for row in historial_rows],
        }
        return JsonResponse(_sanitize_json_payload(payload))
    except Exception as exc:
        return JsonResponse({'error': f'Error al obtener detalle de localización: {exc}'}, status=500)

@padron_interno_admin_o_gestor_required
@ensure_csrf_cookie
def listar_localizaciones(request):
    formato = request.GET.get('formato')

    if formato == 'excel_todo':
        where, params = '', []
    else:
        where, params = _build_where(request)

    orden_key = request.GET.get('orden', 'cue')
    col_orden = CAMPO_SQL.get(orden_key, 'vl.cue::text')

    if orden_key == 'cue':
        order_clause = "vl.cue::text, COALESCE(vl.anexo, '') DESC, vl.id_localizacion"
    elif orden_key == 'anexo':
        order_clause = "COALESCE(vl.anexo, '') DESC, vl.cue::text, vl.id_localizacion"
    else:
        order_clause = f"{col_orden}, vl.cue::text, COALESCE(vl.anexo, '') DESC, vl.id_localizacion"

    data_sql = f"{_SELECT_FIELDS} {_BASE_SQL} {where} ORDER BY {order_clause}"

    if formato in {'excel_pagina', 'excel_todo'}:
        datos_base = _fetch_all(data_sql, params)
        datos_exportar = [_serialize_list_item(row) for row in datos_base]
        return _exportar_excel(datos_exportar, formato, request)

    count_sql = f"SELECT COUNT(*) {_BASE_SQL} {where}"

    try:
        current_page_size = int(request.GET.get('page_size', PAGE_SIZE))
    except (TypeError, ValueError):
        current_page_size = PAGE_SIZE

    raw = _RawPage(data_sql, count_sql, params, PADRON_DB)
    paginator = Paginator(raw, current_page_size)

    try:
        page_number = request.GET.get('page', 1)
        page_obj = paginator.page(page_number)
    except (EmptyPage, PageNotAnInteger):
        page_obj = paginator.page(1)

    total = len(raw)
    desde = (page_obj.number - 1) * current_page_size + 1 if total else 0
    hasta = min(page_obj.number * current_page_size, total)

    lista_items = [_serialize_list_item(row) for row in page_obj.object_list]

    context = {
        'lista_items': lista_items,
        'page_obj': page_obj,
        'resultado_total': total,
        'resultado_desde': desde,
        'resultado_hasta': hasta,
        'username': getattr(request.user, 'username', ''),
        'request': request,
        'filter_options_json': json.dumps(_get_filter_options(), ensure_ascii=False),
    }
    context.update(get_contexto_fecha_padron(request))
    return render(request, 'padroninterno/localizaciones.html', context)
