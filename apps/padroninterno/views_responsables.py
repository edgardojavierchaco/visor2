import json
import re
from datetime import date, datetime
from functools import lru_cache
from io import BytesIO
from .permisos import padron_interno_admin_o_gestor_required
from .views_fecha import get_contexto_fecha_padron
import openpyxl
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import connections
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from .views_fecha import get_contexto_fecha_padron
from .views_establecimientos import (
    OBSERVACIONES_SQL as EST_OBSERVACIONES_SQL,
    TIPO_OFERTAS_SQL as EST_TIPO_OFERTAS_SQL,
    _BASE_SQL as EST_BASE_SQL,
    _serialize_establecimiento as _serialize_establecimiento_detalle,
    _serialize_localizacion as _serialize_localizacion_detalle,
)

PADRON_DB = 'Padron'
PAGE_SIZE = 10


def _normalize_text(value, keep_linebreaks=False):
    if value is None:
        return ''

    text = str(value)
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    if text.strip().upper() == 'NULL':
        return ''

    if keep_linebreaks:
        lines = [re.sub(r'\s+', ' ', line).strip() for line in text.split('\n')]
        return '\n'.join([line for line in lines if line]).strip()

    return re.sub(r'\s+', ' ', text).strip()


def _normalize_optional_text(value, placeholders=None, keep_linebreaks=False):
    normalized = _normalize_text(value, keep_linebreaks=keep_linebreaks)
    if not normalized:
        return ''

    invalid_values = {
        '-',
        '.',
        '0',
        'null',
        's/inf.',
        's/inf',
        's/info',
        's/info.',
        's/datos',
        's/i',
        's/i.',
        'sin dato',
        'sin dato-',
        'sin información',
        'sin informacion',
        'no declara',
    }

    if placeholders:
        invalid_values.update({str(item).strip().lower() for item in placeholders if str(item).strip()})

    if normalized.lower() in invalid_values:
        return ''

    return normalized


def _format_date(value):
    normalized = _normalize_optional_text(value)
    if not normalized:
        return ''

    if isinstance(value, datetime):
        dt_value = value
    elif isinstance(value, date):
        dt_value = datetime.combine(value, datetime.min.time())
    else:
        raw = normalized.split(' ')[0]
        dt_value = None
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y', '%d-%m-%Y', '%d-%m-%y', '%d/%m/%y'):
            try:
                dt_value = datetime.strptime(raw, fmt)
                break
            except ValueError:
                continue
        if dt_value is None:
            return normalized

    meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    return f"{meses[dt_value.month - 1]} {dt_value.day:02d} {dt_value.year}"


def _dictfetchall(cursor):
    columns = [column[0] for column in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _fetch_one(sql, params):
    with connections[PADRON_DB].cursor() as cursor:
        cursor.execute(sql, params)
        rows = _dictfetchall(cursor)
    return rows[0] if rows else None


def _fetch_all(sql, params):
    with connections[PADRON_DB].cursor() as cursor:
        cursor.execute(sql, params)
        return _dictfetchall(cursor)


def _sanitize_json_payload(value):
    if isinstance(value, dict):
        return {key: _sanitize_json_payload(inner_value) for key, inner_value in value.items()}
    if isinstance(value, list):
        return [_sanitize_json_payload(item) for item in value]
    if value is None:
        return ''
    return value


LIST_CUEANEXO_SQL = (
    "COALESCE(BTRIM(locs.cueanexo), '')"
)
LIST_CODIGO_JURISDICCIONAL_SQL = (
    "COALESCE(BTRIM(locs.codigo_jurisdiccional), '')"
)

CAMPO_SQL = {
    'apellido': "COALESCE(BTRIM(r.apellido), '')",
    'nombre': "COALESCE(BTRIM(r.nombre), '')",
    'tipo_documento': "COALESCE(BTRIM(tdt.descripcion), '')",
    'nro_documento': "COALESCE(r.nro_documento::text, '')",
    'telefono': "COALESCE(BTRIM(r.telefono), '')",
    'sexo': "COALESCE(BTRIM(st.descripcion), '')",
    'email': "COALESCE(BTRIM(r.email), '')",
    'cueanexo': LIST_CUEANEXO_SQL,
    'codigo_jurisdiccional': LIST_CODIGO_JURISDICCIONAL_SQL,
    'cuil_cuit': "COALESCE(BTRIM(r.cuil_cuit), '')",
}

VISIBLE_NAME_MAP = {
    'apellido': 'Apellido',
    'nombre': 'Nombre',
    'tipo_documento': 'Tipo Documento',
    'nro_documento': 'Nro Documento',
    'telefono': 'Telefono',
    'sexo': 'Sexo',
    'email': 'Email',
    'cueanexo': 'Cueanexo',
    'codigo_jurisdiccional': 'Codigo Jurisdiccional',
    'cuil_cuit': 'Cuil/Cuit',
}

COLUMNAS_EXPORTACION = [
    ('Apellido', 'apellido'),
    ('Nombre', 'nombre'),
    ('Tipo Documento', 'tipo_documento'),
    ('Nro Documento', 'nro_documento'),
    ('Telefono', 'telefono'),
    ('Sexo', 'sexo'),
    ('Email', 'email'),
    ('Cueanexo', 'cueanexo'),
    ('Codigo Jurisdiccional', 'codigo_jurisdiccional'),
    ('Cuil/Cuit', 'cuil_cuit'),
]

_SELECT_FIELDS = f"""
    SELECT
        r.id_responsable AS id,
        COALESCE(BTRIM(r.apellido), '') AS apellido,
        COALESCE(BTRIM(r.nombre), '') AS nombre,
        COALESCE(BTRIM(tdt.descripcion), '') AS tipo_documento,
        COALESCE(r.nro_documento::text, '') AS nro_documento,
        COALESCE(BTRIM(r.telefono), '') AS telefono,
        COALESCE(BTRIM(st.descripcion), '') AS sexo,
        COALESCE(BTRIM(r.email), '') AS email,
        {LIST_CUEANEXO_SQL} AS cueanexo,
        {LIST_CODIGO_JURISDICCIONAL_SQL} AS codigo_jurisdiccional,
        COALESCE(BTRIM(r.cuil_cuit), '') AS cuil_cuit
"""

_BASE_SQL = """
    FROM responsable r
    LEFT JOIN tipo_documento_tipo tdt
      ON tdt.c_tipo_documento = r.c_tipo_documento
    LEFT JOIN sexo_tipo st
      ON st.c_sexo = r.c_sexo
    LEFT JOIN LATERAL (
        SELECT
            STRING_AGG(
                e.cue::text || LPAD(COALESCE(l.anexo::text, ''), 2, '0'),
                ', '
                ORDER BY e.cue::text ASC, COALESCE(l.anexo::text, '') DESC
            ) AS cueanexo,
            STRING_AGG(
                NULLIF(BTRIM(l.codigo_jurisdiccional), ''),
                ', '
                ORDER BY e.cue::text ASC, COALESCE(l.anexo::text, '') DESC
            ) AS codigo_jurisdiccional
        FROM localizacion l
        JOIN establecimiento e
          ON e.id_establecimiento = l.id_establecimiento
        WHERE l.id_responsable = r.id_responsable
    ) locs ON TRUE
"""


class _RawPage:
    def __init__(self, data_sql, count_sql, params, db_alias):
        self._data_sql = data_sql
        self._count_sql = count_sql
        self._params = params
        self._db_alias = db_alias
        self._length = None

    def __len__(self):
        if self._length is None:
            with connections[self._db_alias].cursor() as cursor:
                cursor.execute(self._count_sql, self._params)
                self._length = cursor.fetchone()[0]
        return self._length

    def __getitem__(self, item):
        if not isinstance(item, slice):
            raise TypeError('Los elementos deben pedirse como slice.')

        start = item.start or 0
        stop = item.stop or start
        limit = max(stop - start, 0)
        sql = f"{self._data_sql} LIMIT {limit} OFFSET {start}"

        with connections[self._db_alias].cursor() as cursor:
            cursor.execute(sql, self._params)
            cols = [d[0] for d in cursor.description]
            return [dict(zip(cols, row)) for row in cursor.fetchall()]


OPERADOR_SQL = {
    '0': ('ILIKE', lambda value: f'%{value}%'),
    '1': ('NOT ILIKE', lambda value: f'%{value}%'),
    '2': ('=', lambda value: value),
    '3': ('>', lambda value: value),
    '4': ('>=', lambda value: value),
    '5': ('<', lambda value: value),
    '6': ('<=', lambda value: value),
    '7': ('!=', lambda value: value),
}

_ACCENTED_CHARS = '\u00c1\u00c9\u00cd\u00d3\u00da\u00dc\u00d1\u00e1\u00e9\u00ed\u00f3\u00fa\u00fc\u00f1'
_UNACCENTED_CHARS = 'AEIOUUNaeiouun'
_ACCENT_TRANSLATION = str.maketrans(_ACCENTED_CHARS, _UNACCENTED_CHARS)


def _fold_filter_text(value):
    return _normalize_text(value).translate(_ACCENT_TRANSLATION).lower()


def _folded_sql(expr):
    return f"LOWER(TRANSLATE(BTRIM(COALESCE(({expr})::text, '')), '{_ACCENTED_CHARS}', '{_UNACCENTED_CHARS}'))"


def _filter_tokens(value):
    folded = _fold_filter_text(value)
    tokens = [token for token in re.split(r'[^a-z0-9]+', folded) if token]
    return list(dict.fromkeys(tokens))


def _build_text_search_clause(expr, value):
    tokens = _filter_tokens(value)
    if not tokens:
        return '', []

    folded_expr = _folded_sql(expr)
    return ' AND '.join([f"{folded_expr} LIKE %s" for _ in tokens]), [f'%{token}%' for token in tokens]


def _build_where(request):
    clauses = []
    params = []
    grouped_positive_filters = {}

    campos = request.GET.getlist('campo_filtro')
    opers = request.GET.getlist('operador_filtro')
    valores = request.GET.getlist('valor_filtro')

    for index, campo in enumerate(campos):
        campo = campo.strip()
        oper = opers[index].strip() if index < len(opers) else '0'
        valor = valores[index].strip() if index < len(valores) else ''

        if not campo or not valor or campo not in CAMPO_SQL:
            continue

        if oper in {'0', '2'}:
            grouped_positive_filters.setdefault((campo, oper), [])
            if valor not in grouped_positive_filters[(campo, oper)]:
                grouped_positive_filters[(campo, oper)].append(valor)
            continue

        col = CAMPO_SQL[campo]
        op_str, val_fn = OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])
        if oper == '0':
            token_clause, token_params = _build_text_search_clause(col, valor)
            if token_clause:
                clauses.append(token_clause)
                params.extend(token_params)
            continue

        clauses.append(f"{col}::text {op_str} %s")
        params.append(val_fn(valor))

    for (campo, oper), values in grouped_positive_filters.items():
        col = CAMPO_SQL[campo]
        op_str, val_fn = OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])

        if len(values) == 1:
            if oper == '0':
                token_clause, token_params = _build_text_search_clause(col, values[0])
                if token_clause:
                    clauses.append(token_clause)
                    params.extend(token_params)
                continue

            clauses.append(f"{col}::text {op_str} %s")
            params.append(val_fn(values[0]))
            continue

        subclauses = []
        for value in values:
            if oper == '0':
                token_clause, token_params = _build_text_search_clause(col, value)
                if token_clause:
                    subclauses.append(token_clause)
                    params.extend(token_params)
                continue

            subclauses.append(f"{col}::text {op_str} %s")
            params.append(val_fn(value))
        clauses.append(f"({' OR '.join(subclauses)})")

    q = request.GET.get('q', '').strip()
    if q:
        token_groups = []
        for token in _filter_tokens(q):
            global_search_clauses = [f"{_folded_sql(sql_expr)} LIKE %s" for sql_expr in CAMPO_SQL.values()]
            token_groups.append(f"({' OR '.join(global_search_clauses)})")
            params.extend([f'%{token}%'] * len(global_search_clauses))
        if token_groups:
            clauses.append(f"({' AND '.join(token_groups)})")

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ''
    return where, params


def _armar_texto_filtros(request):
    partes = []
    operadores_txt = {
        '0': 'parecido a',
        '1': 'no parecido a',
        '2': 'igual a',
        '3': 'mayor a',
        '4': 'mayor o igual a',
        '5': 'menor a',
        '6': 'menor o igual a',
        '7': 'distinto de',
    }

    if request.GET.get('formato') == 'excel_todo':
        return 'Sin filtros aplicados'

    q = request.GET.get('q', '').strip()
    if q:
        partes.append(f'Búsqueda: {q}')

    campos = request.GET.getlist('campo_filtro')
    opers = request.GET.getlist('operador_filtro')
    valores = request.GET.getlist('valor_filtro')

    for index, campo in enumerate(campos):
        campo = campo.strip()
        valor = valores[index].strip() if index < len(valores) else ''
        oper = opers[index].strip() if index < len(opers) else '0'

        if not campo or not valor:
            continue

        partes.append(f"{VISIBLE_NAME_MAP.get(campo, campo)} {operadores_txt.get(oper, 'parecido a')}: {valor}")

    return ' | '.join(partes) if partes else 'Sin filtros aplicados'


def _formatear_fecha_estilo_seguimiento():
    ahora = datetime.now()
    dia = ahora.day
    mes = ahora.month
    anio = ahora.year
    hora_12 = ahora.strftime('%I').lstrip('0') or '0'
    minuto = ahora.strftime('%M')
    ampm = 'p. m.' if ahora.hour >= 12 else 'a. m.'
    return f'{dia}/{mes}/{anio} a las {hora_12}:{minuto} {ampm}'


def _resolver_columnas_exportar(request, formato):
    if formato == 'excel_todo':
        return COLUMNAS_EXPORTACION

    visibles = request.GET.getlist('visible_col')
    if not visibles:
        return COLUMNAS_EXPORTACION

    visibles_set = {value.strip() for value in visibles if value.strip()}
    columnas_filtradas = [col for col in COLUMNAS_EXPORTACION if col[1] in visibles_set]
    return columnas_filtradas or COLUMNAS_EXPORTACION


def _exportar_excel(datos_exportar, formato, request):
    columnas_mapeo = _resolver_columnas_exportar(request, formato)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Responsables'

    total_columnas = len(columnas_mapeo)
    ultima_columna = get_column_letter(total_columnas)

    ws.merge_cells(f'A1:{ultima_columna}1')
    ws['A1'] = 'Informe Responsables'
    ws['A1'].font = Font(bold=True, size=10)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'A2:{ultima_columna}2')
    ws['A2'] = f'Informe generado el: {_formatear_fecha_estilo_seguimiento()}'
    ws['A2'].font = Font(size=9)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'A3:{ultima_columna}3')
    ws['A3'] = f'Filtros aplicados: {_armar_texto_filtros(request)}'
    ws['A3'].font = Font(size=9)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20

    header_row = 4
    for col_idx, (titulo_col, _) in enumerate(columnas_mapeo, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=titulo_col)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for item in datos_exportar:
        fila = []
        for _, clave in columnas_mapeo:
            fila.append(_normalize_text(item.get(clave), keep_linebreaks=True))
        ws.append(fila)

        ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{ultima_columna}{ws.max_row}'

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{ultima_columna}{ws.max_row}'
    for col_idx in range(1, total_columnas + 1):
        length = 0
        column_letter = get_column_letter(col_idx)

        for row_idx in range(header_row, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)

            try:
                value_length = len(str(cell.value or ''))
            except Exception:
                value_length = 0

            length = max(length, value_length)

        ws.column_dimensions[column_letter].width = min(max(length + 2, 12), 45)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_archivo = datetime.now().strftime('%Y%m%d_%H%M')
    sufijo = 'Filtros' if formato == 'excel_pagina' else 'Todo'
    filename = f'Responsables_{sufijo}_{fecha_archivo}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _make_option(value, label=None):
    value_txt = _normalize_optional_text(value)
    label_txt = _normalize_optional_text(label if label is not None else value)
    if not value_txt or not label_txt:
        return None
    return {'value': value_txt, 'label': label_txt}


@lru_cache(maxsize=1)
def _get_filter_options():
    with connections[PADRON_DB].cursor() as cursor:
        cursor.execute("SELECT descripcion FROM tipo_documento_tipo ORDER BY c_tipo_documento")
        tipo_documento = [_normalize_text(row[0]) for row in cursor.fetchall() if _normalize_text(row[0])]

        cursor.execute("SELECT descripcion FROM sexo_tipo ORDER BY c_sexo")
        sexo = [_normalize_text(row[0]) for row in cursor.fetchall() if _normalize_text(row[0])]

    return {
        'tipo_documento': tipo_documento,
        'sexo': sexo,
    }


def _serialize_list_item(row):
    return {
        'id': row.get('id'),
        'apellido': _normalize_text(row.get('apellido')),
        'nombre': _normalize_text(row.get('nombre')),
        'tipo_documento': _normalize_text(row.get('tipo_documento')),
        'nro_documento': _normalize_optional_text(row.get('nro_documento')),
        'telefono': _normalize_optional_text(row.get('telefono'), placeholders={'00000'}),
        'sexo': _normalize_text(row.get('sexo')),
        'email': _normalize_optional_text(row.get('email')),
        'cueanexo': _normalize_optional_text(row.get('cueanexo')),
        'codigo_jurisdiccional': _normalize_optional_text(row.get('codigo_jurisdiccional')),
        'cuil_cuit': _normalize_optional_text(row.get('cuil_cuit')),
    }


def _serialize_responsable(row):
    return {
        'id_responsable': row.get('id_responsable'),
        'apellido': _normalize_text(row.get('apellido')),
        'nombre': _normalize_text(row.get('nombre')),
        'tipo_documento': _normalize_text(row.get('tipo_documento')),
        'nro_documento': _normalize_optional_text(row.get('nro_documento')),
        'nacionalidad': _normalize_text(row.get('nacionalidad')),
        'fecha_nacimiento': _format_date(row.get('fecha_nacimiento')),
        'sexo': _normalize_text(row.get('sexo')),
        'telefono': _normalize_optional_text(row.get('telefono'), placeholders={'00000'}),
        'email': _normalize_optional_text(row.get('email')),
        'cuil_cuit': _normalize_optional_text(row.get('cuil_cuit')),
        'fecha_actualizacion': _format_date(row.get('fecha_actualizacion')),
    }

@padron_interno_admin_o_gestor_required
def detalle_responsable_json(request, id_responsable):
    responsable_sql = """
        SELECT
            r.id_responsable,
            r.apellido,
            r.nombre,
            tdt.descripcion AS tipo_documento,
            r.nro_documento,
            ot.descripcion AS nacionalidad,
            r.fecha_nacimiento,
            st.descripcion AS sexo,
            r.telefono,
            r.email,
            r.cuil_cuit,
            r.fecha_actualizacion
        FROM responsable r
        LEFT JOIN tipo_documento_tipo tdt ON tdt.c_tipo_documento = r.c_tipo_documento
        LEFT JOIN sexo_tipo st ON st.c_sexo = r.c_sexo
        LEFT JOIN origen_tipo ot ON ot.c_origen = r.c_nacionalidad
        WHERE r.id_responsable = %s
    """

    establecimientos_sql = f"""
        SELECT
            ve.*,
            {EST_OBSERVACIONES_SQL} AS observaciones,
            {EST_TIPO_OFERTAS_SQL} AS tipo_ofertas
        {EST_BASE_SQL}
        WHERE ve.id_establecimiento IN (
            SELECT e.id_establecimiento
            FROM establecimiento e
            WHERE e.id_responsable = %s
        )
        ORDER BY ve.cue::text, ve.id_establecimiento
    """

    localizaciones_sql = """
        SELECT
            vl.*,
            tel_sup.valor AS tel_supervisor,
            email_sup.valor AS email_supervisor,
            ofres.ofertas_resumen
        FROM vp_localizaciones vl
        LEFT JOIN loc_campo_prov_valor tel_sup
            ON tel_sup.id_localizacion = vl.id_localizacion
           AND tel_sup.id_campo_prov = 1019638042
        LEFT JOIN loc_campo_prov_valor email_sup
            ON email_sup.id_localizacion = vl.id_localizacion
           AND email_sup.id_campo_prov = 1019638043
        LEFT JOIN LATERAL (
            SELECT STRING_AGG(BTRIM(ot.descripcion), ', ' ORDER BY ot.c_oferta, ol.id_oferta_local) AS ofertas_resumen
            FROM oferta_local ol
            JOIN oferta_tipo ot ON ot.c_oferta = ol.c_oferta
            WHERE ol.id_localizacion = vl.id_localizacion
        ) ofres ON TRUE
        WHERE vl.id_responsable = %s
        ORDER BY vl.cue::text, COALESCE(vl.anexo::text, '') DESC, vl.id_localizacion
    """

    try:
        responsable_row = _fetch_one(responsable_sql, [id_responsable])
        if not responsable_row:
            return JsonResponse({'error': 'Responsable no encontrado.'}, status=404)

        establecimientos_rows = _fetch_all(establecimientos_sql, [id_responsable])
        localizaciones_rows = _fetch_all(localizaciones_sql, [id_responsable])

        payload = {
            'responsable': _serialize_responsable(responsable_row),
            'establecimientos': [_serialize_establecimiento_detalle(row) for row in establecimientos_rows],
            'localizaciones': [_serialize_localizacion_detalle(row) for row in localizaciones_rows],
        }
        return JsonResponse(_sanitize_json_payload(payload))
    except Exception as exc:
        return JsonResponse({'error': f'Error al obtener detalle de responsable: {exc}'}, status=500)

@padron_interno_admin_o_gestor_required
def listar_responsables(request):
    formato = request.GET.get('formato')

    if formato == 'excel_todo':
        where, params = '', []
    else:
        where, params = _build_where(request)

    orden_key = request.GET.get('orden', 'apellido')
    col_orden = CAMPO_SQL.get(orden_key, "COALESCE(BTRIM(r.apellido), '')")

    data_sql = f"{_SELECT_FIELDS} {_BASE_SQL} {where} ORDER BY {col_orden}, r.id_responsable"

    if formato in {'excel_pagina', 'excel_todo'}:
        datos_base = _fetch_all(data_sql, params)
        datos_exportar = [_serialize_list_item(row) for row in datos_base]
        return _exportar_excel(datos_exportar, formato, request)

    count_sql = f"SELECT COUNT(*) {_BASE_SQL} {where}"

    try:
        current_page_size = int(request.GET.get('page_size', PAGE_SIZE))
    except (TypeError, ValueError):
        current_page_size = PAGE_SIZE

    raw = _RawPage(data_sql, count_sql, params, PADRON_DB)
    paginator = Paginator(raw, current_page_size)

    try:
        page_number = request.GET.get('page', 1)
        page_obj = paginator.page(page_number)
    except (EmptyPage, PageNotAnInteger):
        page_obj = paginator.page(1)

    total = len(raw)
    desde = (page_obj.number - 1) * current_page_size + 1 if total else 0
    hasta = min(page_obj.number * current_page_size, total)

    lista_items = [_serialize_list_item(row) for row in page_obj.object_list]

    context = {
        'lista_items': lista_items,
        'page_obj': page_obj,
        'resultado_total': total,
        'resultado_desde': desde,
        'resultado_hasta': hasta,
        'username': getattr(request.user, 'username', ''),
        'request': request,
        'filter_options_json': json.dumps(_get_filter_options(), ensure_ascii=False),
    }
    context.update(get_contexto_fecha_padron(request))
    return render(request, 'padroninterno/listadoresponsables.html', context)
