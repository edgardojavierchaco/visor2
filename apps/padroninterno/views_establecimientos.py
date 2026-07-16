import re
from datetime import date, datetime
from functools import lru_cache
from io import BytesIO
from .permisos import padron_interno_admin_o_gestor_required
from .views_fecha import get_contexto_fecha_padron
import openpyxl
from django.db import connections
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

MATERIALIZADAS_DB = 'default'
PAGE_SIZE = 10
OFERTA_ESTADO_FILTER_OPTIONS = ('Activo', 'Inactivo', 'Baja', 'Inactivo sin Docentes')

# Helpers SQL para separar campos guardados como "codigo-descripcion".
# Se reutilizan en columnas visibles, filtros y exportacion.

def _cp_desc_expr(column_name):
    return (
        f"COALESCE(BTRIM(CASE "
        f"WHEN POSITION('-' IN {column_name}) > 0 "
        f"THEN SUBSTRING({column_name} FROM POSITION('-' IN {column_name}) + 1) "
        f"ELSE {column_name} END), '')"
    )


def _cp_code_expr(column_name):
    return (
        f"COALESCE(BTRIM(CASE "
        f"WHEN POSITION('-' IN {column_name}) > 0 "
        f"THEN SUBSTRING({column_name} FROM 1 FOR POSITION('-' IN {column_name}) - 1) "
        f"ELSE {column_name} END), '')"
    )


DIRECTOR_SQL = """
    BTRIM(
        CONCAT(
            COALESCE(ve.responsable_apellido, ''),
            CASE
                WHEN COALESCE(ve.responsable_nombre, '') <> '' AND COALESCE(ve.responsable_apellido, '') <> ''
                    THEN ', ' || ve.responsable_nombre
                ELSE COALESCE(ve.responsable_nombre, '')
            END,
            CASE
                WHEN ve.documento_responsable IS NOT NULL
                    THEN '(' || ve.documento_responsable::text || ')'
                ELSE ''
            END
        )
    )
"""


# Fragmentos SQL compartidos por listado, detalle y exportacion.
OBSERVACIONES_SQL = "COALESCE(BTRIM(ve.observaciones), '')"
TIPO_OFERTAS_SQL = "COALESCE(BTRIM(ve.tipo_ofertas), '')"

# Mapa de campos permitidos para ordenar y filtrar desde parametros GET.
CAMPO_SQL = {
    'cue': 've.cue::text',
    'cantidad_localizaciones': 'COALESCE(ve.cantidad_localizaciones, 0)',
    'codigo_jurisdiccional': "COALESCE(BTRIM(ve.codigo_jurisdiccional_sede), '')",
    'nombre': "COALESCE(BTRIM(ve.nombre), '')",
    'sector': "COALESCE(BTRIM(ve.sector), '')",
    'dependencia': "COALESCE(BTRIM(ve.dependencia), '')",
    'confesional': "COALESCE(BTRIM(ve.confesional), '')",
    'arancelado': "COALESCE(BTRIM(ve.arancelado), '')",
    'categoria': "COALESCE(BTRIM(ve.categoria), '')",
    'director': DIRECTOR_SQL,
    'estado': "COALESCE(BTRIM(ve.estado), '')",
    'tipo_ofertas': TIPO_OFERTAS_SQL,
    'localidad': "COALESCE(BTRIM(ve.localidad_sede), '')",
    'departamento': "COALESCE(BTRIM(ve.departamento_sede), '')",
    'nro_establecimiento': "COALESCE(BTRIM(ve.cp_numeroestablecimiento), '')",
    'tipo_educacion': _cp_desc_expr('ve.cp_est_tipo_ed'),
    'nivel': _cp_desc_expr('ve.cp_est_nivel'),
    'cargo_director': _cp_desc_expr('ve.cp_est_cargo_director'),
    'observaciones': OBSERVACIONES_SQL,
    'fecha_inst_legal': "COALESCE(BTRIM(ve.cp_est_fecha_inst_legal), '')",
    'nro_inst_legal': "COALESCE(BTRIM(ve.cp_est_nro_inst_legal), '')",
    'anio_creacion': "COALESCE(BTRIM(ve.cp_est_anio_inst_legal), '')",
    'descrip_inst_legal': _cp_code_expr('ve.cp_est_descrip_inst_legal'),
}

# Nombres legibles para mostrar filtros aplicados en reportes.
VISIBLE_NAME_MAP = {
    'cue': 'Cue',
    'cantidad_localizaciones': 'Cantidad de Localizaciones',
    'codigo_jurisdiccional': 'Codigo Jurisdiccional',
    'nombre': 'Nombre',
    'sector': 'Sector',
    'dependencia': 'Dependencia',
    'confesional': 'Confesional',
    'arancelado': 'Arancelado',
    'categoria': 'Categoría',
    'director': 'Director',
    'estado': 'Estado',
    'tipo_ofertas': 'Tipo de Ofertas',
    'localidad': 'Localidad',
    'departamento': 'Departamento',
    'nro_establecimiento': 'Nro. de Establecimiento',
    'tipo_educacion': 'Establecimiento TIPO EDUCACION',
    'nivel': 'Establecimiento NIVEL',
    'cargo_director': 'Establecimiento CARGO DIRECTOR',
    'observaciones': 'Observaciones',
    'fecha_inst_legal': 'FECHA Inst. Legal Creacion del Establecimiento',
    'nro_inst_legal': 'NRO. Inst. Legal Creacion del Establecimiento',
    'anio_creacion': 'AÑO Creacion del Establecimiento',
    'descrip_inst_legal': 'DESCRIP. Inst. Legal Creacion del Establecimiento',
}

# Columnas disponibles al generar el Excel.
COLUMNAS_EXPORTACION = [
    ('Cue', 'cue'),
    ('Cantidad de Localizaciones', 'cantidad_localizaciones'),
    ('Codigo Jurisdiccional', 'codigo_jurisdiccional'),
    ('Nombre', 'nombre'),
    ('Sector', 'sector'),
    ('Dependencia', 'dependencia'),
    ('Confesional', 'confesional'),
    ('Arancelado', 'arancelado'),
    ('Categoría', 'categoria'),
    ('Director', 'director'),
    ('Estado', 'estado'),
    ('Tipo de Ofertas', 'tipo_ofertas'),
    ('Localidad', 'localidad'),
    ('Departamento', 'departamento'),
    ('Nro. de Establecimiento', 'nro_establecimiento'),
    ('Establecimiento TIPO EDUCACION', 'tipo_educacion'),
    ('Establecimiento NIVEL', 'nivel'),
    ('Establecimiento CARGO DIRECTOR', 'cargo_director'),
    ('Observaciones', 'observaciones'),
    ('FECHA Inst. Legal Creacion del Establecimiento', 'fecha_inst_legal'),
    ('NRO. Inst. Legal Creacion del Establecimiento', 'nro_inst_legal'),
    ('AÑO Creacion del Establecimiento', 'anio_creacion'),
    ('DESCRIP. Inst. Legal Creacion del Establecimiento', 'descrip_inst_legal'),
]

# SELECT principal del listado de establecimientos.
_SELECT_FIELDS = f"""
    SELECT
        ve.id_establecimiento AS id,
        COALESCE(BTRIM(ve.cue::text), '') AS cue,
        COALESCE(ve.cantidad_localizaciones, 0) AS cantidad_localizaciones,
        COALESCE(BTRIM(ve.codigo_jurisdiccional_sede), '') AS codigo_jurisdiccional,
        COALESCE(BTRIM(ve.nombre), '') AS nombre,
        COALESCE(BTRIM(ve.sector), '') AS sector,
        COALESCE(BTRIM(ve.dependencia), '') AS dependencia,
        COALESCE(BTRIM(ve.confesional), '') AS confesional,
        COALESCE(BTRIM(ve.arancelado), '') AS arancelado,
        COALESCE(BTRIM(ve.categoria), '') AS categoria,
        {DIRECTOR_SQL} AS director,
        COALESCE(BTRIM(ve.estado), '') AS estado,
        {TIPO_OFERTAS_SQL} AS tipo_ofertas,
        COALESCE(BTRIM(ve.localidad_sede), '') AS localidad,
        COALESCE(BTRIM(ve.departamento_sede), '') AS departamento,
        COALESCE(BTRIM(ve.cp_numeroestablecimiento), '') AS nro_establecimiento,
        {_cp_desc_expr('ve.cp_est_tipo_ed')} AS tipo_educacion,
        {_cp_desc_expr('ve.cp_est_nivel')} AS nivel,
        {_cp_desc_expr('ve.cp_est_cargo_director')} AS cargo_director,
        {OBSERVACIONES_SQL} AS observaciones,
        COALESCE(BTRIM(ve.cp_est_fecha_inst_legal), '') AS fecha_inst_legal,
        COALESCE(BTRIM(ve.cp_est_nro_inst_legal), '') AS nro_inst_legal,
        COALESCE(BTRIM(ve.cp_est_anio_inst_legal), '') AS anio_creacion,
        {_cp_code_expr('ve.cp_est_descrip_inst_legal')} AS descrip_inst_legal
"""

# FROM principal sobre la materializada validada.
_BASE_SQL = """
    FROM padroninterno.mv_establecimientos ve
"""


_COUNT_LIGHT_SQL = """
    FROM padroninterno.mv_establecimientos ve
"""

_COUNT_HEAVY_FILTER_FIELDS = {'tipo_ofertas', 'observaciones'}
_COUNT_HEAVY_ALIASES = ('est_obs.', 'otipos.')


# Decide si el conteo necesita joins pesados segun filtros o busqueda.
def _request_requires_heavy_count(request, where):
    if request.GET.get('q', '').strip():
        return True

    if any(
        campo.strip() in _COUNT_HEAVY_FILTER_FIELDS
        for campo in request.GET.getlist('campo_filtro')
    ):
        return True

    return any(alias in where for alias in _COUNT_HEAVY_ALIASES)


def _build_count_sql_establecimientos(request, where, params):
    base_sql = _BASE_SQL if _request_requires_heavy_count(request, where) else _COUNT_LIGHT_SQL
    return f"SELECT COUNT(*) {base_sql} {where}"


def _build_order_clause(request):
    orden_key = request.GET.get('orden', 'cue')
    col_orden = CAMPO_SQL.get(orden_key, 've.cue::text')
    return f"{col_orden}, ve.id_establecimiento"


def _build_data_sql(request, where):
    return f"{_SELECT_FIELDS} {_BASE_SQL} {where} ORDER BY {_build_order_clause(request)}"


def _parse_positive_int(value, default):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _get_page_size(request):
    return min(_parse_positive_int(request.GET.get('page_size'), PAGE_SIZE), 100)


def _get_page_number(request):
    return _parse_positive_int(request.GET.get('page'), 1)


# Paginador liviano para SQL crudo: calcula total solo cuando Django lo pide.
class _RawPage:
    def __init__(self, data_sql, count_sql, params, db_alias):
        self._data_sql = data_sql
        self._count_sql = count_sql
        self._params = params
        self._db_alias = db_alias
        self._length = None

    def __len__(self):
        if self._length is None:
            with connections[self._db_alias].cursor() as cursor:
                cursor.execute(self._count_sql, self._params)
                self._length = cursor.fetchone()[0]
        return self._length

    def __getitem__(self, item):
        if not isinstance(item, slice):
            raise TypeError('Los elementos deben pedirse como slice.')

        start = item.start or 0
        stop = item.stop or start
        limit = max(stop - start, 0)
        sql = f"{self._data_sql} LIMIT {limit} OFFSET {start}"

        with connections[self._db_alias].cursor() as cursor:
            cursor.execute(sql, self._params)
            cols = [d[0] for d in cursor.description]
            return [dict(zip(cols, row)) for row in cursor.fetchall()]


# Operadores permitidos en filtros avanzados.
OPERADOR_SQL = {
    '0': ('ILIKE', lambda v: f'%{v}%'),
    '1': ('NOT ILIKE', lambda v: f'%{v}%'),
    '2': ('=', lambda v: v),
    '3': ('>', lambda v: v),
    '4': ('>=', lambda v: v),
    '5': ('<', lambda v: v),
    '6': ('<=', lambda v: v),
    '7': ('!=', lambda v: v),
}

# Normalizacion para busquedas tolerantes a acentos.
_REPEATED_PAIR_RE = re.compile(r'^(.+)-\1$')
_ACCENTED_CHARS = '\u00c1\u00c9\u00cd\u00d3\u00da\u00dc\u00d1\u00e1\u00e9\u00ed\u00f3\u00fa\u00fc\u00f1'
_UNACCENTED_CHARS = 'AEIOUUNaeiouun'
_ACCENT_TRANSLATION = str.maketrans(_ACCENTED_CHARS, _UNACCENTED_CHARS)


def _fold_filter_text(value):
    return _normalize_text(value).translate(_ACCENT_TRANSLATION).lower()


def _folded_sql(expr):
    return f"LOWER(TRANSLATE(BTRIM(COALESCE(({expr})::text, '')), '{_ACCENTED_CHARS}', '{_UNACCENTED_CHARS}'))"


def _filter_tokens(value):
    folded = _fold_filter_text(value)
    tokens = [token for token in re.split(r'[^a-z0-9]+', folded) if token]
    return list(dict.fromkeys(tokens))


def _like_token_values(value):
    return [f'%{token}%' for token in _filter_tokens(value)]


def _tipo_ofertas_filter_values(oper, value):
    folded = _fold_filter_text(value)
    return _like_token_values(value) if oper in {'0', '1'} else [folded]


def _is_oferta_estado_filter_value(value):
    folded = _fold_filter_text(value)
    return any(folded == _fold_filter_text(option) for option in OFERTA_ESTADO_FILTER_OPTIONS)


def _split_tipo_ofertas_filter_params(oper, values):
    tipo_param_groups = []
    estado_params = []

    for value in values:
        filter_values = _tipo_ofertas_filter_values(oper, value)
        if _is_oferta_estado_filter_value(value):
            estado_params.extend(filter_values)
        else:
            tipo_param_groups.append(filter_values)

    return tipo_param_groups, estado_params


def _append_tipo_ofertas_filter(clauses, params, oper, values):
    tipo_param_groups, estado_params = _split_tipo_ofertas_filter_params(oper, values)
    if not tipo_param_groups and not estado_params:
        return

    clauses.append(_build_tipo_ofertas_clause(
        oper,
        [len(group) for group in tipo_param_groups],
        len(estado_params),
    ))
    for group in tipo_param_groups:
        params.extend(group)
    params.extend(estado_params)


def _build_text_search_clause(expr, value):
    tokens = _filter_tokens(value)
    if not tokens:
        return '', []

    folded_expr = _folded_sql(expr)
    return ' AND '.join([f"{folded_expr} LIKE %s" for _ in tokens]), [f'%{token}%' for token in tokens]


def _tipo_ofertas_clause_operator(oper):
    if oper in {'0', '1'}:
        return 'LIKE'
    if oper in {'2', '7'}:
        return '='
    return OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])[0]


def _build_tipo_ofertas_clause(oper, tipo_value_counts=None, estado_count=0):
    tipo_value_counts = tipo_value_counts or []
    descripcion = _folded_sql('olf.oferta')
    estado = _folded_sql('olf.estado_ofertalocal')
    op_str = _tipo_ofertas_clause_operator(oper)
    conditions = []

    if tipo_value_counts:
        tipo_checks = []
        for token_count in tipo_value_counts:
            token_checks = ' AND '.join([f"{descripcion} {op_str} %s" for _ in range(max(token_count, 1))])
            tipo_checks.append(f"({token_checks})")
        conditions.append(f"({' OR '.join(tipo_checks)})")

    if estado_count:
        estado_checks = ' OR '.join([f"{estado} {op_str} %s" for _ in range(max(estado_count, 1))])
        conditions.append(f"({estado_checks})")

    exists_operator = 'NOT EXISTS' if oper in {'1', '7'} else 'EXISTS'
    return (
        f"{exists_operator} ("
        "SELECT 1 "
        "FROM padroninterno.mv_ofertaslocales olf "
        "WHERE olf.id_establecimiento = ve.id_establecimiento "
        f"AND {' AND '.join(conditions)}"
        ")"
    )


# Normaliza valores provenientes de la base antes de mostrarlos.
def _normalize_text(value, keep_linebreaks=False):
    if value is None:
        return ''

    text = str(value)
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    if text.strip().upper() == 'NULL':
        return ''

    if keep_linebreaks:
        lines = [re.sub(r'\s+', ' ', line).strip() for line in text.split('\n')]
        return '\n'.join([line for line in lines if line]).strip()

    return re.sub(r'\s+', ' ', text).strip()


def _normalize_optional_text(value, placeholders=None, keep_linebreaks=False):
    normalized = _normalize_text(value, keep_linebreaks=keep_linebreaks)
    if not normalized:
        return ''

    normalized_lower = normalized.lower()
    invalid_values = {
        '-',
        '.',
        's/inf.',
        's/inf',
        's/info',
        's/info.',
        's/datos',
    }

    if placeholders:
        invalid_values.update({str(item).strip().lower() for item in placeholders if str(item).strip()})

    if normalized_lower in invalid_values:
        return ''

    return normalized


def _format_date(value):
    normalized = _normalize_optional_text(value)
    if not normalized:
        return ''

    if isinstance(value, datetime):
        dt_value = value
    elif isinstance(value, date):
        dt_value = datetime.combine(value, datetime.min.time())
    else:
        raw = normalized.split(' ')[0]
        dt_value = None
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%d/%m/%Y'):
            try:
                dt_value = datetime.strptime(raw, fmt)
                break
            except ValueError:
                continue
        if dt_value is None:
            return normalized

    meses = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
    return f"{meses[dt_value.month - 1]} {dt_value.day:02d} {dt_value.year}"


def _bool_to_si_no(value):
    if value is None:
        return ''

    if isinstance(value, bool):
        return 'Si' if value else 'No'

    normalized = _normalize_text(value).lower()
    if normalized in {'true', 't', '1', 'si', 'sí'}:
        return 'Si'
    if normalized in {'false', 'f', '0', 'no'}:
        return 'No'
    return _normalize_text(value)


def _build_persona(apellido, nombre, documento=''):
    apellido_txt = _normalize_text(apellido)
    nombre_txt = _normalize_text(nombre)
    documento_txt = _normalize_text(documento)

    if apellido_txt and nombre_txt:
        persona = f'{apellido_txt}, {nombre_txt}'
    else:
        persona = apellido_txt or nombre_txt

    if documento_txt:
        persona = f'{persona}({documento_txt})' if persona else documento_txt

    return persona


def _build_domicilio(calle, nro, cod_postal, localidad, departamento):
    calle_txt = _normalize_text(calle)
    nro_txt = _normalize_text(nro)
    cod_postal_txt = _normalize_text(cod_postal)
    localidad_txt = _normalize_text(localidad)
    departamento_txt = _normalize_text(departamento)
    localidad_con_cp = localidad_txt

    if cod_postal_txt and localidad_txt:
        localidad_con_cp = f'({cod_postal_txt}) {localidad_txt}'
    elif cod_postal_txt:
        localidad_con_cp = f'({cod_postal_txt})'

    primera_parte = ' '.join([value for value in [calle_txt, nro_txt] if value]).strip()
    return ' - '.join([value for value in [primera_parte, localidad_con_cp, departamento_txt] if value])

def _build_domicilio_con_referencia(calle, nro, referencia, cod_postal, localidad, departamento):
    calle_txt = _normalize_text(calle)
    nro_txt = _normalize_text(nro)
    referencia_txt = _normalize_optional_text(referencia, keep_linebreaks=True)

    domicilio_base = _build_domicilio(calle, nro, cod_postal, localidad, departamento)

    if referencia_txt and not calle_txt and not nro_txt:
        if domicilio_base:
            return f"Referencia: {referencia_txt} - {domicilio_base}"
        return f"Referencia: {referencia_txt}"

    return domicilio_base

def _build_domicilio_establecimiento(calle, nro, referencia, cod_postal, localidad, departamento):
    domicilio_base = _build_domicilio(calle, nro, cod_postal, localidad, departamento)
    referencia_txt = _normalize_optional_text(referencia, keep_linebreaks=True)

    if referencia_txt:
        if domicilio_base:
            return f"Referencia: {referencia_txt} - {domicilio_base}"
        return f"Referencia: {referencia_txt}"

    return domicilio_base

def _build_cue_anexo(cue, anexo):
    cue_txt = _normalize_text(cue)
    anexo_txt = _normalize_text(anexo)

    if not cue_txt:
        return ''
    if not anexo_txt:
        return cue_txt

    if anexo_txt.isdigit():
        anexo_txt = anexo_txt.zfill(2)

    return f'{cue_txt}-{anexo_txt}'


@lru_cache(maxsize=None)
def _get_campo_codigo_pairs(id_campo_prov):
    return tuple()


def _split_campo_codigo(value, id_campo_prov=None):
    normalized = _normalize_text(value)
    if not normalized:
        return '', ''

    if id_campo_prov is not None:
        for code, description in _get_campo_codigo_pairs(id_campo_prov):
            if not code:
                continue
            if normalized == code:
                return code, description or code
            prefix = f'{code}-'
            if normalized.startswith(prefix):
                return code, _normalize_text(normalized[len(prefix):]) or description or code

    repeated_match = _REPEATED_PAIR_RE.match(normalized)
    if repeated_match:
        half = _normalize_text(repeated_match.group(1))
        return half, half

    if '-' in normalized:
        code, description = normalized.split('-', 1)
        return _normalize_text(code), _normalize_text(description)

    return normalized, ''


def _format_desc_only(value, id_campo_prov=None):
    code, description = _split_campo_codigo(value, id_campo_prov)
    return description or code


def _format_desc_with_code(value, id_campo_prov=None):
    normalized = _normalize_text(value)
    code, description = _split_campo_codigo(value, id_campo_prov)
    if normalized == '-' and not code and not description:
        return '()'
    if description and code:
        return f'{description} ({code})'
    return description or code


def _format_code_then_desc(value, id_campo_prov=None):
    code, description = _split_campo_codigo(value, id_campo_prov)
    if code and description and code.casefold() == description.casefold():
        return code
    if code and description:
        return f'{code}, {description}'
    return code or description


def _format_desc_then_code(value, id_campo_prov=None):
    code, description = _split_campo_codigo(value, id_campo_prov)
    if code and description:
        return f'{description}, {code}'
    return description or code


# Construye el WHERE a partir de filtros individuales y busqueda global.
def _build_where(request):
    clauses = []
    params = []
    grouped_positive_filters = {}
    geo_positive_filters = []
    cantidad_localizaciones_iguales = []

    campos = request.GET.getlist('campo_filtro')
    opers = request.GET.getlist('operador_filtro')
    valores = request.GET.getlist('valor_filtro')

    for index, campo in enumerate(campos):
        campo = campo.strip()
        oper = opers[index].strip() if index < len(opers) else '0'
        valor = valores[index].strip() if index < len(valores) else ''

        if not campo or not valor or campo not in CAMPO_SQL:
            continue

        if campo == 'cantidad_localizaciones':
            if oper in {'0', '1'}:
                oper = '2'
            if oper not in {'2', '3', '4', '5', '6', '7'}:
                continue
            try:
                valor_int = int(valor)
            except ValueError:
                continue
            if oper == '2':
                if valor_int not in cantidad_localizaciones_iguales:
                    cantidad_localizaciones_iguales.append(valor_int)
                continue
            op_str = OPERADOR_SQL.get(oper, OPERADOR_SQL['2'])[0]
            clauses.append(f"{CAMPO_SQL[campo]} {op_str} %s")
            params.append(valor_int)
            continue

        if oper in {'0', '2'}:
            grouped_positive_filters.setdefault((campo, oper), [])
            if valor not in grouped_positive_filters[(campo, oper)]:
                grouped_positive_filters[(campo, oper)].append(valor)
            continue

        if campo == 'tipo_ofertas' and oper in {'1', '7'}:
            _append_tipo_ofertas_filter(clauses, params, oper, [valor])
            continue

        col = CAMPO_SQL[campo]
        op_str, val_fn = OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])
        if oper == '0':
            token_clause, token_params = _build_text_search_clause(col, valor)
            if token_clause:
                clauses.append(token_clause)
                params.extend(token_params)
            continue

        clauses.append(f"{col}::text {op_str} %s")
        params.append(val_fn(valor))

    if cantidad_localizaciones_iguales:
        col = CAMPO_SQL['cantidad_localizaciones']
        if len(cantidad_localizaciones_iguales) == 1:
            clauses.append(f"{col} = %s")
            params.append(cantidad_localizaciones_iguales[0])
        else:
            clauses.append(f"({' OR '.join([f'{col} = %s'] * len(cantidad_localizaciones_iguales))})")
            params.extend(cantidad_localizaciones_iguales)

    for (campo, oper), values in grouped_positive_filters.items():
        if campo in {'localidad', 'departamento'}:
            geo_positive_filters.append((campo, oper, values))
            continue

        if campo == 'tipo_ofertas':
            _append_tipo_ofertas_filter(clauses, params, oper, values)
            continue

        col = CAMPO_SQL[campo]
        op_str, val_fn = OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])

        if len(values) == 1:
            if oper == '0':
                token_clause, token_params = _build_text_search_clause(col, values[0])
                if token_clause:
                    clauses.append(token_clause)
                    params.extend(token_params)
                continue

            clauses.append(f"{col}::text {op_str} %s")
            params.append(val_fn(values[0]))
            continue

        subclauses = []
        for value in values:
            if oper == '0':
                token_clause, token_params = _build_text_search_clause(col, value)
                if token_clause:
                    subclauses.append(token_clause)
                    params.extend(token_params)
                continue

            subclauses.append(f"{col}::text {op_str} %s")
            params.append(val_fn(value))
        clauses.append(f"({' OR '.join(subclauses)})")

    if geo_positive_filters:
        geo_subclauses = []

        for campo, oper, values in geo_positive_filters:
            col = CAMPO_SQL[campo]
            op_str, val_fn = OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])

            for value in values:
                if oper == '0':
                    token_clause, token_params = _build_text_search_clause(col, value)
                    if token_clause:
                        geo_subclauses.append(token_clause)
                        params.extend(token_params)
                    continue

                geo_subclauses.append(f"{col}::text {op_str} %s")
                params.append(val_fn(value))

        if geo_subclauses:
            clauses.append(f"({' OR '.join(geo_subclauses)})")

    q = request.GET.get('q', '').strip()
    if q:
        token_groups = []
        for token in _filter_tokens(q):
            global_search_clauses = [f"{_folded_sql(sql_expr)} LIKE %s" for sql_expr in CAMPO_SQL.values()]
            token_groups.append(f"({' OR '.join(global_search_clauses)})")
            params.extend([f'%{token}%'] * len(global_search_clauses))
        if token_groups:
            clauses.append(f"({' AND '.join(token_groups)})")

    where = f"WHERE {' AND '.join(clauses)}" if clauses else ''
    return where, params


# Arma el resumen de filtros que aparece en el encabezado del Excel.
def _armar_texto_filtros(request):
    partes = []
    operadores_txt = {
        '0': 'parecido a',
        '1': 'no parecido a',
        '2': 'igual a',
        '3': 'mayor a',
        '4': 'mayor o igual a',
        '5': 'menor a',
        '6': 'menor o igual a',
        '7': 'distinto de',
    }

    if request.GET.get('formato') == 'excel_todo':
        return 'Sin filtros aplicados'

    campos = request.GET.getlist('campo_filtro')
    opers = request.GET.getlist('operador_filtro')
    valores = request.GET.getlist('valor_filtro')

    q = request.GET.get('q', '').strip()
    if q:
        partes.append(f'Búsqueda: {q}')

    for index, campo in enumerate(campos):
        campo = campo.strip()
        valor = valores[index].strip() if index < len(valores) else ''
        oper = opers[index].strip() if index < len(opers) else '0'

        if not campo or not valor:
            continue

        partes.append(
            f"{VISIBLE_NAME_MAP.get(campo, campo)} {operadores_txt.get(oper, 'parecido a')}: {valor}"
        )

    return ' | '.join(partes) if partes else 'Sin filtros aplicados'


def _formatear_fecha_estilo_seguimiento():
    ahora = datetime.now()
    dia = ahora.day
    mes = ahora.month
    anio = ahora.year
    hora_12 = ahora.strftime('%I').lstrip('0') or '0'
    minuto = ahora.strftime('%M')
    ampm = 'p. m.' if ahora.hour >= 12 else 'a. m.'
    return f'{dia}/{mes}/{anio} a las {hora_12}:{minuto} {ampm}'


def _resolver_columnas_exportar(request, formato):
    if formato == 'excel_todo':
        return COLUMNAS_EXPORTACION

    visibles = {value.strip() for value in request.GET.getlist('visible_col') if value.strip()}
    if not visibles:
        return COLUMNAS_EXPORTACION

    columnas = [item for item in COLUMNAS_EXPORTACION if item[1] in visibles]
    return columnas or COLUMNAS_EXPORTACION


# Genera el archivo Excel del listado.
def _exportar_excel(datos_exportar, formato, request):
    columnas_mapeo = _resolver_columnas_exportar(request, formato)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Establecimientos'

    total_columnas = len(columnas_mapeo)
    ultima_columna = get_column_letter(total_columnas)

    ws.merge_cells(f'A1:{ultima_columna}1')
    ws['A1'] = 'Informe Establecimientos'
    ws['A1'].font = Font(bold=True, size=10)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'A2:{ultima_columna}2')
    ws['A2'] = f'Informe generado el: {_formatear_fecha_estilo_seguimiento()}'
    ws['A2'].font = Font(size=9)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'A3:{ultima_columna}3')
    ws['A3'] = f'Filtros aplicados: {_armar_texto_filtros(request)}'
    ws['A3'].font = Font(size=9)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20

    header_row = 4
    for col_idx, (titulo_col, _) in enumerate(columnas_mapeo, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=titulo_col)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for item in datos_exportar:
        fila = []
        for _, clave in columnas_mapeo:
            valor = item.get(clave, '')
            fila.append(_normalize_text(valor, keep_linebreaks=True))
        ws.append(fila)

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{ultima_columna}{ws.max_row}'

    if formato == 'excel_todo':
        for col_num in range(1, total_columnas + 1):
            titulo_col = columnas_mapeo[col_num - 1][0]
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = min(max(len(str(titulo_col)) + 4, 14), 28)
    else:
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.font = Font(size=9)

        for col_num in range(1, total_columnas + 1):
            col_letter = get_column_letter(col_num)
            max_length = 0
            for row_num in range(1, ws.max_row + 1):
                value = ws[f'{col_letter}{row_num}'].value
                max_length = max(max_length, len(str(value)) if value is not None else 0)
            ws.column_dimensions[col_letter].width = min(max_length + 2, 42)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_archivo = datetime.now().strftime('%Y%m%d_%H%M')
    sufijo = 'Filtros' if formato == 'excel_pagina' else 'Todo'
    nombre_archivo = f'Establecimientos_{sufijo}_{fecha_archivo}.xlsx'

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response

# Convierte el resultado de un cursor SQL en diccionarios.
def _dictfetchall(cursor):
    columns = [column[0] for column in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def _fetch_one(sql, params):
    with connections[MATERIALIZADAS_DB].cursor() as cursor:
        cursor.execute(sql, params)
        rows = _dictfetchall(cursor)
    return rows[0] if rows else None


def _fetch_all(sql, params):
    with connections[MATERIALIZADAS_DB].cursor() as cursor:
        cursor.execute(sql, params)
        return _dictfetchall(cursor)


# Limpia None recursivamente para que el frontend reciba cadenas vacias.
def _sanitize_json_payload(value):
    if isinstance(value, dict):
        return {key: _sanitize_json_payload(inner_value) for key, inner_value in value.items()}
    if isinstance(value, list):
        return [_sanitize_json_payload(item) for item in value]
    if value is None:
        return ''
    return value


@lru_cache(maxsize=1)
# Opciones cacheadas para combos de filtros.
def _get_filter_options():
    with connections[MATERIALIZADAS_DB].cursor() as cursor:
        def fetch_distinct(column_name):
            cursor.execute(
                f"""
                SELECT DISTINCT BTRIM({column_name}) AS valor
                FROM padroninterno.mv_establecimientos
                WHERE COALESCE(BTRIM({column_name}), '') <> ''
                ORDER BY valor
                """
            )
            return [_normalize_text(row[0]) for row in cursor.fetchall() if _normalize_text(row[0])]

        sectores = fetch_distinct('sector')
        dependencias = fetch_distinct('dependencia')
        confesionales = fetch_distinct('confesional')
        arancelados = fetch_distinct('arancelado')
        categorias = fetch_distinct('categoria')
        estados = fetch_distinct('estado')

        cursor.execute(
            """
            SELECT DISTINCT BTRIM(localidad_sede) AS valor
            FROM padroninterno.mv_establecimientos
            WHERE COALESCE(BTRIM(localidad_sede), '') <> ''
            ORDER BY valor
            """
        )
        localidades = [_normalize_text(row[0]) for row in cursor.fetchall()]

        cursor.execute(
            """
            SELECT DISTINCT BTRIM(departamento_sede) AS valor
            FROM padroninterno.mv_establecimientos
            WHERE COALESCE(BTRIM(departamento_sede), '') <> ''
            ORDER BY valor
            """
        )
        departamentos = [_normalize_text(row[0]) for row in cursor.fetchall()]

        return {
            'sector': sectores,
            'dependencia': dependencias,
            'confesional': confesionales,
            'arancelado': arancelados,
            'categoria': categorias,
            'estado': [value for value in estados if value],
            'localidad': localidades,
            'departamento': departamentos,
            'tipo_educacion': fetch_distinct('cp_est_tipo_ed'),
            'cargo_director': fetch_distinct('cp_est_cargo_director'),
        }


# Serializador compacto para filas del listado.
def _serialize_list_item(row):
    return {
        'id': row.get('id'),
        'cue': _normalize_text(row.get('cue')),
        'cantidad_localizaciones': _normalize_text(row.get('cantidad_localizaciones')),
        'codigo_jurisdiccional': _normalize_text(row.get('codigo_jurisdiccional')),
        'nombre': _normalize_text(row.get('nombre')),
        'sector': _normalize_text(row.get('sector')),
        'dependencia': _normalize_text(row.get('dependencia')),
        'confesional': _normalize_text(row.get('confesional')),
        'arancelado': _normalize_text(row.get('arancelado')),
        'categoria': _normalize_text(row.get('categoria')),
        'director': _normalize_text(row.get('director')),
        'estado': _normalize_text(row.get('estado')),
        'tipo_ofertas': _normalize_text(row.get('tipo_ofertas')),
        'localidad': _normalize_text(row.get('localidad')),
        'departamento': _normalize_text(row.get('departamento')),
        'nro_establecimiento': _normalize_text(row.get('nro_establecimiento')),
        'tipo_educacion': _normalize_text(row.get('tipo_educacion')),
        'nivel': _normalize_text(row.get('nivel')),
        'cargo_director': _normalize_text(row.get('cargo_director')),
        'observaciones': _normalize_text(row.get('observaciones'), keep_linebreaks=True),
        'fecha_inst_legal': _normalize_text(row.get('fecha_inst_legal')),
        'nro_inst_legal': _normalize_text(row.get('nro_inst_legal')),
        'anio_creacion': _normalize_text(row.get('anio_creacion')),
        'descrip_inst_legal': _normalize_text(row.get('descrip_inst_legal')),
    }


# Serializadores de detalle: formatean datos para los modales.
def _serialize_establecimiento(row):
    return {
        'cue': _normalize_text(row.get('cue')),
        'nombre': _normalize_text(row.get('nombre')),
        'sector': _normalize_text(row.get('sector')),
        'dependencia': _normalize_text(row.get('dependencia')),
        'estado': _normalize_text(row.get('estado')),
       'domicilio_principal': _build_domicilio_con_referencia(
            row.get('calle_sede'),
            row.get('nro_sede'),
            row.get('referencia_sede'),
            row.get('cod_postal_sede'),
            row.get('localidad_sede'),
            row.get('departamento_sede'),
        ),
        'fecha_creacion': _format_date(row.get('fecha_creacion')),
        'fecha_alta': _format_date(row.get('fecha_alta')),
        'fecha_baja': _format_date(row.get('fecha_baja')),
        'fecha_actualizacion': _format_date(row.get('fecha_actualizacion')),
        'confesional': _normalize_text(row.get('confesional')),
        'arancelado': _normalize_text(row.get('arancelado')),
        'categoria': _normalize_text(row.get('categoria')),
        'director': _build_persona(
            row.get('responsable_apellido'),
            row.get('responsable_nombre'),
            row.get('documento_responsable'),
        ),
        'ofertas': _normalize_text(row.get('tipo_ofertas')),
        'nro_establecimiento': _normalize_text(row.get('cp_numeroestablecimiento')),
        'fecha_inst_legal': _format_date(row.get('cp_est_fecha_inst_legal')),
        'nro_inst_legal': _normalize_text(row.get('cp_est_nro_inst_legal')),
        'anio_creacion': _normalize_text(row.get('cp_est_anio_inst_legal')),
        'descrip_inst_legal': _format_desc_with_code(row.get('cp_est_descrip_inst_legal'), 1019638057),
        'tipo_educacion': _format_desc_with_code(row.get('cp_est_tipo_ed'), 1019638074),
        'nivel': _format_desc_with_code(row.get('cp_est_nivel'), 1019638075),
        'cargo_director': _format_desc_with_code(row.get('cp_est_cargo_director'), 1019638076),
        'observaciones': _normalize_text(row.get('observaciones'), keep_linebreaks=True),
    }


def _serialize_localizacion(row):
    return {
        'id_localizacion': row.get('id_localizacion'),
        'cue_anexo': _build_cue_anexo(row.get('cue'), row.get('anexo')),
        'nombre': _normalize_text(row.get('nombre')),
        'estado': _normalize_text(row.get('estado_localizacion')),
        'ofertas_resumen': _normalize_text(row.get('ofertas_resumen')),
        'ambito': _normalize_text(row.get('ambito')),
        'codigo_jurisdiccional': _normalize_text(row.get('codigo_jurisdiccional')),
        'telefono': _normalize_text(row.get('telefono')),
        'telefono_cod_area': _normalize_text(row.get('telefono_cod_area')),
        'domicilio_principal': _build_domicilio_con_referencia(
            row.get('calle'),
            row.get('nro'),
            row.get('referencia'),
            row.get('cod_postal'),
            row.get('localidad_nombre'),
            row.get('departamento_nombre'),
        ),
        'alternancia': _normalize_text(row.get('alternancia')),
        'periodo_funcionamiento': _normalize_text(row.get('periodo_funcionamiento')),
        'email': _normalize_text(row.get('email')),
        'sitio_web': _normalize_text(row.get('sitio_web')),
        'cooperadora': _normalize_text(row.get('cooperadora')),
        'responsable': _build_persona(
            row.get('responsable_apellido'),
            row.get('responsable_nombre'),
            row.get('documento_responsable'),
        ),
        'sede': _bool_to_si_no(row.get('sede')),
        'permanencia': _normalize_text(row.get('permanencia')),
        'sede_administrativa': _bool_to_si_no(row.get('sede_administrativa')),
        'observaciones': _normalize_text(row.get('observaciones'), keep_linebreaks=True),
        'zona_provincial': _format_desc_then_code(_normalize_optional_text(row.get('cp_zonaprovincial'))),
        'cui': _normalize_optional_text(row.get('cp_esvar2')),
        'cui_localizacion': _normalize_optional_text(row.get('cp_esvar2')),
        'regional_actual': _format_code_then_desc(row.get('cp_esvat5'), 1019638011),
        'regional_hasta_2015': _format_desc_then_code(row.get('cp_p8104_localizacion_1019638033')),
        'regional_hasta_2020': _format_desc_then_code(row.get('cp_reg_hasta_2020'), 1019638079),
        'supervisor_tecnico': _format_desc_then_code(row.get('cp_supervisortecnico')),
        'microregion': _format_desc_then_code(row.get('cp_esvat4'), 1019638010),
        'udt': _format_desc_then_code(row.get('cp_esvat6'), 1019638012),
        'cuof_localizacion': _normalize_text(row.get('cp_esvar4')),
        'director_regional': _normalize_text(
            row.get('director_regional_detalle') or row.get('director_regional')
        ),
        'telefono_director_regional': _normalize_text(row.get('cp_tedirregional')),
        'email_director_regional': _normalize_text(row.get('cp_emaildirregional')),
        'telefono_supervisor': _normalize_text(row.get('tel_supervisor')),
        'email_supervisor': _normalize_text(row.get('email_supervisor')),
        'patrimonio_edilicio': _format_desc_then_code(_normalize_optional_text(row.get('cp_patrimonioedilicio'))),
        'fecha_creacion_edificio': _format_date(row.get('cp_estfechacreacionedificio')),
        'instr_legal_creacion_edificio': _normalize_optional_text(row.get('cp_edif_instlegal_creaciondeestablecimiento')),
        'anterior_regional_educativa': _format_desc_then_code(row.get('cp_esvat3')),
    }


def _serialize_oferta(row):
    return {
        'id_oferta_local': row.get('id_oferta_local'),
        'localizacion': _build_cue_anexo(row.get('cue'), row.get('anexo')),
        'oferta': _normalize_text(row.get('oferta')),
        'estado': _normalize_text(row.get('estado_ofertalocal')),
        'codigo_jurisdiccional': _normalize_text(row.get('codigo_jurisdiccional_oferta_local')),
        'subvencion': _normalize_text(row.get('subvencion_oferta_local')),
        'jornada': _normalize_text(row.get('jornada_ofertalocal')),
        'cuof': _normalize_text(row.get('cp_efvar4')),
        'cuof_ryc': _normalize_text(row.get('cp_cuof_ryc')),
        'cui': _normalize_optional_text(row.get('cp_efvar2')),
        'cua': _normalize_optional_text(row.get('cp_of_cua')),
        'udt': _format_desc_only(row.get('cp_efvar6')),
        'regional': _format_desc_only(row.get('cp_efvar5')),
        'acronimo': _format_desc_only(row.get('cp_acronimo')),
        'supervisor_tecnico': _format_desc_only(row.get('cp_supervisortecnico_oferta')),
        'tel_supervisor': _normalize_optional_text(row.get('cp_tesupervisor_oferta'), placeholders={'0'}),
        'email_supervisor': _normalize_optional_text(row.get('cp_mailsupervisor_oferta')),
        'fecha_inst_legal': _format_date(row.get('cp_of_fecha_inst_legal')),
        'nro_inst_legal': _normalize_optional_text(row.get('cp_of_nro_inst_legal')),
        'anio_creacion': _normalize_optional_text(row.get('cp_of_anio_inst_legal')),
        'descrip_inst_legal': _format_desc_only(row.get('cp_of_descrip_inst_legal')),
        'cabecera_anexo': _format_desc_only(row.get('cp_of_cab_anexo')),
        'ambito': _format_desc_only(row.get('cp_of_ambito')),
        'tipo_ed': _format_desc_only(row.get('cp_of_tipo_ed')),
        'nivel': _format_desc_only(row.get('cp_of_nivel')),
        'sector': _format_desc_only(row.get('cp_of_sector')),
        'categoria': _format_desc_only(row.get('cp_of_categoria')),
    }


def _format_historial_instr_legal(numero, descripcion):
    numero_txt = _normalize_text(numero)
    descripcion_txt = _normalize_text(descripcion)

    if numero_txt and descripcion_txt:
        return f'({numero_txt}) {descripcion_txt}'
    return descripcion_txt or numero_txt


def _serialize_historial(row):
    cue = _normalize_text(row.get('cue'))
    nombre = _normalize_text(row.get('nombre'))
    fecha_vigencia = _format_date(row.get('fecha_vigencia'))
    establecimiento_label = ' - '.join([value for value in [cue, nombre] if value])

    return {
        'caption': f'{establecimiento_label} ({fecha_vigencia})' if establecimiento_label and fecha_vigencia else establecimiento_label or fecha_vigencia,
        'establecimiento': establecimiento_label,
        'movimiento': ' - '.join(
            [
                value
                for value in [
                    _normalize_text(row.get('id_movimiento')),
                    _normalize_text(row.get('cod_tipo_mov')),
                    _normalize_text(row.get('tipo_movimiento')),
                ]
                if value
            ]
        ),
        'estado': _normalize_text(row.get('estado')),
        'fecha_instr_legal': _format_date(row.get('fecha_inst_legal')),
        'fecha_vigencia': fecha_vigencia,
        'observacion': _normalize_text(row.get('observacion'), keep_linebreaks=True),
        'instr_legal': _format_historial_instr_legal(row.get('nro_instr_legal'), row.get('instr_legal')),
        'motivo': _normalize_text(row.get('motivo')),
        'usuario': _normalize_text(row.get('usuario')),
    }

@padron_interno_admin_o_gestor_required
def detalle_establecimiento_json(request, id_establecimiento):
    # Endpoint de detalle: reune establecimiento, localizaciones, ofertas e historial.
    establecimiento_sql = f"""
        SELECT
            ve.*,
            {OBSERVACIONES_SQL} AS observaciones,
            {TIPO_OFERTAS_SQL} AS tipo_ofertas
        {_BASE_SQL}
        WHERE ve.id_establecimiento = %s
    """

    localizaciones_sql = """
        SELECT
            vl.*,
            vl.tel_supervisor,
            vl.email_supervisor,
            ofres.ofertas_resumen
        FROM padroninterno.mv_localizaciones vl
        LEFT JOIN LATERAL (
            SELECT STRING_AGG(BTRIM(ol.oferta), ', ' ORDER BY ol.c_oferta, ol.id_oferta_local) AS ofertas_resumen
            FROM padroninterno.mv_ofertaslocales ol
            WHERE ol.id_localizacion = vl.id_localizacion
        ) ofres ON TRUE
        WHERE vl.id_establecimiento = %s
        ORDER BY vl.anexo, vl.id_localizacion
    """

    ofertas_sql = """
        SELECT *
        FROM padroninterno.mv_ofertaslocales
        WHERE id_establecimiento = %s
        ORDER BY c_oferta, anexo, id_oferta_local
    """

    historial_sql = """
        SELECT *
        FROM padroninterno.mv_establecimiento_historial
        WHERE id_establecimiento = %s
        ORDER BY fecha_vigencia DESC, id_movimiento DESC
    """

    try:
        establecimiento_row = _fetch_one(establecimiento_sql, [id_establecimiento])
        if not establecimiento_row:
            return JsonResponse({'error': 'Establecimiento no encontrado.'}, status=404)

        localizaciones_rows = _fetch_all(localizaciones_sql, [id_establecimiento])
        ofertas_rows = _fetch_all(ofertas_sql, [id_establecimiento])
        historial_rows = _fetch_all(historial_sql, [id_establecimiento])

        payload = {
            'establecimiento': _serialize_establecimiento(establecimiento_row),
            'localizaciones': [_serialize_localizacion(row) for row in localizaciones_rows],
            'ofertas': [_serialize_oferta(row) for row in ofertas_rows],
            'historial': [_serialize_historial(row) for row in historial_rows],
        }
        return JsonResponse(_sanitize_json_payload(payload))
    except Exception as exc:
        return JsonResponse({'error': f'Error al obtener detalle de establecimiento: {exc}'}, status=500)

@padron_interno_admin_o_gestor_required
def listar_establecimientos(request):
    # Sin formato Excel, solo renderiza la pantalla; los datos llegan por AJAX.
    formato = request.GET.get('formato')

    if formato == 'excel_todo':
        where, params = '', []
    elif formato == 'excel_pagina':
        where, params = _build_where(request)
    else:
        context = {
            'lista_items': [],
            'page_obj': None,
            'resultado_total': None,
            'resultado_desde': 0,
            'resultado_hasta': 0,
            'username': getattr(request.user, 'username', ''),
            'request': request,
            'filter_options_json': '{}',
            'establecimientos_async_loading': True,
        }
        context.update(get_contexto_fecha_padron(request))
        return render(request, 'padroninterno/establecimientos.html', context)

    data_sql = _build_data_sql(request, where)

    if formato in {'excel_pagina', 'excel_todo'}:
        datos_exportar = _fetch_all(data_sql, params)
        return _exportar_excel(datos_exportar, formato, request)


@padron_interno_admin_o_gestor_required
def establecimientos_datos_json(request):
    # Endpoint paginado: trae una fila extra para saber si existe pagina siguiente.
    where, params = _build_where(request)
    page_size = _get_page_size(request)
    page = _get_page_number(request)
    offset = (page - 1) * page_size
    data_sql = _build_data_sql(request, where)
    rows = _fetch_all(f"{data_sql} LIMIT %s OFFSET %s", params + [page_size + 1, offset])
    page_rows = rows[:page_size]
    items = [_serialize_list_item(row) for row in page_rows]

    return JsonResponse(_sanitize_json_payload({
        'items': items,
        'has_next': len(rows) > page_size,
        'has_previous': page > 1,
        'desde': offset + 1 if items else 0,
        'hasta': offset + len(items) if items else 0,
        'page': page,
        'page_size': page_size,
        'total': None,
        'total_pending': True,
    }))


@padron_interno_admin_o_gestor_required
def establecimientos_total_json(request):
    # Conteo separado para no bloquear la carga inicial del listado.
    where, params = _build_where(request)
    count_sql = _build_count_sql_establecimientos(request, where, params)

    with connections[MATERIALIZADAS_DB].cursor() as cursor:
        cursor.execute(count_sql, params)
        total = cursor.fetchone()[0]

    return JsonResponse({'total': total})


@padron_interno_admin_o_gestor_required
def establecimientos_filtros_json(request):
    # Catalogos usados por los filtros avanzados del listado.
    return JsonResponse(_get_filter_options())
