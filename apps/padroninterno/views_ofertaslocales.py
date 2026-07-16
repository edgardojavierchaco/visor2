from django.shortcuts import render
from django.db import connections
from .permisos import padron_interno_admin_o_gestor_required
from .views_fecha import get_contexto_fecha_padron
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse, JsonResponse
from datetime import datetime
from io import BytesIO
import re

MATERIALIZADAS_DB = 'default'
PAGE_SIZE = 10
OFERTA_ESTADO_FILTER_OPTIONS = ('Activo', 'Inactivo', 'Baja', 'Inactivo sin Docentes')

# Mapa cerrado de campos que pueden usarse en filtros y ordenamientos.
CAMPO_SQL = {
    'cue': 'e.cue::text',
    'anexo': "COALESCE(l.anexo, '')",
    'codigo_jurisdiccional': 'ol.codigo_jurisdiccional_oferta_local',
    'localizacion': 'l.nombre',
    'tipo_oferta': "BTRIM(COALESCE(ol.oferta, ''))",
    'nombre_titulo': "BTRIM(COALESCE(ol.oferta, ''))",
    'estado': "BTRIM(COALESCE(ol.estado_ofertalocal, ''))",
    'subvencion': "BTRIM(COALESCE(ol.subvencion_oferta_local, ''))",
    'jornada': "BTRIM(COALESCE(ol.jornada_ofertalocal, ''))",
    'matricula_total': 'ol.matricula_total_ofertalocal::text',
    'mod_compl_planes': 'ol.modalidad_basica',
    'cabecera_anexo': "BTRIM(COALESCE(ol.cp_of_cab_anexo, ''))",
    'ambito': "BTRIM(COALESCE(ol.cp_of_ambito, ''))",
    'tipo_ed': "BTRIM(COALESCE(ol.cp_of_tipo_ed, ''))",
    'nivel': "BTRIM(COALESCE(ol.cp_of_nivel, ''))",
    'fecha_creacion': 'ol.fecha_creacion_ofertalocal::text',
    'sector': "BTRIM(COALESCE(ol.cp_of_sector, ''))",
    'acronimo': "BTRIM(COALESCE(ol.cp_acronimo, ''))",
    'categoria': "BTRIM(COALESCE(ol.cp_of_categoria, ''))",
    'fecha_inst_legal': 'ol.cp_of_fecha_inst_legal',
    'nro_inst_legal': 'ol.cp_of_nro_inst_legal',
    'anio_creacion': 'ol.cp_of_anio_inst_legal',
    'descrip_inst_legal': "BTRIM(COALESCE(ol.cp_of_descrip_inst_legal, ''))",
    'cui': 'ol.cp_efvar2',
    'tel_supervisor': 'ol.cp_tesupervisor_oferta',
    'email_supervisor': 'ol.cp_mailsupervisor_oferta',
    'supervisor_tecnico': "BTRIM(COALESCE(ol.cp_supervisortecnico_oferta, ''))",
    'udt': "BTRIM(COALESCE(ol.cp_efvar6, ''))",
    'cuof': 'ol.cp_efvar4',
    'cua': 'ol.cp_of_cua',
    'regional': "BTRIM(COALESCE(ol.cp_efvar5, ''))",
    'cuof_ryc': 'ol.cp_cuof_ryc',
}

# Columnas disponibles para el Excel de ofertas locales.
COLUMNAS_EXPORTACION = [
    ('CUE', 'cue'),
    ('Anexo', 'anexo'),
    ('Cód. Jurisdiccional', 'codigo_jurisdiccional'),
    ('Localización', 'localizacion'),
    ('Tipo Oferta', 'tipo_oferta'),
    ('Estado', 'estado'),
    ('Subvención', 'subvencion'),
    ('Jornada', 'jornada'),
    ('Matrícula', 'matricula_total'),
    ('Mod. Compl. Planes', 'mod_compl_planes'),
    ('Oferta Cabecera', 'oferta_cabecera'),
    ('Ámbito', 'oferta_ambito'),
    ('Tipo Ed.', 'oferta_tipo_ed'),
    ('Nivel', 'oferta_nivel'),
    ('Fecha Creación', 'fecha_creacion'),
    ('Sector', 'oferta_sector'),
    ('Acrónimo', 'acronimo'),
    ('Categoría', 'oferta_categoria'),
    ('Fecha Inst. Legal', 'fecha_inst_legal'),
    ('Nro Inst. Legal', 'nro_inst_legal'),
    ('Año Creación', 'anio_creacion'),
    ('Descrip. Legal', 'descrip_inst_legal'),
    ('CUI', 'cui'),
    ('Tel. Supervisor', 'tel_supervisor'),
    ('Email Supervisor', 'email_supervisor'),
    ('Sup. Técnico', 'supervisor_tecnico'),
    ('CUOF', 'cuof'),
    ('CUA', 'cua'),
    ('Regional', 'regional'),
    ('UDT', 'udt'),
    ('CUOF RyC', 'cuof_ryc'),
]

# SELECT principal del listado de ofertas locales.
_SELECT_FIELDS = """
    SELECT 
        ol.id_oferta_local AS id,
        e.cue, 
        COALESCE(l.anexo, '') AS anexo, 
        UPPER(COALESCE(BTRIM(ol.codigo_jurisdiccional_oferta_local), '')) AS codigo_jurisdiccional,
        COALESCE(BTRIM(l.nombre), '') AS localizacion,
        COALESCE(BTRIM(ol.oferta), '') AS tipo_oferta,
        COALESCE(BTRIM(ol.estado_ofertalocal), '') AS estado,
        COALESCE(BTRIM(ol.subvencion_oferta_local), '') AS subvencion,
        COALESCE(BTRIM(ol.jornada_ofertalocal), '') AS jornada,
        COALESCE(ol.matricula_total_ofertalocal::text, '') AS matricula_total,
        COALESCE(BTRIM(ol.modalidad_basica), '') AS mod_compl_planes,
        COALESCE(BTRIM(ol.cp_of_cab_anexo), '') AS oferta_cabecera,
        COALESCE(BTRIM(SUBSTRING(ol.cp_of_ambito FROM 1 FOR 1)), '') AS oferta_ambito,
        COALESCE(BTRIM(ol.cp_of_ambito), '') AS ambito,
        COALESCE(BTRIM(ol.cp_of_tipo_ed), '') AS oferta_tipo_ed,
        COALESCE(BTRIM(ol.cp_of_nivel), '') AS oferta_nivel,
        COALESCE(BTRIM(ol.fecha_creacion_ofertalocal::text), '') AS fecha_creacion,
        COALESCE(BTRIM(ol.cp_of_sector), '') AS oferta_sector,
        COALESCE(BTRIM(ol.cp_acronimo), '') AS acronimo,
        COALESCE(BTRIM(ol.cp_of_categoria), '') AS oferta_categoria,
        COALESCE(BTRIM(ol.cp_of_fecha_inst_legal), '') AS fecha_inst_legal,
        COALESCE(BTRIM(ol.cp_of_nro_inst_legal), '') AS nro_inst_legal,
        COALESCE(BTRIM(ol.cp_of_anio_inst_legal), '') AS anio_creacion,
        COALESCE(BTRIM(ol.cp_of_descrip_inst_legal), '') AS descrip_inst_legal,
        COALESCE(BTRIM(ol.cp_efvar2), '') AS cui,

        COALESCE(BTRIM(ol.cp_tesupervisor_oferta), '') AS tel_supervisor,
        COALESCE(BTRIM(ol.cp_mailsupervisor_oferta), '') AS email_supervisor,
        COALESCE(BTRIM(ol.cp_supervisortecnico_oferta), '') AS supervisor_tecnico,
        COALESCE(BTRIM(ol.cp_efvar6), '') AS udt,

        COALESCE(BTRIM(ol.cp_efvar4), '') AS cuof,
        COALESCE(BTRIM(ol.cp_of_cua), '') AS cua,
        COALESCE(BTRIM(ol.cp_efvar5), '') AS regional,
        COALESCE(BTRIM(ol.cp_cuof_ryc), '') AS cuof_ryc
"""

# FROM principal sobre materializadas; mv_ofertaslocales concentra los campos EAV.
_BASE_SQL = """
    FROM padroninterno.mv_ofertaslocales ol
    LEFT JOIN padroninterno.mv_localizaciones l ON ol.id_localizacion = l.id_localizacion
    LEFT JOIN padroninterno.mv_establecimientos e ON ol.id_establecimiento = e.id_establecimiento
    

    


"""

# Detecta que filtros estan activos para decidir estrategia de conteo.
def _active_count_filter_fields(request):
    campos = request.GET.getlist('campo_filtro')
    valores = request.GET.getlist('valor_filtro')
    active_fields = []

    for i, campo_raw in enumerate(campos):
        campo = campo_raw.strip()
        valor = valores[i].strip() if i < len(valores) else ''

        if campo and valor:
            active_fields.append(campo)

    return active_fields


def _request_requires_heavy_count(request):
    return True


def _count_light_join_keys(request):
    return set()


def _build_count_light_sql(request, where):
    return f"SELECT COUNT(*) {_BASE_SQL} {where}"


# Elige entre conteo liviano o completo segun filtros solicitados.
def _build_count_sql(request, where):
    return f"SELECT COUNT(*) {_BASE_SQL} {where}"


def _build_order_clause(request):
    orden_key = request.GET.get('orden', 'cue')
    col_orden = CAMPO_SQL.get(orden_key, 'e.cue::text')
    return f"{col_orden}, l.anexo"


def _build_data_sql(request, where):
    return f"{_SELECT_FIELDS} {_BASE_SQL} {where} ORDER BY {_build_order_clause(request)}"


def _parse_positive_int(value, default):
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _get_page_size(request):
    return min(_parse_positive_int(request.GET.get('page_size'), PAGE_SIZE), 100)


def _get_page_number(request):
    return _parse_positive_int(request.GET.get('page'), 1)


def _serialize_list_item(row):
    return {key: '' if value is None else value for key, value in row.items()}

# Paginador liviano para SQL crudo.
class _RawPage:
    def __init__(self, data_sql, count_sql, params, db):
        self._data_sql = data_sql
        self._count_sql = count_sql
        self._params = params
        self._db = db
        self._len = None

    def __len__(self):
        if self._len is None:
            with connections[self._db].cursor() as c:
                c.execute(self._count_sql, self._params)
                self._len = c.fetchone()[0]
        return self._len

    def __getitem__(self, s):
        if not isinstance(s, slice):
            raise TypeError
        sql = f"{self._data_sql} LIMIT {s.stop - s.start} OFFSET {s.start or 0}"
        with connections[self._db].cursor() as c:
            c.execute(sql, self._params)
            cols = [d[0] for d in c.description]
            return [dict(zip(cols, row)) for row in c.fetchall()]


# Operadores permitidos por los filtros avanzados.
OPERADOR_SQL = {
    '0': ('ILIKE', lambda v: f'%{v}%'),
    '1': ('NOT ILIKE', lambda v: f'%{v}%'),
    '2': ('=', lambda v: v),
    '3': ('>', lambda v: v),
    '4': ('>=', lambda v: v),
    '5': ('<', lambda v: v),
    '6': ('<=', lambda v: v),
    '7': ('!=', lambda v: v),
}

_ACCENTED_CHARS = '\u00c1\u00c9\u00cd\u00d3\u00da\u00dc\u00d1\u00e1\u00e9\u00ed\u00f3\u00fa\u00fc\u00f1'
_UNACCENTED_CHARS = 'AEIOUUNaeiouun'
_ACCENT_TRANSLATION = str.maketrans(_ACCENTED_CHARS, _UNACCENTED_CHARS)


def _fold_filter_text(value):
    return _normalize_text(value).translate(_ACCENT_TRANSLATION).lower()


def _folded_sql(expr):
    return f"LOWER(TRANSLATE(BTRIM(COALESCE(({expr})::text, '')), '{_ACCENTED_CHARS}', '{_UNACCENTED_CHARS}'))"


def _filter_tokens(value):
    folded = _fold_filter_text(value)
    tokens = [token for token in re.split(r'[^a-z0-9]+', folded) if token]
    return list(dict.fromkeys(tokens))


def _build_text_search_clause(expr, value):
    tokens = _filter_tokens(value)
    if not tokens:
        return '', []

    folded_expr = _folded_sql(expr)
    return ' AND '.join([f"{folded_expr} LIKE %s" for _ in tokens]), [f'%{token}%' for token in tokens]


def _tipo_oferta_filter_values(oper, value):
    folded = _fold_filter_text(value)
    return [f'%{token}%' for token in _filter_tokens(value)] if oper in {'0', '1'} else [folded]


def _is_oferta_estado_filter_value(value):
    folded = _fold_filter_text(value)
    return any(folded == _fold_filter_text(option) for option in OFERTA_ESTADO_FILTER_OPTIONS)


def _split_tipo_oferta_filter_params(oper, values):
    tipo_param_groups = []
    estado_params = []

    for value in values:
        filter_values = _tipo_oferta_filter_values(oper, value)
        if _is_oferta_estado_filter_value(value):
            estado_params.extend(filter_values)
        else:
            tipo_param_groups.append(filter_values)

    return tipo_param_groups, estado_params


def _tipo_oferta_clause_operator(oper):
    if oper in {'0', '1'}:
        return 'LIKE'
    if oper in {'2', '7'}:
        return '='
    return OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])[0]


def _tipo_oferta_filter_clause(oper, tipo_value_counts=None, estado_count=0):
    tipo_value_counts = tipo_value_counts or []
    tipo_expr = _folded_sql('ol.oferta')
    estado_expr = _folded_sql('ol.estado_ofertalocal')
    op_str = _tipo_oferta_clause_operator(oper)
    conditions = []

    if tipo_value_counts:
        tipo_checks = []
        for token_count in tipo_value_counts:
            token_checks = ' AND '.join([f"{tipo_expr} {op_str} %s" for _ in range(max(token_count, 1))])
            tipo_checks.append(f"({token_checks})")
        conditions.append(f"({' OR '.join(tipo_checks)})")

    if estado_count:
        estado_checks = ' OR '.join([f"{estado_expr} {op_str} %s" for _ in range(max(estado_count, 1))])
        conditions.append(f"({estado_checks})")

    clause = ' AND '.join(conditions)
    return f"NOT ({clause})" if oper in {'1', '7'} else clause


def _append_tipo_oferta_filter(clauses, params, oper, values):
    tipo_param_groups, estado_params = _split_tipo_oferta_filter_params(oper, values)
    if not tipo_param_groups and not estado_params:
        return

    clauses.append(_tipo_oferta_filter_clause(
        oper,
        [len(group) for group in tipo_param_groups],
        len(estado_params),
    ))
    for group in tipo_param_groups:
        params.extend(group)
    params.extend(estado_params)


# Construye el WHERE de filtros, agrupando positivos repetidos como OR.
def _build_where(request):
    clauses, params = [], []
    grouped_positive_filters = {}

    campos = request.GET.getlist('campo_filtro')
    opers = request.GET.getlist('operador_filtro')
    valores = request.GET.getlist('valor_filtro')

    for i in range(len(campos)):
        campo = campos[i].strip()
        oper = opers[i] if i < len(opers) else '0'
        valor = valores[i].strip() if i < len(valores) else ''

        if campo and valor and campo in CAMPO_SQL:
            if oper in {'0', '2'}:
                group_key = (campo, oper)
                grouped_positive_filters.setdefault(group_key, [])
                if valor not in grouped_positive_filters[group_key]:
                    grouped_positive_filters[group_key].append(valor)
            elif campo == 'tipo_oferta':
                _append_tipo_oferta_filter(clauses, params, oper, [valor])
            else:
                col = CAMPO_SQL[campo]
                op_str, val_fn = OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])
                if oper == '0':
                    token_clause, token_params = _build_text_search_clause(col, valor)
                    if token_clause:
                        clauses.append(token_clause)
                        params.extend(token_params)
                    continue

                clauses.append(f"{col}::text {op_str} %s")
                params.append(val_fn(valor))

    for (campo, oper), grouped_values in grouped_positive_filters.items():
        if campo == 'tipo_oferta':
            _append_tipo_oferta_filter(clauses, params, oper, grouped_values)
            continue

        col = CAMPO_SQL[campo]
        op_str, val_fn = OPERADOR_SQL.get(oper, OPERADOR_SQL['0'])

        if len(grouped_values) == 1:
            if oper == '0':
                token_clause, token_params = _build_text_search_clause(col, grouped_values[0])
                if token_clause:
                    clauses.append(token_clause)
                    params.extend(token_params)
                continue

            clauses.append(f"{col}::text {op_str} %s")
            params.append(val_fn(grouped_values[0]))
            continue

        or_clauses = []
        for valor in grouped_values:
            if oper == '0':
                token_clause, token_params = _build_text_search_clause(col, valor)
                if token_clause:
                    or_clauses.append(token_clause)
                    params.extend(token_params)
                continue

            or_clauses.append(f"{col}::text {op_str} %s")
            params.append(val_fn(valor))

        clauses.append(f"({' OR '.join(or_clauses)})")

    q = request.GET.get('q', '').strip()
    if q:
        token_groups = []
        for token in _filter_tokens(q):
            global_search_clauses = [f"{_folded_sql(sql_expr)} LIKE %s" for sql_expr in CAMPO_SQL.values()]
            token_groups.append(f"({' OR '.join(global_search_clauses)})")
            params += [f'%{token}%'] * len(global_search_clauses)
        if token_groups:
            clauses.append(f"({' AND '.join(token_groups)})")

    where = ('WHERE ' + ' AND '.join(clauses)) if clauses else ''
    return where, params


# Arma el resumen de filtros aplicado al reporte Excel.
def _armar_texto_filtros(request):
    if request.GET.get('formato') == 'excel_todo':
        return 'Sin filtros aplicados'

    partes = []

    q = request.GET.get('q', '').strip()
    if q:
        partes.append(f'Búsqueda: {q}')

    campos = request.GET.getlist('campo_filtro')
    opers = request.GET.getlist('operador_filtro')
    valores = request.GET.getlist('valor_filtro')

    operadores_txt = {
        '0': 'contiene',
        '1': 'no contiene',
        '2': 'igual a',
        '3': 'mayor a',
        '4': 'mayor o igual a',
        '5': 'menor a',
        '6': 'menor o igual a',
        '7': 'distinto de',
    }

    _nombres_campos_legacy = {
        'cue': 'CUE',
        'anexo': 'Anexo',
        'codigo_jurisdiccional': 'Cód. Jurisdiccional',
        'localizacion': 'Localización',
        'tipo_oferta': 'Tipo Oferta',
        'estado': 'Estado',
        'subvencion': 'Subvención',
        'jornada': 'Jornada',
        'matricula_total': 'Matrícula',
        'mod_compl_planes': 'Mod. Compl. Planes',
        'cabecera_anexo': 'Cabecera-Anexo',
        'ambito': 'Ámbito',
        'tipo_ed': 'Tipo Ed.',
        'nivel': 'Nivel',
        'fecha_creacion': 'Fecha Creación',
        'sector': 'Sector',
        'acronimo': 'Acrónimo',
        'categoria': 'Categoría',
        'fecha_inst_legal': 'Fecha Inst. Legal',
        'nro_inst_legal': 'Nro Inst. Legal',
        'anio_creacion': 'Año Creación',
        'descrip_inst_legal': 'Descrip. Legal',
        'cui': 'CUI',
        'tel_supervisor': 'Tel. Supervisor',
        'email_supervisor': 'Email Supervisor',
        'supervisor_tecnico': 'Sup. Técnico',
        'cuof': 'CUOF',
        'cua': 'CUA',
        'regional': 'Regional',
        'udt': 'UDT',
        'cuof_ryc': 'CUOF RyC',
    }

    nombres_campos = {
        'cue': 'Cue',
        'anexo': 'Anexo',
        'codigo_jurisdiccional': 'Codigo Jurisdiccional',
        'localizacion': 'Localizacion',
        'tipo_oferta': 'Tipo Oferta',
        'nombre_titulo': 'Nombre Titulo',
        'estado': 'Estado',
        'subvencion': 'Subvencion',
        'jornada': 'Jornada',
        'matricula_total': 'Matricula Total',
        'mod_compl_planes': 'Mod.Compl.Planes',
        'cabecera_anexo': 'Oferta CABECERA-ANEXO',
        'ambito': 'Oferta AMBITO',
        'tipo_ed': 'Oferta TIPO ED.',
        'nivel': 'Oferta NIVEL',
        'fecha_creacion': 'Fecha de creación de la Oferta',
        'sector': 'Oferta SECTOR',
        'acronimo': 'Acrónimo',
        'categoria': 'Oferta CATEGORIA',
        'fecha_inst_legal': 'FECHA Inst. Legal Creacion de la Oferta',
        'nro_inst_legal': 'NRO. Inst. Legal Creacion de la Oferta',
        'anio_creacion': 'AÑO Creacion de la Oferta',
        'descrip_inst_legal': 'DESCRIP. Inst. Legal Creacion de la Oferta',
        'cui': 'CUI de la Oferta',
        'tel_supervisor': 'Teléfono Supervisor Oferta',
        'email_supervisor': 'Email Supervisor Oferta',
        'supervisor_tecnico': 'Supervisor Técnico Oferta',
        'cuof': 'CUOF Código Único de la Oficina de la Oferta',
        'cua': 'CUA de la Oferta',
        'regional': 'Regional Educativa de la Oferta',
        'udt': 'UDT Universidad de Desarrollo Territorial UDT de la Oferta',
        'cuof_ryc': 'CUOF_RyC',
    }

    nombres_campos['fecha_creacion'] = 'Fecha de creaci\u00f3n de la Oferta'
    nombres_campos['acronimo'] = 'Acr\u00f3nimo'
    nombres_campos['anio_creacion'] = 'A\u00d1O Creacion de la Oferta'
    nombres_campos['tel_supervisor'] = 'Tel\u00e9fono Supervisor Oferta'
    nombres_campos['supervisor_tecnico'] = 'Supervisor T\u00e9cnico Oferta'
    nombres_campos['cuof'] = 'CUOF C\u00f3digo \u00danico de la Oficina de la Oferta'

    for i in range(len(campos)):
        campo = campos[i].strip() if i < len(campos) else ''
        oper = opers[i].strip() if i < len(opers) else '0'
        valor = valores[i].strip() if i < len(valores) else ''

        if campo and valor:
            campo_txt = nombres_campos.get(campo, campo)
            oper_txt = operadores_txt.get(oper, 'contiene')
            partes.append(f'{campo_txt} {oper_txt}: {valor}')

    if not partes:
        return 'Sin filtros aplicados'

    return ' | '.join(partes)

     
def _formatear_fecha_estilo_seguimiento():
    ahora = datetime.now()
    dia = ahora.day
    mes = ahora.month
    anio = ahora.year
    hora_12 = ahora.strftime('%I').lstrip('0') or '0'
    minuto = ahora.strftime('%M')
    ampm = 'p. m.' if ahora.hour >= 12 else 'a. m.'
    return f'{dia}/{mes}/{anio} a las {hora_12}:{minuto} {ampm}'


def _resolver_columnas_exportar(request, formato):
    if formato == 'excel_todo':
        return COLUMNAS_EXPORTACION

    visibles = request.GET.getlist('visible_col')
    if not visibles:
        return COLUMNAS_EXPORTACION

    visibles_set = {v.strip() for v in visibles if v.strip()}
    columnas_filtradas = [col for col in COLUMNAS_EXPORTACION if col[1] in visibles_set]

    return columnas_filtradas or COLUMNAS_EXPORTACION


# Genera el Excel de ofertas locales.
def _exportar_excel(datos_exportar, formato, request):
    columnas_mapeo = _resolver_columnas_exportar(request, formato)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ofertas Locales'

    total_columnas = len(columnas_mapeo)
    ultima_columna = get_column_letter(total_columnas)

    titulo = 'Informe Ofertas Locales'
    fecha_str = _formatear_fecha_estilo_seguimiento()
    filtros_str = _armar_texto_filtros(request)

    ws.merge_cells(f'A1:{ultima_columna}1')
    ws['A1'] = titulo
    ws['A1'].font = Font(bold=True, size=10)
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'A2:{ultima_columna}2')
    ws['A2'] = f'Informe generado el: {fecha_str}'
    ws['A2'].font = Font(size=9)
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells(f'A3:{ultima_columna}3')
    ws['A3'] = f'Filtros aplicados: {filtros_str}'
    ws['A3'].font = Font(size=9)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 20

    fila_headers = 4
    for col_idx, (titulo_col, _) in enumerate(columnas_mapeo, start=1):
        celda = ws.cell(row=fila_headers, column=col_idx, value=titulo_col)
        celda.font = Font(bold=True, size=9)
        celda.alignment = Alignment(horizontal='left', vertical='center')

    for item in datos_exportar:
        fila = []
        for _, clave in columnas_mapeo:
            valor = item.get(clave, '')
            if valor is None:
                valor = ''
            fila.append(str(valor).replace('\r', ' ').replace('\n', ' '))
        ws.append(fila)

    ws.freeze_panes = 'A5'
    ws.auto_filter.ref = f'A4:{ultima_columna}{ws.max_row}'

    if formato == 'excel_todo':
        for col_num in range(1, total_columnas + 1):
            titulo_col = columnas_mapeo[col_num - 1][0]
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = min(max(len(str(titulo_col)) + 4, 14), 28)
    else:
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                cell.font = Font(size=9)

        for col_num in range(1, total_columnas + 1):
            col_letter = get_column_letter(col_num)
            max_length = 0
            for row_num in range(1, ws.max_row + 1):
                cell_value = ws[f'{col_letter}{row_num}'].value
                cell_len = len(str(cell_value)) if cell_value is not None else 0
                if cell_len > max_length:
                    max_length = cell_len
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    fecha_archivo = datetime.now().strftime('%Y%m%d_%H%M')
    sufijo = 'Filtros' if formato == 'excel_pagina' else 'Todo'
    nombre_archivo = f'OfertasLocales_{sufijo}_{fecha_archivo}.xlsx'

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
    return response

# Convierte resultados SQL en diccionarios.
def _dictfetchall(cursor):
    cols = [col[0] for col in cursor.description]
    return [dict(zip(cols, row)) for row in cursor.fetchall()]


# Limpia None antes de responder JSON.
def _sanitize_json_payload(value):
    if isinstance(value, dict):
        return {k: _sanitize_json_payload(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_sanitize_json_payload(item) for item in value]
    return '' if value is None else value


def _fetch_one(sql, params):
    with connections[MATERIALIZADAS_DB].cursor() as cursor:
        cursor.execute(sql, params)
        rows = _dictfetchall(cursor)
    return rows[0] if rows else None


def _fetch_all(sql, params):
    with connections[MATERIALIZADAS_DB].cursor() as cursor:
        cursor.execute(sql, params)
        return _dictfetchall(cursor)


# Normalizacion local usada por la version final del detalle de oferta.
def _normalize_text(value):
    if value is None:
        return ''
    return ' '.join(str(value).split())


def _collapse_repeated_hyphen_value(value):
    normalized = _normalize_text(value)

    if '-' not in normalized:
        return normalized

    for index, char in enumerate(normalized):
        if char != '-':
            continue

        left = normalized[:index].strip()
        right = normalized[index + 1:].strip()

        if left and right and left == right:
            return left

    return normalized


def _clean_code_desc_value(value):
    normalized = _collapse_repeated_hyphen_value(value)

    if '-' not in normalized:
        return normalized

    prefix, suffix = normalized.split('-', 1)
    prefix = prefix.strip()
    suffix = suffix.strip()

    if (
        prefix
        and suffix
        and ' ' not in prefix
        and all(char.isalnum() or char in './' for char in prefix)
    ):
        return suffix

    return normalized


def _format_responsable(apellido, nombre, documento):
    apellido_txt = _normalize_text(apellido)
    nombre_txt = _normalize_text(nombre)
    documento_txt = _normalize_text(documento)

    if apellido_txt and nombre_txt:
        base = f'{apellido_txt}, {nombre_txt}'
    else:
        base = apellido_txt or nombre_txt

    if base:
        return f'{base}({documento_txt})'

    return base or documento_txt


# Limpia valores repetidos/codificados antes de devolver el detalle.
def _postprocess_oferta_payload(oferta):
    if not oferta:
        return oferta

    cleaned = dict(oferta)

    for key in (
        'cabecera_anexo',
        'ambito',
        'tipo_ed',
        'nivel',
        'sector',
        'categoria',
        'acronimo',
        'regional',
        'udt',
        'supervisor_tecnico',
        'descrip_inst_legal',
    ):
        cleaned[key] = _clean_code_desc_value(cleaned.get(key, ''))

    cleaned['responsable'] = _format_responsable(
        cleaned.pop('apellido_responsable', ''),
        cleaned.pop('nombre_responsable', ''),
        cleaned.pop('documento_responsable', ''),
    )

    cleaned['mod_compl_planes'] = _normalize_text(cleaned.get('mod_compl_planes', ''))

    return cleaned

@padron_interno_admin_o_gestor_required
def detalle_oferta_local_json(request, id_oferta):
    # Version final efectiva del endpoint de oferta local.
    oferta_sql = """
        SELECT
            v.id_oferta_local AS id,
            COALESCE(v.cue::text, '') AS cue,
            COALESCE(v.anexo::text, '') AS anexo,
            COALESCE(v.oferta, '') AS carrera,
            COALESCE(v.oferta, '') AS oferta,
            COALESCE(v.codigo_jurisdiccional_oferta_local, '') AS codigo_jurisdiccional,
            COALESCE(v.estado_ofertalocal, '') AS estado,
            COALESCE(v.subvencion_oferta_local, '') AS subvencion,
            COALESCE(v.fecha_creacion_ofertalocal::text, '') AS fecha_creacion,
            COALESCE(v.fecha_alta_ofertalocal::text, '') AS fecha_alta,
            COALESCE(v.fecha_baja_ofertalocal::text, '') AS fecha_baja,
            COALESCE(v.fecha_actualizacion_ofertalocal::text, '') AS fecha_actualizacion,
            COALESCE(v.jornada_ofertalocal, '') AS jornada,
            COALESCE(v.matricula_total_ofertalocal::text, '') AS matricula_total,
            COALESCE(v.modalidad_basica, '') AS mod_compl_planes,
            COALESCE(v.cp_efvar4, '') AS cuof,
            COALESCE(BTRIM(v.cp_cuof_ryc), '') AS cuof_ryc,
            COALESCE(v.cp_efvar6, '') AS udt,
            COALESCE(v.cp_efvar5, '') AS regional,
            COALESCE(v.cp_acronimo, '') AS acronimo,
            COALESCE(v.cp_supervisortecnico_oferta, '') AS supervisor_tecnico,
            COALESCE(v.cp_tesupervisor_oferta, '') AS tel_supervisor,
            COALESCE(v.cp_mailsupervisor_oferta, '') AS email_supervisor,
            COALESCE(v.cp_efvar2, '') AS cui,
            COALESCE(v.cp_of_cua, '') AS cua,
            COALESCE(v.cp_of_nro_inst_legal, '') AS nro_inst_legal,
            COALESCE(v.cp_of_anio_inst_legal, '') AS anio_creacion,
            COALESCE(v.cp_of_fecha_inst_legal, '') AS fecha_inst_legal,
            COALESCE(v.cp_of_descrip_inst_legal, '') AS descrip_inst_legal,
            COALESCE(v.cp_of_cab_anexo, '') AS cabecera_anexo,
            COALESCE(v.cp_of_ambito, '') AS ambito,
            COALESCE(v.cp_of_tipo_ed, '') AS tipo_ed,
            COALESCE(v.cp_of_nivel, '') AS nivel,
            COALESCE(v.cp_of_sector, '') AS sector,
            COALESCE(v.cp_of_categoria, '') AS categoria,
            COALESCE(r.apellido, '') AS apellido_responsable,
            COALESCE(r.nombre, '') AS nombre_responsable,
            COALESCE(r.nro_documento::text, '') AS documento_responsable,
            COALESCE(v.c_oferta_base::text, '') AS c_oferta_base
        FROM padroninterno.mv_ofertaslocales v
        LEFT JOIN padroninterno.mv_responsables r ON r.id_responsable = v.id_responsable
        WHERE v.id_oferta_local = %s
        LIMIT 1
    """

    titulos_sql = """
        SELECT
            vpe.id_plan_estudio_local AS id_titulo,
            COALESCE(vpe.c_titulo::text, '') AS c_titulo,
            COALESCE(vpe.titulo, '') AS titulo,
            COALESCE(
                NULLIF(BTRIM(vpe.titulo_completo_plan), ''),
                NULLIF(BTRIM(vpe.titulo_completo), ''),
                BTRIM(COALESCE(vpe.titulo, '')),
                ''
            ) AS titulo_completo,
            COALESCE(vpe.duracion, '') AS duracion,
            COALESCE(vpe.norma, '') AS norma,
            COALESCE(vpe.dictado, '') AS dictado,
            COALESCE(vpe.requisito, '') AS requisito,
            COALESCE(vpe.condicion_ingreso, '') AS condicion_ingreso
        FROM padroninterno.mv_oferta_titulos vpe
        WHERE vpe.id_oferta_local = %s
        ORDER BY COALESCE(vpe.titulo, ''), vpe.id_plan_estudio_local
    """

    historial_sql = f"""
        SELECT
            *
        FROM padroninterno.mv_oferta_historial
        WHERE id_oferta_local = %s
        ORDER BY fecha_vigencia DESC, id_movimiento DESC
    """

    try:
        oferta = _fetch_one(oferta_sql, [id_oferta])

        if not oferta:
            return JsonResponse({'error': 'Oferta local no encontrada.'}, status=404)

        titulos = _fetch_all(titulos_sql, [id_oferta])
        historial = _fetch_all(historial_sql, [id_oferta])

        return JsonResponse(_sanitize_json_payload({
            'oferta': _postprocess_oferta_payload(oferta),
            'titulos': titulos,
            'historial': historial,
        }))
    except Exception as exc:
        return JsonResponse({'error': f'Error al obtener detalle de oferta local: {exc}'}, status=500)

@padron_interno_admin_o_gestor_required
def detalle_titulo_json(request, id_titulo):
    # Version final efectiva del endpoint de titulo.
    titulo_sql = """
        SELECT
            COALESCE(BTRIM(t.titulo), '') AS titulo,
            COALESCE(
                NULLIF(BTRIM(t.titulo_completo), ''),
                CASE
                    WHEN COALESCE(t.c_titulo::text, '') <> '' AND COALESCE(BTRIM(t.titulo), '') <> '' THEN
                        t.c_titulo::text || ' - ' || BTRIM(t.titulo)
                    ELSE COALESCE(BTRIM(t.titulo), '')
                END
            ) AS titulo_completo,
            COALESCE(t.duracion, '') AS duracion,
            COALESCE(t.norma, '') AS norma,
            COALESCE(t.dictado, '') AS dictado,
            COALESCE(t.requisito, '') AS requisito,
            COALESCE(t.condicion_ingreso, '') AS condicion_ingreso,
            COALESCE(t.carrera, '') AS carrera,
            COALESCE(t.nivel, '') AS nivel,
            COALESCE(t.disciplina, '') AS disciplina,
            COALESCE(t.rama, '') AS rama,
            COALESCE(t.orientacion, '') AS orientacion,
            COALESCE(t.c_titulo::text, '') AS c_titulo
        FROM padroninterno.mv_oferta_titulos t
        WHERE t.id_plan_estudio_local = %s
        LIMIT 1
    """

    try:
        titulo = _fetch_one(titulo_sql, [id_titulo])

        if not titulo:
            return JsonResponse({'error': 'Titulo no encontrado.'}, status=404)

        return JsonResponse(_sanitize_json_payload({'titulo': titulo}))
    except Exception as exc:
        return JsonResponse({'error': f'Error al obtener detalle del titulo: {exc}'}, status=500)

@padron_interno_admin_o_gestor_required
def listar_ofertas_locales(request):
    # Renderiza la pantalla; los datos se cargan por AJAX salvo exportacion Excel.
    formato = request.GET.get('formato')

    if formato == 'excel_todo':
        where, params = '', []
    elif formato == 'excel_pagina':
        where, params = _build_where(request)
    else:
        context = {
            'lista_items': [],
            'page_obj': None,
            'resultado_total': None,
            'resultado_desde': 0,
            'resultado_hasta': 0,
            'username': getattr(request.user, 'username', ''),
            'request': request,
            'ofertaslocales_async_loading': True,
        }
        context.update(get_contexto_fecha_padron(request))
        return render(request, 'padroninterno/ofertaslocales.html', context)

    data_sql = _build_data_sql(request, where)

    if formato in ['excel_pagina', 'excel_todo']:
        datos_exportar = _fetch_all(data_sql, params)
        return _exportar_excel(datos_exportar, formato, request)


@padron_interno_admin_o_gestor_required
def ofertas_locales_datos_json(request):
    # Endpoint paginado: obtiene una fila extra para calcular has_next.
    where, params = _build_where(request)
    page_size = _get_page_size(request)
    page = _get_page_number(request)
    offset = (page - 1) * page_size
    data_sql = _build_data_sql(request, where)
    rows = _fetch_all(f"{data_sql} LIMIT %s OFFSET %s", params + [page_size + 1, offset])
    page_rows = rows[:page_size]
    items = [_serialize_list_item(row) for row in page_rows]

    return JsonResponse(_sanitize_json_payload({
        'items': items,
        'has_next': len(rows) > page_size,
        'has_previous': page > 1,
        'desde': offset + 1 if items else 0,
        'hasta': offset + len(items) if items else 0,
        'page': page,
        'page_size': page_size,
        'total': None,
        'total_pending': True,
    }))


@padron_interno_admin_o_gestor_required
def ofertas_locales_total_json(request):
    # Conteo separado del listado para mantener agil la carga inicial.
    where, params = _build_where(request)
    count_sql = _build_count_sql(request, where)

    with connections[MATERIALIZADAS_DB].cursor() as cursor:
        cursor.execute(count_sql, params)
        total = cursor.fetchone()[0]

    return JsonResponse({'total': total})
