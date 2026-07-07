import json
import logging
import time
import unicodedata
from datetime import datetime
from io import BytesIO

from django.core.cache import cache
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import connections
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import ensure_csrf_cookie
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .models import (
    get_cefs_visualizacion_usuario,
    normalizar_cueanexo,
    usuario_es_admin_cef,
)
from .permisos import cef_required


logger = logging.getLogger(__name__)


PAGE_SIZE = 10
PAGE_SIZE_OPTIONS = [10, 25, 50, 100]
CACHE_TTL_LOCALIZACIONES_CEF = 60 * 5
CACHE_VERSION_LOCALIZACIONES_CEF = "v3_columnas_vista_20260520"

# Columnas que se leen desde Padron y se exponen en tabla, filtros y Excel.
COLUMNAS_LOCALIZACIONES_CEF = [
    "cueanexo",
    "nom_est",
    "oferta",
    "ambito",
    "sector",
    "region_loc",
    "ref_loc",
    "calle",
    "numero",
    "localidad",
    "departamento",
    "estado_loc",
    "est_oferta",
    "estado_est",
    "resploc_cuitcuil",
    "resploc_doc",
    "apellido_resp",
    "nombre_resp",
    "resploc_email",
    "resploc_telefono",
    "sup_tecnico",
    "email_suptecnico",
    "tel_suptecnico",
    "categoria",
    "jornada",
]

FIELD_MAP = {
    "cueanexo": ("cueanexo",),
    "nom_est": ("nom_est",),
    "oferta": ("oferta",),
    "ambito": ("ambito",),
    "sector": ("sector",),
    "region_loc": ("region_loc",),
    "ref_loc": ("ref_loc",),
    "calle": ("calle",),
    "numero": ("numero",),
    "localidad": ("localidad",),
    "departamento": ("departamento",),
    "estado_loc": ("estado_loc",),
    "est_oferta": ("est_oferta",),
    "estado_est": ("estado_est",),
    "resploc_cuitcuil": ("resploc_cuitcuil",),
    "resploc_doc": ("resploc_doc",),
    "apellido_resp": ("apellido_resp",),
    "nombre_resp": ("nombre_resp",),
    "resploc_email": ("resploc_email",),
    "resploc_telefono": ("resploc_telefono",),
    "sup_tecnico": ("sup_tecnico",),
    "email_suptecnico": ("email_suptecnico",),
    "tel_suptecnico": ("tel_suptecnico",),
    "categoria": ("categoria",),
    "jornada": ("jornada",),
}

# Etiquetas visibles para encabezados, filtros activos y exportaciones.
LABELS_COLUMNAS = {
    "cueanexo": "CUE-Anexo",
    "nom_est": "Establecimiento",
    "oferta": "Oferta",
    "ambito": "Ámbito",
    "sector": "Sector",
    "region_loc": "Región",
    "ref_loc": "Referencia localización",
    "calle": "Calle",
    "numero": "Número",
    "localidad": "Localidad",
    "departamento": "Departamento",
    "estado_loc": "Estado localización",
    "est_oferta": "Estado oferta",
    "estado_est": "Estado establecimiento",
    "resploc_cuitcuil": "CUIL/CUIT responsable",
    "resploc_doc": "Documento responsable",
    "apellido_resp": "Apellido responsable",
    "nombre_resp": "Nombre responsable",
    "resploc_email": "Email responsable",
    "resploc_telefono": "Teléfono responsable",
    "sup_tecnico": "Supervisor técnico",
    "email_suptecnico": "Email supervisor técnico",
    "tel_suptecnico": "Teléfono supervisor técnico",
    "categoria": "Categoría",
    "jornada": "Jornada",
}

# Subconjunto inicial que se muestra al entrar a la pantalla.
COLUMNAS_VISIBLES_DEFAULT = [
    "cueanexo",
    "nom_est",
    "oferta",
    "ambito",
    "sector",
    "region_loc",
    "localidad",
    "departamento",
    "estado_loc",
    "est_oferta",
    "estado_est",
    "apellido_resp",
    "nombre_resp",
    "sup_tecnico",
    "categoria",
    "jornada",
]

OPERADORES_TXT = {
    "0": "parecido a",
    "1": "no parecido a",
    "2": "igual a",
    "3": "mayor a",
    "4": "mayor o igual a",
    "5": "menor a",
    "6": "menor o igual a",
    "7": "distinto de",
}

# Limita la consulta ORM a los campos usados por la pantalla para reducir carga.
ONLY_FIELDS_LOCALIZACIONES_CEF = [
    "cueanexo",
    "nom_est",
    "oferta",
    "ambito",
    "sector",
    "region_loc",
    "ref_loc",
    "calle",
    "numero",
    "localidad",
    "departamento",
    "estado_loc",
    "est_oferta",
    "estado_est",
    "resploc_cuitcuil",
    "resploc_doc",
    "apellido_resp",
    "nombre_resp",
    "resploc_email",
    "resploc_telefono",
    "sup_tecnico",
    "email_suptecnico",
    "tel_suptecnico",
    "categoria",
    "jornada",
]

GLOBAL_SEARCH_LIST_FIELDS = list(COLUMNAS_LOCALIZACIONES_CEF)

FILTER_OPTIONS_LIST_FIELDS = set(COLUMNAS_LOCALIZACIONES_CEF)


def _log_perf(label, started):
    # CEF_PATCH_LOCALIZACIONES_REFRESH_COMO_RESPONSABLES_20260515
    logger.debug(
        "CEF_PATCH_LOCALIZACIONES_REFRESH_COMO_RESPONSABLES_20260515 %s %.1fms",
        label,
        (time.perf_counter() - started) * 1000,
    )


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _serialize_item(item):
    return {
        "cueanexo": _clean(getattr(item, "cueanexo", "")),
        "nom_est": _clean(getattr(item, "nom_est", "")),
        "oferta": _clean(getattr(item, "oferta", "")),
        "ambito": _clean(getattr(item, "ambito", "")),
        "sector": _clean(getattr(item, "sector", "")),
        "region_loc": _clean(getattr(item, "region_loc", "")),
        "ref_loc": _clean(getattr(item, "ref_loc", "")),
        "calle": _clean(getattr(item, "calle", "")),
        "numero": _clean(getattr(item, "numero", "")),
        "localidad": _clean(getattr(item, "localidad", "")),
        "departamento": _clean(getattr(item, "departamento", "")),
        "estado_loc": _clean(getattr(item, "estado_loc", "")),
        "est_oferta": _clean(getattr(item, "est_oferta", "")),
        "estado_est": _clean(getattr(item, "estado_est", "")),
        "resploc_cuitcuil": _clean(getattr(item, "resploc_cuitcuil", "")),
        "resploc_doc": _clean(getattr(item, "resploc_doc", "")),
        "apellido_resp": _clean(getattr(item, "apellido_resp", "")),
        "nombre_resp": _clean(getattr(item, "nombre_resp", "")),
        "resploc_email": _clean(getattr(item, "resploc_email", "")),
        "resploc_telefono": _clean(getattr(item, "resploc_telefono", "")),
        "sup_tecnico": _clean(getattr(item, "sup_tecnico", "")),
        "email_suptecnico": _clean(getattr(item, "email_suptecnico", "")),
        "tel_suptecnico": _clean(getattr(item, "tel_suptecnico", "")),
        "categoria": _clean(getattr(item, "categoria", "")),
        "jornada": _clean(getattr(item, "jornada", "")),
    }


def _cache_key_localizaciones_cef(request):
    user_id = getattr(request.user, "pk", None) or "anon"
    return f"cef:localizaciones:{CACHE_VERSION_LOCALIZACIONES_CEF}:user:{user_id}"


def _get_items_base_cached(request):
    """
    Obtiene y serializa los CEF visibles una sola vez por usuario.

    La pantalla filtra y ordena en memoria sobre esta lista cacheada para evitar
    repetir consultas pesadas a Padron en cada cambio de filtros o paginacion.
    """

    global _GEO_FILTER_OPTIONS_CEF_CACHE
    started = time.perf_counter()
    cache_key = _cache_key_localizaciones_cef(request)
    if request.GET.get("refresh") == "1":
        cache.delete(cache_key)
        _GEO_FILTER_OPTIONS_CEF_CACHE = None

    sentinel = object()
    cached_items = cache.get(cache_key, sentinel)
    if cached_items is not sentinel:
        _log_perf("_get_items_base_cached hit", started)
        return cached_items

    qs = _base_cef_queryset(request)
    try:
        qs = qs.only(*ONLY_FIELDS_LOCALIZACIONES_CEF)
        items = [_serialize_item(item) for item in qs]
    except Exception:
        items = [_serialize_item(item) for item in _base_cef_queryset(request)]

    cache.set(cache_key, items, CACHE_TTL_LOCALIZACIONES_CEF)
    _log_perf("_get_items_base_cached miss", started)
    return items


def _normalize_text(value):
    """
    Normaliza texto para comparar sin distinguir mayusculas ni acentos.
    """

    text = _clean(value).casefold()
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char))


def _contains(value, needle):
    return _normalize_text(needle) in _normalize_text(value)


def _iexact(value, needle):
    return _normalize_text(value) == _normalize_text(needle)


def _compare_text(value, operator, needle):
    left = _normalize_text(value)
    right = _normalize_text(needle)
    if operator == "3":
        return left > right
    if operator == "4":
        return left >= right
    if operator == "5":
        return left < right
    if operator == "6":
        return left <= right
    return False


def _item_matches_operator(item, field_key, operator, value):
    item_value = item.get(field_key, "")
    if operator == "1":
        return not _contains(item_value, value)
    if operator == "2":
        return _iexact(item_value, value)
    if operator in {"3", "4", "5", "6"}:
        return _compare_text(item_value, operator, value)
    if operator == "7":
        return not _iexact(item_value, value)
    return _contains(item_value, value)


def _apply_filters_list(items, request):
    """
    Aplica todos los filtros de la pantalla sobre la lista ya cacheada.

    Combina busqueda global, busqueda rapida, filtros por columna y filtros
    avanzados agrupados por campo/operador.
    """

    started = time.perf_counter()
    q = request.GET.get("q", "").strip()
    if q:
        items = [
            item
            for item in items
            if any(_contains(item.get(field, ""), q) for field in GLOBAL_SEARCH_LIST_FIELDS)
        ]

    smart_col = request.GET.get("smart_ui_col", "").strip()
    smart_val = request.GET.get("smart_ui_val", "").strip()
    if smart_col in COLUMNAS_LOCALIZACIONES_CEF and smart_val:
        items = [item for item in items if _contains(item.get(smart_col, ""), smart_val)]

    for campo in COLUMNAS_LOCALIZACIONES_CEF:
        value = request.GET.get(campo, "").strip()
        if not value:
            continue
        items = [item for item in items if _contains(item.get(campo, ""), value)]

    campos = request.GET.getlist("campo_filtro")
    operadores = request.GET.getlist("operador_filtro")
    valores = request.GET.getlist("valor_filtro")
    grouped_filters = {}

    for index, campo in enumerate(campos):
        campo = campo.strip()
        valor = valores[index].strip() if index < len(valores) else ""
        operador = operadores[index].strip() if index < len(operadores) else "0"
        if not campo or not valor or campo not in COLUMNAS_LOCALIZACIONES_CEF:
            continue
        grouped_filters.setdefault((campo, operador), [])
        if valor not in grouped_filters[(campo, operador)]:
            grouped_filters[(campo, operador)].append(valor)

    for (campo, operador), valores_grupo in grouped_filters.items():
        items = [
            item
            for item in items
            if any(_item_matches_operator(item, campo, operador, valor) for valor in valores_grupo)
        ]

    _log_perf("_apply_filters_list", started)
    return items


def _sort_key_default(item):
    return (
        _normalize_text(item.get("region_loc", "")),
        _normalize_text(item.get("departamento", "")),
        _normalize_text(item.get("localidad", "")),
        _normalize_text(item.get("cueanexo", "")),
    )


def _apply_order_list(items, request):
    """
    Ordena con un criterio geografico estable y permite cambiar columna desde la UI.
    """

    started = time.perf_counter()
    orden_param = request.GET.get("orden", "region_loc").strip()
    raw_field = orden_param.lstrip("-")
    reverse = orden_param.startswith("-")

    ordered_items = sorted(items, key=_sort_key_default)
    if raw_field in COLUMNAS_LOCALIZACIONES_CEF:
        ordered_items = sorted(
            ordered_items,
            key=lambda item: _normalize_text(item.get(raw_field, "")),
            reverse=reverse,
        )
        _log_perf("_apply_order_list", started)
        return ordered_items, orden_param

    _log_perf("_apply_order_list", started)
    return ordered_items, ""


def _unique_sorted_options(items, field):
    values = {_clean(item.get(field, "")) for item in items}
    values.discard("")
    return sorted(values, key=_normalize_text)


# CEF_PATCH_GEO_FILTROS_COMPLETOS_20260515
# Opciones geograficas obligatorias para Localizaciones CEF.
# Se mezclan siempre con BD real + items visibles para que los filtros no queden incompletos.
GEO_FILTER_OPTIONS_CEF_REQUIRED = {
    "region_loc": [
        "R.E. 1",
        "R.E. 10-A",
        "R.E. 10-B",
        "R.E. 10-C",
        "R.E. 2",
        "R.E. 3",
        "R.E. 4-A",
        "R.E. 4-B",
        "R.E. 5",
        "R.E. 6",
        "R.E. 7",
        "R.E. 8-A",
        "R.E. 8-B",
        "R.E. 9",
        "SUB. R.E. 1-A",
        "SUB. R.E. 1-B",
        "SUB. R.E. 2",
        "SUB. R.E. 3",
        "SUB. R.E. 5",
    ],
    "departamento": [
        "1§ DE MAYO",
        "12 DE OCTUBRE",
        "2 DE ABRIL",
        "25 DE MAYO",
        "9 DE JULIO",
        "ALMIRANTE BROWN",
        "BERMEJO",
        "CHACABUCO",
        "COMANDANTE FERNANDEZ",
        "FRAY JUSTO SANTA MARIA DE ORO",
        "GENERAL BELGRANO",
        "GENERAL DONOVAN",
        "GENERAL GUEMES",
        "INDEPENDENCIA",
        "LIBERTAD",
        "LIBERTADOR GENERAL SAN MARTIN",
        "MAIPU",
        "MAYOR LUIS J FONTANA",
        "O HIGGINS",
        "PRESIDENCIA DE LA PLAZA",
        "QUITILIPI",
        "SAN FERNANDO",
        "SAN LORENZO",
        "SARGENTO CABRAL",
        "TAPENAGA",
    ],
    "localidad": [
        "AVIA TERAI",
        "BAJO HONDO CHICO",
        "BARRANQUERAS",
        "BASAIL",
        "CAMPO LARGO",
        "CAPITAN SOLARI",
        "CHARADAI",
        "CHARATA",
        "CHOROTIS",
        "CIERVO PETISO",
        "COLONIA ABORIGEN",
        "COLONIA ABORIGEN NAPALPI",
        "COLONIA BARANDA",
        "COLONIA BENITEZ",
        "COLONIA CAMPO HERMOSO",
        "COLONIA EL CURUPI",
        "COLONIA ELISA",
        "COLONIA JUAN JOSE PASO",
        "COLONIA POPULAR",
        "COLONIA PRESIDENCIA ROQUE SAENZ PEÑA",
        "COLONIA SAN EDUARDO",
        "COLONIA URIBURU",
        "COLONIAS UNIDAS",
        "CONCEPCION DEL BERMEJO",
        "CORONEL DU GRATY",
        "CORZUELA",
        "COTE LAI",
        "EL CABURE",
        "EL ESPINILLO",
        "EL PALMAR",
        "EL SAUZALITO",
        "EL VIZCACHERAL",
        "ENRIQUE URIEN",
        "FONTANA",
        "FUERTE ESPERANZA",
        "GANCEDO",
        "GENERAL CAPDEVILA",
        "GENERAL JOSE DE SAN MARTIN",
        "GENERAL PINEDO",
        "GENERAL VEDIA",
        "HERMOSO CAMPO",
        "HORQUILLA",
        "ISLA DEL CERRITO",
        "JUAN JOSE CASTELLI",
        "LA CLOTILDE",
        "LA EDUVIGIS",
        "LA ESCONDIDA",
        "LA LEONESA",
        "LA SABANA",
        "LA TIGRA",
        "LA VERDE",
        "LAGUNA BLANCA",
        "LAGUNA LIMPIA",
        "LAPACHITO",
        "LAS BREÑAS",
        "LAS GARCITAS",
        "LAS PALMAS",
        "LAS VERTIENTES",
        "LOS FRENTONES",
        "MACHAGAI",
        "MAKALLE",
        "MARGARITA BELEN",
        "MIRAFLORES",
        "NAPENAY",
        "NO FIGURA EN TABLA",
        "NUEVA POMPEYA",
        "PAMPA ALMIRON",
        "PAMPA CEJAS",
        "PAMPA DEL INDIO",
        "PAMPA DEL INFIERNO",
        "PRESIDENCIA DE LA PLAZA",
        "PRESIDENCIA ROCA",
        "PRESIDENCIA ROQUE SAENZ PEÑA",
        "PUERTO BERMEJO NUEVO",
        "PUERTO BERMEJO VIEJO",
        "PUERTO EVA PERON",
        "PUERTO TIROL",
        "PUERTO VILELAS",
        "QUITILIPI",
        "RESISTENCIA",
        "RIO MUERTO",
        "SAMUHU",
        "SAN BERNARDO",
        "SANTA SYLVINA",
        "TACO POZO",
        "TRES ISLETAS",
        "TRES MOJONES",
        "VICENTINI",
        "VILLA ANGELA",
        "VILLA BERTHET",
        "VILLA RIO BERMEJITO",
    ],
}


_GEO_FILTER_OPTIONS_CEF_CACHE = None


def _geo_option_key(value):
    text = _normalize_text(value)
    for marker in ("§", "°", "º", "ª"):
        text = text.replace(marker, " ")
    text = "".join(char if char.isalnum() else " " for char in text)
    return " ".join(text.split())


def _geo_option_sort_key(value):
    return _geo_option_key(value)


def _merge_filter_option_values(*groups):
    values_by_key = {}

    for group in groups:
        for value in group or []:
            value_txt = _clean(value)
            if not value_txt:
                continue
            key = _geo_option_key(value_txt)
            if not key:
                continue
            values_by_key.setdefault(key, value_txt)

    return sorted(values_by_key.values(), key=_geo_option_sort_key)


def _fetch_geo_filter_values(sql, db_alias="default"):
    try:
        with connections[db_alias].cursor() as cursor:
            cursor.execute(sql)
            return [_clean(row[0]) for row in cursor.fetchall() if row and _clean(row[0])]
    except Exception as exc:
        logger.exception("CEF_PATCH_GEO_FILTROS_COMPLETOS_20260515 fallo consulta geo con alias '%s': %s", db_alias, exc)
        return []


def _get_global_geo_filter_options(force_refresh=False):
    """
    Construye opciones globales para filtros geograficos.

    Mezcla valores obligatorios con valores reales de las vistas/tablas de Padron
    para que region, departamento y localidad no queden incompletos.
    """

    global _GEO_FILTER_OPTIONS_CEF_CACHE

    if force_refresh:
        _GEO_FILTER_OPTIONS_CEF_CACHE = None

    if _GEO_FILTER_OPTIONS_CEF_CACHE is not None:
        return _GEO_FILTER_OPTIONS_CEF_CACHE

    _GEO_FILTER_OPTIONS_CEF_CACHE = {
        "region_loc": _merge_filter_option_values(
            GEO_FILTER_OPTIONS_CEF_REQUIRED.get("region_loc", []),
            _fetch_geo_filter_values(
                """
                SELECT DISTINCT BTRIM(region_loc::text) AS valor
                FROM v_capa_unica_ofertas_ant
                WHERE region_loc IS NOT NULL
                  AND LENGTH(BTRIM(region_loc::text)) > 0
                ORDER BY valor
                """
            ),
            _fetch_geo_filter_values(
                """
                SELECT DISTINCT BTRIM(region_loc::text) AS valor
                FROM vp_localizaciones
                WHERE region_loc IS NOT NULL
                  AND LENGTH(BTRIM(region_loc::text)) > 0
                ORDER BY valor
                """
            ),
        ),
        "departamento": _merge_filter_option_values(
            GEO_FILTER_OPTIONS_CEF_REQUIRED.get("departamento", []),
            _fetch_geo_filter_values(
                """
                SELECT DISTINCT BTRIM(departamento::text) AS valor
                FROM v_capa_unica_ofertas_ant
                WHERE departamento IS NOT NULL
                  AND LENGTH(BTRIM(departamento::text)) > 0
                ORDER BY valor
                """
            ),
            _fetch_geo_filter_values(
                """
                SELECT DISTINCT BTRIM(departamento_nombre::text) AS valor
                FROM vp_localizaciones
                WHERE departamento_nombre IS NOT NULL
                  AND LENGTH(BTRIM(departamento_nombre::text)) > 0
                ORDER BY valor
                """
            ),
            _fetch_geo_filter_values(
                """
                SELECT DISTINCT BTRIM(nombre::text) AS valor
                FROM departamento_tipo
                WHERE nombre IS NOT NULL
                  AND LENGTH(BTRIM(nombre::text)) > 0
                ORDER BY valor
                """
            ),
        ),
        "localidad": _merge_filter_option_values(
            GEO_FILTER_OPTIONS_CEF_REQUIRED.get("localidad", []),
            _fetch_geo_filter_values(
                """
                SELECT DISTINCT BTRIM(localidad::text) AS valor
                FROM v_capa_unica_ofertas_ant
                WHERE localidad IS NOT NULL
                  AND LENGTH(BTRIM(localidad::text)) > 0
                ORDER BY valor
                """
            ),
            _fetch_geo_filter_values(
                """
                SELECT DISTINCT BTRIM(localidad_nombre::text) AS valor
                FROM vp_localizaciones
                WHERE localidad_nombre IS NOT NULL
                  AND LENGTH(BTRIM(localidad_nombre::text)) > 0
                ORDER BY valor
                """
            ),
            _fetch_geo_filter_values(
                """
                SELECT DISTINCT BTRIM(nombre::text) AS valor
                FROM localidad_tipo
                WHERE nombre IS NOT NULL
                  AND LENGTH(BTRIM(nombre::text)) > 0
                ORDER BY valor
                """
            ),
        ),
    }

    logger.info(
        "CEF_PATCH_GEO_FILTROS_COMPLETOS_20260515 global geo options region=%s departamento=%s localidad=%s",
        len(_GEO_FILTER_OPTIONS_CEF_CACHE.get("region_loc", [])),
        len(_GEO_FILTER_OPTIONS_CEF_CACHE.get("departamento", [])),
        len(_GEO_FILTER_OPTIONS_CEF_CACHE.get("localidad", [])),
    )

    return _GEO_FILTER_OPTIONS_CEF_CACHE


def _get_filter_options_from_items(items):
    """
    Prepara las listas de opciones que usa el modal de filtros en el frontend.
    """

    started = time.perf_counter()

    options = {
        campo: _unique_sorted_options(items, campo) if campo in FILTER_OPTIONS_LIST_FIELDS else []
        for campo in COLUMNAS_LOCALIZACIONES_CEF
    }

    geo_options = _get_global_geo_filter_options()

    for campo in ("region_loc", "departamento", "localidad"):
        options[campo] = _merge_filter_option_values(
            GEO_FILTER_OPTIONS_CEF_REQUIRED.get(campo, []),
            geo_options.get(campo, []),
            options.get(campo, []),
        )

    logger.info(
        "CEF_PATCH_GEO_FILTROS_COMPLETOS_20260515 final filter options region=%s departamento=%s localidad=%s",
        len(options.get("region_loc", [])),
        len(options.get("departamento", [])),
        len(options.get("localidad", [])),
    )

    _log_perf("_get_filter_options_from_items", started)
    return options


def _get_page_size(request):
    try:
        page_size = int(request.GET.get("page_size", PAGE_SIZE))
    except (TypeError, ValueError):
        return PAGE_SIZE
    return page_size if page_size in PAGE_SIZE_OPTIONS else PAGE_SIZE


def _build_active_filters_text(request, cef_options=None):
    """
    Genera un resumen legible de los filtros aplicados para UI y Excel.
    """

    partes = []
    q = request.GET.get("q", "").strip()
    if q:
        partes.append(f"Búsqueda: {q}")

    smart_col = request.GET.get("smart_ui_col", "").strip()
    smart_val = request.GET.get("smart_ui_val", "").strip()
    if smart_col in FIELD_MAP and smart_val:
        partes.append(
            f"Búsqueda rápida en {LABELS_COLUMNAS.get(smart_col, smart_col)}: {smart_val}"
        )

    for campo in COLUMNAS_LOCALIZACIONES_CEF:
        valor = request.GET.get(campo, "").strip()
        if valor:
            partes.append(f"{LABELS_COLUMNAS.get(campo, campo)}: {valor}")

    campos = request.GET.getlist("campo_filtro")
    operadores = request.GET.getlist("operador_filtro")
    valores = request.GET.getlist("valor_filtro")
    for index, campo in enumerate(campos):
        valor = valores[index].strip() if index < len(valores) else ""
        operador = operadores[index].strip() if index < len(operadores) else "0"
        if campo in FIELD_MAP and valor:
            partes.append(
                f"{LABELS_COLUMNAS.get(campo, campo)} {OPERADORES_TXT.get(operador, 'parecido a')}: {valor}"
            )

    cefs = request.GET.getlist("cefs")
    if cefs:
        partes.append(f"CEF: {_summarize_cef_selection(cefs, cef_options)}")

    return " | ".join(partes) if partes else "Sin filtros aplicados"


def _get_cef_selector_options(source):
    """
    Arma las opciones del selector multiple de CEF, sin CUE-Anexos duplicados.
    """

    started = time.perf_counter()
    options = []
    seen = set()

    if isinstance(source, (list, tuple)):
        rows = sorted(
            (
                (item.get("cueanexo", ""), item.get("nom_est", ""))
                for item in source
            ),
            key=lambda row: (_normalize_text(row[0]), _normalize_text(row[1])),
        )
    else:
        rows = (
            source.order_by("cueanexo", "nom_est")
            .values_list("cueanexo", "nom_est")
            .distinct()
        )

    for cueanexo, nombre in rows:
        cueanexo_normalizado = normalizar_cueanexo(cueanexo)
        if not cueanexo_normalizado or cueanexo_normalizado in seen:
            continue
        seen.add(cueanexo_normalizado)
        options.append(
            {
                "cueanexo": cueanexo_normalizado,
                "nom_est": _clean(nombre) or "CEF sin nombre",
            }
        )

    _log_perf("_get_cef_selector_options", started)
    return options


def _summarize_cef_selection(cueanexos, cef_options=None):
    if not cueanexos:
        return ""
    options_map = {
        _clean(option.get("cueanexo")): _clean(option.get("nom_est"))
        for option in (cef_options or [])
        if _clean(option.get("cueanexo"))
    }

    def label(cueanexo):
        nombre = options_map.get(_clean(cueanexo))
        return f"{cueanexo} - {nombre}" if nombre else cueanexo

    if len(cueanexos) == 1:
        return label(cueanexos[0])

    preview = ", ".join(label(cueanexo) for cueanexo in cueanexos[:6])
    if len(cueanexos) > 6:
        preview = f"{preview}, ..."
    return f"{len(cueanexos)} CEF seleccionados ({preview})"


def _column_slug(column):
    return column.replace("_", "-")


def _columnas_config():
    return [
        {
            "key": column,
            "label": LABELS_COLUMNAS[column],
            "slug": _column_slug(column),
            "default": column in COLUMNAS_VISIBLES_DEFAULT,
        }
        for column in COLUMNAS_LOCALIZACIONES_CEF
    ]


def _resolver_columnas_exportar(request, formato):
    """
    Decide que columnas exportar segun el formato y las columnas visibles elegidas.
    """

    if formato == "excel_todo":
        return [(LABELS_COLUMNAS[col], col) for col in COLUMNAS_LOCALIZACIONES_CEF]

    visibles = {value.strip().replace("-", "_") for value in request.GET.getlist("visible_col") if value.strip()}
    if not visibles:
        return [(LABELS_COLUMNAS[col], col) for col in COLUMNAS_LOCALIZACIONES_CEF]

    columnas = [
        (LABELS_COLUMNAS[col], col)
        for col in COLUMNAS_LOCALIZACIONES_CEF
        if col in visibles
    ]
    return columnas or [(LABELS_COLUMNAS[col], col) for col in COLUMNAS_LOCALIZACIONES_CEF]


def _exportar_excel_cef(datos, formato, request, cef_selector_options=None):
    """
    Genera el archivo Excel con encabezado, filtros aplicados y datos resultantes.
    """

    columnas = _resolver_columnas_exportar(request, formato)

    wb = Workbook()
    ws = wb.active
    ws.title = "Localizaciones CEF"

    ultima_columna = get_column_letter(len(columnas))

    ws.merge_cells(f"A1:{ultima_columna}1")
    ws["A1"] = "Informe Localizaciones CEF"
    ws["A1"].font = Font(bold=True, size=10)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    fecha_str = datetime.now().strftime("%d/%m/%Y a las %I:%M %p").lstrip("0")
    fecha_str = fecha_str.replace("AM", "a. m.").replace("PM", "p. m.")
    ws.merge_cells(f"A2:{ultima_columna}2")
    ws["A2"] = f"Informe generado el: {fecha_str}"
    ws["A2"].font = Font(size=9)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"A3:{ultima_columna}3")
    ws["A3"] = f"Filtros aplicados: {_build_active_filters_text(request, cef_selector_options) if formato != 'excel_todo' else 'Sin filtros aplicados'}"
    ws["A3"].font = Font(size=9)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    header_row = 4
    for col_idx, (label, _) in enumerate(columnas, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for item in datos:
        ws.append([item.get(field, "") for _, field in columnas])

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{ultima_columna}{ws.max_row}"

    for col_num in range(1, len(columnas) + 1):
        col_letter = get_column_letter(col_num)
        max_length = 0
        for row in ws.iter_rows(min_row=4, max_col=col_num, max_row=ws.max_row):
            value = row[0].value
            if value:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 42)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    sufijo = "Filtros" if formato == "excel_pagina" else "Todo"
    nombre_archivo = f"Localizaciones_CEF_{sufijo}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response


def _base_cef_queryset(request):
    started = time.perf_counter()
    qs = get_cefs_visualizacion_usuario(request.user)
    _log_perf("get_cefs_visualizacion_usuario", started)
    return qs


def _get_cueanexos_visibles_usuario(user):
    cueanexos = []

    queryset = get_cefs_visualizacion_usuario(user)

    for cueanexo in (
        queryset
        .values_list("cueanexo", flat=True)
        .distinct()
    ):
        cueanexo_normalizado = normalizar_cueanexo(cueanexo)

        if (
            cueanexo_normalizado
            and cueanexo_normalizado not in cueanexos
        ):
            cueanexos.append(cueanexo_normalizado)

    return cueanexos


def _resolver_cueanexos_visualizacion_desde_request(request):
    """
    Resuelve la seleccion de CEF desde GET y la cruza con lo permitido al usuario.
    """

    cueanexos_visibles = set(
        _get_cueanexos_visibles_usuario(request.user)
    )

    if not cueanexos_visibles:
        return []

    seleccion_raw = list(request.GET.getlist("cefs"))

    cueanexo_unico = request.GET.get("cueanexo")
    if cueanexo_unico:
        seleccion_raw.append(cueanexo_unico)

    seleccion_normalizada = []

    for cueanexo in seleccion_raw:
        cueanexo_normalizado = normalizar_cueanexo(cueanexo)

        if (
            cueanexo_normalizado
            and cueanexo_normalizado in cueanexos_visibles
            and cueanexo_normalizado not in seleccion_normalizada
        ):
            seleccion_normalizada.append(cueanexo_normalizado)

    if seleccion_normalizada:
        return seleccion_normalizada

    return sorted(cueanexos_visibles)


def _resolver_cueanexos_seleccionados_explicitamente(request, cefs_visualizacion):
    """
    Distingue entre seleccion explicita del usuario y visualizacion completa.
    """

    started = time.perf_counter()
    seleccion_raw = request.GET.getlist("cefs")
    cueanexo_unico = request.GET.get("cueanexo")
    if cueanexo_unico:
        seleccion_raw.append(cueanexo_unico)

    if not seleccion_raw:
        _log_perf("_resolver_cueanexos_seleccionados_explicitamente", started)
        return [], False

    if isinstance(cefs_visualizacion, (list, tuple)):
        visibles = {
            normalizar_cueanexo(item.get("cueanexo", ""))
            for item in cefs_visualizacion
        }
    else:
        visibles = {
            normalizar_cueanexo(cueanexo)
            for cueanexo in cefs_visualizacion.values_list("cueanexo", flat=True).distinct()
        }
    visibles.discard(None)
    visibles.discard("")

    seleccion_raw_normalizada = {normalizar_cueanexo(raw) for raw in seleccion_raw}
    seleccion_raw_normalizada.discard(None)
    seleccion_raw_normalizada.discard("")
    seleccion_resuelta = _resolver_cueanexos_visualizacion_desde_request(request)
    seleccion_valida = [
        cueanexo
        for cueanexo in seleccion_resuelta
        if cueanexo in visibles
        and cueanexo in seleccion_raw_normalizada
    ]
    _log_perf("_resolver_cueanexos_seleccionados_explicitamente", started)
    return seleccion_valida, bool(seleccion_valida)


@cef_required
@ensure_csrf_cookie
def visualizacion_localizaciones(request):
    """
    Vista principal de Localizaciones CEF.

    Flujo general:
    1. carga la base cacheada de CEF visibles;
    2. resuelve selector, filtros, orden y paginacion;
    3. exporta Excel o renderiza la tabla HTML segun el parametro solicitado.
    """

    view_started = time.perf_counter()
    formato = request.GET.get("formato")

    base_items = _get_items_base_cached(request)
    cef_selector_options = _get_cef_selector_options(base_items)
    total_cefs_visibles = len(cef_selector_options)

    cueanexos_seleccionados, seleccion_cef_explicita = _resolver_cueanexos_seleccionados_explicitamente(
        request,
        base_items,
    )

    if formato == "excel_todo":
        cueanexos_seleccionados = []
        items = list(base_items)
        orden = ""
        return _exportar_excel_cef(items, formato, request, cef_selector_options)

    items = list(base_items)
    if formato != "excel_todo":
        if seleccion_cef_explicita:
            seleccion_set = set(cueanexos_seleccionados)
            items = [item for item in items if item.get("cueanexo") in seleccion_set]
        items = _apply_filters_list(items, request)

    cueanexos_selector = cueanexos_seleccionados if seleccion_cef_explicita else []
    total_cefs_seleccionados = (
        len(cueanexos_selector) if seleccion_cef_explicita else total_cefs_visibles
    )

    items, orden = _apply_order_list(items, request)

    if formato == "excel_pagina":
        return _exportar_excel_cef(items, formato, request, cef_selector_options)

    page_size = _get_page_size(request)
    lista_items_total = items

    paginator = Paginator(lista_items_total, page_size)
    page_number = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page_number)
    except (EmptyPage, PageNotAnInteger):
        page_obj = paginator.page(1)

    lista_items = list(page_obj.object_list)
    total = paginator.count
    desde = (page_obj.number - 1) * page_size + 1 if total else 0
    hasta = min(page_obj.number * page_size, total)

    smart_col = request.GET.get("smart_ui_col", "").strip()
    smart_val = request.GET.get("smart_ui_val", "").strip()
    if smart_col not in FIELD_MAP:
        smart_col = "all" if request.GET.get("q", "").strip() else "cueanexo"
    smart_search_value = smart_val or request.GET.get("q", "").strip()
    filter_options_started = time.perf_counter()
    filter_options_json = json.dumps(_get_filter_options_from_items(base_items), ensure_ascii=False)
    _log_perf("filter_options_json", filter_options_started)

    context = {
        "title": "Localizaciones CEF",
        "active_menu": "localizaciones",
        "es_admin_cef": usuario_es_admin_cef(request.user),
        "cefs_visualizacion": [],
        "cef_selector_options": cef_selector_options,
        "cueanexos_seleccionados": cueanexos_selector,
        "cueanexos_seleccionados_explicitos": cueanexos_selector,
        "selector_cef_usa_explicitos": True,
        "seleccion_cef_explicita": seleccion_cef_explicita,
        "total_cefs_visibles": total_cefs_visibles,
        "total_cefs_seleccionados": total_cefs_seleccionados,
        "lista_items": lista_items,
        "localizaciones": lista_items,
        "total_localizaciones": total,
        "resultado_total": total,
        "resultado_desde": desde,
        "resultado_hasta": hasta,
        "page_size": page_size,
        "page_size_options": PAGE_SIZE_OPTIONS,
        "page_obj": page_obj,
        "paginator": paginator,
        "columnas": COLUMNAS_LOCALIZACIONES_CEF,
        "columnas_config": _columnas_config(),
        "columnas_config_json": json.dumps(_columnas_config(), ensure_ascii=False),
        "labels_columnas": LABELS_COLUMNAS,
        "columnas_visibles_default": COLUMNAS_VISIBLES_DEFAULT,
        "filtros": _build_active_filters_text(request, cef_selector_options),
        "active_filters_text": _build_active_filters_text(request, cef_selector_options),
        "orden": orden,
        "smart_search_col": smart_col,
        "smart_search_value": smart_search_value,
        "limpiar_filtros_url": request.path,
        "request": request,
        "filter_options_json": filter_options_json,
        "cef_options_json": json.dumps(cef_selector_options, ensure_ascii=False),
    }

    render_started = time.perf_counter()
    response = render(request, "cef/localizaciones_cef.html", context)
    _log_perf("render final", render_started)
    _log_perf("visualizacion_localizaciones total", view_started)
    return response




