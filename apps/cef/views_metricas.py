# -*- coding: utf-8 -*-

import logging
from io import BytesIO

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.http import require_GET
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .permisos import cef_metricas_required, get_permisos_cef_request
from .services_metricas import (
    MetricasValidationError,
    construir_configuracion_metricas,
    ejecutar_consulta_metricas,
)


logger = logging.getLogger(__name__)


def _contexto_pagina_metricas(request):
    """Contexto global que deliberadamente no resuelve ni escribe contexto operativo."""
    permisos = get_permisos_cef_request(request)
    return {
        "title": "Consultas CEF",
        "active_menu": "metricas",
        "es_admin_cef": permisos.get("es_admin", False),
        "puede_metricas": permisos.get("puede_metricas", False),
        "metricas_config": construir_configuracion_metricas(),
    }


@cef_metricas_required
@require_GET
def metricas(request):
    return render(
        request,
        "cef/metricas_cef.html",
        _contexto_pagina_metricas(request),
    )


@cef_metricas_required
@require_GET
def metricas_consulta(request):
    try:
        resultado = ejecutar_consulta_metricas(request.GET, limite_detalle=500)
    except MetricasValidationError as exc:
        return JsonResponse(
            {"ok": False, "message": str(exc)},
            status=400,
            json_dumps_params={"ensure_ascii": False},
        )
    except Exception:
        logger.exception("Error al calcular una consulta de Métricas CEF")
        return JsonResponse(
            {
                "ok": False,
                "message": "No se pudo calcular la consulta. Revisá los filtros e intentá nuevamente.",
            },
            status=500,
            json_dumps_params={"ensure_ascii": False},
        )

    payload = dict(resultado)
    payload["ok"] = True
    return JsonResponse(
        payload,
        json_dumps_params={"ensure_ascii": False},
    )


def _lista_texto(valor, vacio="Sin selección"):
    if valor is None:
        return vacio
    if isinstance(valor, (list, tuple, set)):
        textos = []
        for item in valor:
            if isinstance(item, dict):
                item = item.get("label") or item.get("nombre") or item.get("value")
            texto = str(item or "").strip()
            if texto:
                textos.append(texto)
        return ", ".join(textos) or vacio
    if isinstance(valor, dict):
        valor = valor.get("label") or valor.get("nombre") or valor.get("value")
    return str(valor or "").strip() or vacio


def _consulta_valor(consulta, *claves, vacio="—"):
    for clave in claves:
        if clave in consulta and consulta[clave] not in (None, "", []):
            return _lista_texto(consulta[clave], vacio=vacio)
    return vacio


def _cefs_consulta_texto(consulta):
    if consulta.get("todos_cef"):
        return "Todos los CEF"
    return _consulta_valor(
        consulta,
        "cef_labels",
        "cefs_etiquetas",
        "cefs",
        vacio="Todos los CEF",
    )


def _filtros_consulta_texto(consulta):
    filtros = consulta.get("filters") or consulta.get("filtros") or []
    partes = []
    for filtro in filtros:
        if not isinstance(filtro, dict):
            continue
        etiqueta = str(filtro.get("label") or filtro.get("key") or "Filtro")
        resumen = str(filtro.get("summary") or "").strip()
        if resumen:
            partes.append(f"{etiqueta}: {resumen}")
    return "; ".join(partes) or "Sin filtros adicionales"


def _valor_excel(valor):
    if isinstance(valor, dict):
        if "value" in valor:
            valor = (
                valor["value"]
                if valor["value"] is not None
                else valor.get("formatted") or valor.get("display") or ""
            )
        elif "valor" in valor:
            valor = (
                valor["valor"]
                if valor["valor"] is not None
                else valor.get("valor_formateado") or ""
            )
        else:
            valor = valor.get("formatted") or valor.get("display") or ""
    if isinstance(valor, str) and valor.startswith(("=", "+", "-", "@")):
        return "'" + valor
    return valor


def _tabla_resultado(resultado):
    tabla = resultado.get("table") or resultado.get("tabla") or {}
    columnas = tabla.get("columns") or tabla.get("columnas") or []
    filas = tabla.get("rows") or tabla.get("filas") or []
    columnas_normalizadas = []
    for indice, columna in enumerate(columnas):
        if isinstance(columna, dict):
            clave = str(
                columna.get("key")
                or columna.get("value")
                or columna.get("id")
                or indice
            )
            etiqueta = str(
                columna.get("label")
                or columna.get("nombre")
                or clave
            )
        else:
            clave = str(indice)
            etiqueta = str(columna)
        columnas_normalizadas.append((clave, etiqueta))
    return columnas_normalizadas, filas


def _crear_excel_metricas(resultado):
    consulta = resultado.get("query") or resultado.get("consulta") or {}
    definicion = resultado.get("definition") or resultado.get("definicion") or ""
    notas = resultado.get("notes") or resultado.get("notas") or []
    if isinstance(notas, str):
        notas = [notas]
    columnas, filas = _tabla_resultado(resultado)

    wb = Workbook()
    ws = wb.active
    ws.title = "Consultas CEF"
    ancho = max(2, len(columnas))
    ultima_columna = get_column_letter(ancho)

    ws.merge_cells(f"A1:{ultima_columna}1")
    ws["A1"] = "Consulta CEF"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="17365D")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    generado = timezone.localtime().strftime("%d/%m/%Y %H:%M")
    metadatos = [
        ("Fecha de generación", generado),
        ("Área", _consulta_valor(consulta, "area_label", "area_etiqueta", "area")),
        ("Información", _consulta_valor(consulta, "indicator_label", "indicador_label", "indicador_etiqueta", "indicador")),
        ("Ciclos", _consulta_valor(consulta, "cycle_labels", "ciclos_etiquetas", "ciclos")),
        ("CEF", _cefs_consulta_texto(consulta)),
        ("Filtros", _filtros_consulta_texto(consulta)),
        ("Definición", definicion or "—"),
    ]
    if notas:
        metadatos.append(("Notas metodológicas", " ".join(str(nota) for nota in notas if nota)))

    fila = 3
    for etiqueta, valor in metadatos:
        ws.cell(row=fila, column=1, value=etiqueta).font = Font(bold=True, size=9)
        ws.cell(row=fila, column=2, value=_valor_excel(valor)).alignment = Alignment(wrap_text=True, vertical="top")
        if ancho > 2:
            ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=ancho)
        fila += 1

    fila += 1
    encabezado = fila
    if not columnas:
        columnas = [("resultado", "Resultado")]
        filas = []
    for indice, (_, etiqueta) in enumerate(columnas, start=1):
        celda = ws.cell(row=encabezado, column=indice, value=etiqueta)
        celda.font = Font(bold=True, color="FFFFFF", size=9)
        celda.fill = PatternFill("solid", fgColor="2F75B5")
        celda.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for fila_resultado in filas:
        valores = []
        for indice, (clave, _) in enumerate(columnas):
            if isinstance(fila_resultado, (list, tuple)):
                valor = fila_resultado[indice] if indice < len(fila_resultado) else ""
            else:
                valor = fila_resultado.get(clave, "")
            valores.append(_valor_excel(valor))
        ws.append(valores)

    ws.freeze_panes = f"A{encabezado + 1}"
    if filas:
        ws.auto_filter.ref = f"A{encabezado}:{get_column_letter(len(columnas))}{ws.max_row}"

    for indice in range(1, max(2, len(columnas)) + 1):
        letra = get_column_letter(indice)
        longitud = 0
        for celdas in ws.iter_rows(min_col=indice, max_col=indice):
            valor = celdas[0].value
            if valor is not None:
                longitud = max(longitud, len(str(valor)))
        ws.column_dimensions[letra].width = min(max(longitud + 2, 12), 48)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


@cef_metricas_required
@require_GET
def metricas_exportar(request):
    try:
        resultado = ejecutar_consulta_metricas(request.GET, limite_detalle=None)
    except MetricasValidationError as exc:
        return HttpResponse(str(exc), status=400, content_type="text/plain; charset=utf-8")
    except Exception:
        logger.exception("Error al exportar una consulta de Métricas CEF")
        return HttpResponse(
            "No se pudo generar el archivo Excel.",
            status=500,
            content_type="text/plain; charset=utf-8",
        )

    contenido = _crear_excel_metricas(resultado)
    nombre = f"Consulta_CEF_{timezone.localtime().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        contenido,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return response
