# apps/especial/views_localizaciones.py
# -*- coding: utf-8 -*-

import json
import logging
import time
import unicodedata
from datetime import datetime
from hashlib import sha256
from io import BytesIO

from django.core.cache import cache
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.http import HttpResponse
from django.shortcuts import render
from django.views.decorators.csrf import ensure_csrf_cookie
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .models import normalizar_cueanexo
from .permisos import especial_required, get_permisos_especial_request
from .views_contexto import contexto_base


logger = logging.getLogger(__name__)


PAGE_SIZE = 10
PAGE_SIZE_OPTIONS = [10, 25, 50, 100]

# Columnas que se leen desde Padrón y se exponen en tabla, filtros y Excel.
COLUMNAS_LOCALIZACIONES_ESPECIAL = [
    "cueanexo",
    "nom_est",
    "oferta",
    "ambito",
    "sector",
    "region_loc",
    "ref_loc",
    "calle",
    "numero",
    "localidad",
    "departamento",
    "estado_loc",
    "est_oferta",
    "estado_est",
    "resploc_cuitcuil",
    "resploc_doc",
    "apellido_resp",
    "nombre_resp",
    "resploc_email",
    "resploc_telefono",
    "sup_tecnico",
    "email_suptecnico",
    "tel_suptecnico",
    "categoria",
    "jornada",
]

LABELS_COLUMNAS = {
    "cueanexo": "CUE-Anexo",
    "nom_est": "Establecimiento",
    "oferta": "Oferta",
    "ambito": "Ámbito",
    "sector": "Sector",
    "region_loc": "Región",
    "ref_loc": "Referencia localización",
    "calle": "Calle",
    "numero": "Número",
    "localidad": "Localidad",
    "departamento": "Departamento",
    "estado_loc": "Estado localización",
    "est_oferta": "Estado oferta",
    "estado_est": "Estado establecimiento",
    "resploc_cuitcuil": "CUIL/CUIT responsable",
    "resploc_doc": "Documento responsable",
    "apellido_resp": "Apellido responsable",
    "nombre_resp": "Nombre responsable",
    "resploc_email": "Email responsable",
    "resploc_telefono": "Teléfono responsable",
    "sup_tecnico": "Supervisor técnico",
    "email_suptecnico": "Email supervisor técnico",
    "tel_suptecnico": "Teléfono supervisor técnico",
    "categoria": "Categoría",
    "jornada": "Jornada",
}

COLUMNAS_VISIBLES_DEFAULT = [
    "cueanexo",
    "nom_est",
    "oferta",
    "region_loc",
    "localidad",
    "departamento",
    "apellido_resp",
    "nombre_resp",
    "sup_tecnico",
    "categoria",
    "jornada",
]

ONLY_FIELDS_LOCALIZACIONES_ESPECIAL = COLUMNAS_LOCALIZACIONES_ESPECIAL
SORTABLE_FIELDS_LOCALIZACIONES_ESPECIAL = {col: col for col in COLUMNAS_LOCALIZACIONES_ESPECIAL}
DEFAULT_ORDER_LOCALIZACIONES_ESPECIAL = (
    "region_loc",
    "departamento",
    "localidad",
    "cueanexo",
)
CACHE_TTL_LOCALIZACIONES_ESPECIAL = 60 * 5
CACHE_VERSION_LOCALIZACIONES_ESPECIAL = "v2_cache_admin_all_20260810"
_CACHE_MISS = object()

def _log_perf(label, started):
    logger.debug(
        "ESPECIAL_LOCALIZACIONES %s %.1fms",
        label,
        (time.perf_counter() - started) * 1000,
    )


def _clean(value):
    if value is None:
        return ""
    return str(value).strip()


def _serialize_item(item):
    return {
        "cueanexo": _clean(getattr(item, "cueanexo", "")),
        "nom_est": _clean(getattr(item, "nom_est", "")),
        "oferta": _clean(getattr(item, "oferta", "")),
        "ambito": _clean(getattr(item, "ambito", "")),
        "sector": _clean(getattr(item, "sector", "")),
        "region_loc": _clean(getattr(item, "region_loc", "")),
        "ref_loc": _clean(getattr(item, "ref_loc", "")),
        "calle": _clean(getattr(item, "calle", "")),
        "numero": _clean(getattr(item, "numero", "")),
        "localidad": _clean(getattr(item, "localidad", "")),
        "departamento": _clean(getattr(item, "departamento", "")),
        "estado_loc": _clean(getattr(item, "estado_loc", "")),
        "est_oferta": _clean(getattr(item, "est_oferta", "")),
        "estado_est": _clean(getattr(item, "estado_est", "")),
        "resploc_cuitcuil": _clean(getattr(item, "resploc_cuitcuil", "")),
        "resploc_doc": _clean(getattr(item, "resploc_doc", "")),
        "apellido_resp": _clean(getattr(item, "apellido_resp", "")),
        "nombre_resp": _clean(getattr(item, "nombre_resp", "")),
        "resploc_email": _clean(getattr(item, "resploc_email", "")),
        "resploc_telefono": _clean(getattr(item, "resploc_telefono", "")),
        "sup_tecnico": _clean(getattr(item, "sup_tecnico", "")),
        "email_suptecnico": _clean(getattr(item, "email_suptecnico", "")),
        "tel_suptecnico": _clean(getattr(item, "tel_suptecnico", "")),
        "categoria": _clean(getattr(item, "categoria", "")),
        "jornada": _clean(getattr(item, "jornada", "")),
    }


def _get_items_base_authorized(permisos):
    """Devuelve el QuerySet autorizado con solo las columnas de Localizaciones."""
    started = time.perf_counter()
    queryset = permisos["escuelas_visualizacion"].only(*ONLY_FIELDS_LOCALIZACIONES_ESPECIAL)
    _log_perf("_get_items_base_authorized", started)
    return queryset


def _normalizar_orden_localizaciones(orden_param):
    orden = (orden_param or "").strip()
    if not orden:
        return ""
    signo = "-" if orden.startswith("-") else ""
    campo = orden[1:] if signo else orden
    if campo not in SORTABLE_FIELDS_LOCALIZACIONES_ESPECIAL:
        return ""
    return f"{signo}{campo}"


def _normalize_text(value):
    """Normaliza texto para comparar sin distinguir mayúsculas ni acentos."""
    text = _clean(value).casefold()
    text = unicodedata.normalize("NFKD", text)
    return "".join(char for char in text if not unicodedata.combining(char))


def _get_filter_options(items):
    """Obtiene opciones desde el snapshot autorizado en memoria."""
    return {
        field: sorted(
            {
                _clean(item.get(field))
                for item in items
                if _clean(item.get(field))
            }
        )
        for field in ("region_loc", "departamento", "localidad")
    }


def _get_establecimientos_options(items):
    """Obtiene opciones de establecimientos sin consultar la base."""
    options_by_cue = {}
    for item in items:
        cueanexo = _clean(item.get("cueanexo"))
        if cueanexo:
            options_by_cue[cueanexo] = {
                "cueanexo": cueanexo,
                "nom_est": _clean(item.get("nom_est")),
            }
    return sorted(options_by_cue.values(), key=lambda item: item["cueanexo"])


def _resolver_establecimientos_autorizados(items, valores_solicitados):
    """Devuelve, en el orden recibido, los CUE del snapshot autorizado."""
    solicitados = []
    for value in valores_solicitados:
        cueanexo = normalizar_cueanexo(value)
        if cueanexo and cueanexo not in solicitados:
            solicitados.append(cueanexo)

    autorizados = {
        normalizar_cueanexo(item.get("cueanexo"))
        for item in items
    }
    autorizados.discard("")
    return [cueanexo for cueanexo in solicitados if cueanexo in autorizados]


def _cache_key_localizaciones_especial(request, permisos):
    """Construye una clave privada y estable para el alcance autorizado."""
    user_id = getattr(getattr(request, "user", None), "pk", None)
    if user_id is None or not str(user_id).strip():
        return None

    if permisos.get("es_admin"):
        authorized_scope = "ALL"
    else:
        authorized_scope = sorted(
            {
                normalizar_cueanexo(value)
                for value in permisos.get("cueanexos_visualizacion", [])
                if normalizar_cueanexo(value)
            }
        )
    payload = json.dumps(
        {
            "user_id": str(user_id),
            "rol": str(permisos.get("rol") or ""),
            "authorized_scope": authorized_scope,
            "version": CACHE_VERSION_LOCALIZACIONES_ESPECIAL,
        },
        ensure_ascii=True,
        separators=(",", ":"),
    )
    fingerprint = sha256(payload.encode("utf-8")).hexdigest()
    return (
        "especial:localizaciones:"
        f"{CACHE_VERSION_LOCALIZACIONES_ESPECIAL}:{fingerprint}"
    )


def _iter_serialized_items(items):
    """Itera datos serializados o materializa un QuerySet autorizado una vez."""
    if hasattr(items, "iterator"):
        items = items.only(*ONLY_FIELDS_LOCALIZACIONES_ESPECIAL).iterator()
    for item in items:
        yield item if isinstance(item, dict) else _serialize_item(item)


def _is_serialized_items(value):
    return isinstance(value, list) and all(
        isinstance(item, dict)
        and all(
            isinstance(key, str) and isinstance(field_value, str)
            for key, field_value in item.items()
        )
        for item in value
    )


def _get_items_base_cached(request, permisos):
    """Obtiene el snapshot autorizado, usando cache solo por usuario y alcance."""
    key = _cache_key_localizaciones_especial(request, permisos)
    if key is None:
        items = list(_iter_serialized_items(_get_items_base_authorized(permisos)))
        logger.debug("ESPECIAL_LOCALIZACIONES cache bypass sin user.pk")
        return items

    cached = cache.get(key, _CACHE_MISS)
    if cached is not _CACHE_MISS and _is_serialized_items(cached):
        logger.debug("ESPECIAL_LOCALIZACIONES cache hit")
        return cached

    items = list(_iter_serialized_items(_get_items_base_authorized(permisos)))
    cache.set(key, items, CACHE_TTL_LOCALIZACIONES_ESPECIAL)
    logger.debug("ESPECIAL_LOCALIZACIONES cache miss")
    return items


def _normalize_item_field(item, field):
    if field == "cueanexo":
        return normalizar_cueanexo(item.get(field))
    return _normalize_text(item.get(field))


def _matches_filter_condition(actual, operator, expected):
    if operator == "1":
        return expected not in actual
    if operator == "2":
        return actual == expected
    if operator == "3":
        return actual > expected
    if operator == "4":
        return actual >= expected
    if operator == "5":
        return actual < expected
    if operator == "6":
        return actual <= expected
    if operator == "7":
        return actual != expected
    return expected in actual


def _apply_filters_items(items, request, establecimientos=None):
    """Aplica a una lista las mismas reglas de filtrado de Localizaciones."""
    filtered = list(items)

    q = request.GET.get("q", "").strip()
    if q:
        q_normalized = _normalize_text(q)
        filtered = [
            item
            for item in filtered
            if any(
                q_normalized in _normalize_item_field(item, field)
                for field in COLUMNAS_LOCALIZACIONES_ESPECIAL
            )
        ]

    smart_col = request.GET.get("smart_ui_col", "").strip()
    smart_val = request.GET.get("smart_ui_val", "").strip()
    if smart_col in COLUMNAS_LOCALIZACIONES_ESPECIAL and smart_val:
        smart_normalized = _normalize_text(smart_val)
        filtered = [
            item
            for item in filtered
            if smart_normalized in _normalize_item_field(item, smart_col)
        ]

    establecimientos_normalizados = {
        normalizar_cueanexo(value)
        for value in (establecimientos or [])
        if normalizar_cueanexo(value)
    }
    if establecimientos_normalizados:
        filtered = [
            item
            for item in filtered
            if _normalize_item_field(item, "cueanexo")
            in establecimientos_normalizados
        ]

    for field in COLUMNAS_LOCALIZACIONES_ESPECIAL:
        value = request.GET.get(field, "").strip()
        if value:
            normalized_value = _normalize_text(value)
            filtered = [
                item
                for item in filtered
                if normalized_value in _normalize_item_field(item, field)
            ]

    fields = request.GET.getlist("campo_filtro")
    operators = request.GET.getlist("operador_filtro")
    values = request.GET.getlist("valor_filtro")
    grouped_filters = {}
    for index, field in enumerate(fields):
        field = field.strip()
        value = values[index].strip() if index < len(values) else ""
        operator = operators[index].strip() if index < len(operators) else "0"
        if not field or not value or field not in COLUMNAS_LOCALIZACIONES_ESPECIAL:
            continue
        grouped_filters.setdefault((field, operator), [])
        if value not in grouped_filters[(field, operator)]:
            grouped_filters[(field, operator)].append(value)

    for (field, operator), group_values in grouped_filters.items():
        normalized_values = [_normalize_text(value) for value in group_values]
        filtered = [
            item
            for item in filtered
            if any(
                _matches_filter_condition(
                    _normalize_item_field(item, field), operator, expected
                )
                for expected in normalized_values
            )
        ]

    return filtered


def _apply_order_items(items, request):
    """Ordena en memoria y conserva los defaults como desempates ascendentes."""
    orden = _normalizar_orden_localizaciones(request.GET.get("orden", ""))
    ordered = list(items)
    for field in reversed(DEFAULT_ORDER_LOCALIZACIONES_ESPECIAL):
        ordered = sorted(
            ordered,
            key=lambda item, field=field: _normalize_item_field(item, field),
        )

    if orden:
        descending = orden.startswith("-")
        field = orden[1:] if descending else orden
        ordered = sorted(
            ordered,
            key=lambda item: _normalize_item_field(item, field),
            reverse=descending,
        )
    return ordered, orden


def _get_page_size(request):
    try:
        page_size = int(request.GET.get("page_size", PAGE_SIZE))
    except (TypeError, ValueError):
        return PAGE_SIZE
    return page_size if page_size in PAGE_SIZE_OPTIONS else PAGE_SIZE


def _resolver_columnas_exportar(request, formato):
    if formato == "excel_todo":
        return [(LABELS_COLUMNAS[col], col) for col in COLUMNAS_LOCALIZACIONES_ESPECIAL]

    visibles = {
        value.strip().replace("-", "_")
        for value in request.GET.getlist("visible_col")
        if value.strip()
    }
    columnas = [
        (LABELS_COLUMNAS[col], col)
        for col in COLUMNAS_LOCALIZACIONES_ESPECIAL
        if col in visibles
    ]
    return columnas or [
        (LABELS_COLUMNAS[col], col) for col in COLUMNAS_LOCALIZACIONES_ESPECIAL
    ]


def _exportar_excel_especial(datos, request, formato):
    """Genera el archivo Excel."""
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.cell.cell import Cell
    
    columnas = _resolver_columnas_exportar(request, formato)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Localizaciones Especial"

    num_columnas: int = len(columnas)
    ultima_columna: str = get_column_letter(num_columnas)

    # Encabezado del informe
    rango_titulo: str = f"A1:{ultima_columna}1"
    ws.merge_cells(rango_titulo)
    celda_titulo: Cell = ws["A1"]
    celda_titulo.value = "Informe Localizaciones Educación Especial"
    celda_titulo.font = Font(bold=True, size=10)
    celda_titulo.alignment = Alignment(horizontal="left", vertical="center")

    fecha_str: str = datetime.now().strftime("%d/%m/%Y a las %I:%M %p").lstrip("0")
    fecha_str = fecha_str.replace("AM", "a. m.").replace("PM", "p. m.")
    
    rango_fecha: str = f"A2:{ultima_columna}2"
    ws.merge_cells(rango_fecha)
    celda_fecha: Cell = ws["A2"]
    celda_fecha.value = f"Informe generado el: {fecha_str}"
    celda_fecha.font = Font(size=9)
    celda_fecha.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(f"A3:{ultima_columna}3")
    ws["A3"] = (
        "Filtros aplicados: Sin filtros aplicados"
        if formato == "excel_todo"
        else "Filtros aplicados desde la vista"
    )
    ws["A3"].font = Font(size=9)
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Encabezados de columnas
    header_row: int = 4
    for col_idx, (label, _) in enumerate(columnas, start=1):
        cell: Cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Datos
    for item in datos:
        fila: list = [item.get(field, "") for _, field in columnas]
        ws.append(fila)

    ws.freeze_panes = "A5"
    max_row: int = ws.max_row or header_row
    rango_filtro: str = f"A{header_row}:{ultima_columna}{max_row}"
    ws.auto_filter.ref = rango_filtro

    # Ajuste de ancho de columnas
    for col_num in range(1, num_columnas + 1):
        col_letter: str = get_column_letter(col_num)
        max_length: int = 0
        for row in ws.iter_rows(min_row=header_row, max_col=col_num, max_row=max_row):
            if not row:
                continue
            cell_value = row[0].value
            if cell_value is not None:
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 42)

    # Generar respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    sufijo = "Filtros" if formato == "excel_pagina" else "Todo"
    nombre_archivo: str = f"Localizaciones_Especial_{sufijo}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    return response

@especial_required
@ensure_csrf_cookie
def visualizacion_localizaciones(request):
    """Vista principal de Localizaciones de Educación Especial."""
    view_started = time.perf_counter()
    formato = request.GET.get("formato")

    permisos = get_permisos_especial_request(request)
    base_items = _get_items_base_cached(request, permisos)
    is_fragment_request = (
        request.GET.get("fragmento") == "resultados"
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    )

    if formato == "excel_todo":
        return _exportar_excel_especial(base_items, request, formato)

    establecimientos_seleccionados = _resolver_establecimientos_autorizados(
        base_items,
        request.GET.getlist("establecimientos"),
    )

    if not is_fragment_request and formato != "excel_pagina":
        filter_options = _get_filter_options(base_items)
        establecimientos_options = _get_establecimientos_options(base_items)

    seleccion_establecimientos_explicita = bool(establecimientos_seleccionados)

    filtered_items = _apply_filters_items(
        base_items,
        request,
        establecimientos=establecimientos_seleccionados,
    )
    ordered_items, orden_actual = _apply_order_items(filtered_items, request)

    if formato == "excel_pagina":
        return _exportar_excel_especial(ordered_items, request, formato)

    page_size = _get_page_size(request)
    paginator = Paginator(ordered_items, page_size)
    page_number = request.GET.get("page", 1)
    try:
        page_obj = paginator.page(page_number)
    except (EmptyPage, PageNotAnInteger):
        page_obj = paginator.page(1)

    lista_items = [dict(item) for item in page_obj.object_list]
    total = paginator.count
    desde = (page_obj.number - 1) * page_size + 1 if total else 0
    hasta = min(page_obj.number * page_size, total)

    if is_fragment_request:
        context = {
            "lista_items": lista_items,
            "resultado_total": total,
            "resultado_desde": desde,
            "resultado_hasta": hasta,
            "page_size": page_size,
            "page_size_options": PAGE_SIZE_OPTIONS,
            "page_obj": page_obj,
            "paginator": paginator,
            "request": request,
        }
        render_started = time.perf_counter()
        response = render(
            request,
            "especial/componentes/localizaciones_resultados_especial.html",
            context,
        )
        _log_perf("render final", render_started)
        _log_perf("visualizacion_localizaciones total", view_started)
        return response

    columnas_config = [
        {
            "key": col,
            "label": LABELS_COLUMNAS[col],
            "slug": col.replace("_", "-"),
            "default": col in COLUMNAS_VISIBLES_DEFAULT,
        }
        for col in COLUMNAS_LOCALIZACIONES_ESPECIAL
    ]
    columnas_config_json = json.dumps(columnas_config)

    context = contexto_base(request, "localizaciones")
    especial_context = context["especial_context"]

    context.update({
        "establecimientos_options": establecimientos_options,
        "establecimientos_seleccionados": establecimientos_seleccionados,
        "total_establecimientos_seleccionados": len(establecimientos_seleccionados),
        "seleccion_establecimientos_explicita": seleccion_establecimientos_explicita,
        "lista_items": lista_items,
        "localizaciones": lista_items,
        "total_localizaciones": total,
        "resultado_total": total,
        "resultado_desde": desde,
        "resultado_hasta": hasta,
        "page_size": page_size,
        "page_size_options": PAGE_SIZE_OPTIONS,
        "page_obj": page_obj,
        "paginator": paginator,
        "filter_options": filter_options,
        "columnas": COLUMNAS_LOCALIZACIONES_ESPECIAL,
        "labels_columnas": LABELS_COLUMNAS,
        "columnas_visibles_default": COLUMNAS_VISIBLES_DEFAULT,
        "columnas_config": columnas_config,
        "columnas_config_json": columnas_config_json,
        "orden": orden_actual,
        "mostrar_contexto": False,
        "region_loc": filter_options["region_loc"],
        "departamento": filter_options["departamento"],
        "localidad": filter_options["localidad"],
        "limpiar_filtros_url": request.path,
        "request": request,
        # ✅ AGREGAR ESTO PARA QUE LOS FILTROS FUNCIONEN:
        **filter_options,
        "smart_search_col": request.GET.get("smart_ui_col", "") or ("all" if request.GET.get("q") else "cueanexo"),
        "smart_search_value": request.GET.get("smart_ui_val", "") or request.GET.get("q", ""),
    })

    render_started = time.perf_counter()
    response = render(request, "especial/localizaciones_especial.html", context)
    _log_perf("render final", render_started)
    _log_perf("visualizacion_localizaciones total", view_started)
    return response
